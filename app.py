from functools import wraps
from flask import Flask, request, redirect, render_template_string, jsonify, session, send_file, Response
import psycopg2
import os, random, io, json, threading, traceback, qrcode

try:
    from pywebpush import webpush, WebPushException
    PUSH_AKTIF = True
except ImportError:
    PUSH_AKTIF = False
    
from datetime import datetime
from zoneinfo import ZoneInfo
import barcode
from barcode.writer import ImageWriter

# ==================== BARKOD/RESİM CACHE (RAM, ücretsiz) ====================
_BARKOD_CACHE = {}
_CACHE_KILIT = threading.Lock()  

TR_TZ = ZoneInfo("Europe/Istanbul")

# ==================== VERİTABANI (Postgres) ====================
DATABASE_URL = os.environ.get("DATABASE_URL", "")

# ==================== PUSH BİLDİRİM (VAPID) ====================
VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY", "GnlhR4K1blYAg-Unc57v-tAIP-G7rfiQqetV1otak4Q")
VAPID_PUBLIC_KEY = os.environ.get("VAPID_PUBLIC_KEY", "BHVDYu-AltC6T-LUrITOY3toLuHEec99bDk5Uokzj5NrtK8fV2mJwctTmGDxE5ANtavoLlOod9_27Jrm-HgnJ1k")
VAPID_CLAIMS_EMAIL = os.environ.get("VAPID_EMAIL", "mailto:admin@heris-stok.local")

app = Flask(__name__, static_folder=os.path.join(os.path.dirname(os.path.abspath(__file__)) or ".", "static"))
app.secret_key = os.environ.get("SECRET_KEY", "bu-anahtari-canliya-almadan-once-degistir")
SITE_PAROLA = os.environ.get("SITE_PAROLA", "") 

@app.before_request
def site_girisi_kontrol():
    if not SITE_PAROLA:
        return
    if request.endpoint in ("site_giris", "manifest", "service_worker"):
        return
    if request.path.startswith("/static/") or request.path.startswith("/barkod/"):
        return
    if not session.get("site_giris"):
        return redirect("/site_giris")


@app.route("/site_giris", methods=["GET", "POST"])
def site_giris():
    hata = None
    if request.method == "POST":
        if request.form.get("parola", "") == SITE_PAROLA:
            session["site_giris"] = True
            return redirect("/")
        hata = "❌ Yanlış şifre"
    icerik = f"""
    <h2>🔒 Giriş</h2>
    <p style="text-align:center;color:var(--muted);">Devam etmek için şifreyi gir</p>
    {'<p class="hata">' + hata + '</p>' if hata else ''}
    <form method="post">
    <input type="password" name="parola" placeholder="Şifre" autofocus autocomplete="off">
    <button class="btn mavi">Gir</button>
    </form>
    """
    return sayfa(icerik, "Giriş")

if os.path.exists(app.static_folder) and not os.path.isdir(app.static_folder):
    os.remove(app.static_folder)
if not os.path.exists(app.static_folder):
    os.makedirs(app.static_folder)

DEPOLAR = [
    "MDF SATIŞ DEPOSU",
    "LAMİNANT DEPOSU",
    "KAPI DEPOSU",
    "HGLOSS DEPOSU (MORAY YANI)",
    "SÜTÇÜ YANI",
    "HELVACI YANI",
    "RÖTBALANSÇI YANI",
    "KESİMHANE",
    "OFİS",
]

ROLLER = {
    "Ramazan": "depocu",
    "Behiç": "depocu",
    "Orhan": "depocu",
    "Berke": "muhasebeci",
    "İrem": "muhasebeci",
    "Hatice": "patron",
    "Ahmet": "patron",
}

def bugunku_ozet(kullanici):
    bugun = tr_simdi().date()
    try:
        con = db()
        try:
            with con.cursor() as cur:
                cur.execute("""
                    SELECT tip, COUNT(*) FROM hareket
                    WHERE kullanici=%s AND tarih::date=%s
                    GROUP BY tip
                """, (kullanici, bugun))
                satirlar = dict(cur.fetchall())
        finally:
            con.close()
        return satirlar.get("giris", 0), satirlar.get("cikis", 0)
    except Exception:
        traceback.print_exc()
        return 0, 0

def _pin_yukle():
    varsayilan = {
        "Ramazan": "1111", "Behiç": "2222", "Orhan": "3333",
        "Berke": "4444", "İrem": "4444", "Hatice": "4444", "Ahmet": "4444",
    }
    ozel_json = os.environ.get("PIN_KODLARI_JSON", "")
    if ozel_json:
        try:
            varsayilan.update(json.loads(ozel_json))
        except Exception:
            pass
    return varsayilan

PIN_KODLARI = _pin_yukle()

def db():
    return psycopg2.connect(DATABASE_URL)

def tr_simdi():
    return datetime.now(TR_TZ).replace(tzinfo=None)



def tablolari_olustur():
    con = db()
    try:
        with con:
            with con.cursor() as cur:
                cur.execute("""
                CREATE TABLE IF NOT EXISTS urun(
                    id SERIAL PRIMARY KEY,
                    ad TEXT, cins TEXT, ebat TEXT, kalinlik TEXT,
                    yuzey TEXT, sinif TEXT, renk TEXT,
                    adet INTEGER, depo TEXT, barkod TEXT UNIQUE
                )
                """)
                cur.execute("""
                CREATE TABLE IF NOT EXISTS hareket(
                    id SERIAL PRIMARY KEY,
                    barkod TEXT,
                    ad TEXT,
                    tip TEXT,
                    adet INTEGER,
                    kullanici TEXT,
                    tarih TIMESTAMP
                )
                """)
                cur.execute("""
                ALTER TABLE urun ADD COLUMN IF NOT EXISTS min_stok INTEGER DEFAULT 5
                """)
                cur.execute("""
                CREATE TABLE IF NOT EXISTS urun_barkod(
                    id SERIAL PRIMARY KEY,
                    urun_id INTEGER REFERENCES urun(id) ON DELETE CASCADE,
                    barkod TEXT UNIQUE
                )
                """)
                cur.execute("""
                INSERT INTO urun_barkod (urun_id, barkod)
                SELECT u.id, u.barkod FROM urun u
                WHERE u.barkod IS NOT NULL
                AND NOT EXISTS (
                    SELECT 1 FROM urun_barkod ub WHERE ub.barkod = u.barkod
                )
                """)
                cur.execute("ALTER TABLE hareket ALTER COLUMN tarih DROP DEFAULT")
    finally:
        con.close()


if DATABASE_URL:
    tablolari_olustur()
    
def rol_gerekli(*izinli_roller):
    TAM_YETKILI = ("patron", "muhasebeci")
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            rol = session.get("rol")
            if not rol:
                return redirect("/")
            if rol not in TAM_YETKILI and rol not in izinli_roller:
                return sayfa("""
                <h2>⛔ Erişim Yetkiniz Yok</h2>
                <p style="text-align:center;color:var(--muted);">
                Bu işlem senin rolüne kapalı. Yanlış kişi olarak girdiysen
                sağ üstten kullanıcı değiştir.
                </p>
                """, "Yetkisiz Erişim")
            return f(*args, **kwargs)
        return wrapper
    return decorator

def bugunku_ozet(kullanici):
    bugun = tr_simdi().date()
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("""
                SELECT tip, COUNT(*) FROM hareket
                WHERE kullanici=%s AND tarih::date=%s
                GROUP BY tip
            """, (kullanici, bugun))
            satirlar = dict(cur.fetchall())
    finally:
        con.close()
    return satirlar.get("giris", 0), satirlar.get("cikis", 0)


def barkod_uret():
    while True:
        kod = str(random.randint(100000000000, 999999999999))
        con = db()
        try:
            with con.cursor() as cur:
                cur.execute("SELECT 1 FROM urun WHERE barkod=%s", (kod,))
                var = cur.fetchone()
        finally:
            con.close()
        if not var:
            return kod


def barkod_png_bytes(kod):
    from barcode.writer import ImageWriter
    CODE128 = barcode.get_barcode_class("code128")
    yazici_ayarlari = {
        "module_width": 0.32,     # XP-470B termal kafası için biraz kalın (net basım)
        "module_height": 14.0,    # 40mm etiket yüksekliğine göre optimize
        "quiet_zone": 3.0,
        "font_size": 7,
        "text_distance": 2.5,
        "write_text": True,
        "dpi": 203,               # XP-470B çözünürlüğü genelde 203 dpi
    }
    bio = io.BytesIO()
    CODE128(kod, writer=ImageWriter()).write(bio, options=yazici_ayarlari)
    bio.seek(0)
    return bio


def ikon_uret():
    try:
        from PIL import Image, ImageDraw
    except ImportError:
        return
    for boyut in (192, 512):
        yol = os.path.join(app.static_folder, f"icon-{boyut}.png")
        if os.path.exists(yol):
            continue
        img = Image.new("RGB", (boyut, boyut), "#2196F3")
        d = ImageDraw.Draw(img)
        for y in range(boyut):
            oran = y / boyut
            r = int(0x21 + (0x00 - 0x21) * oran)
            g = int(0x96 + (0xBC - 0x96) * oran)
            b = int(0xF3 + (0xD4 - 0xF3) * oran)
            d.line([(0, y), (boyut, y)], fill=(r, g, b))
        pad = int(boyut * 0.22)
        cizgi = max(2, boyut // 35)
        d.rectangle([pad, int(pad * 1.3), boyut - pad, boyut - pad], outline="white", width=cizgi)
        d.line([(pad, int(boyut * 0.46)), (boyut - pad, int(boyut * 0.46))], fill="white", width=cizgi)
        d.line([(boyut / 2, int(pad * 1.3)), (boyut / 2, int(boyut * 0.46))], fill="white", width=max(2, cizgi // 2))
        img.save(yol)


ikon_uret()


BASE_HEAD_EXTRA = """
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#111111">
<link rel="manifest" href="/manifest.json">
<link rel="icon" href="/static/icon-192.png">
<link rel="apple-touch-icon" href="/static/icon-192.png">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="mobile-web-app-capable" content="yes">
"""

ORTAK_CSS = """
:root {
  --bg:#0b0d12; --bg2:#12151c; --card:#171a22; --card2:#1c202b;
  --border:#262b38; --text:#f4f5f7; --muted:#8b93a3; --accent:#2196F3;
  --radius:16px; --radius-sm:10px;
}
* { box-sizing: border-box; }
html { scrollbar-color: #333 transparent; }
body {
  background:
    radial-gradient(1200px 600px at 50% -10%, rgba(33,150,243,.08), transparent 60%),
    var(--bg);
  color: var(--text); margin:0; padding:0;
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
  -webkit-tap-highlight-color: transparent;
}
input, select, button, textarea { font-size: 16px; font-family: inherit; }
a { color: inherit; }
::selection { background: rgba(33,150,243,.35); }

.topbar {
  position: fixed; top:0; left:0; right:0; z-index: 9999;
  height: 56px; padding-top: env(safe-area-inset-top);
  display:flex; align-items:center; justify-content:space-between;
  padding-left:12px; padding-right:12px;
  background: rgba(18,21,28,.75); backdrop-filter: blur(14px) saturate(140%);
  -webkit-backdrop-filter: blur(14px) saturate(140%);
  border-bottom: 1px solid rgba(255,255,255,.06);
}
.topbar-btn {
  text-decoration:none; color:white; font-size:19px; padding:8px 12px;
  border-radius:10px; transition: background .15s ease, transform .1s ease;
}
.topbar-btn:active { background:rgba(255,255,255,.08); transform: scale(0.94); }
.topbar-title { font-weight:700; font-size:14.5px; color:#eee; letter-spacing:.2px; display:flex; align-items:center; gap:7px; }
.topbar-logo { height:26px; width:26px; border-radius:6px; object-fit:contain; background:white; padding:2px; }

.sayfa {
  max-width: 520px; margin: 0 auto;
  padding: calc(56px + env(safe-area-inset-top) + 22px) 16px 60px;
  animation: belir .25s ease;
}
@keyframes belir { from { opacity:0; transform: translateY(6px);} to { opacity:1; transform:none; } }

h1, h2, h3 { text-align:center; letter-spacing:-.01em; }
h1 { font-size: 24px; }
h3.alt { color: var(--muted); font-weight:500; margin-top:-6px; font-size:14px; }

.btn {
  display:block; width:100%; margin:10px 0; padding:17px;
  font-size:17px; border-radius: var(--radius); text-decoration:none;
  color:white; font-weight:600; text-align:center; border:none; cursor:pointer;
  box-shadow: 0 4px 14px rgba(0,0,0,.28);
  transition: transform .1s ease, filter .15s ease, box-shadow .15s ease;
  letter-spacing:.1px;
}
.btn:active { filter:brightness(0.88); transform: scale(0.985); box-shadow: 0 2px 8px rgba(0,0,0,.3); }
.mavi   { background: linear-gradient(135deg, #2196F3, #00BCD4); }
.yesil  { background: linear-gradient(135deg, #00C853, #64DD17); }
.turuncu{ background: linear-gradient(135deg, #FF6F00, #FF9800); }
.mor    { background: linear-gradient(135deg, #5E35B1, #7E57C2); }
.kirmizi{ background: linear-gradient(135deg, #D50000, #FF1744); }
.turkuaz{ background: linear-gradient(135deg, #00838F, #00BFA5); }
.gri    { background: linear-gradient(135deg, #3a3f4b, #2a2e38); }

.btn-kucuk {
  display:inline-block; padding:10px 16px; font-size:14px; font-weight:600;
  border-radius: 10px; text-decoration:none; color:white; border:none; cursor:pointer;
  transition: transform .1s ease, filter .15s ease; margin-right:8px; margin-top:4px;
}
.btn-kucuk:active { filter:brightness(0.85); transform: scale(0.96); }

.kart {
  background: linear-gradient(180deg, var(--card2), var(--card));
  border:1px solid var(--border); border-radius: var(--radius);
  padding:16px; margin:12px 0; text-align:left;
  box-shadow: 0 2px 10px rgba(0,0,0,.18);
  transition: border-color .15s ease;
}

input[type=text], input[type=number], input[type=password], input[type=file], select, textarea {
  width:100%; padding:14px; margin:6px 0 14px; border-radius: var(--radius-sm);
  border:1px solid var(--border); background:#12151c; color:white;
  transition: border-color .15s ease, box-shadow .15s ease;
}
input:focus, select:focus, textarea:focus {
  outline:none; border-color: var(--accent);
  box-shadow: 0 0 0 3px rgba(33,150,243,.18);
}
label { color: var(--muted); font-size:13px; font-weight:600; letter-spacing:.2px; text-transform:uppercase; }

.arama {
  width:100%; padding:14px 16px; margin: 4px 0 16px; border-radius:999px;
  border:1px solid var(--border); background:#12151c; color:white;
  transition: border-color .15s ease, box-shadow .15s ease;
}
.arama:focus { outline:none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(33,150,243,.18); }

.pin-input {
  font-size: 34px; text-align:center; width:200px; padding:12px;
  border-radius:14px; border:1px solid var(--border); letter-spacing:10px;
  background:#12151c; color:white; display:block; margin:20px auto;
}
.pin-input:focus { outline:none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(33,150,243,.18); }

.video-alan { display:flex; justify-content:center; margin:16px 0; }
.video-alan video { border-radius: var(--radius); box-shadow:0 8px 24px rgba(0,0,0,.5); }


.hata { color:#FF6B6B; font-weight:600; text-align:center; }
.basari { color:#2FE686; font-weight:600; text-align:center; }

.rozet {
  display:inline-block; padding:3px 10px; border-radius:999px; font-size:11.5px;
  font-weight:700; letter-spacing:.3px; background:rgba(33,150,243,.15); color:#64B5F6;
}

.ozet-satir { display:flex; gap:10px; margin: 4px 0 18px; }
.ozet-kutu {
  flex:1; border-radius: var(--radius); padding: 14px 10px; text-align:center;
  border:1px solid var(--border);
  background: linear-gradient(180deg, var(--card2), var(--card));
}
.ozet-sayi { font-size: 26px; font-weight:800; line-height:1.1; }
.ozet-etiket { font-size:11.5px; color:var(--muted); margin-top:2px; text-transform:uppercase; letter-spacing:.3px; }
.ozet-yesil .ozet-sayi { color:#64DD17; }
.ozet-turuncu .ozet-sayi { color:#FF9800; }

.okut-kart {
  display:flex; align-items:center; gap:14px;
  text-decoration:none; color:white; border-radius: var(--radius);
  padding: 20px 18px; margin: 12px 0; border:1px solid var(--border);
  box-shadow: 0 4px 16px rgba(0,0,0,.25);
  transition: transform .12s ease, filter .15s ease;
}
.okut-kart:active { transform: scale(0.97); filter:brightness(0.92); }
.okut-yesil { background: linear-gradient(135deg, rgba(0,200,83,.22), rgba(100,221,23,.10)); border-color: rgba(100,221,23,.35); }
.okut-turuncu { background: linear-gradient(135deg, rgba(255,111,0,.22), rgba(255,152,0,.10)); border-color: rgba(255,152,0,.35); }
.okut-mavi { background: linear-gradient(135deg, rgba(33,150,243,.22), rgba(0,188,212,.10)); border-color: rgba(33,150,243,.35); }
.okut-mor { background: linear-gradient(135deg, rgba(94,53,177,.28), rgba(126,87,194,.12)); border-color: rgba(126,87,194,.4); }
.okut-kirmizi { background: linear-gradient(135deg, rgba(213,0,0,.22), rgba(255,23,68,.10)); border-color: rgba(255,23,68,.35); }
.okut-turkuaz { background: linear-gradient(135deg, rgba(0,131,143,.28), rgba(0,191,165,.12)); border-color: rgba(0,191,165,.4); }
.okut-ikon { font-size: 30px; width:46px; text-align:center; flex-shrink:0; }
.okut-metin { flex:1; }
.okut-baslik { font-size:17px; font-weight:700; }
.okut-alt { font-size:12.5px; color: var(--muted); margin-top:2px; }
.okut-ok { font-size:24px; color: var(--muted); }

.bolum-baslik {
  color: var(--muted); font-size:12.5px; font-weight:700; text-transform:uppercase;
  letter-spacing:.5px; margin: 22px 4px 8px;
}
.rapor-satir { display:flex; gap:8px; margin-bottom:6px; }
.rapor-pil {
  flex:1; text-align:center; padding:12px 6px; border-radius:999px;
  background:#12151c; border:1px solid var(--border); color:var(--text);
  text-decoration:none; font-weight:700; font-size:13px;
  transition: transform .1s ease, filter .15s ease;
}
.rapor-pil:active { transform: scale(0.96); filter:brightness(1.15); }

.kisi-kart {
  display:flex; align-items:center; gap:14px; text-decoration:none; color:white;
  border-radius: var(--radius); padding:14px 16px; margin:10px 0;
  border:1px solid var(--border);
  background: linear-gradient(180deg, var(--card2), var(--card));
  transition: transform .1s ease, filter .15s ease;
}
.kisi-kart:active { transform: scale(0.98); filter:brightness(0.92); }
.kisi-avatar {
  width:44px; height:44px; border-radius:50%; flex-shrink:0;
  display:flex; align-items:center; justify-content:center;
  font-weight:800; font-size:17px; color:white;
}
.kisi-yesil { background: linear-gradient(135deg, #00C853, #64DD17); }
.kisi-mavi { background: linear-gradient(135deg, #2196F3, #00BCD4); }
.kisi-mor { background: linear-gradient(135deg, #5E35B1, #7E57C2); }
.kisi-metin { flex:1; }
.kisi-ad { font-size:16px; font-weight:700; }
.kisi-rol { font-size:12px; color:var(--muted); margin-top:1px; }

.ozet-kutu:not(.ozet-yesil):not(.ozet-turuncu) .ozet-sayi { color: #64B5F6; }
.urun-kart {
  background: linear-gradient(180deg, var(--card2), var(--card));
  border:1px solid var(--border); border-radius: var(--radius);
  padding:14px 16px; margin:12px 0; text-align:left;
  box-shadow: 0 2px 10px rgba(0,0,0,.18);
}
.urun-kart.urun-kritik { border-color: rgba(255,152,0,.5); }
.etiket-sec-satir {
  display:flex; align-items:center; gap:8px; cursor:pointer; margin-bottom:10px;
  text-transform:none; font-size:13px; color:var(--muted); font-weight:500;
}
.etiket-sec-satir input { width:19px; height:19px; margin:0; }
.urun-ust { display:flex; align-items:center; justify-content:space-between; gap:8px; }
.urun-ad { font-size:16.5px; font-weight:700; }
.urun-adet-rozet {
  flex-shrink:0; padding:4px 11px; border-radius:999px; font-size:12.5px; font-weight:700;
  background: rgba(100,221,23,.15); color:#9CCC65;
}
.urun-adet-rozet.kritik { background: rgba(255,152,0,.18); color:#FFB74D; }
.urun-ozellik { font-size:13px; color:var(--muted); margin-top:6px; line-height:1.5; }
.urun-depo, .urun-barkod { font-size:12.5px; color:var(--muted); margin-top:3px; }
.urun-gorseller { display:flex; align-items:center; gap:14px; margin-top:10px; }
.urun-gorseller img:first-child { flex:1; max-width:180px; }
.urun-gorseller img:last-child { width:64px; }
.urun-aksiyonlar { margin-top:10px; }

.filtre-satir { display:flex; gap:8px; margin: 0 0 14px; flex-wrap:wrap; }
.filtre-cip {
  flex:1; text-align:center; padding:9px 6px; border-radius:999px;
  background:#12151c; border:1px solid var(--border); color:var(--muted);
  font-weight:700; font-size:13px; cursor:pointer; transition: all .15s ease;
  min-width: 80px;
}
.filtre-cip.aktif { background: var(--accent); border-color: var(--accent); color:white; }
.hareket-kart {
  display:flex; gap:12px; align-items:flex-start;
  background: linear-gradient(180deg, var(--card2), var(--card));
  border:1px solid var(--border); border-left:3px solid var(--border);
  border-radius: var(--radius); padding:13px 15px; margin:10px 0;
}
.hareket-kart.hareket-giris { border-left-color: #64DD17; }
.hareket-kart.hareket-cikis { border-left-color: #FF9800; }
.hareket-kart.hareket-iade { border-left-color: #B388FF; }
.hareket-ikon { font-size:19px; flex-shrink:0; margin-top:1px; }
.hareket-govde { flex:1; min-width:0; }
.hareket-ust { display:flex; align-items:center; justify-content:space-between; gap:8px; }
.hareket-ad { font-size:15px; font-weight:700; }
.hareket-adet { font-size:13.5px; font-weight:800; flex-shrink:0; }
.hareket-adet.giris { color:#9CCC65; }
.hareket-adet.cikis { color:#FFB74D; }
.hareket-adet.iade { color:#B388FF; }
.hareket-detay { font-size:12.5px; color:var(--muted); margin-top:3px; }
.hareket-alt { font-size:11.5px; color:var(--muted); margin-top:5px; opacity:.85; }

/* --- Toplu Excel Yükleme --- */
.yukleme-alan {
  border:2px dashed var(--border); border-radius: var(--radius);
  padding: 28px 16px; text-align:center; margin: 14px 0;
  background: linear-gradient(180deg, var(--card2), var(--card));
}
.yukleme-ikon { font-size:40px; margin-bottom:6px; }
.on-izleme-satir {
  display:flex; justify-content:space-between; align-items:center;
  padding:10px 14px; border-bottom:1px solid var(--border); font-size:13.5px;
}
.on-izleme-satir:last-child { border-bottom:none; }
.on-izleme-hata { color:#FF6B6B; font-size:12px; margin-top:2px; }
.sablon-link { display:inline-block; margin-top:8px; font-size:13px; color:var(--accent); text-decoration:none; }

/* --- Depo Özet --- */
.depo-ozet-kart {
  display:flex; align-items:center; justify-content:space-between; gap:10px;
  background: linear-gradient(180deg, var(--card2), var(--card));
  border:1px solid var(--border); border-radius: var(--radius);
  padding:14px 16px; margin:10px 0;
}
.depo-ozet-ad { font-size:14.5px; font-weight:700; }
.depo-ozet-sayilar { display:flex; gap:14px; flex-shrink:0; }
.depo-ozet-tekli { text-align:center; }
.depo-ozet-tekli .sayi { font-size:17px; font-weight:800; color:#64B5F6; }
.depo-ozet-tekli .etiket { font-size:10px; color:var(--muted); text-transform:uppercase; }

/* --- Sipariş --- */
.siparis-kart {
  display:block; text-decoration:none; color:white;
  background: linear-gradient(180deg, var(--card2), var(--card));
  border:1px solid var(--border); border-radius: var(--radius);
  padding:15px 17px; margin:10px 0;
  box-shadow: 0 2px 10px rgba(0,0,0,.18);
}
.siparis-ust { display:flex; justify-content:space-between; align-items:center; gap:8px; }
.siparis-no { font-size:16px; font-weight:800; }
.siparis-durum {
  padding:3px 10px; border-radius:999px; font-size:11px; font-weight:700; text-transform:uppercase;
}
.siparis-durum.acik { background: rgba(255,152,0,.18); color:#FFB74D; }
.siparis-durum.tamamlandi { background: rgba(100,221,23,.15); color:#9CCC65; }
.siparis-durum.iptal { background: rgba(255,23,68,.15); color:#FF6B6B; }
.siparis-detay { font-size:13px; color:var(--muted); margin-top:6px; line-height:1.6; }
.siparis-ilerleme-bar {
  height:6px; border-radius:999px; background:#12151c; margin-top:10px; overflow:hidden;
}
.siparis-ilerleme-dolu { height:100%; background: linear-gradient(90deg, #2196F3, #00BCD4); }

.kalem-satir {
  display:flex; justify-content:space-between; align-items:center; gap:8px;
  padding:10px 0; border-bottom:1px solid var(--border); font-size:14px;
}
.kalem-satir:last-child { border-bottom:none; }
.kalem-ad { flex:1; }
.kalem-miktar { font-weight:800; font-size:13.5px; flex-shrink:0; }
.kalem-miktar.tam { color:#9CCC65; }
.kalem-miktar.eksik { color:#FFB74D; }

.urun-arama-kutu { position:relative; }
.urun-arama-sonuc {
  border:1px solid var(--border); border-radius: var(--radius-sm);
  margin-top:-8px; margin-bottom:14px; max-height:260px; overflow-y:auto;
  background:#12151c; display:none;
}
.urun-arama-oge {
  padding:12px 14px; border-bottom:1px solid var(--border); cursor:pointer; font-size:13.5px;
}
.urun-arama-oge:last-child { border-bottom:none; }
.urun-arama-oge:active { background: rgba(33,150,243,.15); }
.urun-arama-oge .ad { font-weight:700; }
.urun-arama-oge .detay { color:var(--muted); font-size:12px; margin-top:2px; }

.sepet-satir {
  display:flex; align-items:center; gap:8px;
  background:#12151c; border:1px solid var(--border); border-radius: var(--radius-sm);
  padding:10px 12px; margin-bottom:8px;
}
.sepet-ad { flex:1; font-size:13.5px; font-weight:600; }
.sepet-adet-input { width:70px !important; margin:0 !important; padding:8px !important; text-align:center; }
.sepet-sil { background:none; border:none; color:#FF6B6B; font-size:20px; cursor:pointer; padding:4px 8px; }

.uyari-kutu {
  border-radius: var(--radius); padding:14px 16px; margin: 10px 0;
  background: rgba(255,152,0,.12); border:1px solid rgba(255,152,0,.35);
  color:#FFB74D; font-size:13.5px; line-height:1.6;
}
"""

LOGO_URL = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAeEAAAG7CAIAAAB7PbrIAAABWGlDQ1BJQ0MgUHJvZmlsZQAAeJx9kLFLw1AQxr9WpaB1EB0cHDKJQ5SSCro4tBVEcQhVweqUvqapkMZHkiIFN/+Bgv+BCs5uFoc6OjgIopPo5uSk4KLleS+JpCJ6j+N+fO+74zggOW5wbvcDqDu+W1zKK5ulLSX1jAS9IAzm8Zyur0r+rj/j/T703k7LWb///43Biukxqp+UGcZdH0ioxPqezyXvE4+5tBRxS7IV8onkcsjngWe9WCC+JlZYzagQvxCr5R7d6uG63WDRDnL7tOlsrMk5lBNYxA48cNgw0IQCHdk//LOBv4BdcjfhUp+FGnzqyZEiJ5jEy3DAMAOVWEOGUpN3ju53F91PjbWDJ2ChI4S4iLWVDnA2Rydrx9rUPDAyBFy1ueEagdRHmaxWgddTYLgEjN5Qz7ZXzWrh9uk8MPAoxNskkDoEui0hPo6E6B5T8wNw6XwBA6diE8HYWhMAAQAASURBVHic7L13nGRXcf79VJ1zbug4eaNWu8oZIYksEEgiZ5GTsU2ywSRjsm0wtnEAk4Mx0RhjbLDI0SCyJEAgCeW82rw7eTrdcE7V+0fP7M5KAr+/F78/duX7/fSn1ZqduX3v7e6n69apeopUFRUVFRUVhyAK/m3vQkVFRUXFr6TS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDl0qjKyoqKg5dKo2uqKioOHSpNLqioqLi0KXS6IqKiopDF/vb3oGKuxu66jHdxT8JgGFwcId//TVbozv/aNVz3OEZdeX+wD4oQAJAh897Fxtd2fKv3aeDtllR8X+FSqMrfjMOFkhd0eAhDKEDv8gCUQQACiUwKxFAqgCURAnL2q0EgBRK6ilA2QiTQgkEkHooAAITiES1hAeRIUsgQhAwKwkhAAw4Ge6UCCFACWyFoFCGEohAECgfOBY6+Ith1Q/3H9qdLz9/5bdIJecVvxlVrqPiN+OuNEjvFOzeiYPfeCQAAxyAACgNVXv1v+JOv6+AgogZzExEK9LKAISW92H/XypBgOH2h/tMuhxQ6/6nOPhY9Fcf4H8DrfxVJdAVvzFVHF3xG7OiRPs1cfgDvqNGCQOkyxqK/bHpyi/JsvQqAFrOR4CHSkrLkq2Q5Wg6+JVnEAsVVRqG38oMBIIABDCgBFJe+dpgxXL4jINTIssyrbyyM8swDvwyQ35VlubO30lUCXTF/wSVRlf8RtxlvMwHPTwQBa9kD5iW5VEAKC2nGlgRhvoLYSCAh7+vxFAIgVfyxQqICitIRQGoqCqIAOKhhtKyfvOB/WNaUW1A96ssgXR5w8NQ+kC6HHeM9uX/SHV/VdK7ouL/iEqjK/4nGYraAWFSALzy/6KA0lCg90fHMswtD4NgAwCwimF8LYCAlZa1mxWkHAhKKB0RGYIB1IkhCEGZGaKkZBTLT7G8dbCASA2wLLUHdtEc2NMVpSaIAd8pE7gqZ407HOdB/Jq0dUXF/ymVRlf8D7A6+DxIuO6Q7ACTrtRdHIhkD4ScrCtlGACRMDgsp6pXPwMDsGQVw/CaoEpgMrocZ5Ou3B30TLS/umNVgpvU0HLOhRlycGXHQTF1RcVvhUqjK34j6M6RpQIkuiJty/mK/anq4TIdq6ykpwkg5eE6IQGA6MqaG0GMKiBmeVN2pfADFgoRSBm8UDBimYwDsXJYfioCg0kBWlVvh/1ribqcLh9mWmi4uwzw/kK95VTM8t/xitL/9+ekEvWK/0Eqja74n2V5DVAguqJrRpfXB/cXO+gBIePlAg14YH9EPfxHhQp5hQpUQAxWwDAzBMhK+AJlyaIAhyhSgCNjSKCyvB1VEMCi+/MtCiIB1BxISATQMNe9skfLJYD74/flB/r/uqa7ouJ/kEqjK/5foXrX1XRBvRBK7yMXMShIySAhKNmFQaeVNoNXo2CGR9nr9UYbY3k2SGpxng3iqB6Ckig70jynyGBmCTZGPQErDNDrYe9CuW/vJd//3rU331TbtOk5L/5Dak9gsZtdefUV3/nejVdc/tNrrqSN61/ypj894X73gzpki1DG9HzW6e6ZmdnX7558v/vUJkeFDDP35hfq7ZHQ69hmTULpVbrd7tjYBGDyIg9KUZQwwRdeRZIkKoqCrWEOCgoQBhswAIUOT4jq8gNDdx09E1dRdcVvRKXRFf/fURIv3rJzLhYNQYUJUGXiQcgaaW2pPyBFDG7UYxKM1usMSaKoPzMTJfEgX0rrTTDLzCybkG+dufzHl95wyy3P/dM3aG9AkcGuvR/9s7/YetU1zbGRK7be8s/f/ga1mti184ovffNH//zppmLfzN5U9cEPO/eE+5wFw+j30e3/21v/bs8t27Zu25Wn0dmPe+S9H3U+IFQGBKp3MgzmbCj33nDTmjNOiZwZGxsty9y5JHGm3+mBYGxsouHnQqIoAuClVOZ+MXDOJZQe3NgIIvpVX2AVFb85lUZX/EZ4CWKIYImGhRZa+iwUwSWpQOLIOuuyxU7I1aj3vS5zBGabF3akbYMi6yMPey7/5ec+/KHLfvTDIzZtOf+CJ2BxgWKn1978D696jduzb7B9e0FHPPLpT3FHbkA5mL700m++691u7+yOPLNHb7zng8999MtfCM7RGcjlV739JS+rCdeTdr5t9+S9T3/2834fZYky4LadP/+Pr9x80aVbb7o1byZz9Sgcs/b1735bWquNTkwgz9Dtu7xw1sCXIFXnsqKwUWSMKctgY2OJLVmF6spVBa8INB1cC02VYlf8z1FpdMVvhHNxHorYOAKVWjhSw5DCI8vipKYkWbcz0qyh34eKDYy9+y655OJrb7/14U98woajjqI4veY/Pn/h+z/k5hePi5tzM/tOPeEYOEXW+/Tb3h5uu329de1NRwyOOfo5L30x8hxCV3zlK+3F+U1j7X1lMlg3/sxXvRQcYBxuv/lTb/mL9uzcRGN0pljauH7da9/yZkxNoFjCYPAvb/6LzpU3jQ1wQrP1y7npM848++l/+fpo/UYI0Cv2/Ozni3v31pJ0EMrGujXrjz+W4lZai5bbZIJ3iNglpXpdqdXbz7JAV7pc8f8PlUZX/CZw0EDKZZEzs7NsoAhB8sI1YnR6Sa2WJNHczh1j4xO47savf/Jff/aDH+/tLJ722IevWzNBZY65hZ//23+4627eMjIxXRQnP+rsiQffD0v7dlz689t+8KPxftELWTbSfPJTngQbQbj/k5/99HMXPuCITVmnOz0/fd75z0WaIigWZz/2qlfn118/odRfmO61x576ohe5E49Zuu2mVrv+nY9+tHPzza4zH6ftHfO77vPo8x72uldhahyzc3M/v/Yz7//QzLZti7MzNrZlzSVHrH3iC37vXuc9JDhHsJw2UzLISxM7IiOy0iJOtP8ev7bisKLiN6HS6IrfiIgsrC2yPLLGgAaDjisLZx0Wewhy2423bDn2mLHJKezd969vf1fnptt4emayXX/84x5j6jX0+7d996KFy395QtqI8mJJwsMedj7m96G39PNvfmOc6OTNR8x15q5dmF136ikoA3bedvW3vrM2MrO335IFOfaEY7asW4tsgO7ie1/8Urr++s1RxIVmCBuPOfK4hzwQ+/a06vVbv/5fl/7b55Lte6ZGxucHSzPov+T5z0I7Rr+H2cX3vfSP0342mdQ2rVm7bXFG4vgFL3nRhnucBiIDg8yju4hBjrSGNOHIgCUYwyAdtswMO2SAOwTSSlXGo+J/hkqjK34jytInLkpswkTBZ0Zgk5ru2nPhxz/1o4svefDjn7DllJMxP/vdz/7HLT/+yQQ4IZx0nzPW3uNkqEeeX/61r9d63UZc7wThZlQ//VR0Ojd/41u3fe/Htbx/w03XRZMjR51wHPIcW7d+6f0fvPq/vn36SD0shkaztXXnztG4Jt//8Uff9d41eSadfr1pe0QDNg8554FQj12zv/iv7/zgX/91MyiJajt2bT/q7LOf9NynY9MUHC/+8NJP/OXfHSua9/ogzM/20nXjr/6n92K0hakpLA2yy6/e+fPrbrnyuqXF7vpjt7SP2bTxjHu0TzkpDEu5mUgPxNF36BOXlYLqKqSu+A2pNPpuznLB8SqDzf9ONQ401612nFhp1hOs+CIRhBSpY8APlSqUWRRHuG3XB970FzdfeR3V4oc/5IGY3Q3R6euuHS1lxDFFdP55D0KRwzrcdOv2yy7faLmR2Pl+Z3LNOuybxuz0V977EbuwsLY9GmLbDUVNy6+++U3XXHtjS/SodrM7NxuTapm7YnDhW/+q9KG4dVs6NspRtNTralKfnBi75L++Vbv8F9dffU1LebyXr6nXbJSWzfaOrbe0Nx2BmdlbfvqNb33kU1NZKdN7UmuXytCcWvP0V78EY3W061hcmPvxFf/+rn9cvG7bRKPdh/z85xfXj97wBye/CZpZNQQDIcCAzLCeen8/o6w6jwzocnM59p/Vu3yNKimv+FVUGn13Zn93n1nlRnGgY1tXTOaWvfdVSXyZW2vLgoitOA7i62SJUXgPVmV1MIY0lKVlyntLjpUtB2eLrEiTBHNLH3zl6xauu2Udu2RtO2lZ9R3aOzdz5RVJt5fUa9OyFBmP2Vn0in/723eMk80G3T1hoI3a3PTOf37JK03A+IKPysj1S0eqZSe76YZgtm2A4UCskhgXEYpOZ8pa3b2Tide10tznHCecIpRFObM7m9nTAY2TjY2txSaXYse+fWvWr9+yceMlb/37fQtzs/um417fq65ZNznX7yDCBc99QuMeW5AK9mwrb59578tftYHq64xVLRalN3XipvN//+kTZ22BG5APOigJDjbRAGq2JZTsYu/F67AOmxmAlEE1IBDBkjlQQC0HSkGUSAEhZYCXvxwrua44iEqj7+bcOUaTX9usTARSRNYFobLwbACCBsTWlVr64GG49GVEgPdxP7v80ks2bDly6uTj4D0k3PaDH8nW3enMkmvVTz/2GISMRurXfOqb8WKnoUK9XuSyn33pG/dKx7/yoY/He2ak2x8fG9s3uyckrGpT6kel6iB3EigiRdGCYWUV9UqFD6pkiBhUs5EhgWgBnwGeTMnWALG1iUhDFcqqCl8qgmeKkyjrd+e2biVSU5bjIURRbCJ3+23bRjdMjNTjsWOOwsISZue/8clPX/fDX0z1slQlqTd29ObGTt70vD9/ZXLaCYgY+/aB3M++9M3ZfQuNRvvEe5wZTU401qwZmMw0Uhe5wpdZVjBpM6kRq4isvBAKVSKjNHRXhQ59QvY7u1Yp7Iq7otLouzkHDI+H9smr/+1gr43hpbohQ0LExAQmNoYBiITBYq9eTyPjur1uo95E4bG4dNU3v3fhRz/52KdcMLXlGM5LlPrtz3w227ZznYtz+Kl2CsfYuePyr30j7nTiUJgiTDLv+MZP5Oa5+euuH2NXLC1onNQoypRCKKOiH3uAvDVKJKrEBZxhYSMiVIQggYiDobQWA+JVEPxyzbJ4KKuKCmjo0SHqvS+89xKiJA551i1zY4xzzhgOIZS9cqyW1j3Nb5v++iv/bOf8bC/LE5essbUkQFA2JhsY+Je8+2+wYQKxw7Y9O7596Yfe8rZ8MRsfnWiOjn9x74exZrxx9BEv/Zu3tBobCRpZGwHQQNC8l5k0Hb4MAvDQmIRICQod+lUD4JVEklKV9Ki4I5VG/++Afs3/AStjShhs2CEAhULVxASIlKUGqdca8AXAJihU4AvdvuuL//Sxuetv6913L8oQJzF2T++86po1Wb9to4Ixc9MNx88u/PLC/+xv2z5SBEjJbEajpByU1/zs+8duOCqbWxxtNjqdpXqj4aWINDjJnKqhwCBR8kKGY2Zj2CoksKqqGX5zKIuKCFSJCY7IAEY1UWNBvJzLkZLUkAYiy9ZrUCZrDBsWkdyXPi98EZzXUZh0vn9KOt5BP66l3V4/lNJHMTOz5+WvexWMwVwH2/Z85n0f2f7tS45EGk2MmLS+a3p2fKR50gPu++jf/x27drK/tBQ36lC1bGBMNr8AZxnkMexEDApiZSICJCy7poLBw2WCym+64i6pNPpuzoo7EegOY/qGP6fh2MCV626FUYaSBh/UWzFQUfFGCXunwTzXXRiZGkORY3b2a//8qXo329hqZdPT6C5gcmzx+mtGEKbqcZifSWxz5qpriq99++cXfnVUYDWoM4G4b0RZN06sme3tcixjo+u65aBfZGCxEgyVRtSUgWA8maAoGDIM5SEwygwmkKEslBANKkpqiRxIAatscxgdDlMRVXXgwDawQomZSpWyyLQADDtna3Gj7A3Wjox2du9ru/i2nTtrSTKTdTlJ6uONvbv3bmyuR6eff/7b191089e/+vVYzLhNFhb31NdMzmu/fcrRz/vT10X3OAXMQKjVEnQz+AAoIpvUa7CmoBCIh3EzgRRhxbF61QSYZds/EjowZquiYkil0Xdz6M6PD5bq/dUIw3IxDUpEYDXGgAQGBlb2zNzy0yu+9a1vJetGn/fHL0c2gIuu/9GPw+59I/VGMb0LocT89GVf+4brdBMjZCX0uqz2h5/5XHbrzsnmCJSMi4JISSFQ6fNuQE6Gbp++3VEDho1qbIxTYRIYEpBAYV0eyiyUVgmAh1+ZWKXqV4ZpERswAxoCRMpAXpkkAGBSZgKRMdTLuyZyseOgyH0ppQIxO7KOu70lQRgU+WSzQWlKFDpZNjMzMzU+ViwsvvP1bzxq05buYvekqAbmNK1n3aW4kTzsMeef8fzfQT2F78IrKN558c8v+uLXF2YX1h658ZjTTznh/me5yTHTHhFWAhjEQ0O/5ZExK7u/PI6AlMBaTdiquCOVRt+dodVVHEPutColKz9WiIBhWH0QJhOxD7lVwkL34i99+b/+44u33r71uX/6x4BgcXHHj3/S37b9yKhmJUt6C5ifLa+9+tbvfS/tdzu9bKzRCCGUeT574+2jsL43SNiaUiyTMZxROTOXbzqylnezQaFiQyOqSZE7VgZUTVmGoMYbo4bhnGrwUCJ1bIdzDlUpEx2mcYepZw0SfO5FqdYMZEl46D3NpAZKrEWmKYXYRcwc55znufTznPP6WGuQZa3J1t75pVqrvWvf9MTEqBXby/qxU0/FvbccMzMzM5mmeb8TuWQsGaEyn9w4dcbTngAbIAVu3VNML376PR/b+csbm8GK16u/97NffOHrU2eddP5znn7qwx+u1hgiItpf/KgErPbDW71UUGl0xcFUGn0356CPvB78+K7kgAzyIogWKcWDQa8JWrj++i997BMy3z1ibOT8xz86zO01ub/mhz/Y1KjH80s2oWL37fjRj77/ta/J9l1bptbumFss06TeaHT37J0cH+/2ev3BIG00Q54lcW3QGxTSP+6IsVtvm2u3acO6qfm5/uzcbDtNAVLDpYacOId48myYQSpACMRiiYceRqUAzqkSkWGohWoQL5p7jY0JRCBiEImQeg5KwXPNBMOFBJJAqtYYAgKrLzKwTneXGmPNXpm3RuqDbi+GO2X95h27tjca7d233TI6NtZfmm83m74stt5y49rx0XZi5370Qx1t/+znv5jdtmf7lTeMhHhtKZoNktERhp21MtKuL83NIHi7PGfR7p8pEJSEAcCspKGWRw9QVSxdcUcqW8W7OcPSL1o9PgoAkGVZnCZKUkowbAN0kA8acc0o+TJ3kQFKhBLzS1/7+/de+sWv18mOH7Pp+e//OyQG2/e8/XeeP7V3yc3PNxq2gB/EjoC052MhxE4J+aAgD5ZgI4N6ZNm4XLTwgTjAg8oQfCGI4tSa1Ge5BQwpDJUGPdWeek1cUqtzYMuGSYLkZVkIFNaSTXyAsCMyRGpUoJ40QHQ4q4UhEGEJ0JJECSFNEvWlhJJEjRrLbAQASi0UCCGUZfACa11i4oitdAaGWCEBGiIiAwnBl2LqrdzYgdG+o9ywKFnhKLAtORB3LDaeccojXvC7jZOOxXgLxIhqcA7sAAPRIL40jpNEgQCQVwIiIiYgqFexdn+Vx0FQlQT534lWcfTdneV0551+7pwriqIscxu54MiQqUVxyHNrXXffntH1k95nWgycL2ZuvTXqduCV5uvodWEbs5f9wnW7TsNIIyXJST2XlhXOwwxT28r1qMZWNXgPn+WDnCT1xiiXPgCwLrI2IcDamKOYbVRk/UBUCHJl30hNM+lovqfXN4goKEhUWV0cp0nabEW1Rn1kXMmSsRhOm1VRKUh1z67dJKo+SFki5FoaCt6olbxgCQZITJRwxEFDkfsiq9fiWhzFzFHhCYiMzcvQ7S3V4ijLc1JJGrV6M8nzPOsWxjIgLGVNEQkCsSrZQAD3ct9j3XT/e17wihdg0xpsGkNWwhDKLNx62+LcQnN0wm3aZBopIRQhh3EEYgPeL79EFkYrA72Kg6k0+m4OrQqeV18zKUIcuziyIBr4wllmAXvB3t0XfuhD5zzpURuPOTJpNLB3rje9b9xFfrBY7txV/vRyNzV25de+GfX6hq1JY9/PWdiGoT8ng8gIG6Do9mMXRbFxjsmQaHDKhiOCCUIWrOSCLwcAVDRiG4+QiwLFS0UxI0Uhpr1x09T6dUcec2p7bHzdujWTayeT0RHUa0hiuAQC8LAVmwCBDMcYlmgm8AUGGRYXe9PTi3v2Ls7MFEudG6+8ejA/35ueWVjs2lJiIZMq+1BGMt1ZcoVvuyhlLorCsyTjzX6ZmVrkjO2X5cLcXC2K6vX6YJCLejOc2RVA4o0QK5fMRaQbNq2/4MW/h42TaNawcyeS+tYfX/r59/xjNj3fHWQ+TjecdsrDnvmME85/SBTFlJKyAbGq+CAifvi6RDb+v/r+qDjkqXIdd3dWvDSXZzvtn8YdPJnhN7T2Bv3IOVMELuXHH/7wJz7+4Yf+3lOf+kcvgggW+29/6OO3iPULCwu9xVPOfeDY5NgPv/AVW/hWFMUKLnIKAjYAI/dERIljBXIxzGqlZF+gCAjGq8DYWnOxmxWlcFKTWr2MeGC0JHhB2hrbfMzJp9zr3kefcU9aP4W6hbOoNUEWhgAFCQzDRnAOpQet0mhVaAgUSvLK3ig5VfIK7+EFPiDLMMix1CtmZvZs3b71uhtvue6GvTu39RenW4wRmFopNBigLIzRKHbdrGcdp2nqjJVBrnkZqXXOZVos92yLGoURQKg0nNeifQijpxz3qGc/a64o5nuDb3/z23tvuPmM0alifrErWjQbe0nnk+jej3vMs17yh9HIBFw03FJQCaEkIgYx33XYVOU6/pdS5Tru9gxFWYfTVwGBDjsPrbFFp+usFVVWdcrwHrv3fedjn1pfquzci8VOKArjtViYN66RhqCl3/6Ty+ZbdbO42KjVIVpKiEFQYmWAAw2nbUsgNo7LEIpQEiiK09SyVROUKa1xY3RJacYX84bc+sktZ5121OmnnHbW/ZKxKTTGQBECYBnOwgJBQQSD4QjaoCJBURTOxSBavsFCAWXAWlcTBBb1QdgqcyAHqCL2GAE2mAi66b5hU1E+qMjh8xsuu/iWX15+4yWX7dq6rW5pVBpRf5Av9qYm1xrHi4NOUfh22vKU9zu91DkZ1v7JsARbPSkMCXHZydpJ4m/a/vHXvbk2Pnnr1h2NtL4lSpdu2BbHcYjNYDBI1oxv2Xzk5OTkwuzCeNKiAHYRGEaZKV4edat6l/noiv+1VBr9vwUlqKoAAiUFVCPrEMdGNQ1ei5KE5m6+Ld++e+1Ivdy+G3tmzfjo/C8vrytHZUG97qjhQZGVC8VUrc7M3ofhloUAAqsoiRIRsZJ0iwEbZ5zjJCLDuWKhLLu+XBjk1G6NbjnqrAfc97SHPLB+0tFo1WEIjRGoCSWFoFBrrWU2gPqQMykRKw2LpMEOpBAROrAKqlBSJSFIICLLhP3xqApUlWsryidefZCaZ60RyfGPfsTxDzv3Ub3Mb9t+/cU/vfKi789eewNbV+TeZJ5ALITOoBkltYbpdBbjFKpqAIIoFIRACKTtdpNctH33vjXt8XpP1q/bnPUK8do+4qirt2/dcMoJT3jJC+197wVWuATtNsgAgCoUUIGIBBGojaL/m++KikOfSqPv5uiqQU5DgQ4IAGwhZOyuG2/qDwbHnHoyiaII1/7iyiPabe0s7Lv6+vKWbS5KrvruxbUi1Kz1EgRl6hoDX8SWHDgbeDCJ0cAqHAhQCSASg0DM9YhNDLYDCYtFWcYujI36Vuv8Cy7YeNo9xk49GVMTsECRKZTSOhCLFw9QZEQkkyLmxDBZmwAIEnzQADXGOGYm+EIMBMt9eQySQAQFS4CoVyJRBhERMxMvnwUBAiCsRCwMS4aUYW1prDv1pFPOuucpz39u95fXbrv8yu/+++eyHbtHShk3TntdX/iIyTIDIiQEsMiwRTMwSgowfnrfYjOJxlys872l3p46t6Vdu27P3nN//7mnPuvJOHYTWCGAcfAe/YFahosodogsGRoaMCn9es+riv91VBp9mPHfujocZPq83w16peFbICxsRUiBfTNf/MgnukX2qrf8OTeamJ+55rKftGPmgpPgb/r+D05au/667343KgaDYlCvu7m5fhzKMMgDqFGLy5AX3muNPXuLaPi8SgDYEw8MeYNMQ2FsY2rtsWedecbDzmvd+94YaaGeqmouwVhn47ZKIGUpAjsXM5UQEsRkWAO8IpQgMmSMYRBDBT5A1GpY9uQgAwpQWChUVYSIYe1yoCqQ4H0RnDpPqhrAZK0lUpUQRK1xiJ2Na6LKRGjEjbPOPOmkk0969KOWLr/q4gs/f/0PL7EqdZioCGncUMoYBQAlMkNvaFEY1Bp1nu/VakliTSnlkc0pFdq+uJBMNZpjKRoxygxZhkG2NLs0t3c6sc64OGrV44mxZGIM7QaSyBpWhIMXiA7o9YE20f09olUt9f8CqjXDw4k7+0Hv/5RqEDCBaPgLNGwcVF3oLTbqqfNq4qgn5SDPmjaNYXHrrd/+zGc//8+f2nLqiX/y0Q8gdtg1/edPfMKxS3OtcpAJ9xVCNnHc0OC0RFkykRaQPKQc1dMk9/nCYMmM2ILUiVVPWT8PYNds8ujI9qVO3qiPbTnmnuedf8ZDH4bNW2AtHFBPuv1eo1YXkV6Zq+XEJqLeBRgQWJVWxk8VpQwGPGzuUF1uoV5u+SD4cvmMLBeuDe8FsQXrynKixbB2gpjYBgkhKCARGxBBVCHibHfQb7maFuK9j+o1sCKUyLswBlmGm2+94r++d9nXL+rdvG2MeW3Lzey4bdDtj9XTujF51idSju3CoEwbkQTmQCNxkwry/dwnZk+aFw07Nr7G1Vp7FpZmF5cINomizsx8o9XsQ+c4TJ1xj3Of8ZR7nneeHRtVmAAqSh9UjLFBSQTW2pgdK3il1UWHNlgEU8n03ZtqzfCwQ+8qnroDvCqabjXrWa+fxLX+Use10kwRuwjTc5hduPQLXzqqVmuXBcTDC3qLiRZU9l1ZGnGxiJIYZWLv4YuyMMLIYTxrTKSUWNNK4k6Wj02MZYMw2+nYWsM2mjNlsXdu9j6PefQx973Plnvdh9dvRL0VjAnsEFkhiuptVQ1BGkldoADsMIb1AUWgQQ9FBhCCcAiYm5PeYHFhYXZ2dn52ZmlpKR8MvPeWDTMbY4yz1rK11hgDgy1HH5WkaX1kxLRH0BhBmsA5IoaQscYYCyjyDCJgJWuh7NgRRxTBhEJKKiBZmTebLZHMmJhPPOb0Izef/rgnbv32jy/+0pd/dsUla0fHJtZt1KVuWZat0YbP+qXP142OegqlF0/SK3oRWZOQD/0x5/KsjHfsY5qfCKhLyLQs0Jtot+fzzDpqN5uR5aw/yPsD22oOfM7GMZiNM8ZClQxZcqscmJb/K5U2/++g0ujDDFpl7KAr9tDLQaQCKxUQpApVUXVkOI3LPIuc8UXZjCJAYM1Pv/5NTM8YJurVUWYowr5rrq6VYjLYYEFkwIEEIiqliKdSGIZdbCIrhKzIuBxQWa5pj83syHoQNzKxYHlvmW++71kvevbT15x6EqbWoD0K4xRWAvmgUDjlxblFVhkbG0WZF4vzEQIbg24P3Qw7d26/9rqbr7p6+y23ze7dWy51is6CIzhjLJuhLjOEiPI8B8DMw4wzDBMpK1/WD0HgmVGPk5GR9tqp9tRUMtI8+oSTGxOjY+vXm4lR1GKkDrUaDJedxVpjFEXW7w5q9XYg5P282WobgpJ2QsnONqdaaLU2jz1+8wWPvvkrX/jWZz59+TU3Hj82mS8szO6bnRptqg+9ha6ziJw1KoX40rrSogwes75unC0HPvQcOVOLk0Ytb9Zv7y+Nn3rCfR9+3kkPfEB982aMjsGYstdP4hpTpKreIwQpfFAmk9gDs21p/+LCEPN/9f1X8X+dKtdxmLF6PuHQ9BlD24cgQ7+65V9ShRAQCi0j61AGKMEwVBA8yvC2hzxiQ9A9W2+urZt43lv/wh137Jff8le7LvnZyL7ZMSUyTlkCKUwIoShLH1kXhJUi62ILRZlHZWZg5gpMbDxm3uqNC/vapx7/qBf8zrqHnAMHTK3xov0yEEfG1VQpeCVR9lKvxQiK+TmUHgzM7rvlB9+/7Lvf3XHDDf3d+9bW21P1ps+LrNc3CJGWLCIiWCkbGTrFWWuBYf5ZVDWoqAYSa7wFWCwHx4Wj0lAwpjDIVSSKtJ7GYyMjmzZsPunEY+5x8tSRWzA6hqSJ3JdeXGMESVKWHga+yAKJbSQAB/jQHTQQsTXYtQ0+v+lb3/38P354cPv2TbV61OtOpHG51GnEzvvC+4LTKBAGZQFwHKLUxCh88ELG9A0vWO62ay9+7ztxzBFYMwVr4GKwRSEIChchqcOQKoRJGGrsqgujZdvVlSZ/WJhqjfHujFYafZiiABBW8tEM4I4aHaAEDbkvY+vgA7zuuub69UdtQd1hafEdj3nC2GKHB91cw3H3v/85T3z8O1/zmnGlxuxSE0zGKQEkpCFIXnpfbzUWur0sUFSrp3ECX1pfqkuLZPL66dnGcUc88vnP2fyER2C8MRDxaSxqbZQklAAsXoR4OColtRZFjulZdPtLP7/8i//0kT1XX2Pzfj129chRWQ6W+iglTVKb1ojVaDE0vjYgIjLEhpSI+v0+ACJSpmE0DVYjLINgmNmyMHkp+2XeKwd5kKieFsze2BA5imOJIrGmYB494ogHPOwRpz70odh4BETLoK7RQJKgLBDZAAQg97lRxJ4l67uGRdYFCLOzP/nkv1382Qvrs4tpp7uxOepCmfe7vsjIUoAWpCaKJZjIxiwqPqhqTjCT4/VjNp/7mpdh43rUawCBLDrZ/M2379iz9/gH3DcaH0c9BbNY40lLBZONncMqxxWl1Rpdcfel0ujDFQUIAXfSaKKh8TJ0OC5FNYhlg4Xermuu+8+Pf+opz3zq2nPOxMLMN179hlu/ddEYUZHlpt448eSTfv6TS9rWxIM8hSFYETEKIlGEImRJqzG31Amg+tg4u2S2u9QdZFmUYnLTQ576tNMf9zBsWYeYFnxZxlGzOQmYUrwRTowlQEQJxKS+u2TFX/aFL133X9/d+YMfp3OLG5PY5FmrXe91l8q8SFxiTZyX2pcgJNaUhsWRI1ZSDqEMpRcR5xwxCEwMEBMBTKTw/ZJEVQNImWGM4diAOfehUC0BWEdRpODC6yCUjanJ3f3ejGD9aafe/4lPPPPhD6N1GxYX5pqtEY4cgCyUUBbvI7U2NoKMWQBBr4Olfv+aa3/ymQu3X/rzfPuueJCNp2nd2sXuYhnKWqsZJbH3oSgKCUji2BD3srxLWrRTrJ86/uz7HXfWmYtL/Ysv+tE1P7uSs5A0W71G7fRzH3TuYx619qwzEDvRMCjLOKkxW9L9Gk37SyqrEbV3cyqNPlxZpdHDlX0NINJljQZU1ZOqKklwpWD7ng+95W9uuOzyJz7naQ98zfMwWLzqHR/4+af+nffORT44m5Y+T1OnUloER5ZKE0oYYbIUrAbNbcKDwcDZOGq1F5T3aKB1a1rHHfeEF7+0tXkLGjVwQK0GNr4IJayLk/6gIKI0cexVfGHVAFLO773ye9/70Fv+ap0P0fbdp01NxVmvOzcbJ8N4kCmwqhGKJIpcRCxd0gICkA6PjPevmpGokKgXJREZdtRIgHNxGsVxZEnFe58XWe49W+tJfVARQNkwWxM7NoOs1weZdet3G/r5/Nypj3rE773+9WuOP14NZ+XyyqSqWmshNL8wOzrWLnyfodZalANIwPZd133921d87kv5bTuSXt5gyvr9Uoo4TdOIs4VZDWpiF0WJeg1BwCyRkWYy0+/1y5LJRVxjWOLIu2RapGjU1p9ywn0f+4hTz32QnRr3bEyS6vDlXV3ujtWrExV3U6q6jsOd5dFKK+EVCAomQCkMy7TKonBk9txw4xXf/e5U3Kh5j9kZxGg0aszUSGI/vWhY6knUX1qqNWKyBKh4HdoUMTkwQIlogLE58fygv5ctH3nkvZ/6pAc8/WmYmsy9NwzrEoQgWWE4tnGcZ6EeR2CIwvsisgaifqnjGo3jjj16tNkY9ZLTzlYt6S7N1GIiDlEUEZlQqgqxMb1Qdhe6EeeW4SwZY3joWRGCqhKp0LAGUZlULRygYNesl2XZK7qLXSGFM3DOpmmS50U9jqNaJCL5oPBFaYJGTFP1xu7FhazbbY60N01MHrNp85qJSTCXXpityHB8ogml9xKaI6OLnU6zWScg95mAWYKbHD/xWU8//pRTLr3wyz//yjd27ZufqKcR0n6elf286SKYUozxQX1eMIxjKQcDF8pNUVQqEay18WIps4M82GhqaurqXTsWrr12w+knn1CWlg0B0DtNo9VVD6pQ+m5NpdGHGYIAgMDQlZX+YQpa0c8GST31wUfGzi92onqc2rgIAVH83a9+ZfP4RDEzN33LzVBBVlBZ+rJcmluasBGLOjYRE7wv+r5eS3OUA1+apJ7UYogUed5f6LfXjS8x7QnlsQ89/3F/9Ed02ulQnysy8e1aHSHAGDZR0cmiqB5bI4CqMIRiUhUi2JGG73VbJ578rBf+wZff//6JtWv3LsyNpsmg6LLCF1ni2DpXZIX3hYVNHKyJhr5QPoB06IJNAIkEXU5v0HBqKxEJoZ8PhIhjZ5LhaeE8CAVijpiiQS8DoBrYknVmMMiKbkH11MdmZtDZfK97PfnZz8bIONSwcQAxgRUSFGzImABK6yOD3DPDchpxSuxhMljme51+/9NP2Xi/M7/y4Y9tv/qGSROPNFqp+MH8dDOtm6S2uNRfLIqxdHRkfKzbXfIhSznpLe5NGiOFo+1LnXjL5vYJxx3/gPs/5+z71Y46EkmMZh3EKAsSVQlgVqiqEhkygIgvSxtXPnl3cyqNPpzYX2yHlUq7/ZQ+JGla+uCs7fd6I/WasTbL+820jvnF/sJc3pnb0GrsuPpqbNuFNaN7brvFhjDSHI3LUPR7gDrnDKMsfdYdUBTFjTiABj5jNZ4tj4xv7RZ+zcgjf++FJz/5yRidQBFCHEVxBMMhKAkxNB/04tbIoLOYNpqkAIbNzcs+QURQF6Es73HOOdO33HzpJ//FqnA+aNdq1kiWZ0WQ3OehQBqrswYhQMz+lp3l4FEBwBhzIBdLJAQiMgCJJ4KADIwSVInIKiAaghLIqAZRUgas4SQqysCtxhzEN5uPePrTtF4jJhUiJqXl9vFhHnhovxoKTSLLgAZ47y1b2NogW3JJ4v3Spkec9/yjtnznE5/+xVe/2e9nm9qtYilSG2V5IDUbx49AkL1796ZpnGWZIRppjxTKu+ZnT7//fc575ctw7FFIHMbHAYHPkDOi2Aghy8tQBmKbxKoaQpEkCTG8LyqNvttTafThh6zUxNJ+zSIIVBmkZAAuMhP8vn37ptavhVFIEJ8bDZGE3Tdv3frlb248ct0NP/6Z9nNrayEPYq0aUg/nokxzS5Eya/AgYY6KoAsFakcese6YLee98LntB9wHaYqkgX4wnnLxcVoX9SGUHEWsDOa03RQRrLJCBQQEVSYXwTpsXHv+7zzztst/0fnl1S2KpSwNMStgSQhaau5DQgWpqFroSunC8rwCEYIKQ1a2TwQgEIyqzQoLqDFKIcAGpiAQKDMHoQAWJqXADDEEikJk5pXmnTnnKU868uz7Y+0GzYVSHvZrLjfQ04HUgrW0tJSNtRIhgGwA5udnJsbH8rJv66PFoB8dd9wjX/HyDcec+IPPfO6qa68/fmw0K4vQzWIykff9zhJpFkdJt1dGaSMbFIuDLAuyb3bHnpt+uXYiwto1mO4gqqGfT++ebo6OJUccCfHR6JhAvFUioz7kZebYJGn6//u7reK3TbVmeDihwNARyeiwSvhAy2EZQhF8PXEocirkxksv+d63vv3CV70CsYXK5//4VTsv+Ynb13EqI5vXpqm7+aorJqN6W1Pfz0zCziDvL9TjejkQr6IRkkYaLM12BgveFu2xkx/+iPs89YnNe56KyXFflIPMN+sjC3tnR9aNzy4stVpNa+m2rVs3b95MhKLInXNKQ9chWem2AcAqLKF0zmBxYeHHl376LX+VbN3V7vXHE1PkPTGiBpa53804wFqjGgstr4vx8ghtLK+LYjm/sX+0uVNNSx+pKhuQEeIAFKBAWqook5Iws2EyUCbNwf1a4/pu74RHP+xpf/lmTEyiOSqFsomEhvaA0JWOPiVASRXOoOyJYzbx8gDZXlYaB2gwJDTou8xjUOz+4aUXffrTg5tvkH17WyWNuyibn3VAe6K9OOgEaL3dWlrqIoqoWdvaWZSx1rrTTlh7wom54BeXXrb91m3NeruX+fVHHf343/2dE899CMZHYWmgcDYB4Isyso65WjS8W1OtGd49UAIMQpmTEsoci/2L//NLV/3wBzvOPHPjA+8D0iOnxmfygrN8w8TEvltuny8HGxqNFMaWwpGl2IqULBh0ByNrjtyxZzflZdTkxX5/VsLoaacc+cAHnf26N8AYNJqDft+ZqNls9Ob7I2vHRTA+2hJg587OFZffFEJ87DHrDMcASDFsslHQfp1VKUXFkulrMfLg+z/glqd+/wMfZ9EGOLUmy3uDvJ+Mt63TUnIlIntgkIwqQMoKITLGKCCqwzh3eBKCSMFMQsMBVASwkh2uuKmW4odNiY6tBWRQlNBZi4mTTn3s778AU+uQRPMLc6Oja3yeD38FAEiMMgASVuJub6BRTEE1qM/JWKYaoEaIAxnDHIwW5VK9Xl/38HOfeeKxF779HdOX/2Jxz6wfdDeOjdYMpmf3ceyYaG56ziVpLUkHpV+bJFk+2PXdH2794SVQ23DJveKmdkK3lMH1Wy9867vOuPLaBzzh8a0z7pFACi3IJRTFHnDVkuHdnepL+PBmeDEOUkFIEufzHlTmr7rm2h/8YDz3F/37fyAEFNkR4xN+sTOWpOj2a17rXkddJIO+DnoJsy9zX+aJi8oQQpC00XLN1rbu4q2DwdS9Tjvvhb979u8+G6nTJJlfWEobDZtEg04Rt2qDDEWB665Z+PM//cATHveM177mzR9830c7i7AGJCBlDG/goWIrCGSiOJ6en09Hx4Iv7vGUJ53xxMf02q2d/RyuPtIYd2K1gIios8EQWImFhkbOLERKpIaCaoCWICEM+zxKaCkIGaSPMGDJSAr4nIpCS48ShlQFDDAJkzCXxEUUF2NjT3jxS2pnnqUhZF44iQeDLjsGeVBQDiAFBYIwhBTORZbMa1/7+sc/7kkX//gn/X65NFsycVmCwAtLvV4R6lOTmia5A5107JNe/fJ7PP4xC3WXt9KOC9sX9rlWnZi9F0POqJ3bN7+4Z6ZWypTyWk+bNT42aowN1M53W2pSZvUlM775zW/++Ec/8tMzZF2ZF0WRYVU/eMXdmCqOPuw4cHG7fDFOEEggXyNDpCiL63/2E56b3zQ6sfvyq+Z+cfnYycdz4SNCSla6/UbkgtV+dwneoyBmwxTKUKjl2Nj5vbNupF3UbRlFR55+3P2f/vS155yLxrgGFWB0pJ11B2pscDw73x1pNV73x3931RXXz88v9Hr5wmL/Fz+99srLr73XmSfF6XAPhzkZMBiQQChZARNRhIIlSk2tffZznrkwu3DT574y5+UIFzeihop671G3qso+WAFUQbJqNKMQQRB06BytysMJBgSyzhMJiIZrfsu2JQoIWTFRpEwZCbENtagYGb3vU5+y4eEPhTFFEveLstVoDTqd1BivQUhWWoQYQACgVIvtxz/2+a9/43tRlLzoj155/Eknvve97x2vuVB4tRhp1QF0806SmDhN88XF+NhN93vpC9cdMfW5d76zNzt9zJa1i/vmapG1FNWTZuG154t2rR174zud9fW2cDrfySDcZw716NRzzp067ZS9Gm7tLGDDxM7e0kaarDXqZVAS9WWJuJoJcDen0ujDDAUCYJavb5d9SRkwApiAiLHQ27X15vEkigb9UV9efOGFj9n4h73dO0fjOHQHrXpCVkqhufnO5OhI0MyH4GLO85CpT5ojg8LMlPmefnHEGWed94Lfm7zvfeFqYh1TpKX2Oz1j4yhxCx1845s/fM+7PiR9zM90W42RRjo50py64bqtb/mzv/3aNz65ep9J99tAodvrNur1kfbY4txMe3wEGvJ262HPe+62y67efdNW6vZrIaQ154ylKMrzwXAwARBwoMsOAJwjwAYoYAwJACENRP3ghckMC/QUpAGiEMQ2AuCMLYA+cRkZH0Xl2rH7vuD5ufcwiZBrNOq9waDVbJehUD5QRUJKUCa1qrj8Z9vf/c6PMJrdbh7ZdPv22fMe8piXvfIlT3vak0anOAQMijJNawSfl72o3eh1uvVatPmRD3tuu/WNj33k8p/8dGO9kXKs4hcW5tK03qrVRD18QBaWsoW4jti5uV5n9JgjH/WHv1976ENQd5utu8/4OAKUHSlQlpqXal0tSQDoXVjgrXyPr+52qWqpD0+qXMfhhAAFylx9Hnw5/MAp4IlLjoW7MzOQAjWeXNOQ0Eko19m90z+7BD/76dXf+W4igSnnRJYGC7tnZhqNWl74wpqB1dJw2m42R8b39JZm27a7eWzjuec8581/OXmfc9FYV5oGR81uL7dxYqLURO7W25b+5I/f8mevfVt/nntLRbs5Cu/KPhc9m7qx+bn+S178RjAG/Ux8oaEssr74wMRaSrvZcmxU0B4ZC8oBJl67ITr1tGe/8XXRUZv3aEjXTCW1plEs7u3VTWQsYAkMJQ2sYgBL5Djzeal+ue5CtCjKMi/zvFAdDghXiA59h4wxSeTypcJkwc93TTAF8Syb5MxTX/SBd2O0GY+OOps6ROqRxGmhGoxdGvQV1Ol2DKyhWAvOO+X8nuwVL3l9f4kiGkvMpMPI/D6vRfzed3zsWU97wY8uutEp6tblHR9yiV0zy32I46xWw9T6ycc/+dEve/3aez94XqMicqUWaR1lvjDoz6YRM4iSZGL9+sWiV1DZGKudct9Ta6cdhaZgNEI7wtI8AJ2Zxcwcehn3BwSfh0GpeallUC/qVbyKVwwTQcPb8DwAAtHlxtSKw4uqruNwQoGBFgI1Yg2xY4KAPCAeRR81g8EsKFz9sX/5z7995wmU2KzYvjR31ln3vOkXV9Stc6qteqM76Fpru0uDNHatVisv/KAsm+3Wzj17aHJifs1EPjXx6rf/AybXYGrtwsJS2hyxxrEAiqD43Of+6x3v+Mf+wITMkrKEMnFJkVPwcM6Vvk+mNz4V/clrXvjox54bxZTl/aReH3T6zjkbu5IEgBXG8iBEAcSUATNz13ziUz/68Mcbi0v1vJdQObl2dNfuHUSM4RyZlTfq8CeGWFVFMPS9s9ZGUWQjI87kRVEUhWgwxhhDFkRe6iYNntS4HVk+267VTjnx2W95Y7xpE8bWg2JlI8QCiIEQlIIPhWFY4pD52NThwR5/9qf/cOF/fk81TdN6v983xhCRsZrn/fZIvdub/Yu/fMOTn/IgWICxOD9Ta9e8ZVFJYYvphWTgFy//xeff8bZtl/74pPFWPBjQoJgaHZ+ZXlDmtF6b7yy5ej1pNZaKYuT4LWP3PbUYaxXWZUWoIe0uDUDRhs1Hbzn+xMaJx2KinYWCowiABWF/LSYZ7L92WSn+UV524LJVKH14Ufl1HHZ4KUUEapmZmQigskSRf+kTn7j3vc5Ye+ox8EV55S/f+YcvWze91C79QHIF4ijioBTCSKu9tDjvnOl1OuvWrZvdM+MFrYnJbTPTdmJ0a5FP3O8+f/A3b8XEBEbHoEZszGy0RChhY7zjHz75jx/8uLWt3lIZm4ZhZ21EZHxJEuCcCzLIigWlpc1HTb3rPX99wklb4jotzs+1R1q+LNlZkFE6oCDDqjmSUqb3srXfedNf/uyzFx4bx/V8UPSXamlMpKoaQhhqMQCAiWjoIUVkRESDAGBmQyxFaS1TZAOFPHiR4Ihj2LLvbb1RRrUd4ufHR1/49r9un/dARAkQg6yCAzEIgYffHErQvMjrUYqA2X1LayZbH/nAZz/4/k/Oz4VGfYJger2etZG1HCeuLLN90zuPP2HLbbdf/4xnPuG1r39FvWHjusnynkvTftavmcQoY3ERzPqLn/3HO96x4+KLp0SPmZia3baj0WgoYbGzFKAKbo2N9rNB36gfb/YNgqDIAufSqI9yXL99ekZGWuc+62n3esYT6sceFdgFshZ0IBHEwOra+ZWF5WEPTqXRhxla5ToONwyxIbYEJhJRVQ8tMeh97iMf+/InPoV+QCdzUxsmNxwZFOrDSFyzha+7GKohSJ4X/UG+tNQZHR/fvXt37KI0SmYXl7TZ2kvmuIed9wdvezvGxjA6lvcGiCOAB52SGCHDHz7vzz/4nk+k8UQoTKs2lsa1YdhZ5OWyER0RkYmjWpqMbL1tzxte99eWaWmuqNdGghcbx+rDsLxD75AbZeK1EyA57wXPPfkxD52N3bRKbuLcU5EjG8igH7q9stf3/UKLwF7tQDmDyWAKdgW7nGwGkysrTF7oYnew1MvzUpUiqCvV+iiZC9iputRqP/7Ff9g+6z4QI+xgDJiUaTiBi3S5CtuAYxtLQPCYmmxdevHNH/jAx8sCcVxjsllWGBNbGxWF5JkyxRs3bLnpxttH2pOf+bfPv/Llr9u7Z8FnsFxXL42koUqwrLUYGuj0ezztta/FhiNoat3lW2+pTUwWIp1edzhKxjDIlzFrVOTYN51Oz47MLEzML55Ub65TGZdyypk4G1x20UXf+Ox/9PbuNgo+OFOPlZpuXVHn/eWJFYcjlUYfZgwrfJmZVEQ9NIjkOugk/d7NF/9k18U/R0boh3VHHJkHMcYV3V6NXW9mnooQm9h7z9ZEtXhmbpaZfVG6JJW4Nq1Ye7/7POVPXoMNGzsmQZTEo2Mz0/NMxOp23dp5+tNedM1Vt6e1Nf0ei48JUbeTARARhRAFUCFaEKlzqTV1SH3n9vmX/OFfxDZisKFopQiP97/rlkcVEgdmZVc2Ehx95GNf/Qps3rCL2U6tpaTJplaPR0baayanNk2uPbI9vj5ujaHWtM0xaoz4Wr1Ma77RoNHRaGoqXb++GB31I6MyMobWqG2NRPV21Bjh5ihNTvk1a3ck7t7PfPpxz36WRImkzSyQLO+MLOvcclU1VNQx57lKgC/xpj9/a54hy5XgvPchhCRJ4qjG5PLM55nkGZqNiUEPE2NHfP97P33h81/+s59cZxTW2F63ayMSLYsIRTPqhQHuedoL/+av55ppesTmJauSGLJIUtOqRY3EloOOhMxyWdeyTaGFsmnDUnff7sXtuzq7QwNR09182/UXf/972667EVoyCoIneCzPQdBAGkg9IzACAwArjMJoFUQfflR1HYcTCghYyDsiSEAIMABEy2J9rVHOLP7os1986imnIUrWrVl/m1IpnsqwYe3U7t27jcAQvGhcS0MoXRyNNEa6893p/qDbHDni1BOf/id/Qscc48lEY+O5Vw06MTEpBeZnui/5gz/esWN+YbFc6vmJyXW1Wn3b1u2TY6OqqhDnjKqWPiMy1lrvfb/fn5hcW5S9L37+2/e855nPe/5j2AzdjiNdbdm2fzQf7PzS3HiribREMfG7b/rTD77xTTfddHs984lXCxUfxJAaVkOBEJjy4K1zbBMwvA9FkZedAcoQgZ011rJRUFEaX8Rk1MXzeVmOJfd+6lPPfuHzwcxj47kXa4hIlTwNZ0CuBNEC+Nwru9iQBrz5TR+4/rpbW41JCYVAhl2OIQSVMooS772qZoNiZHQ8y7uLCwv1dOLaq29/4+vf+kcvfd4FT39wEqVQ9poPfFFP67W169DrNh9wv2e+8XX/+ld/vbjj9kmRsUZc5nnIi6nxddOzM5aMNTaEMuW4DGUfOpN1j7nvfU588IPXnnDCAO6Sa66bDsHGMZR55YQSDQ9i+RCGPq77E9X7vaf/L7xRK/4HqTT6cEKAEhoAg6FbMphZDLNCFzqTwtsuuTT86BJz4tELu3ZF1pVFZqHe+xishS+9H6hHarv9/KSjj7n51lunNh41MxiEDeue8aY/w0knlABcIgGLi521U+NFD9N7Fv/w+a/YtX0+cmNZNjs5eYRLa7v3Ta/ZeMSg17EGEoIhVhVAiEVEvC+J1JpoYWFhfGz9e9714ampiXMecu+xKVYBDIQOSOF+mR5tjQ0GS5FjOzYRnYhn/fkbPvaWv/O53z0zIyKBwZGrtZqtybGR8bGk1RiZHI+TJK7VyaDIfae72Fnq+X6259bbpZcNuh3Js6gMzmvuQ65kJidHjt7yyFe+DKOjpbEsMMpMIBYlLO8/GKoAGEwK8cgG4Xvf/cn73vuhY48+ZWa6U0ua6peXKEUky/v1epOZVSiKeHZmqT3SGB9tdLpzI63187PFH7/iTT688UlPeZgaDAZ5UovzMldXM7W64XziEQ999MLMp//ub0ZFDdHi9sV2VMs9eoU0kRIEhRKiImCRwgkPOfucF78QZ5wOw/XayPmPfRz6RT8vALdfipfz+wqDlRHqqxV7mWpuy2FGpdGHGR4aQhjKnFEBGTYORJTl61tju+emv/axjzz29579y4t/OGaNNZGLeXp2pkam0WwU5EN3KUrSNK1fd9MtkxuPvGLn7sl7nPacN7wG69dA4VrtIIaDrp0aV4/F+exP3/CX226f1pB0e72pyQ2Z17m5BWY7tzDfbtVCWahQKUFVrTXEWhQ5wGNjY51Oz5oojtJ90/Nvf9t7Tz7l/WOTI8x3Xf61bFWvzCYqsn7UGpk45ZQXv+2tS7un282mc86lNarHiBNEDDagAOvACmYQIAJZWR0rgaxEv+vn5gezc4OFpbyXZyJ7+oMHPvmJZRS5Ro3ZLHX6qY2TyCgJIESsqiABFMoQimPHhBuv3/2Xb/mb8bG1iwv9sdE1i4sdq6QakqTmXFwWHRFfFCURJXGtXudBPx8MAnNcZJ5YG7XJv3rzO3qd7u/9wQXtxshSZymu17LSq6q1UU3lmHPPfXxn8evvfu/svr2nTK0tF7tOWWFiF1MIBcSaxBnrUO7cvgudPgYZajWEgP4SbFwbHYUIGFACMejAdx7r8DtnP6KEamzL4UhV13E4EYDFbBA50zBUdJaipAHmrLuQ9POPP/HpbvuuCGG+v9TjUEvSxsBjsTfWrHFZRqA0ibr9Ti8r0kZN42SuX5at0b2Je+xLX3zic5+N2HnigrhWa5aFWPC+vd0/e+Nbv/rli0ZH1ouPoU7goKwEJRESJSWFFbNyrS2gsD9kG1bFAeIidHr77nO/U9757r9eu7FJCbIihBDqaVSWyPr9JEnimHXYhgIZNqNABSKQFUtQZpCAVt2DQQLdf22//wEBAglQxXBeWCAQIUphGWxKy4EgyyV9omXmrCGisixDGeI4tTYeDubN+nj2M//g9q37ej1dmBscfdRJe/fuS+x+3WPo6jkovHLsK0ErKaEwnPcGM29/118+8SkPCoKszJNWXAgWut2Rekqz+xKEi/76rZd/7gubKErzIuSDlNl3+rXI2ShazHoUO27Ut87P2ImxyWOP2XKPe4xt2uyt4aTWWLdh7F73Qq0BZxG5Xr/vooiMNWx1xSNQCRg6BQIEWNhKpg8nqtq7wwsF+mXJElIpZWnx6iuvPu3B5yDvo9f/2otfOX/pT9rs8363zxpFNVeAsjIybBHEl3nWHxtp1uvNmfmlLjivtW4uske9/OVnPf/3C9YijevtESLbWeo2m42yiz/7s7//zrcv9j5dmB/EcRvCwDD1KUNPuMACwAbHurLstkqmjXEiyLKs1U66vZlBPvfoxz7oPR/8m5IFLCEEFUlMNKyrq9fjobDRikwrKSCkTAIID+fN7K9VkAPRohAAUdblhh5howSCQGTourS8YSYFC1h4mGOR4bVIzCwSmFlEGIbIFnlZ5GpN9PrX/vVF375kdqafpuOW68GzKjmSA1UUulrueJW59zKEMOjMHX/ikQvdXW9751/d/+wTOUE/h60hB/Ism0js4Lbb0j17vvb+f7zii18/bqTdoIBuF51+LYp7vig1pCOtrCzJOIrjjLkPGhB6KqWxvtXstEee8qI/OOPh52Kk1V1aSsfHBESwqmAdRs2y3xcQgKUqlD6sqGrvDi8IsEopGfSzS7/y9a9//J87V1yJ2MHRxmM39XzXhEHqs1FfNvLM+dI5k+eDwpeITdxMldAb9EshH6e3DXonPeyhZz3n2Rhtu5ExV2uCrRKSpAbgk//yuf/83JdnZgdZjihp+wAFD7MrVoOTYNVb8UYFK83aywUbaoYPFhaWCEywEuzU5IYkbv74R5d94P2fyPM8iW0tjRPjWMSoNuqxCGQowUMVBiuoIFOyEUSAgw5vFrBQy2KNOquO4VidgSM4wBE5kBN2Qk5MXJiocHHh4sJF3jhhY4icUqwaqUYQC4WSChFMKGVpsTvoZ1CXJtFn/+Or3/j6RfXaSBw1pETs0n6/kMCAhZoVL6v9L4yuVIbIwTeanFx343Vbs054w2vedNutc/0u0gS+gANi4woNyeQ4Tjju/Oc+Kzn6yO3FILPLJtqlBC/BQ0EcQqAyhIVBsXde9y3Gs/3GYpHM9+I987Xbtn/5bW//5rvfF264oZHYQXe+V/QKeE8SltvZmYVZ2AobrQT68KPS6MMM9SVUsdS76N/+Y+bKa771yU9i9y5oceTm9R6FlrktfN1LEjwjOIsQSmWQ4bTZWBwMdi8shUZ9j/eNE4998qtejlatP+irtVEcZ6XvDgrn+MLP/df73/dP1tSTuNnvlUwOw845gFRJheGNBqNqNDBKUAn4A0qtBmqajZEkqUVRkmdFr5fFUX10ZM0H3v+hK6+8KsvyhdkFa0yUuEG/CxlOOxmaQ0GJlSDECg7g5Ylgy0E06fIEx5VyZmIlVjIgA7CAS4hXCYqgEJAnKoc35rDslCoQsMIKWBFKEa9QdjaN4zRNEwKuufr2j37knyVQZylv1EfKUotcJsenInsXFYQrRXsr97S6XFmNMcbYQb/M+v6FL/ij2bl+nkMVg26WOtPv96meSJlFZ97zoc99VqcR7+jMDdh7+DIUUWLYSJ71EmfKQT/vLznFeLM+2mxYhpKy48RGszt2f+U///Of/vEfEUKj0YyjOPf5/t2j6jr5MKfS6MMMYkU+6N10a7Z995Hsbr7oop99+pOY29uuGRfxcF5UBE6UHAkbjRNu1KKsP1ha6oqJfa2+WE/mW+mTX/kyHLul74vayKgELMx1mC3UXP7Lbe9+94fnZgat+kQ5QLM2mg+EKVqtSsOmCQYAAZWgHFSChqUmw0gzCp7yLBhjfCjn5+f7/UFRlNnAf/KTn+z3+5MTI6HIESStJzBgBkNIZeXtyFAm8LAEQQ6ahKKACqkQhNRDAoln8Ry8EW+8kAc8DX9FYUBMREQKCQxhCA99ThhKrMzKzLYshBjWxBJw/fW3vu99H+h181qttXv3PqiJo5r33hijwf/qV+bgCHpZqWUw6NUatWZrdHEx27Fj5vVveEtQMENVCJLG0fzcNLfqMHLqEx55z8c8dN74pTDIOcBJknDEwr5fs6IoXI1MHYu+N50vFO0oOX5949RjRk8/5QFPf+oZj3oUj050stKr8ZkYbyJhG5brZ4bZIc8SSCrFPuyo6joOLySOHIpixy231LznbjeR7Eef+dd7bV47u+1Wi6ACKIegUC+szJpGRnwpReFcM2qPbJ2Z9oyzn3bBhvMeOBh0a5PrYawhaH8QGSz0ije+4S3z84OR9prF2a4JSciknjZ6/YFzFhAlAlggShAaVgrLcip2WLSmdiUJoN1uN06oVoviNBpkCzfffMMDz733P77/fcbK/PxiGkdwrKX6srSR0bBSP7ZyfyAO1AN5heEiGGMYtS/npXV5UWU4CJyw4gs49Cdd8VgihgyrhcEAGDBQZgtW5KUPnlV1aSm/+uprL7rou/2u1pLRycnJoigajbYvMbtvOk3jlUPGndyb+cCp2N+kQ+gV/VajvtTpJPXRge9defn1b3v7h1//Zy9oNGqlFBJCrdFSCNVqgDz6+c+dufHqzpVXmZhbLip9aY0hRfBeFElsC1A+yNdsOep+T3js+HkPwNgEmlPwglYDo+3Ml/1enqZ1a6yG1YXQAhLQ8IxVXqaHGVUcfZih4hG7kWaDvS/mZsd8aXft/sln/+OWn/7UFCUZKy4aeOkXRSjLkA+0zKTMYuI0qQ0U3SjecNYZ5z/3WTAct1tg+FBkRT421prZs3jhv1/4nW//qFmfCgWFXJtxw/ez1EbD6+VAPLwJcVg1/2plzwg6zEdbwCZxygbWaV4uzC/sqjfx5Kc++p8+/N5+r9ftdtvtpnF2qbPk0shGJu91SYVFWGVY4bFqq6I0jJs9wRPEiljxJpQmBBuC8d4Gb4M3PhhfOl84X7B4Vk9SkgQjwUGcipXlXImCl9cPlxMniAyzwkU0MpKsWbv2UY9+9Jq1a42zSZLkeS6+dIasQ62erAqWVx84Ldd4HLhfhqN4ttNL6mNLS74RT0gRf+bTX7zwP39QCAalFF7juNHpFrDJQA0ff/LZT34mrd+cJSOlbXZzY6hBiPqDwlgmcuw50riBesu2UZtAYwSRxVgL9RTO9vv9RlpzbFVWXGsJw+JCHq4iyB1ftIpDn6qu47fM/4mlrwChuzjbSBLcsPUfnvXcdQs9zO9rTcXbds5s2jzRnV+aikYol6zbC5Izq3XIc2010rwnA1u7nZCcetKjXv3yjeecjWYDNur3syRtZiXiiC+7+NpnPu0FUTyxOJeP1ceaUSvrlnHanFtcihu1UkolIQgjkMqw/A6krALwyppetCxSFNj40nfTui529kyuaTzzWU96/gt+tzFKvoSNkWUFIHEch1ASxLCBKpQBKC2XXoCEFVC/4qEhEEBlOesRBEM115XWRdmfrAZAYANiGFKwMJHhFRNlBlixPHxg0M9qcUSWe0tddlGaREudstV01107885/+MB3vvXDOGr7TEWo2RjtdrvOOQDLwY3+2peOFJDcD0bHJ+amu87GztBSf3ZyQ12j3he++ql2y7Zq0aDXia0REYqMybo0s/efXvGa7LIbpnKhxcXxmvP50qDIbCMyNom43u/7pULMusn0HsfwhqljTr/nkcceW1+3HuMTcAZpClEkyYqbEiuWxz+yKpRgbBWZHU5UtXe/XYYzZBnMQ00igJQFIMl9MOwsU7+Xp2k87OLrd+d6g7nJyUnsmP3RB//5lx/9t+PTeM/t16c10rpRobiMjRgiEi2Dz4KEej1KXNqZzmhi/fXMpz/nmee+5Q0SmSzLrIl6vUGrPdLzGPTxhPOeOru3y6YJdaxMalhYiKGsLMOSO9Y7Vp71B73JiSmAp6dnnY1Hx9reZ/3BYtB+nIhS/8STN7/kj57/kPPv68tAlnildXm4EDi0mAAgAmOteGVrAqBSap47S/2yV0sciqCdLnmFiJ+Z27dr18LcfL/b6yzO97s9hjRq9ZFWu9ZoNtdMja9ba0bHERkAiIxGUcEoFOwiZxMmq8PoPAxrpO8sWAxFlikTfe2r33/H2993y807RtqTEjh47i0NNh2xZWFhKcuKVrM9GORxHPd6vXR5RPfBUSoFIQ8wNIZGLATOwV0xnS3HjH/hS/8CLeOag0peFFFsSTP4Hm65/V1Pfsnang/T02OxZr09UYpcxVpnfRK8KcFlYvuJ78JPdzsTE5NWjcbx2nuedq/HP3rTeeeg3YZxQARyEMpVvYDIGFAcVenNw4pq5uyhgQw7dGnZ71cAWGsNEQSxdUSAqvb6P/zGN3pL0xc865mI7NlPfsKt3/z+bdded/zGTb2lvWUIqmAZGi2zwBBbJiG2RSCJa7l17S1bznrEw4uslDQxzkB0dHRkdqbXnKy/7e8/0espNGY1w5SFYtmOZyUDK6wrVQLLvRsMgClmjjpLPWvt2Hir051f6kyvWTuSFcViZ/rZv/OkP3n1y8Ynagvz82ktjl0ist+SdOgnIQDrylqWgIZx+XC5D1les4rpfTtuvOnyH1161cWX7rr5tqLTd1DLRpf7XMIwKnbEQm4x+HR0vD7SWrNp473OffAZD3pAsmFdHCWWWdkRwUsQERJ1SmBe6Uvff6TDl4ONDVDzxCedc9/73uttf/+uz/zbhY366MTUWobcfMsNGzZsjKJG8OGII4644YYbJiamyjL/b19mUoZYkDMSze5e/Py/f+2CJz8KQL+b15opCBKYY4dm/RHPecY33vWRI0ZGy2yRorgMAxuRZWJWocAa4rKEFFbLM4898fbbt9lgbcte97Vv/fIXvzhv764HPfOZaDRhGWQAwxRZw0RV+8phSRVH/5YRBFKQGGBY2oZhLFYUhTMRBZAxMIB4LHX+7k9eteO6q9/7Tx/EEeuQFd96/V9M//gntbm9vZm50ZEIQUzhWI0aK0aVSlFviD05tCd2cvzwP3nFlgsuwEQ7s9YFGGUIigI/veLGP3zJq5fmfWKaFCwvp5Vx8EXx6giR999bE83NzdUb6Zo141tvv6Uoe5uOXLNj101bjlr/6te+9HGPfzBblGVwiQGQ55mLE8VKDZ0CNJylIipkjClLBRFbRihNUcjc9L+/5x9uufwXt119LWfleNKoMft+FvLMGWsMG2OMJVKoD+JDIerjel+DhwZnqdUIteTIU05+1NOffuqDHyQuUuMU7EEMROBhgYWSDNMsKxoNAES2yAPURBFm9oWPfuQTn/qX/5idWWrXpxr10d279jYa7Vramp1ZWLt2/dLSEg1XPJdHie//TMmqONoaMUAA5+CejbIoCZ+78F8m17ZrdVuEMopdUXQNZ6bXx67uJ//gjwdXXbcxcd2FnZYKZ4mZRa0EWDCYA/lSqd/x7dZ4kjQQR2Zi/Cc7bs02TL3+n95TO/nE4fFCI/KEYTHJULQrpT6MqHpYfuuszM448E05XH9yzrEhsgYEDGuiF5Z2XX19a6G47MKvwAsSe+r972XH6tfMzI20I2CYsNWAAMCAmK01UbeX+yjeC11/79O2POp8OMBZs6JFWTcr+v4jH/xYb6FvOAK7g3frv19fEpEkSZIk2rb9tjjB1NrmwtLu+9zvtH94518/5NyzOQIYxpil+U4xKOMkWX2MSqIHiuqWJ62oCFSgCqL+3NwlX/7K/FXXTBThyDgezXMzPdfKis315phiREIz+HpZ1HzRkLIJ3yYd8X6DjTan9TVkmv2sd9u2Gy/56TWX/AR5AR9YYYyz1hl2SgbyK6uHfVmqBtFybq47Omb+5NXPe9WrX7Jh4wRxsXvPtiM3byCSXbt2xElUlnkI4b89T8vVJQBAUFcWds/u+bf//XsjZyWAyEgIzJx5Qb2Jyfb5v/OU2ZCVzmYCF9VZDQIF9UpCBAs4JadmbHydbU7c1u39ct++y7beNjcoKdj//NSFyBQlpKTgFQEcVl2+VBxWVLmO3zKrqsr2214AgGXG/omFJAhSzM7ZxZ7rDq798c/O+v1noBFPbVjjrR453kTwy/W/Zvi3cEwWJDBqk56xe40843efiXYdaay+MCZisB+EpJ786ye+eMWlV9jg0ma70yui/W+J5bTGr5FpAVgIrdHW3MLupGGimBc7ex587n1e/oo/OPa4zc22ybMQxNfqcXOkGTy8B5nlijk+WDKIhpd0y1aapAIvi9PT/X37RvNBzUQNQxGTxkwSUHS1HKgBGzO89BiOatEA8ZnL00JDkQ9q42Pramk3L/bdcivIMBs1NhCJQgUsgIodHi6teHmuPA5SRlFCzEnq8oHGCT3jmY9bu3bN6171pqk17Vtuvfboo04EYJj7g561DnfkQNpk+DqD/KoRwZFhqiXj3/zGD770he9d8LQHq3CQzEVOXRsQcf31j3jIcV+4z+5LL2ulTRc5Lk0IhbIHFOI1KAI8qFunI8886exHPKR92klw7torfvnDH1269fZt6OZUKsWAJSLA8PIIlkqnDzcqjf5tsr8SiglKB7WEBRUt1Qxd2ayBF58XjSgZobDv2hu/9rfvfMQFjzat9s7rbzwxilve9rKeEnlDIPasDLIBAZRMTmzN+ied97D6mfdAYoMlDUHFcxQF4oWZ7J8/+WkNRnL1sai/0woF6V1VL+xv/hbiMMiWXAQXST+bO/9hD3jzX7z+iM2tPAtZlidJDDJl4VUpioxfFW7KsH9wOSW9rNHMLOIBQBUi83v31q0ZtXECJ0U+6OfwYKBgWAdlUg7LNdLDrREilYhDzdk0bnjSTq8b1RomBBCg6kPwYoSIAWfAy66qd0EcxyH4QXeQJvU4tdmgLEt5xKPug/Dmt/7VO9asHbv+hl8ee8wpnaVclayL/YHuljsV5x24WpWVjnkQjIY8rUd/+zfvPPe8c2ptclEECMH0g48adVZ92HOf8d6fXTYSNcosc7AiJVRgAGNU1QZ2xKZeP/upF+DME5AyRkZP2jS55b5nbNuxBw6wpKzK3oNoaNPBlTPp4UeV6/jto1jpsxgGmMOPMVlENhgSJmhAKEwtaY+PcVlOib3yc1/98t+8e8e/fG4qp2am1MlqJWLPlqCW1bCCEcgLD5I0PvKIh/7OsxFZ1OqlwjrHQAjoZeVH/uVTN27dwZREJi06RcoJL/c637ll7mCWHSpC6fvgfGQ0nZnb+ZjHPvQDH/ybjZtaZSlxaozVohwML+EB5Lm3dmjKIftlf6XfhYcRHjMxMw+n8/lycW5e1IdQ5lKUWlLEaTNOmrFNOFgqWQeig6B90YFqzigtKIJYX4RcpFQpGknUTJO52WmEQtQHFSEQgXm1TefBMq0M5eGwlUazWZTZwvy8tUhrkQge8Zj7vvu9f1drmI1HrNm7b5f3RZJEd53roLDsz6f7Q+kD3iZ55lvN8W4nzOzr/dOH/rnIlYhDUAErWbFxDm3d+6zjHnT/AZGyC0JlUC3VCJSJDBtCDdq9fevHX/+6Wy78LIo+ejPqivSkI44/7yw/7ooRDi0OCUobCir7kEyrPsPDj0qjDxX2f3hIwUpZXgRFUDXM8AKieP2adcdstrVkbnp6HceDG2//9r/8+xi5pfk5S9YoD/uemUkMFQYD4o6h27Le0Q94QOO0U9BoZnkex2nwYqIo8zo9t/ixT3w6TUfLQupJm9VENv5van737+LQ4o6KWt3kxcL2HTc993ef/urXvmxY6Az4PO8bQ9baosyM5Sg23hdYWRFdPuSDkx0KJgIRgQSiCDLo99I4McaEUIYgy7NT1eeleFFRkAEve3VAgKDIRUst1QZy6stCigGT+mwACJjYWWuIGaQIHvprWruVozjOBoOyzOuN1BjDLEVZgnD8CZvf8pdvbI/UggyS1Ip6H8o7Vd3Jwa8qlrMMKzIdx+nSUs/ZWqM+/qUvfHNmugNFnoc8Lx3HuWg8OoHYnPOUJy8xcb1BUWyMIzJERgilig9F8Plku869hc9+4L1/86xnffU971u8+lrs2YMsswgxC0OUVZ2IEzJCRL8mdVVxaFJp9G+T4QRqVlIEIKgOh08RBNZGIYTIWAiQlTCu21l4xhv/JGtabbgIYgb9VCWNTHNkNGo3FvMsipKsPzDGCNMiym7T9UZbC830MS96PqJUvLBxpQ8eJs+FmP78z/5SxZJE0CjLfBKlPve/SqGH82QBWGeKIrOWjVU2odPb1xpxj3jUg1/3hpdtOKJWlCUbqGoURUSGiOI4VvUqUqslIsIr7qN37qX2+/MFRGACdH52zhdlKMUSW+ZQCoOKPFi7bPJkBKwwsjysjwEQAlBKKMsCEMMIeTY3Mw1RhloCDf9KwMPhUivrloqhPdPwBrZGRKI4bjTbxkZKAFMcm6w/qLXpoY+69+vf+Mo1a8dAZZ73AR9CaS2XZVmr1ZrN+vz8nDG0El8fnAZenipA9VqzLJQp2b1r5s1//lYAKlyLY++95TiAMTY+dc9Tt9zrjJmQz3Y6zdaICgdPi0uZixOOLTjMzO3bMFLfIDDX33LdRz/zT8980dse9LgPPPbZP3nfJ7B7jrtdCwkaLDkDsz+krziMqF6y3zIrk04P/iEAILVWizLMzn/wLW+duezyxtopGDn/GReUDdvLuygKn/VDmedFsXd2ztq4zItE2WS+7HRYUURue79zzpOeFOo1BM9pjU2kZIoAE/HFP/7lzm0z7CNLsQpbEyvY/6prYSVVYmZrbZZlIZR50WcTfBi02m7DEWNv+4e/TGsIAYCUpbfW3lENDph7/OpTQaQHB9fiPS9/k60sYaoygZQBGF0e421WyfTBjs5iBCzCslxToaq8cn5pv0G/8p0+CLzqfvVpQFJzIPhCzjzrtNe89uXd3nySGmPBRlWDMTQ9s29pqTs1tWY4HXjVBlcpJMmwGoRgfamEeMf26R987+p6LZICsbXqNSvCoPAYH7vnwx6ys+jVJ6f2zS8GduTiJEn6/X6Spmpk47qRXbfvWNusr4/SM9YesbanR/bNxL7+9z/4L+964R/rngVbhIijbt5lUJmXv+b8VxyaVBp9CLF6vl/W6yNAlrrZnplLP//lS774VczMoBad9sRHTpxy1KLvJ81k/YapyNqiKJrNZhzHUpQ14Ubm417RCs7AuFbzvMc9LsRxXgrAYCts1bg84POf++ru26cjH9lgSS0bF4jLoS/SXe6bSAjBGFOWubFEHETzXn8uTvHu9/z96JgZ9kinaWytXS4WvmPQdtcyTcp05/COCJB80CcVA6X9Vfw09EVaLpoz+8ddDz04Vt+wbM7HEigIRKHLSr/iwLSqmGY/enB+YrUH6fCB0WLQsRFPra8/8YKH/P7znlX4JUVunSp8EG+GbqQmyrKS6C5PpgIhSKEIzjkRWJNu37bnws99rRhAA4zCkIXnMgBpbdODzx458ehF+NIajZ0H1dM6C+fiyZo0SQxgiev1+rYd21nRm56zs72pDs395LrLPv/t/rbpCKYRt3wRanFclXUcdlQafUhwwENy5SftVo0IJvi9t940SfanX/n6Tz//RfR6gD/znAe4ZiQmbNu5a/fsbIkwKApYA+sscUSmCeeIM9XTH/CAaPMmk6YmiiWI914ELsJVv9z+g4t+rBnIWyoNk/OigRiGhfa3ES7v2vB/Vcl78d4TUZpG1ukgWxyfaPzJq1+6+ahJYvgA7wNb5Hn+K7RpOYRlHaY77rQSyatmbbFCNe8PeKXkY6j7AiVaFtnVvs28Emvvf2IlDD1JNYj4gGH9tahCWVVJlFRJ7yjB/y0ibBRaagAZvO4Nzzvnwff1oVf6novgfdZuN40xi4sdCUwY1uQNX93V1R1gBhFcZAC2NsoG8pNLr/jFz2+NIhQFIgNrbZqm4j3GRx/6rGfsGPTM6GhwLqgkxtXiZKnTJ+sW5xYY2LZjz67ZmY6B27R2KeFt3QVO08585/tf/1bD1hf3zmkISWQqL+nDkUqjf8sc3EYk++d7+KKEzwh+cXrfmON4fv6mb323c/FlYL7Hvc/cNTvXKfsjU+2JNe1aq9nxeWYJjSS3EMuulnZD0Y/N2Y99JGzELrX1miqVZSBAPT7/2S/v3TE3UpvgYMWD2eZl6UnhzKpPMa+65B+Ovo7KsjSGjNXBoKMoHv+ERz7xgod2+z3jkBWDODWDQZnU4v1OcIr9c1VYVy72l0M5OkimVUEwwysJhQyNSnv9jmpgUgwT2QQRWenoW/nDVSUiSsuNGn75i4bAJNAQAoa5/lW5kFV5ldWG/Vi1sHkXefPgvY3jXq8jWvhQmhivfd0rTr3HMb3+nCJ3EQPS6w18qbW0SbCr8xurTiyYWXW4CBFUEUe1menFz//n11RRegVAEOciLwHWHv/Q80dPPH4ffFdDUkvFBy29tVbEJq450hyJk3pr46aXfvi9T/vE+17+lU899b1vmTz/rOjodTum9+666ZbR0bFIiAL63ez/1Zuy4lCi0uhDAGVSZj1gMwSIqIc1qLlWO13cu3tjvZ7ddNu3P/ovN37+q0ibZ55wbN0aX+b9fr8sSyLypBS7jCQzKrFdDMX48UeNnn4KVLwEgJiZlWPGtlsWfvCdH0VcS5OWillpP1cQKd3pUnjZbJMBTuIaADYoijxIceKJx/3O7z+n081GRuqLi716PZUVFcrzoRYcXCOsDNhhsmGo2avDOl1JN69+7n6/v2zccUDKdX/R3Op5VfvVVJkCkRACYblPRVVEoMqKYZnDijXer0qO7y86lIPT6ALAuNjnZa2eGstxzXUWB0cdN/Kyl//hhiMmur15Y5d3NE3r0EiHLfV3mYin5d71EMoQfBynKvY7F/3whhvnjCMA3heAuDhCrY6RkXOe/tRphF6Z12pJkQ16vUG7NR6UZxY76uKZpe78IOv2+xhv4aSj1p9z1iNf+oK/+/RH//YD79aIEDyFIvii0Ux+xSFXHLpUGv1bhlbWxFjBKvuVixnQHFSs2bSm1oxq0JGs6Fxx7Xfe87HP/+Fr8xu36b75KC+5KBtRklinpRYhFPA9FD2rg5q99xMeiXYdsYUxEBGRxEUo8PUvfH1+z1y7OZrnoVRSY0tREzkAkIOlZJVAAwghMLMxpvT5li1HPuvZTzv6mLGR0WTP3n2NRt0YzM8vRpGbm1tQXe1TvNy1AQwv+nl1vfX+igdVVSYhhGFrIxNU8zxXJSKjAiJDy/NTWBT7A/NhjcaBkgywDjWaoUxKKxotClEWZYERcFBSMOigIPqgF0YOJKcPFlnvPVkLDSBpNNP5ud6Dzz/l2c95Wq3uQihAErnYcFzkQcKB1MvyFpRXCsFtCIEZoDC8MmCKpvfN/+cXvlIqRAASiCdrYB1gTjv/vCNOP5Vil2d9CaUzHIKYKLWjo9Koj27YsGbDhkt/eMnClddg3zQ2rse6UaxpNu554oZ7nQwn7OAiqsyjD0cqjf4toyRKfrkeCyAMhzyJdc57gXGNY4857YH339td6M53eGmwMartu+rao8fWpp4nWiP1OCkH/V63E6TIisyzHcAsRVHRHj32/PNyUtTrLk1V4EsYRmdev/bFbxgkadLsdPsCJWtKKaPIErz4krDKfXS5DloAUQ29foeNOseAP+XU4y940kMHA+RFmJiYyIvMBxkZac/MTo+NjawYdf4qlhu+7xzJruRElt+WofTLbvw4UNeBofISsDKIZXU0vRwnL/eaC/bH0avPuSrunA0/sI3V+3nHC4s8L5J6K8/yYXkfObRHaoXH7zz3qUcevd6Hfn+wWPrCh2I4a3x/WTRDdRiVKw8va4jIWstsrHUSiGBdVP/6Ny/qD8AMZyKAg6cyC6i3sGbqHg89V0faOzudwpr6xPju+UVu1NccvfnabbdSEt122y3f+sIXX/F7z3vN81/wwZe89KpvfRMgwIMAKXKfgxDUVzJ92FFp9G8BXUEQApUllSUXgbxAVBVB4AVeSRlRA5k+8cV/NHHmmVmSjK1Z2+/Nj7eS+b3TiUl8TiBX+kFaY4+MY6o1R5bEzrjaWU94MmrteHQShEG3GzzFUVT08KUvfHvnzllR2x8UNo5MHJVS2ojKvGekTIwyDh5uTQHwoCJpWBNLHrqFdtZtGnvdn70ybiJO4SLDjCSJmEGMyclxERFZaXpefoPJfvd9IciqFcUVV9LlZbyIiAMFLygB57JBL7amzHNnhoV5EBIhUV5+MNzJlemFwhBBAMQprICCQMWySeN4MD0NBkXWQ7xjsewliA8HHezq2F8Pvq18UlyUiMBFDWMTYyNRIas2Rmscf/v2N6dNCxNsDLJKDjAH6kyAoa/h8nJlXno2Li/EmkTFIFjWiCS6+eYdP/rhz/MBnEl8H97bENcyNksi9332s0dOPzVsWNdvNUKrsfn0U7cvzV581c8a48090zt8v5eW+ZFJ2r/u5hu/9u1/etUbX3bOQ7/x9+8KW3eADDMHUKnL0xDu6lZxiFJp9G+Z/e1oAg60vzCBpZ///AeXFNv3oTmKIzac/bQLeo3atsXZohxoWcCHLC96/VxVo8gxy3C0dmdQmMbIHHjdyScjbSpbKUMURaoUckDx88uuIsRQC/CK89xytzcj8EEf15XHpCDN8l63N19vRKXvv+gPfrfecADojpnWX2/AtGznPzxAWlXntspYCsv98ABURQ4kf1TvIlv+q6BV1dPDZ9+f6RYSJYgh+m82xwffDrD/NVreMnvAg7Fu/dizf/epPvSJfVEOBlkH5Ff7lK7snAKw1jJbEXgvvlQJymSsjUOJi/7rB7FFyGDTJI4TBWVe4vZoWRYXPP/5ebMxeeyxl99821zem+91kkZ9qddRDWnEkYRakJGgG126lqJ4oXvhP370tb//ot71tziKDHCH64mKw4JKo3+7sIExcKRO4QJsydYbCzAPwr9/4CP/+p5/xPQsrBxx9r1Of9z5mBqNm/VGvZbW6y6KV9wnyOeIjGWy/aLgJE7a7RNPPx3OIYj33rgYoBCwa1f/4osvxh2X5u64S6seHJAn1RBFXPrBmWedesGTHt5sOpFfv53/76yUV6uI3Fmaf71YL/t+3OV+3Wlv/zuZvkuEVs0dp2XTb4HqSDt+zrOfdsLxW8qiS1pYDua/a9u54/4oWlH9+9+6aHYaRTn8qgEpWMQo3OhE++ijz374Q6+99Zb6ePOaa653lhdmFhkIvgylRxlicN3Fi9P/D3vvHSfJVV2Pn3vfe1XVaeLORmWhiBBIJBNMzskmYwQGDCZnMNgYgwGTTcZgcjLBJGPAZAMiByGhiFBchc27kzpUeO/d+/ujumdnV7tCK4RX+/3N+dSntqd3prqquvvUrXvPPXcHeW/LuNo13FzvbS99ZXXJlSh80664/B96WOHogwlSGGEjzDLsbVZQYIAEZd7ftOWsr33jqh/9FAJMT9zrkQ/pGj+/uNCdX6iCV8PMIBhrnbWc2YQAk6QLvjrh9Ntgw3qEEKNam4AMlJMM//u9H2zduo2Ir5tm3RPL4sdh6leso85Y0yV42t8+yVgYB5H4R8Vle/at7IMuVWsPvBFj69KDG/+ifyykztYMM+m7KVh9KImxdnXrSU9+fJQ8SZE1TD0DbZ8IIYhI3bppra2T1+LDeHts59adP/7hj51DqEREWaTpUhZFELCZ3HDYtbO7KE0KgbWWHZLEQYlBnbSBKFVVja1alSdUtRNqJJdfecWWK69e3LID7Lpzi/9H52kFNx1WOPogo76Xr407jMCoCgkowBdrrGnPzv/qc1/EpZfhqquwZtUJJx2XWGNAYDXOKlNZlr6sSEECX0XTbHYl/Nl97w1noTDEbBMErTvqvvKV/242xlSJYEbNKXvsy25SHi5m6RPifd4fzN3r3ne5931u3e31RUBExtwIq8vdsfnyklzN0cPmlBEZxxjr7pU6fU9Ey/tZ/iCUhkVFjGoAGGn+SGsHkhsTVtLut2z0ABBfkSIfVA99yD1ue/qtBv15lYIPMI4GGF4n25P//V9ftSmqEIjIKCyD2YV+AWP7YG21dnS74+NZFUqTJrNdb1ySUMqVFvP9xvj4C9/wmtd98ZNv++6XH/T8px3/538WU3fO+ReAuT05cSOOdwUHFyv+0QcTuzs46k4+QFFnllVDAV8cNz1z9ne+uTi/+TZ/fsfb3vXOR62ekVazGY0RApEGDSEYjs65svDeMlpJe2risNufDoqwTXIpwDEqM1184fZzf3vhmtVHDXp/cGhIjWUVP8gg705Muac89a/YYny85X3lnKF6qOCBHG99XdidpFWu1Xi7U9JLvCkSYwTJiKPBPHxwI6h1r+j7RpHzcixLSasAMMYAsdNJfIWnP+MpP/3JM0U7KgrsW5JsrdVRMgdKtbkHsV1cWJycbP/mV2fv2JxPrW0Qw3Bdr1DbGfN5b+0tjhuA2taqagxaacUMFYoxUiUSkToXDdl1qzEzdaf1j7nTIx7+mQ9+Mpkcl7JQ2+Tdl60VHBpYiaMPJpbmRQkJSIyIjWJVQIFWTUSL3vyue590Wrpx+w/f8+GPvfQfNv/2goXN23q9xcFgMBgMRCR1LskazXYnRiXrKqITTjsNE+OwBi6B1m56IODrX/+GMc5XsdYWDPdgn97QwN6FMorG6oMfcr9TbrUhBNgExJH3MVH7hmO3e/OwpXupMUWl9laCxlE+es/GQv0DeeR9JqOHtcfl0myiA6dqRq3F3t08WSepao4WJMhLf7d73OZOd72dUhiK7Q5s+1SVUuTh+/97JtfszIwoiBHGudVrp9Yd5jpj3jhlYsZEs7VqclxVKbEmS9qtZnfX3Gv/4ZVIMpmbw2SrbNvH/+NLbn//u0tC/d7CCkEfcljh6IMMHeoraikvoDCqUIQif9bLXjwbq99feaUu9k6Z2WC3L2w5/8K2cxMTE+3xMeecYwOgKIqFxUVi57JWTnTkKSdDIpIMzFLFIq8IDOBHP/rJzKq13e4AanR5i9717RyPfKJDZyz76yc9rtf3NkFRdK1lkPiquhGHvGcQvZ//Akbd20s//eES3zIhx75wUyayl8kHa9ZWhTGzO3aNTbnK468e/xjRMknNXhZNS9hnPjqE0GpP9POq0xr/8Q9/6isUhQcDkLz0YMAYt3ZdHrUzOVkG32gmJNqdWxANxtBCr68kk62Wy8vn3e3eZ3/nBxjkKTMsN2cmqqoca7duujOwgv8jrHD0QQCNACJlVmKQKX2AMYAJfQ8P2x7bcI97POjZT8eRa+3U5NzOudVpu1FIYmxZlmVZqEYCnLFJ2lBjyyg5dFc+uOXd74GxcbgEArDJWg0RXPL7LVdecXWvVzQbY/1+vmd+YlnOVAmkIYR6bFWSWpGgGlXDXe/6Z0cfsz7LbAh5lqUiMYRg7YHnyvZhBDraDxHUfp61Rx1IQyQaujAzo05PM1+fhkwVDKp9l5ansLvdbh0112OjRG5oUnvv7Y+MPoZuJKMHZA1IJlaNFaXvTOFe97vb5MxYhK/PZH10dd9K3a5ZP67THUu3C0xGyRAnVSm//NlZg65XYQBILJwBAV4gfL8HPbRXlJy4fl75kDdSMMW86CVtjhQkH3R8bF87+8kXv+pND370xu/9BNu2o993HKJUIVb1aVx+2lc0eTdnrHD0QYYCRQxeIhFBgSCu0Qm9AhFoNP/s8Y+421Med0l/tp/AJbaZONZhQnFUsyIBC3E0phAZX78ejRQiShwETFYDjMFFF/0uBGGyzCZxDWa7j7deCYC11jpjjAmhKsvcOmMsRKs73fl2STpsXPbeY3h3/8cc+XWC6D9qcwf++n9UWC3DxpzdP0uInkidMzEiy3DHP7utqFeNy+ucNWXv76WFIFAyluCqUi+9ZKO1ZlhHrccTJAmSdP3hhxc+KNU3SMpMCXNqDTOUpJIqFPkE2xPGZ/LLrn3HS/7+Q699A/p9p2SYrBnG7N77pVj+j0tbreBPi5X35mCiVvJmxmZsEutgAcO62P3Kp78we/m1WOzjyHW3esxD7vv8J/vDpy7cfBU5G2jk/CCjHj0gAuSSfoxHnXIymi0RQp0nZfKVMOHHP/p58CCyKiZNG8Zcd471EMYQM5ghEiqfMwuxHxtv3Ps+d0+SoZYjhMhs6E/3xaYlS6ObLDtBuhQDK+9PQH0DsNT4M1pCvauqGkIQiHHwMaZNPPAh9zV2WPNcsuurPbiv5/IQY7Q2McaWZfj5z35tDUJEVLCzsW7eTNxxJ56Si0YBmAPgRWMloYjRixKZRmrHWpKZS6+5igaV3zZ31te/95YXvhyLg2JuASHWO1P7XC/F8jfydKzgT48Vjj7IoJFJRfSlCGCIVC/89VkfeONb5erNiAHrZ+788AfMWz91+BrKrFLdOzJ0nCBRgUZiOFeCTjz9NDin1ghQj7Yi4qLA2Wf/lqgepSrWJkPt3b4QQlCNCjGWidSHQYj5Kbc6/sijW8QwxoAMkWG2egPlITfwPPyJRc97pbZvLPZqwgzA7qqgMSbGCAgbZYPb3e4209NTAJZz9Gj2+X45sXYxNDYLkc769W/jyPDDMHmJAEC85thjk0bTKzNbo3Agy+SsNcYw4IMMfNlZu7ox0ZnYsIaS1DD/6kc//dKnPpNlLZVhy2WdbMFKruNmjxWOPshQleCrKCFJLJEiVEjsVJZt+fVv3/uyV+CC3+PaTei0H/n4R1cU5ot+XGaXPAIJAdZRmh51y5Ng2TgXohgCIqzF1Vdt2bJ5mzWJKvkq7BnE7S3gLYpBFC8SnDNJykXZCzG/+z3uKoIYQQwoW5NAuU4T642AQBWo9SV12vimI+jlXL838agOHUpHlK0HvvsjQfSShnykklZy1kmIwfvEWY2YnnLHHXds3Ya+lBbX6xV3syL6QEJsEzLp5VdevbA4nAYwPNUEqGB6cu2RRwZBQs56YwIpOBITGahlIShvnp974yc//Dev+fv7Pe0JyfoZM9b+wmf+s7dpGy3LdXjvl+qWN9kbsIKbGiscfZBBpMYxGQLx0K6j7DcsHdZq6cZNn/zH12395vcxO3/chg3NdkMNIoaijCHd1fZvSpHRmppqrlun1sLYYWAUQYyzzjorxmiMU4UqScRoFup190ajeFURCWzEGAZF63DHPzut1wsiQzG0tS7GG02ryxhqmJK+6RPRy4Xne6FmZ6Ebl49e+r6MBhUMp4sxkanrkfWZrwco3va0W9eEeEOS0TUcsYSoRGrM7ELv6qu21CklghJppbX0x5xy+ulLpnoiKKMU0YeoDMqsa6XN4086ESceN3b8UX/x4me/4z8+9oin/nU60fmfb30To6h5adziSq7jZo4Vjj6YUABEQlTEOCjLGAKMAfNRRxwxe+XV0wNfnHXhF17xxk8+9fnf/sR/XnvFtUmjOQqfZS/+iYLJVdPIUk0SBTMzIhAB4Cc/+VmSpCKAsjFOla7nK2kMG0Ns6g4S32iknU7z8MPXWcfMVJWxflnvo7X2AMW/Nwx/im3eZLkOkDKJrQNnUqqHMZIyw8RSUtdMTBJLsQQNOPXU29RCjrq++gc5moBGkmqIEiFqSi+XXH5FfasTKm+MEYKmFqE67Q53VFBUCYzoQJkxqaWEVVWKKnR711x8yffe829YuwYJaE3nsS965ps+8j433a4k1LH8Uj66nlR548/ICv7EWLnHOciI4IBoQC7NjAhigNGTTztl4y2Pmz3/4lsetk5icfXGTVu37Thm9eru7FxmhnfoBInMBFElJVOS6UxOAkypqzSyMVIplDjidxdd4WyrKmCIrbVEKhLM7uh1j+t0kqTGMEAxRu/92ESr1TJZCttgEAZ9JA7ECF6REdSS1IK9A1hjWLKrb+CvIw5RhurIHfRGhthKJKog3T04saZ+VaGR6ZJa6HAHbviej0YSJFh6ZkS5xUBbYwxClYdGmwk4fMMRREZE6mTCEkfvP92hWZKVZRmEoxCp2bJ5u0aAEX3IkiRAiW0pctjxx5fGFhKIiRGjFyISkGHrnGXD+WDwjc9/IU/tQ5/7LExOleVg3TGH33N63DgDZRWJu6Npu8/9WX64KziIWOHog4ClSCoSShEALePq2+P+wmxrvL3+9qfaY9ZWOzZfPbt9XaOxvjG+OL8zFaNsiEEqLJEJwjGSElIyycCmp9z17piYpCQtY5WahktJgUt/312ck35X06QBsDFUVbmxGLXc7X0jNbtrfmxsLE3TIs/bnY7EeMRhJ+3agVYbIBBMf6H+xaQ3pwAkQohqReANWdeCPa4F0EvPMzZv3XjiyUeVRYiVZ2czlyBoGeJ1JXrXDwMWlQiAwUyAGBXLmJ2dXdtpQ30QTh127lxYs2py25ZekmRQZuW99xP72P/rQVUFwOT9Ybq/140AskbibJr3vXOmTgE3Go19EOKy1Ew/zxtZe1DELB0zWfGdb37nWX/z8LSBNG2EGFRRKaVTq2Ds+JFH5puvYZSZmCwI1wbbHEuKoo581VT51We/fNGvzn7xO/81PeowZGWznQ2KvNMc91V0aRrLii0TcVmVLk2WEvgEg9EPf9B/awV/aqxw9MEEA5YNVH0ZJO9bQ61WB/kACT359a+86Mtf+e4HPr64afOt1m4YSxvFQi/NbOS6bCUAhBCJHaDEOVE6OQnrAAhUEIMQK8/P9WMwTBmTjWF4Ww7sWVxbZqKxfv0Rg8HAmNRa313sp5n7xc/OedhDnrTY3Vk3HA7TLLT8+3wA3+KoREQGEURQp0wMUVP97TMev3b9E9pZmjYbqHugXeqSDFi4Ued1yZcbtYno1PQ0YkDaTNQAWL16cmFXeOhDHtmbz6H2uocwuv2/roh736pwGjXIAMvcmtQtzuvU5Lq6b2ViYmIwGBRF0Ww295fxIKo79a0Ba8irwscAKESEHSOKEJQd2aS5anpu+6Y0WBbvlIxIZRCAQKwaxjvtXlFVC4Pq6u0vetQZj33R8+78mIenq2bSZquuFro0My6FqgqSNFWoLDvalfz0zQcrHH0wQQqOEIkRmnVaiBHApvMu3XDC0Ths5uTHPmjmmHX/+eZ/3XTNrk4MFIVFVWpGHTr0L23KS5xes7oe7F3PdY1RSHnTpk0hhLpPwYss3W7rfnqUFxe6ImKNM5xwahNnKz/YuWMhSVMgQh0oAEs5XSLaVzvM9RwyVEVVg0BVTISSinC88oprJMIajlFj0MQSor8JuWJ2dnbVYGBdBjJl6dPUdTp2+/YdLbcaupdanAGYfYgBGcpE+z7YJdWzqqqoaH2O7dRkm4h6vV5dPFTVTqcjIteTlQ4aAGuMGZSh1+sNBkWzlYmKI6ca6/QEmDccceT87863BMtS760SPMOTCKSdpSaqioRe3hb+r7d/6KILLnnGv7xOmkbYpGkKVUSFYe+9279efgUHHSscfbCh0dR9XozQ71O395F3/dvtTrvlg57waBy2dgYnPeDpT/ifV721ndox1zGkAy2XlwtZAbCAxVBz1RQMA2SIDRgamXD55ZeHEIxp1NUhVTWGVTXup27oXGqt9d4DlGUt70uJtt2aLKsBYEAy+sxoHXvG4VDaGwohAAKyTETGGlZDoqZgtp1OYzj8jwHE4MuyLG/sad0bjUbDtloQYcdJwlWl8DQ1uTpfcNB0j19Vxr7t/xmAyr7j6MoLLQPXLedAWcYQylarlaZpr9dTVWNMnufO7YMWa+dDHemXY4z9np/dNb9q1dqlayoRgQnGHHnsMecToFJLSXR0O8SAEuWhmu31Jtdu2LGr23KdXVtnrzzv0vN+9PNb3udeaaeJCIhoFDJMRD74veR3vBJK32ywwtEHG6LWGRCqXs8QDNNFPz9r5wUXarf34Mc8GEevvsU973rYbb5eXXRFb0eXqtJ1MgYLLWULAUAJSbOBTgvECnXERqHMRLj00ktFxDqutXpYclLeD+paP8Axxn4/t5adTXq9QZqmy/s1gCUuMweUsGQyisjqo4goIiRoEPZQQ4QiD9aRSwxUbGKzRoL+TcMVIQR4X0KcSUFsDEnA4kIvpbXQPZ2GaGivsc/t7K/cZ8xQXScyVEXW/9ST1Jm5KAprLRH1+/19EnQNVQWBJKoygDzPt27detwJa4dZlNo0igCmI486RgGVIAIPEEMIDnBKgXhxcXH10YevP+WU8sprF7d216w/6vdXbHz/e9//vnvfG0RlWaRJSomDqnM2YEXXcfPFCkcfZAQVxOiYLTE3UoCOO/wov23rdz/+2fWJO+0v7oVicPe73/2sbfOLm3etGRsrpKrbWEaSB64fd6Yn4QxUVCkhG2M0xiJg48aNS+1kdWtZCJ6I9kesIoHIWMuAeO9VnTGJs0kMtW3zKIocvfyBGhMFjbV9BZiJjGE1xGKkDplFg3UOUCCAwh/t8rwbZVnC2tSmQSREz+yMQaPRCP3rOirXP+47Xlbdd4d6fdkb2kAveWYBxpiqqsqyLIqi0WikaQogSZJqP5aBQcWQARBDSJyrQty0abPqbdiwLFlhEWBozWHrI9WXXsSRzSoBVomUnHMn3+bUB/3j3yP3xUJ50e8v+8p3/vfMc86++JJLjjn5JIapLU9Fla2xZGRfua+VguHNASscfTARCcgSH4INgZm0KHRu7uRbnnjh1ZuO5c533vnRX3zhy2tnJu54/PGbr7xyjMnHWEu+BLz0nar1ZRNrZ0BDx1ECURRyCIqtW7cyWwydIhwzl2UwxrDZ97cvRF/5ksDGmGYzU9WiGIwYZ9mfjNKydGBN3LUwRVhVEEVEVES9iC/LkhnOMYwCAaGU4Pv97oGf1H0jSRIQgdmyNZZVEQskSRLzuu9w+dkYphrqHd5zMwLadxpgdGtSd4To0jPeR2NMs9lstVpVVVVVxcx5nu/PkSrGYK2FivfBZan32LJtqxKs5SpWdX9//TrJ1KQnEmOjVrVFOACOYBEmDHw488wz//yxj+rc8Y7Z0cnpp594+tOf+MkPfKhX9CXEtNWEaqjNQwT1Vvc61JXWiZsJVt6Igw+2rCRILDUSXjV5h3v8Oae2oZisdGxXHi7f9O3PfP7I1evWzKxePtxpFPxR7TrfmZisBzfVz2sECCLo9Xp7xdF7WVPuBWNINbqEFaHX74bok9Q55+weSJYeHehHqG6dsNY655IkSZIkTdMsy6y1zPChlFDl+QBE3Gxm2b6HmNwIOOd8rxerCkOdHIxBnuej/DGIMXxARIw4hI/Rjx7EGKNIkP1gKYlUb7I+0jp27vf7eZ7Xv5MkyfVYBi41JYoPiXMALy4uAgCRhLhb/0dAIwvMahzIKNOotx2qEJEjNqxhiW973evL318CQ0g45PN//YynnnbaaY1WS0KAiE0cOR4U+U3ZjL+CmxorHH0wwUAscoqCxPZjucv3SheOvevtjr/T6Yu+mJmYovl+eyDrsrHelp2D+UULgtYlIiYyda4zqlYSJ2dWIUnrCYdQtc5JxDXXbE+SBICI1MID732WZddvK5okVjUyI8sSIo3RRwmiUTSO6CiMlrpQZm7wQqJBJETxUfyIB2PtWF1VaDabbG2apiCCIkky2g8O9FSrqktT45yIOGdVUZaYnp4OoQIkRh9CBVKR4EOlGoesPeRuXVrvMZB32UJk9nqm9iOpu/icc/U+M3MI4Xr20xhiJudMo9FYmF+cmJjYvHmzsVANSTqcBqAiMAznDjvqqCpKiIRQZ2xYmKM1mpi52dkxTradf+mbn/9353/2S9i5YI0puvO20ZSyYmthDERjkFarUQW/x7mCKnRoE7iCg40Vjj6YIEU7baTGVlUl0DRNowqmxp7wshc2jl1/bVjshWr79u2Dhe7U2DiJzve6WBLhipLWyWhSEKwDEZgIBkqIBEW48T2+uycZjh7sEzeuoLf01V+qfI4+h7t7C2u5yJ82HboX3atG7DN7MxznNcqKHNByYBBiVd19oxOjhhCWxB4AqE7MEMBG2FiTOeMMWQOjSlGl0liFaIlRVScfedT2Sy7/z3d/4B3Pf/HZX/mfrDUBFTYGPgAaoVFF/3gr8BX8KbHC0QcTpCh7fQiyJG0Zl5XSJFd053H46md94K3HPuJe4ZiZ2Uxbq6fmeguLVT42NSGGFLx0bzqUajC5NIExBB7OCiCIDG/qDxD7jhP3wTjDWYhygMt1X4uWDWchgAEDtYCF2gPtM9wfrns7T1RzE9f2FUuB8PDBTcO5Bww2RhAlDgPbut4YRkHt6OWllnawdYjCQUwES633s8zsDEuIi4uDud789MTEeCHnf+3773vxq/77NW/G4gDEIrX34HAemeH9NOasxNE3A6xw9EFGmmaSD2KoWNT2q+La7Rf89Fcochx92AOe+vjJ2xzLR0z1qKogrpV6lUg8zEQLVBXKdficpo3axZLZLsWp+xMP/CEslcv+RB+PfUmPd3PxiKx1vybXfwyWlIfMuz2UsUxUR/hDPiHLw+obsj4Q1B1GIkJQAxO8VGUQGU48oFGvCggwFjVHxz3y4A5siHf1QrOT9SkOWK7ZdO2tbnHCusbY97701c0XX1K3SoFo1NC070NcIeibCVY4+mDDEjeyunkBPm4++8Jvvffjv//v72LTNj76yL96zT8cfodTt/tBZ2ZKaTjxr67h11wjBCEoUy3qApiIZBTw3RiO3jtgXP4J4X0Q98g++QYtWDbPsH6gSzmNP20f8p6+dzWjYWkQyZK/89KDfW2Dh8cLucHrAwfV6Y7RSYgYWvtf91gsszGJc8ayMQxWUYVG8pF8POmY9QVJsmF1NdXaItWOqt/r98t++dvfnC1lxcaAWQnGGgCispyUV9j5ZoUV7d1BRvSVSRLLFqQg5ipu/PV5X90xd/8d20590D2R6CPOOOO/L7n6qrN/d8vDD1vcNQdDUIJXroewMCmxgq21dVGfmLS+KyZcf3lqP1jKRP8pobysHWY56eu+UiI3xQvWiodlYB5xNEbzb+rYl6Ey0rItud4B/2edd0RUX4IZRmrjaYKqEmjJaw8EMMMZr0EQQRASIpCSIQJj05bNPNl53bvejjXrNv7inP/58td/8t0fiaGLL73kvjHutsEejt/d137UnyLd123PCv4PsRJHH0wogZxd7C7mRQ6JaLrJtTPHrF032Zfvv/XDn/qr517zia/s/Ok5Y5yMtxo7duyoQ+bRH+9uDq6VXhDCaMLhaCbhH08rsp/Hex7GDS2g7fPzJiNqxmiGoYICKP5JY2qiWgQtqpGGaQQBZOnBHhHx8luBYTLkhqwPGGEkkayD+qWOGNS2g0NIvfdquOSQc6goKomykoFasDEuS7335/3iLBTlUQ++/3M+/sGXvf+dWDs12+8bYyRGAPXMBhHh/ZiQrODmgJU4+iDDx9AeazFIq4oMedKggsIf256e27j5Ox/5dNYwE8wcxFdVayIrVYTUgEeTnyRSzSwEVYEOW/8IBKiSEPOov2H0mvsxY6hjRtrrznr0V7vDySXs0Y9+w1DP0t7zD0l4yeVZGYigfXa93TDsY0TAfrMoxHtYOavWvv3LC2h7yVqWMa/u0e659zNL6wNMHOgo4aIKJTCzJcvDsmptSQUZesGCDEtiBdb6SEoQUSBGCGniGpnlz3/sU6fe7e4IFay5zb3v+pws/cZ3ziRrQoysWr8VIkIwuvfFePSpWQmiDzZWrp8HE0pS+KIMFaChrODcqiMPP/Kk4+eqwUJ3nnxYbdLxIoZtuyZt0mxmvWIQYqEa1VhymXPOGgJ50bKf5zAObPOiZAARIli3fn0tYWY2ZemJTK3lGhHTPrIKhJJQEsLSqJHdv7k0khyMJXck4AZH0XUcGgHdrRWhYRCdpmmdqoESswMzgrRaHd0P9ndKiXbrmpcpRhgAmEFiDBFRjChKGR8fjzEy2xCkljMTGe/jfsUt9RwWZcKeUw1V9nnXsFuGcUNULkoAskYaondJWnlvrWU2ziTiYU0So3TzQWQIGSKLqpzdtUMce9IgolFSMokyIqpCLJskon/11je/+B/06s0Qoczd+X73eOk/vZwy41ppUQ4YqhKG7S/DlnKhpcvSn65mvIIDwcqbcHDBWdY0YBUlayDA2jV//qiHNW5x+GyDew4Liz0buZM1ow8CzVpZ7d08NDlTBpQQQSIawAQ2NDJuIIJLTIxxmQqYaTTFbhmukxfe97SqfXxUhEC6z1/e/wHv5/m65ln3efxRw6yu7wWXmUpTnSPS3Qy+x3o/2Hf6YinWvu6p0H09eX1Y6uupW35UNcsyW8fQNHRIjHUVkY0vi9od2wAWxKKkGgElVFUVSz9mUr915/Mf+6TFcy5Co22tS7IsEsBU5+JJwSDsnvG+Ui682WGFow8mSCmlxHJjkHubZlJ6lP1jHv7A537+I517nN44/aRkw9pd5WCgkk6MlxJ6gz4AkrrhV5d8jVBLOOqU6TKObqTZklZhlGbFdTh6GZQAA3W17f2w0Wwpu7o7obyc1m/ER2gUgy7D2NjY6D9pmEYfSc1uJPb/p/U1a1+Xqz8EUpBc5xq2R6B9g/ZgfxuvOTpIXewVgkgcHx8f+WlQYi0B4gMAGFMNcgxKmwfrYaIyyBhjUqYmj6+dAVPbptjZXT3A3z3yKed8/D9RyFjWrD8htQkf9nbyW0lt3OywwtEHGRIgEWqsgDhJkDVBEe30KW97w6Nf+rxVp51UjbWvWZi/cvPWhJMGO6NgGc5tUlWpne4UVV7UoRwT1XEoAVmWZFlCNKyPLZWh9s4V7DGTxUJHDtFLhbLdfLosWlSG8oEWxuqpgrs3DgFESCanp5Y+jEulzgPm0P1jDyai+lpV97DcvHyS6/dIVZmhGr34ialJYEj41lgGJHqoIPhq0G/ANMhYpnqGQBVCJVKJXHTJNZHRzfsMg261IR376FvfffH3zoRL6heiUTWSmXVv578V3IywwtEHGT6EqEKZyckDYdfGK8/84n9joY9Gs3Xayfd69lNOf/SDJ447anJmetw1J13LBeUlH2EAI610v7uIKFTz2oiP0tSsXjMlGmr/DUBq1tuH38XuXPNI7EXL0qZ7Chtk6Nw03MgB90XTHjlZIQFketXk6MUligcEw30+UFwv3YwG0BANe1gOdOtKunu5rgL8BrUl7i/uBgBWJNYS19fUKBpWrZoSgaqIyLB5BQRR7XW19NortIxR1ZNGQ7DsnM2ydMOGib4vYzMxUx3TSK/ddPWEa379P/8LQZamwOz9MbiJWjpXcNNi5V05uFCbGJswQQfdHlziBB97+3vf+LyXDM45H5XHscfc+qEPPuGkEzft2LV9587BQp8UkEgjS4dhCU+Rd3sIHirMwzhRBM7i6COOiLEUrYwdyrn+gCBvt0vG8s9GPcYwAAKKoIg9HHcOpBecAhCUAyhILbCjCNKpqanR1ng4MnbYln0TYvcR1RxtLC1TZB/4so8mnevvff/DENHaOIm4zvbQzMx0XZGsRyyqamIsYljYtZNizKxzzlBqg+MKUgbxPkTv8zy/w53v9Op3vvW+j3/kYpNvcbvTts/u/PUvf6X9vDY+1NFc8BXTu5s5VrR3BxscY/CpMWnWxKAYa41Nw86ee8l7X/SKW59y3Omn37rpy0YlR8/MaH9gE1uGPg9DQYnDbkMCUPS68B4Sja3Tx0BUZjryqMN/8uOzVJyzaYQwIwQYw1DFXqpYUmjt7cbQUUxNtd2+yFCtXLcI0u6/PaAvOI3EIVpfSFhpKCQYGxujemvKIgS1wJKHxh+DIUsuixnrJCxAYCYhIQShKAShIASh5VXWvTHy7a6ldctj4ZG+ewl7hKXXOYo6xt4j4S4AfIwtdsyxFm4bQ6tmpoigqlFFFRI0sQ6lX9ix06lWsQqAZx7K/BQOZDjZurPYtH1rcsvjT7/L7U9/2l+/81WvL+Z2ijGLva5rptaqiBjm+prNPLorWwmlb35Y4eiDCzEEAcQHZxMUFZxbN7UmEZtv2nL57OyF3/vfBmInr6g3GE8bVRjO9xMSgqFl3/SqzBFD3cAHACoika1btWoqxkAmLOWjRcRau+/giXTEzksa6jrUrcDLJ8DyspZugGTYN/iH1kKidc1NRv7XtfaOyjTjuo5IBNFhup2Ibky++IapkolAHMH50IzTMGrPT60TSWY/R7G0gSWPERpeqfaqJdZ3JJpAD+BuIEZVJgapggFjqd1uMdf1Q4miEqMxBlGqhUVX20IxYIdTWkgQVRlx/Uxr57adP/j85+/57GdiYvyF73zTy578zN9fdlWvKMZFVDXGaNhhOJ2LVgQdN1uscPRBwO4YTY1UJUcIUFSDLE2RV/d6wP2++f4Pr3GZXex2F3YdfeSGfFC4hmtnia+KvAxJlhpO89IXRWmsde12s9nceNllSC2gzOgvdFtjYxREA06/7alF2Ws3G0SqGquqGk0sXM5lu8mlKAeTk9M7tu+cmJgA6SDvH3OLDZNTWeQBKOT5wBjDbCTCOee9XzavRJavRbDXM/W6bpwwdQ+GshIIAvInnXwUWxX1GkKrnaDqA1WrlR6o4UgMStYAUQRSd0YTqerQukR1qVuvqmT9hpmjjp0IQQ07NgheiNWwC7Ey7PbJ0Uvzv+ur1MjWSpjhnAshjMSORiIS19l81bxKkg/KoqiSJClLPz42URt5X/ctEAKYBoPCqtEQOFECTr3VCVUVbMuYNKliaGYu9BZtnm+68koTo3MuhsqXVVQkAkswbEDUm+831kx/87NfvPVd7zZ1x9MwMfWWT3/i5c9/sW1mTFaFnE1DrFQ1SVNAmK5LBTeiR2kFNz1WOPogIwZyWYYY2VmwYqy56qRbtI5at+2Ci2YsH3XYkZuvueboI9ZrPti+Y7aVcmeik3svMbgspdSVvsrzvLQUEovBAGUBa5PEAlBEgh0fH2u1GgSteWTIBcL76zbsdFoigYi8973+Qmcsudvd/+w5z3vC5BoA8F6X0gDMkN2C472xv6R3nWEecfSIowhRxfsicYYNVd6nqkisaDhAxcFSBob22q16YBhMAkBVe/1Bp9V88Yuf//d/n8Y4bJ0PYfggRtj9fDPM0vOjhHzN2KLDLQBQgQiIQILHP+qVl15yDRFNTIwBNoTFPM+xf8tm5xyICMSWvC83rJ22FmIIEOWRGF0Uzs1u3myisAipWq0nGYIJqiqkY2OtQb9Y25p65yte+0+f+bhrdGDcq9/4RtvskHVL/gEhBKkvKjedhGYFNy1WOPpgQgmVYWYybGOooojJ7BEPvsffHrP6i+94z/azz4vzXWqNV2qY7fi61UbD/MJ8FYRgs3aSNVuJtsvS59B2o4ngQQCRS9MYQv29WzUzMbN6ateOXDUyW4nKTCKyr3SrAJib661atXrVqrUACh+UdMu2HZMzgCIq2KiqgMgYDhFs4Mt9jxGwdt8cNGJ13cO7U8HMKkaVIKxaZ30dU3Kjzus+Lj/OOfEeVeWJk6RhrS1y3+qk3iuhbpgnMrEuJhIk7ufeP1Y1Me9RalOItSxKiuickwhWWIveAq6+ZuOuXfONrBVjJHJpmhKYyMTh+IW9OrDhkkyjQOGMHQzKWxx7tLUQawICEwu4VpPA8FWXXpqqcghWBYBhMqwkiISo0kpTF7l7xZaBpXe9+JVPf/2rxo4+ujk9BUrLIgCRmdmaxBjUuY4V3FyxwtEHGS61/RAblkEwzD4ULKF18vFPetPrcOXVO8467+xvffu33//BievXV4Ne2eu2Op2W46JEXhYIdmZmTWJT66ttRX9+49VjE1ORjc2aIuISC8X4ePuII9bv3H5pjAHEqgzUo7NoWBvcM3XbbDarqsoHgyzLWq1Wf7Dzpz/96VVXPvuIo9MYo3MUJQAGxKBAbNLswKQXGgESqhW5ZIHhYBFBdEkq3vvojTOAwgdrbxxH7wPGEAC2VkMEkGWJePFVIMNM9QjByIaIBABxtNbt6fw2XI+a42vJY71tQX1vQVKP1JI6AQJcfvmmwWAwOTlhOBkMBkRirfWViAS7n0CdmUOIQ320hBNPPB4GVFsYWgNAo7Iogr/myisnABZlATPVtyhKWhubdvNBb6GcSicm2G7+7cUfeN1b/+6TH57bvGVy/VHOOaLd/fTee2MMm5U4+maKlTfmYEKAIvgi5gESjHqVGKOB0YUBGg3c8oSZRz/4/v/6urs/6QlXLXav2byz5TrdvNi2a26hv8CJU8NzCwuDPJ8cn1i/Zm3e6/L4mGs0iLmoyqLIY5QkxQknHqcaQ1xK7e7/TSfpdFrMqGLVL/qApFlSluWuXbtCWLKLUx312hGRj95H72Pcay2AQK67Jq59i5RYiSKxYknPTbuN6utm7ZuowUQw6g0BDY22QwzGsjGGGdayMbTXA5F6vOze66ghqkSNUWNUiVo/4/t5VxGCVD5WUYNNAMKFF16Ypqm1tizLRqPRaDSW9HP731kVQT01MsnS4044DvX9VgyiQkQkSlDs2L64c6cjAJChY8juThQhVI6mjzl8c75g2o1UaeO5F5/7je9NrtlQOyNSPUkgRgB188tNcZ5X8CfBCkcfTDCQWNNMXBDPYGaTJY3BjvnXvvwf5y7diMUefIFW8mcPfUDriA2NyZmF3GetZnOsOTY12eq0qxg279h2ySWXnn/++Vs3b/n5T3+KQa/2Xqipoa4O3uY2t6nF0RiN+SDspWljYEi7s7PzMer09HSr1cjzQa2j/eEPfyQCAtdbqCcr1ZGptdZavu6iGve1DIf1RdThXv0jVMHMdQUySZIYI1ThbLzx8xj3RlVVZK2GUBfr6pbrqiqwZEZKQnWrN4RIjaF9LYaZlwa4MAsbIRYiTTNDUGvZWKoz9YuL8cc//cnc3Nzc3Fye58aYoijqmb/1IOA9MWTJGCODWBGin5qaOOqoI6AAhkPHmanuwtl4+eUSvSJGo9FgpMJUrf03mE69yx3v8ZiHPemVL9lOZRAM5hd//IMzkRchz+tBWUvWVMbalXmGN2escPTBBAGa52lAU02iHCsPsmng3/30rJc/4am//fI34BX9AY4+4g73vZeMtbTTbE1NHX7MUdOrZxYH/fneoksTk7idC3PRl5dc/LvhFOvRFPAYoypudatTsixhBjMZy7Uedt87pExkitwXeRWjeu+Loijy6kdn/jQGEHHU2peJovhaGKcaFfG661GkvOd6Dw8jYJlJUU0ZI2sRQm21Kgc+o0B5nyJfEYG15FwIQVSY2VjOGg3RAKhoEKl3XkSDQqL4OJxfvnztR34dS4ReH10kEiD6UDJRTcH9fv+CCy4wxmRZ1ul0VLUoiroY673fX5OhxGiZa+/vmTWrV69bqxhai5DC1spK0ct+f0lqXRWCT7i0CIyou+u0Yrln5PS/fvSpz3j8c9722qJhVq1f++vfnFUsLtpm0zqz9BkI3qvI7jh6RYF388MKRx9kNBoNIjKCapD7MkDJWLtqYiLM9T74xre9+KGP+J8PfKR3wXm3vOUJ6445nNuNLQtzZ1940dkXXDzbXSg07lro9vL+WDMtu70tV16JfICiGgwG9bfaGAvg8CNmksQZQ8xcdxgTL30X95rDjTRNx8baICmKotVqr12zIcta55xzflkgeJSFD15CCFVVRvE+lLtFwTUzjtZ7GonGpbUxbAwbdmxGsDAGIQRnUwAxxsQlYMZI2HfDLUG4FmuTKGFPQYiMdTooK8RorTVkGeRDRK3/w+4hh6q7H4/6LffwUV46prqjWrXmd2HiqHFpOFlRhLnZxY2XX0MwMUqe5/1+nqbp+NhECGFfPiEMsBBHKCzBAggT483xztD+ybFxxAYgVcS4c8umjFmCF6ZooDq6EZJhreBXP//F4qZrkdlbPOwBr3zXW7Z0F3plPhqoNhyiVl/IidmYUVy/Uju8+WGFow8mFBioREM++DRrdpoNqKLl7vvQBweJLSW+dudPP/mfb3vBy171gudfePE5V+26Zlt3sRsRHaJjSci0yTbJOjWhwKC38bzzEMpWmpR5YdQhCBN8qO53/3uJhqIY1DmHqhqM2ruxe24Iat/pMvc9QLIsUzEL832Ca2TtD/z7xxsNZEkavThjSWGYnbWjuHXvhcgQzHC9tBAJRLVmNxplP6BAliWqCjDBiDAiQNQe60QVGBYATDEOJWu1//1enhgM1PbQgAiBDMigjuuJKM/zOsdt2aiqMdYYIwJmq2KYEsOpiiE4wynUEhzB7bHzMASzdLxEtTk11cfV7XWZuNPqlIUPJRzbT3z0s63mlLXOOZckmbWJYdvv57v16ctpWg3UKHjz3GxMKWppk3i/+92NCWCEUFEMCRlUAaFCmf/yO9+tZhcyUgnBWTYKeIQKrUaLFTQIqzR796veCDJoNNbc+Q4v+JdXe4CMqT0Hs2aDrQEZl2S77fz3xoqB9M0CK+/BwYQAAiVyxqWorZMZaDVOvcsdpo5YnyNmSXr4zFrJ8yajOzcHCFtjnGPrInElsVAdiOZllRKcDxt///tamjvMMBJVoUxTd4c73C4Ebx33+13rTLvdXuYvgWU0jdEolqVAm6AGar/97e+VBfr9Ims0MMxrBxnmi5c+RXu5Be3bkXnPuPi61aqhWXL9v3qD77+X9Cm6rN9vd2ekjhz3wMPht8Mt7+MCcwOWvY+r1WoRuCwC1KqgLPCbsy5QsVA7mq5Le/7hyO5j9xlhBa2amfGxCqhEy1NvfaLLEKpKRRppU2JEVBRluWUTynw8My7oVJo1YDKlpqXEUX8wsCY55rAjOsEuXLn5Q697s+6ahaXbP/A+T3zm07pFdyVSPuSwwtEHEww4MabuPGTAcJAIa4++7a2f8U9/t9A2lxSz1/quNJw1Sdu5Nmy1UIa+r3Jf5j5GTRLT7KTtTssQk4/n/uKXKCsYcMIREYZExGV0+zvcNsuSVitTBNXIZvk3dWT9s9wnur5tXjYW9pyzz/3hD3+dplnNddZagrnOvfGN9xLaB/ZvmvF/gus5iusGmGzIFkVJxNYkhvHD7/9645Wba9eR6yhcZe8fR9dLAsaylhFmpqk1kyefehQSxOiNMQCr1J3ycvGFF8WybGdpIjLYUmCnNwONhbo0U2t7ebHx8o1pCdqxePF3fvLr//4OgsDQwx736M7amRWOPuSwwtEHEwQ4ZghijCJQMLGtfECanXTf+3zqh997zYfff8Rd7nDRzjkZ6+SikUxibZa4xFpT97P5WBRlr9c3ikz5yvMv1Pl5qIAREMAg6wBMTTVOvuXxbKTVTkVCv98f7cFeQSoDdvSpqKNsBRhq2q3JD/77x63B4kJfldgYAHyT6gF2p4GHwjvc1L53Nwn27ZTkfSjykCStWqby2c980ZoG1EF5j4zMEHLdLnyQksL384xZtLrTXe5omoDAOGZjRGBqpTThnF/9ysRoFU5ozMFGJNawRSQMotfMJZ02M6/tTDe7/gvv+9DG35yN6NOJsRtn9rqCg4uV9+zgY1jlFyLAkAVMWXj1AVNTJ9z7ns9425u/9PtzTnvEX/TaY/OgXEXBxhjnTMqwgHr4AsZLi0y5c3bTJZdBo0Aq9QFa90qI4IEPul+vPwdENup9ufvld4+7rq2geZlbUA0G7MyqDb/4+W9+9cuL2+1WDARQ3J3o4H2FljVk5NRxPb8DjPpBlpnT0bI4moHliYKlH/dc9qPouOmxd/JlmKFO00wCDOO3Z2/67TkXGs5iWLrmLT922dcDAYShWnqEWFWD+z74XkGgLpjEAOj3+1Q7H+WDi397TgYJRWmY2522bVjTSEtFN5QxcbGRoN3cPD/n0jQt4o6LLv/k+z4Isr7XXxobv4JDCCscfZAhUZiGw+VCEACJyxLjdu6Yl9zz2BQmp4Tt/Z/1rBe+9rWNteuCNVUMVVVFHymCQWlqxsashpAStcic9/Ofo99D9KQj1wwFCPe5792sQ3+wkCR2WN/fe54shvS3B9NRHQlWVczSzic/8VkoglcVxCj6J+h92O0Legg0KO/OBTnXsCaLEVB88AMfKwvVmKi465je1Ram+55zSCqtJI1lPr1q4rTb3SrCE6tAQgghRhFBrLZceumua67OlMRHEC30B10fBuK7AZ40neikkxNH3OrEVScdu3Fh12Kvu2ps4tLzLly46hqKVBvdreDQwgpHH2RoDImzo0Y7igIQkUk3XXLlq573d2d/7TvYusCtCShtuPUpa44/tjU52Wg0rLVM0Iiy1CJEr1FEWOJk1vjl938Y5uZNCA1ixCC+AmAdNhzeOv22tyKOosGYfd1978Ze/0WA9ZVm6dhPfvzL73z7l0SWYImMCAC+Tlx5U2Sl94ij/4+xvz3f9+EM56/UaSLFr3+18bvf+aGzTVZrye0dQe9Rqr3O9kmsUUK4773vNj4GcFRIFbwSGu1W8CV8de4vfsZ5bqNPGBEqDddndBG4A5MlvXyQjLcf9aLnP/4VLz76Hn9WtrOSaLI9US7mNmtLdeB68xUcbKz4dRxkDL++EhUw1kSJwYuVeNKxt7j47HPe+stfnXzs0Q98wH1npsYu/92FV1+9sZifzXxoMDezhJlCCBWJZVuppyATzv7iwvN9r9eIEUkiIl4IDGIB+M53vsPZZ13Y7w6csYTlrW4j8/7d3LE3OTay9q65zY0mPv2pL9zvfncEYE2G/3/ePBNGNx+jSQhgKDuLQQ9f/tLXVE2SNENORMvPquyfnZcgZTXIGnzf+90zBIBjVPIxJom14LLKEWXjeee2EZ0X55woKo0uYUqtVWFjunOD6ppNyNKZB93rObc84UNveM+vf/6bndu2EDGCqDVLBuMrOFSwEkcfZBjLiL7+wkeJqhEmglQ5PPBe9zRzc8UlV3z6Vf/ysVe85rdf++am837XajRd5myWeo1VVXX7scrVl36y086Iwnx3yqT/8/kvoSxgiRLjMoeoNmEAj3nsI43VRjMJsQrR137/Mca6xVkh1lpWjBZm5aUfq7ya7Ew33PhPfvjrb37tF4MFIKIoqn4vZ0bdRCciZVmWZTmc17Xv+dn7hiqMYRGleiiUCFRbrRbRMPEtIsZw/bi2hN7dInMDEEJY6sOT4Ug/XRam7y/23+t5qSdDMoMgdTtOUVTMrrdYAPjpmef856e/0Eo7ju38/GKapqM9XCJoIVIiCiEwW+9jCJJlWX1ExnBEMbN28q53uTWr1FLutJEF+BiLzJp49RVXnHuOdvsTrUZibF56lKHlUqpCbzZIUbUMmmTf/k+vRlHQicc+/bWvnDjxaLd6pr16TVl5k2YrBH3IYYWjDzaYYBiGQUSkzFDDsJpOtO58rz/3EotBv8lmsG1Hb+vOYw9bP7drMU0bu2bzooq51+Y4T6xqFxW897EsbIyTaXLRL36ByqMsAGGrUXwMFRhphoc89IG93nySGDtqCrc2CUFEhBkh+qXd2msvm82291Lkcf26o9/+tn/r93xVIkuzVqPR6xX5oKzNiLMsaTTSKDeZz8bNDUTqfRlDRcwqUpU+TVqhQrud7dgW3vav7x7rTItQWfg1a9bkeb4s3b/HBcAYx8zWOCIaDAbel8aSaLXQ2/Hgh9wnTcEg7yPABIYIS0Bifv6974SFuYnMhX6fmY1hjbDQsWb78HVjGYwJwKAM3cFFP/wJZuexavwBZzx23QnHNmZWcZb9//Gm59DHCkcfZCiRkioRMROR1h6XhjHePumB93no058822CsmWysnZnr930Rjjr8sPnZxVXTTZvYpJPM9eXa7b2xmVbfl5HVIJqyuvaii64573yoqARAVSMgYFiHxz7u4VnDgLx1UIS6O7w2WiOiEKr9TZkiMlCrylWpV1259QPv/0SaoCwAIMuyNE2JqF6Lyk3ohfSnx/Vkn/fxvK8KghCRxCgCazICq0ArvPedH/jdRZcHz1Up9UiB4ayv6xA0AMOWwHXzYVmWxJokLqKcWt35y8c8uJ61SDCGraiQKvsSvfkzv/F1U+XjjWzQGwjBGDPRboayWpxdLHqDQb9spanNw6YLL/nsuz+Qb59Fmt79YQ9+ygteECyTcft1xV7BzRgrHH1wMZwoFaERWkdcERpAcGml8Qn/8HfPf/Prj7/H3cKqqcaGDVWWbZlfSNrNxvhE2moXZdVssUkAoopgUkuseXehI/j+V74MiRw9fGUtmyQBIIJb3mrD7e9wm7xYUFTDhIGgdp0funfu3rHlncq0uNAFOE1ac7PdNasP/49PfeHnP708TTAYBGvhHC3pMYqi2JeP2qi0eEhPNSURidZZNqYoKolgphjhHL7yXz/+zKe/sHbN4VUVDSdl6bvdbprWOgpZ+nMAQ2sOQQhCRNZaInLOilYi1f0fdM/Djun4IFElyVKCkRCdKqJuO/ucrVdc2jZcT3z03guhDCUbQwa9PFSKtNWcmJiYTJq9jVs+/a4PImnYztitb39722waY1ZS0YciDuUvzP8bECEVhXqgGk545aCmiOCsLc3maQ950JPe/tbXf+WL//D+f7vDw/8yTHQwNva7KzcPKp9ljTKX6bHm4kLPJoCF18oaHU/Mb3/4w3D1RlQlAIwYUzWC8fBHPJitDzE3FqoaglibEDhG3cN4fo8CFxdFRbAAO5vt2DHrbPMtb3rX7C5kqe31yrKMxpiyLKuqajaahnnYa713SvpGYLn2+Tpq6H24hfxpkaYpdDj4tZ6MRcAlF8+++10faDenFxdyZ5tJkpVFZYy5Hl9mVfU+1jcc1jIg/f5is+Ue+diHRQE7qmIAkYBJlEQR40+/9Y00asOaohhkTRdjVNXZXuRWRq3GgMCTjS29xW2Lc9PT081Sz//+T6/4/k+0X1CWCoEUJLribHfIYYWjDy6YRDGyZ9c666HMys6kvbxgl/VKD+tg087JpzzghS94w/vfv/rYYw67xZFbd/Ql8rqZ6au2DBqJtY00hx9IMTHZit1Fv2vX2d//Afo5yEBJgyAqG0jA3e9x51NvfZKiJB56OhMZIlc/2D1pfLdQjAHOsmar1Rn0c1WyJiG488793Xve/SFj4Jyra33MLBoAiP4/6xnPzDHGoiicTQ1bKLZtzd/x9vf97qLLmo2JXrcKHmXpJycnnXOjXqHRt0xHbThKwNAaMERPrCEWPuSn3Obk0253cj8fmJREtfABgCWGRFy76dyf/7JhEL2vKk0aSX0blLW5VN3Rze1k404PvN8t7/Jng9RctWUT5dUq1/yPD3yEyEBQVlXU4CWuRNKHHFY4+mCDbW2cZmAc2CknwlbJAFZNf+fCW/7pded990zkFbpdxIgTjnvs057W935qqqU+Lu6Yv83Ra+bnQ64h5+gTVOq1GHSEf/GN72LXAqJG70MIQO1Mj6nV7vFnPMpYUQ3MbIyLQSWCwMvyyHuLHJxNoaxKIYTp6ZkdO3aOj8188hOf+dpXz4wxpqmtKu+cM8ZUvqpf7pDF9Sm760Oz1gLMzFdctuszn/zcN776rcmJNTt3LHTaEyGICpIkEwn79ekG6o3UhVZAiTAxMfaoRz08TZE0DABOXST2EiARPvzmR2cubt2ihYcEl1IVPIkSUdrI8hgLoLl29fF3vv1fv/xFD3/6k7LVk2vXrt117ZbBzjlEbLv6miRLoyEkvBJGH3JY0UcfbBARkwVi3eRHIECVFmbnJibGEbW/bcfrX/ZyE+Idb3ub0045efVY9vNvf4tikLx01gYfN129bfXq5tb+IG0ZZ9yurcVUJxvEcNV558e5OeMHkazhBIY0REpM5fHAh93nta97S1UMnMlAWlUeysZy5cvEpaM924NfqqqqfNFoNFzCmzdt67Sn80Fwtv2Ot/37sccdc8tTDq/K6JwjGGN1xPW0O3LEkg/zcvrbk8KG5crlVnDLnv5DWPoL0tHYLQUT4nVs6q7XkHp3emdfT4LZAJwkDoKiwHe+fea/v+8TjWwqdWMBFAO3ms00Tbdv39psNq0zpfcAiMA69NsjEJRJYZw1RkUq54JLdPXaiQc86I55FdJGOsh92nAKSBVMWcAX5/70J0kIUhY2dWkrWZzvJwRrMKhyJEkz4V3btv/s57889r73vctfn3Ht1rnfffdXrjOepW0s9tccccR8t+saTedW+gwPPazE0QcZQlEQFdGoEpRUASWKDQMEr72F+97lTmFhziwubD33/C+/5z3v/8dXbvzFL1pFdcT0tPSLhGEMut2BBLBQ7JVNRiy9U51sJv/6+lfDKWthKEI8WS0VwYHbeOkrX5S12GtvfnGrkJ9YNRZDyJJ0d35DzXABA7CO67Spr0KWta1pMLVYxy46f/M//9M7r7qy6LSy4NV7z8QhBO8DlL2PqoYZ3pdVVXifj4577w+eaBjqT4iYGbU3hVD0oS5FOudq8Um9Hp29fZxSVTUgUogHhIhM9DIYDEAWbFW1jl6ZFCojE2oZLcPdE0FVhRi11r1UVQWNEGJJWNzCjqrs4atfOvPd//pRxpjGRlUaiY7IxaiDwaDRaCikipWaGO1QTsHKRgxHZVGSSAjzC7siVZzEbrHzuS/6m7QJk9oKSsZqRHe2mzBxWWz6+U8v+fUvqLvQUE3AvqyMBdvaMlrZyPpGq1PJr3/4M5SCmdWPfctbmieeOGiPHX7CSWh0tCqttU3nyPuVVMchhxWOvlmgJsblAWPaalaL89RMT73NrY48+ghQDL40wGGrVrUsV/mg3+8zwbJJDDlD7RQciSKllptZllmnvcGuq67a/MufUdmHpdDvERGRglUZd7nr7Y84ao3E/vSqTpLywuwua22MS0NVaM8q3J6pD6URgydjnTWXX7r1H//htYsLsIacawCUptmgXxCZJEljjGXprbVZlrhkud5jX4HqsPOFlv3vAX1Eh9shgEcyktHt/Q3ZzvJ4mbMsqcdI1UkJIvJVjIFAGB9PvvSF77/pje9gamTpOCGFupFVdL0HQ72d7O3yOtwTVS3LYnrV+Fg7zfvzD3rwve929zuGCNHICudIIybHO+W2bTD2h1/973zXjo6zrSTzRRm9WGvZGWupkaBt3OLWbqvQ1Vnnwh/+DK1xqPzNq//xiLvc/vHPfiYSI2wajUYM4Qb2+6zgZoUVjj7IYGEWptrBElSHdcKANXa8g062/va3+cunnLH+Vidc0d25yHF7vtiNVWnRixUSw8wJG/KamhQRCkMuZUqsOs613L7w7S98xURCf2Ath1gxw6g44JijVj3pr89g8o3MUPQqMURl465jpLm7xW4fVxLQ5MTU1q3bf3PWuf/yuneUBQhcFMFXcWJiioi8D/XUrtFM8Zs4jLvh6VWtPUOXftz9P7uvPTRKR9T9hFVVpakrikG/3zfGxahplpkEiHj/v33h7W9/V1WGXq+3e6jKaNThHlyvzMoAhCCEyIiESOgXOSDQEEORJvzUp/z1xFRSFgMOwlGpvkhpTA31zz//19//oRXJkpStqaIomMjEKFop9ZFvG6zLkpmkc+3vNn7yXe/D9p3oNFfd+bT7PPWvGiceGdqZpi4SxJDbx6zbFdzcscLRBxV6nQegSIjgXlUFy7mvRONdH/yAN3/6U2/54Pvv+vAHd9tpf6KVT3bmE66aWcGGjEvZmiAaIhHA2q8GlR84xIbiNz84c/Hqa1GVsC4xHKrKklHRQe4f+ah73f8B99q6aaNBlToqQ2lciutkb/fslFsOBlAU1Vhn0nD2ta9+61/f+r58gDTJfEVEVJa+qipjjLUcoxfdK46TZevrwQFKREj2OJ0jLL30DZ3qQhqjHwx6IqHVajHzwnwfisVd+o63fepjH/2PLZt3ELnVq9cuLCws29XltnYC1KOpQAqu8yoYXgSTzGUNW+bdfm/2UY956Om3PXZhtmsZVklDDBHOAoM+kuS/PvEx9PsTWRZ9WRSFsdY5B6HgVRWGkWYG1rFxtz/6FmYQ/uP9HwHbSvzJd7/ztYs7vePaSNaQkZWhsocgVjj6oENGzSs6VN/Vd/tJopTAZdzu8NQM2mPH3u6OT3rNa//9lz99xac/eqenPn5h9cRWh0UH22wSkYnRihB74bLSssQApkioDLO7/veLX4a1qAYwxvuSAY0+eg/GC1/wrFbLBD+IUjnnfAxSj+xa2rdh+Bz3cK7YHUpzt9uXyN6r4ewjH/rMpz7xRQnI0mx+LpcIa5LaEoSImG70h22vEFV2P7P388ssmZeF7Ky74+jlkwSWQuDR7cFSpkJEg3WsiI1GY9DLF+Z601NT3QV597s+8ImPf64sdGJ8Zn5u0VfRDge2Lu3Y0rkCax2b8/J4XwhCMjbVsQl389mTTj7mBc97pgoajSRtpIbYESOGelDDll/87JwfnDltXaJS9osQvHOOlEg0MxnbtErMvNUd7LsWW7Zs8/ODn3/nB4NNO4TQFz82NeU1+uitMUGC393rv4JDBiscfVCxlKwkALJcy0DgSmMQAayvAoTRaCNtoj02fuRRD33Ri9/79a+/8u1vj53277buMmlGZGrrewO4BLYuOebFjEvP/K//6l9yMYq86nbbzVblKybbbGX5wB99wsxzn/d0Cbk10Sacl4NlO7dXSLhX2MsAWGFNoooYuNcNY52Zd73j39/5jk/2e2g1G43MGWNrgjamronu9XnbVyj9Rzci7jUUfPcFZc90x+il93FzACDG6KxLXOZ9JNixTvvqq+Y+8P5Pfvo/vlgMVCVhSlatWrO42Ot0OjH6ZYmgpWhahoejTMNleD6VYl52xYYq9p71vKdOr2mG6JNmCgCGyZiEgLJEd+HzH/6gLXPO+1RVjtDIUmb2ZSlBHIwqxWb2pJe8YN0dTrusP7eoMWu0tl95zeXnX8TKqkiQGCWEaMGsOLjDx1Zw47DC0QcZylBWkOwlCBPRhFwoVYRco7Nj8/Yzv3fmb374c+SMAaM04HT8iGPWHn2L1ngrJ+NhCIaVDThjm6lJK2oFtKpysHnTVz/5CVmYp1ASRMAwzIaT1A0W8yc/5a/ufNfbsvFAYTjuNwWxDx871HLBibHpWClrkrnx7rx/25v/7dOf/OIVl24vcxiqzUcMKdcSi+Ex7xvXN0iFRg58NxKydxw9HNx4XRtWMJRTlwHsy1Dl0mi4Tdf03/zGd//La99mudPvhbKQNGkXeVi1aiaEWge9p0PeHueKAGPEGCEjYERC6BdzVVx82KMedJ8H3W1+sesaTkPQGGEQY0USMOhe8J3vXHbWWW2NLnor0kitZRN9JGGrBkFjkJAkR977rk/6wDtf9J63brjtrRejP+aIYxau2UYL5SrXQlU24CyseHHGpryitT30sMLRBxl1x7TWtf66xQ9gQGI0QDPLBr1c8mrmsCMWF7tvfN0bXvTYJ3zt45+58Ovf/tCrXvfef3njhedekCRJVClDUDDDcDQ2GBdsE27SZXFhcU0zOfMrX+lt3uTa7dnt2xNnK1EvwgbOscvoiU95bOEXoXnWMIy4Z3nwD1j1i6C7OOh0JrO0s3PHQrs1NTmx9tX/9IZPfvw/r7h8KzOM4aqMgKlNo2jfofQ+Nrz8f68rkb4hz+C6/Lv3pbDG0gyapcUCprswyNLWWKf5g/8996Uv/qdvfePM9euODd6sWX041O3aNW842b5tpwgajcYeu703QTOLBYahNCGAy07H5uX8697wTz76ydUdBVSVjAPBlwW8R1F9/8tfnHY2n+92GhmLsmiVF9H7LMtSl3Ekx277jrmzf/lrtLIjH3Cv2z34vjtjVfjQbnZc1iCvmWswyDJrjMOr1EpC+lDDCkcfXDDBKAzIgAyDCFQLko3AABJiu9Op9V/3e8D910xOXvPb877/yU+/44Uvvfjb39913gXjPiRVkKJwzCGEGJSFKZIL5DxLtxx3tiFhHPqef34Ndu6cSDMGquCL0ueld51USe5x/9s/7ZlPGgx2DnrbQ8yJxTlTVUWvtzgxMUFERVEsy94OUbcyW2tVtch9COJsk5BItDOrDvuPT33h5X/3qs9++tsSYdh0F3MoFub7ABYWFohRm03XYup647VXqsS49DplWSZJEqNYa733tUSkJkAiYtB19wqACESWnASJiKwxxSCv/88YE6OmqatCANDr9WpXDWZmhgpi0OB1frbX6bR73fjfX/nx61/39vN+e4kzE6RN0mzQ94bTRtYBuNFoAaiqao89WDZZUUQ6zY73EWrb7XbiTGLVULW4uO1T//HBZsu1Ow5AUZVkHJgq9VknxeL8jz/1qWsvuMDPzzUTKou814WE6KxNXZKYpN/PQxCtgD5+8D//Czaw9i6PfHiyYc2c+ONudStYBwAxiogSWWsRBeH/2R79/4exwtEHHcxgLHXhjeLoZpoAqA3kjLUgpKvXPOM5z11/2Iaqqiab7alWB5UY2KhSidpG5rIsqpYhgFjYllHAZEnSGFy/V+3Y9u2PfpTTZLAw306TVjMl0hCibZleXj3zOU/+87vdrtMxRbk4Mdla7M76kB9xxGHXXHONtUmz2bzORNeRboGI2VprnU2NcSrsKy5yHevMnH/eZW9983s+/MEvdBclSxtViXazNT+32Ol0oOj1enXawRjyfo9almo9zFC99yAxtCxBoUq0j4zHPnMgw/QIgHqw77LfqbcXoc1mk6g2aJWiqGqJIEATE+28j89/7muv/ee3XnzRVY1sKlS2u1gNZ8gOW3uua+S07LEywJZ469atzOycmZ+fB8esYebmtrz0Jc866aSj2y2AkBclAGKEECypdhfi9q3n//hHfteuCZc6QpIkzQZUIT7EGPOqFEKj07ZpcuS6tVdfdtXs5VdDVbPkAY979CKLm+oMB/PWO0IHKo5Zwc0IKxx9kEGjpc56DHMfAIAo0SYOBEoSCQFEpz7wfq/9wL/f/ylP3qa0tQqDJJ31wUxOadbulRLZkXEBVBAGjF5qQispqohSbBV5sffdL3x+4dxzW0zsQyy9MSaqVDG0phKb4WV//7zJiaTZ5CuuvHhm9USaJrNzO9euXbu4uGiM24OSdicMhEQ1xFh5X1Sxigau4ZqdxqRFY7KzujdfvfvtH3jm017wvW/91CiKIo6PjxnDRVG2Wq1GMxWVuj1k6XwMc9YAgKIorr8RvM7HLCfovQqGNVhRlmW9+fpMC4GIoMw0bDwJIUiIUGUgsdi1I7znXR993Wvfsjjv16096srLN1s01kyvh5phg88eQvLrDClf4m7VLEld5uYWZm1GQYvts5se+hf3efKTHzfWshKVub7OcRWUIMi7ZOnLH/3oRT//2brORLG44IwV1Xp7VSGqajJbGemzeObEuMHOxa99+gtAQuPj93rYQ+736Ecka2ZgOBJHYiWOteJ+pVx4aGKFo28euA4TxToDUCvyDEdCnvcBmjj66Ec+81mfPOuXb//cpx/6rKefcJ97zCWm6qTbBr25flcIruECS85VlUqVwGYmseg4zrzPt2z54gf/Hf1+mN2VxBDKIkmdGu4XXq2cfMtbPOe5T1MUhx+xdvv2LeMTrTzPyzLPsmaR+92t4csIum61Zoa11iW2Fm94r0UeQ2XmZ8vuQpCQ/u7CK/7h71/94he9+sILLi7LSFR7ogbv/WAwALCX99CS+KLf79eu1vV66cEN7JdbnnseDAZ79LDocOBWFeLwZsWYZqthLXkvmzctPPc5L/rcZ/9rYmxdr+t3bFs8/rhbJUlzcaHWvSyTkC8x9e71bo4mZQOamZoEBdcylMaFfNetTzvxhS96zth0mhfdEAtodM4wIx/0qKq4qK790Y++9bnP2jzvWKNBRWQwqHxEmqb19YusGWjY7vvXdmd3dXutRvsn3ztTdszBJTNHHfnEZzwdrYaAh3s0iqOlZuoVHGpYedMOMhSie+l/aZQeJY7QKgYF2TRJmq0qRkxMYv36gQZ3qxPv+uwnP/t9b3n9f37sDo944FG3Pl6c9orC2BBNWVLlk7KgoorRkNVKqConoBf84PsXffPraSyp6KfO5FVJhiv1LuV+sfiYJzzkL/7yQZXvE/s87yWJ27Zt29TkdAgC2OFt/vLPzB4WpqPqohLUSHSTE2tnpo8I3hS5Fnn8769884zHP+nZz372VVdtajSzWpBXJxmM2Z1Wrhm8ngve7/cBMEOkrmSKaGSD69VKD3dhLxSDwXDr9S+oKqHOpwMcvdQBe6+X/+53v/+X173prF//dvu2+SKXIw47rpWNb9+ys8hDq9UCsOyll99Y8B7x9ZC7CUBZDUrfH59ubpu75shj177qNS8/9vjDpZJGI01Sl5cDiaU1lBKxaHHZFR95/RublV87ObV986ap8Y6IeAFbKJm0mZCz27vzbnrsiNueEmc6u0KZNJsc6dc/+RXyClmjs2YtiImIYWiEumA5tOxawSGFFY4+mFAAYFnKmu4WggkzAIGIM2bYX2EITGBTVj6bmkaSYGIM69diw5oHPevpD3n4Xxx25GEiiDFEURVYVkuocqiSBHAILQnj0X/+ve/pXnwx0sQ4U5ZlgLYbzW6vPz45BsWLXvrcw49Y02pnedEn0lWrVs/Ozmdpe2Sov4ymSQEoRCSEUHnvaym0cy7L2lk6du3VO/q9MN5ZnaUdw1niGknSuN99H7B69WooyrKMMRrLy9iZVIFl2ef+oDsKnDGKo0e/doAY7MnRo5dTgolRq6qKMQbvY4xzc3Nf+tJ/OZdNT62WaIoiWJMmLkuSZP/xO++53p3riDH2i77N6JrNl8+s6zznBU+9ze2PAoRTtknig1cVaw1BsmYGX33zE5+Yu+jimUbTSrDERVE002bmWIW6vYGxiUntYhlucZuTn/iiZz7w8Y/SRsbWGeHfnXs+jEPwVVkArMs6hoaXFNrHpWsFN3+scPRBRty9jMaLAABURMRbppoaCRJ8qRJ8P1/cOvvFT3zmu1/+anHVVmyexWwPm3cVc72EbTOFKnNECjSjSzy3GynUOsfqfRKqZlVuueiCL33kw5jdIXm/02n380FUJEkChgBr1jTe+a63rV27eny8k6ZpmqaDQUG0zANvz8ohkRpjnDNp6pwzqrEsfb+Xz891N6w/anJizaDvYzDdxTxNWk84468f9aiHM3NRlOPj44lL9khcLB376Jk8z7En/qBPqdDeeaM6yi+KAtgjjiaQKlVVCF6sTYwx1tqxsfYptzz14Q9/RHexn7im42TrtdtE0Gy2sZvo93qF/SWjWWspuiWYMD7V+JunnvGoR9+jKJBkDCBKLMrSJo5YEQP6/V9/5Ws/+6+vHzs+Rfmg7PfGxtqDgQfYmkQEyqbwlVfJxuzMUYc173mXBz7+UXe5+59XZSgWB23bAFuJap2rFXYKRELkYZljBYcoVjTtBx9LnKd7xHlDKpDo2ThAVIRIBTKzetJIeM8b3vS2+dlTjjh8ylDv2mvNfBf93liWNo0zMTeE1FOvL53JbMeuBTgs5lgzifn57u1PPO6X3/yf9Scef7+nPlWsbTcaqiJKeVklnMDg2OM6//Ivr3zxi/7hqq1bJsbXTEx3FuYXs6zJyqQCqpMzKiQASLQe5gKAyBAZ5yw5R0i63S4bLauCfD4xMXbv+9zlmc/6W5ei8grDGoUMM5kQxDlDQlAogwCjAgVUUAXeM3AWwO6nV265eoEUDJJ6b+vFF4gRKqAlKTopwTCnjkHwZRWrGIKsXtN+7rOe9Ysfn7N187VrVh1pV09X3i8u9CYnp41VHwaAgJYKp8suWrR3+49QTBs0KPtVb+Fvnvb4Jz3pccRwDj6GWPm06RJODUXpDxi8ePXVH3jrm2Z83l/sd5K0Et9dWFi1qrMw302z1LhkbGxsx45dgI5PjfUHA8ztwmGHP/CMM779g18mrdaao45GWXFn2gcfyzJJsqUzpkuqIV2pHB56WImjDybq1u16WVZ4Z4CJLRsHMBtXz9BySWbTxLsKTX3w/e+5fry1Rk115ebFS6/O+p7yavX4qsnm2GBHj3rR9KIp0LIYlAtuHNRBZwaFiY0277z2qrWWfvAfH188/1xb5lmoVCOnVpOkVCGHIPJnf3bsG17/D7e+zbE75zeVstiYtLMLW5NUGhmFauCs8eI5SwqplGslsmVOCBZqVElEur35JDWdidb0molBtXCLE4/8p9f+3eQqBIAdO2cAiMBQYk2iQqQwjLIKMYhUOcqie9lVLVgLkhCNQX0pIAMf4949mcs6bYjIGuIIVGIF0BioMi70Z3cgVn7XTg2RDEWlobojYQ+IwNrE2aSRZgg45ZaHv+xFz5lom+7Cdl/1fNVfu251d9D14kEVyAMBANRCbX2Hsdgf2CQZlIPC50IxoDQpytDtyw5JFh7ysHu95CVPm5zE4kJhGKoxoArwCj+/fRsbCldv/NDrX6vzO5QGtgn2fVv5Zoqq7CcZE2lVyaAfMtseSydchcsuvhTT6zC1atVpt/fr1s2mjRPudCeMTUhRAmwt1x3ohtRAGcQgA7NcM7OCQwUrHH2QQcuW64Cvs4ZaVsSk03zsE88YX72qApBms0WZrJrZlvc3L3abq6abnYkkbarsru/F2hiTSAlWJI3RLXbf/apX55dcQoNetTgffJUwRGOedxUVHO5059Oe9rdnnHD8YVu2XZEmMjnVzH2vimWzmcYY161bd/XVV3c643vXzUZIUxfioPSLO3Zee8JJR73mdX/v0pCX5VKYqUPzJgZAinIQJCLLbJ73rbOoyo0XnO985AMZjViXAVWVd9+RCCiQyuUXno/FBddqxbLQqN77Qb/IjCUVQEY2S7okU3/YXzzwHve8izHeujg+1dy+a4tNaFD093xrloQcmJqaijGOj3eS1LbaaZbZHTu2TE43y7jw53e/3Rvf+EqpqpCj1UgXe90kc0kjK6vcqU7PTMerNn77c5/93c9+MpOYVAOprwWYdaYCECHOOmPR2ELiQn+Q94t8cYDcV/0C7c7p973PHe9/3+bqGRCxs/XEslFZdRgH8Cjns4JDDiscfWiBY66EjKfX3POMM+7/nGesvdedLpBBftTaLW3XWzM1mJmYzdwlC/M7yjLapJGNwVv2zBVrZIlGIouAopLXS88979/f8iYszk8yOhqshrYjF9SxDT6mE/jLR9z9Wc960m1OPmbbtZelqaZNM9ebLSWwNRdecPEdTrtD0S1GSeo9LzEkNmNOtNvb2WjhFf/4olufvqEzljCLUeF99ZcbS3ULSepIevOQ6sxvfS0Jnm+ULb0wR4IaEJFRJCI8KHpXXYOyTJQo+PEsaSeZlGIUDCGIkkQW4ShGYdCeMf/02lccdvS6Rsd1e7Mu0ShFljmoJUlJDSnXFz5QBCkzyiovfTkxNb5957bZhZ1H3+LwrTuuutc97/zud7654TDWSfJ+zxhqdzplFBFkJvXdHmYXNv723I++9V8nRVy/lwWxgmBQOhQGFauwiYA3vK3fnROfGySN5vZtO7f+7pKEHNLkoX/58Ic84hHtVTNQwCQYjnKo3xOpj84skwyt4NDCDZWaruAgYvl7FGOM0Xd7C1MTEz4fpIlBnqPfha+QD+Y3brzy3PMu/MnPrzn3ov61W5pKrZaLLIERRloMK1BwSTa021eVxX3/6nFn/NOr0GyADcjBJFDjQxTRNHMo8J+f+eY73/Xvc4vlrtn++nVHL8wPQilrVq8PQYqiMIb2ZR0nWTtd7O4o/fyr/vklT3nqg5QQJc+aTpb9rsJiaQaNQT4oXEJU9kyVX/uzn7zpmc+brKp2NXByABNs64uFgOvIEUwCLp2txsZucec/f+Kb31IpJTNr4TIoYgSz6LC7XESHQ9BJDANQXPjbzQ//y8c3mlP5IHbGpgfdgjUhMXu+IgMoqmp8olOFYsvWq6amx8cnmt7nJ5x02Nvf97rxthtrpTuv3T49s4oarIxBWVmWfHbHRDO75ic/ftNznreaqbt106osSXzlGAWhEAjDssmi8eAFdpOHHdZota+45PfTaXbtoPvof37lvZ/+TEgKoSjRNBtQgbMxBmNt7Xhbh2CjOYr1DyvpjkMKuhJHH1IghWWbUjo5toqNdY3moCy9JayZxhFrcYvDJu5869Oe8fgnvP11d3jkA2XNeDnuSitKgRCsDhdGIEjMBx2Vw13yPx/66Lfe/W4sLKC/CPUwXPnKZTZtuiqvkOEBD7rby1/+3NRWR2xYtWPbtevWrGq1GlVVzW2fJRkNLlnyYKrNpsn70Fvobj3jiX95xhMfZBxAYh2rCF0nvVPPnRFFIzUWwYjH/K7/eO97Xdm3Pj/gNubRrwsbYQNlozBR8p2zP/3fb//2W19PEAHpzs/GKhgL1LtBMvS2oggE5VjGWAmOOXH9a9/0yvnFLTNr2vO7rlUpACjx8hGIdTI8SW0IlTU0PtZhChLzO9zhlFe/5u/Xrm4bK4OFxVWrp6jB3od+P2+liRbFRKvd/93v/+2V/9zo9saqaiygEaMTqFLdGagKVjYwAgqt9OFPf8ojnvE3Znpi+8Li+NjkxkuvRF5AgUbTNBoKwCUA15PL61Kh1G7VKnVK5yb5EK7g/xgrHH2IoRoU3ntmjlGjarMzXvqAMvi5eQBIGygDiGxqqWGjRWAIgRVWYAWsw1bp1JhibjHLwwlTq/77Ix/71Re+gLKA5bLf9xAd6jfKqhiMr2795aPv+6Y3/XPm5KjDZ665+rJWO52b3TEzM506A5JlDSwChLqqtnP2moc89N4veskzjYUoXMLWuSXvIdKh0k5omCsvy1zFo98D4Wsf/9glZ/26EX2qwRxIPpprtbkMh3Ab1HNpYKK66Med+8h73zXYtQ2h7DSzsroeYZ+YhIQlaPW4M+7zkr979pVXXJBab7SuLy77g9Gxq0YfihDzdicJsXfHO5366le99JRT1kC1KHppM0Vm5+fnyIAkoswzm2DH7Bte+NLulRunyfQ275hMGT4CiFAlMGAjuchWWcDJxPhx97nHhvvc/bb3uVewSVTesmWb5BUMAyhj8Cr1SHlgP2K7lRvmQxMr2rtDCUrgxBljBFFVu93u1MS4n+895Wl/6/r9kzesM73BjisvM2WpZcnVoJ02tKyI2UidBUYkKFkFd7Lm7Ox8w+iYTUMIn3vPv6/ZcNjaU2+NmXWm4cqqLxJd6qw1EA+4Bz7wbicef+IZZ/ztYRumu72Fqel28IMkSYNfPkNLQFEogMvTb3vS37/iBWvWuhBRVYOGy/I8bzSasmyKy1KbMkOYQWXld+288te/+s6nP7OKKfHBHLhQzAhIoabu9GOCQNSwrJ2a3rg4bxrJe9/8+uf98+sbazY02xODfJBlmRIvjW1ZmmcIhWWynQTAM5/1xAt/e+55vz4/BDMog8LJSEUyWmneW5xZPREl7+e7nviER77oJc+amEC/lyetZHp6piwGg978+NT4/OzOyfFJ7NiGXfOvfuaz47Wbp4Vkbj5VTHY6m7cvuBZ5VatsBFyhYYxhCwNuZWg10Gre/1GP+cU3ftjNq9XNNlsHZh8rNQTDAg3BE9d9RlEAg2XqaMaK9O5QxEocfYjBJsbHwGysTaYmp7Twk9Orn/i4J8xes/WKsy/aev6l6aKn2XxMTEtctdg3sqz3RHcrPfJysGrVFMVixzVXJUW/UQw+8uY3XfPbs9MkySSktXlFQlFDGYqy7KKBo49d/dnPf+KoW6wrqoXSL/i4OMjnQEvJ4robO4A8qHzxS5+14fApEELMG62kLAbGUn0bvjTWb2R+JARJmWCNGxT/9rrXtXzMQmw7htwYTiElo7ACp2QUpJFjnN22q82cDPLffOd7//He92LQR1E0MwdAwEKsMACRUj0zRaPEGC1hbr6bNvCud7/1iCPXlcUCyIP8Hg5YUFBYvWZy0+YrFrpbn/f8v/27lz9rfBIgpE1rnRlUZZJlnbG29/nkRAcLsyB969+9dOHS32eD3qosSRnrZiZ27lhoj7lIFLkuG7AVMsIMowpmizSFyNStT+0x+2Z27IknwTkYApNNEmYuiqKoSgBR4p5nZIWbD2GscPQhAFqGKHDWkgIRCERswO5ej3r03R70oMq6ikwUK16rboVcmpSFMiTk0qRB1imTMQaQsiqiVN3eTnDZbpOVXlot6OaN7/37v9v+ox9i1y4yOujNLfa6atg4m7YbRT5AA0eeOPHO977pb5/5hMkJW/rZ0i+MjbXyvB9UYNimzkulVLz6NS+7011u3R6jiJg0TJTAlhPnFDF6IeLalc3HiqAp4EJEUeCaTS964pMGV2+ZYJuJIsqN0IoJQVUpCPlIwXNUBhnmjFHN+7EQj8iaP/+v//7Cu96NbdvR7UnliVDkoTcomNPKK5NldhI0c2lZFFOTncGgbIzRxz75wSOOXh9DXhSLKnlikViWUGksU4dtu64+9bYnvPp1L3/m8x7TmsDs4iySYBMaeK/WSi0xLHMszFeXX/amx5+x87xz2sUgkxCrnJlnF7uukfQKD+tckuV9ER+UXGd8ejAo2432tZdfvf38i9AeR+o23OpWrcMPe+STn4Q0g2HnTH3L0MjSTrvlEmuYCMbAUG1NzkOD8hV99KGIFV3HIQaJy4xxBKpBpWIfkffDpi3f+I/PnPuDH1Szs66qMoO8201EY1kE7xlIEzjHxibKGr0nUsMEIBJVkFKox84ddou/ecUrbnHvu2G8rUxFlEbSLKsAtRLZGJc4lF18+Qvf+fiHPjE/l1+7Zf6II47fNT8XQmWshth/zvOe8qSnPGp6JgHFKubOWVWtqqrRaBBsmQcFU8JRQ1EMmomT3qDpHLbu+OdnPTNZmOtdeQX3F8cck0ZnIQc4y5qVWWEERgGwkgRoJNjUlFGKiNyaojE2fvRRd3/ko+/2uMdjapW3iXUmAnletBtZWVQSYrPZCFUZEdO0WZXekmODay6de/CDnjgxvmbb1p1KnLgMgCL2B/N3vcftn/jXj7nTXW5tE2m2GSQ7d22bml7VrcQ5h3yQijexwpatH/nnf952wfmyYycG/dRZSwhFaYmdTfOqROaChliG8c5YqMz8/MK6ybXXLM711oy/6J1vOeqB94exP/7KN/K8vN8j/xKtJqy7buVfgKXwayjt+GM+cys4iNAVjj7UoHGZHAIIKiKB1Pui6LQzLMzDWczOnfWdb53z85/tvOzycM1W0+1xCBmzE/VVIUGZwdao1kM6IAQ1Vp2LNt3cLXVy+lmvfsUtH/1wSKiylJKGsAFZIlPmgQQJJYnDRb+58r3/9uEf/uhsl7Y3Xbtl/Ya11uIWxx/+7x98x9Qq+KjEXjQkLlGlqqqstYY5FELW5Vq5xBrEYtfOdlQs9j/3r2/71mc/c0Snk1W5X5ifmWjHUPjSH9CdnhLAlgETlCUCHCGRACBJrDXJwtxAEuOmVl05Pz99wvGPfe4LbvmYxyoYjaROYQQv7dTFUhLHVZE7Z8gaBXsfJVKWmMsumDvjsU9eWOi3OmPdbr8sy6lVEw980L2f8rd/ffgR42kTIESR0hfOOTBXwTcTV83tygDM7nzD0/92cNnlYyFwMQg+p4SJSEq1alJ23nsx5DWqhVh3+l3ueenFV/Y27+oVRVjVftQLn3fHpz8NmQtVZLacOCQWxApQPWsNAFDH7Nc9bSs0fUhihaMPOWgc2noo1ZMHNagIomPq5wuTjYzFI3hUlW7fQcbikkvmzjrnx9/97oW/+k3sx4l2kjpXFSXb4W1vVI0gD6mgQY3YtAeSybFHPP3pd33SGWg3CiZqtaMQm8TB+tKrhwaFULPjPvT+/3rLv75rcmJ667bNa9as+sznPn7c8dOVh1CV1AQkAgzN7ULlE3YBKonxmjcBWuhiduEnn/zMR9761tOPPvqaS343M9Y20aeWAOn1cucO5OQQRzYArAajQ115nS9nIDNp2S/ZOjs2trnb7aXJ6lNu9dRXvXryuOPTww6rQjWI6pUnW00z8t4IVamsxibKLIq8W7U4uejcK5/yN0/r9/NWZ3L16lWvfNU/3vlux4uCDaJibn7WJUmr3chL325kBB8W5q1LsPHKVzz1qeU1V7e7g6mEEcpSfHQAGyljgiQTE73AoiINKW/Ji3d+8hPbt82/5gUvO3ztYTvy7j3/6rF/8YbXBmtMmhLboiiMJWtTLONfHR7rXs8AKxx9iGJFH33IQVgiS2TIqBA0cgg2xqaVYG5u0XcL2AYFE6/d9MNvffVL3/7S2Rf+Rl2cmE5AoaoGacMCYaTH4NHQbwaJjdXaLGsu9P7ngx8684MfwUI/SzPkBULoLczlvUVrqNFyaStxGYHxuCf+xVe//pmgs9Or3cc++d7jTpieXxgkGZIkIaot/6P3sdbFWZuEGJlBiInK3NVXw5hf/+cXvvCe95wwOVFu37am1Uoglrm72M/zMsvSAyqZCBChERqgFWkwCAZwBMs2SUsfrDWkMZ+fm8zstONrzv3NO//x79NqgO1bE0Ync+1WWlTVsIc6qvceyj6GyldEsJkxDRx/0uHvfd9bTjz58Kc89RHf+O4nTjh5A4AoKIrAjPHx8Xa7TTBVXlT9RVS57c0tnvXzVz/9af2rrra9wdqxTihy1ThscIQSESugakgTTpxzefTTR641d7n9uofed+ZWJ3QdBWM4cbBOjekWvu+jGhsFo0b8pYLwyjf6/zWsaO8OJSgQEUG05NbO9VpNhGQmddC0MfGT//nGtz7/hdkrrtq18ZL1bW1I4Ug8YRAqZ1hU+/0iTQlRpU4IEBnDDWvIMAu68ztWTc3MD/qfe9/75rqLf/nC56Zr10mQ5vgEYGpdc4zepHZQFp3x7MTO+s99/iNXXXXVhsMnwUhSA4AJla/tpFMiVY0hBOdcfUEJ5aBhdKrdOe/zn//Khz/sugvrp6c2btq0ZtVEWRYAkiQpJTCx0gG4TDDAMhxJWN9qEKEeFxCVQoiJdcxc5CWVRcu1Jg3PXXHZPz39b173xS+j3+Vmmy354CMZEdUQkyQhY4zloqoWe4tZ0igGVdZpnHjKsV/+6ueKqqy8n17Tmt0xNz41aawJPjhnokYinZ4aQzUoL//dwsZLX/uCl/ldu9al7YlOZ+e2bROdRhBfq8QhakEEUQUBZFgJeYinn3ZrMGDxwDMe+/5/fefE+NhRJ54AUXJpktoITQzt1m/QMGDeHSwPe4tWwudDHitX3UMMI6krDGAVVmFqhVZFxjMPtNjR3XTBFZf9+oJi88JRkxsakriYSLRVpNy4Kk1jIwkNDIyWCUnDaeZsYhnQqgj9ngXGmll3bidVZVvCNz/96Y++5g398y9ik1Rzi1VvMOh2e92+S52PlSAGkSjxqKM33PcBd280zfbtW5rtdHZ2URUSh8O260W0CqHi1BRV3kjT3rYdvSuu/Ng73tnddO1xa9buvPqq1Z1m0V0IZRFDSJIkSbMqHkAXOABSJGpTsFPiJeJSgagvSihFAqzJMkMQKQdjjjuotl980Qde9mKURXfHNinyLLHGEjPIkkkSVSJQI0mbSZolNm02YNGaaFZaNsYSTgQsU6snNXgNUXzI+13DylrN79qM0DvnzG+/+lnP6Pi87T3yvlRllqUL/TyAADZCTmBFjAgjKCNAS40mc56BKgfF0+9792TdjIy1jjn1FAAKdgYJEQDHy3Qay8l4WTPOdZYVHGJY4ehDEbzbKLlOVRBSZ0IIZExjYvpxT3zSM579wnRyeq6SPjd6SaufdXqNzi6T7ICZt65oNAbEA2g/xkIqL544Zgm3Gk7IlyG3ThtWJw03B4Pzv/Od9/zDKwdn/SbJ88RQs5k0My6rnjGaNZNBlXuKJnNVKPMqn1mzJkicnBojgnOWmYqiKMuSSK01ylHgnSX4spUX73zFP4YtO6bTrL8wK7ESrUjVMso8dOd71lraj1X0fs+LIgEnai3IKhkBR1AARaGoxpgYY1mWxphG6lii5oN1rWZWDs774Q+/+eEPj41NNAwnBFIxzrI1IJBBiEGhaeIA2bGwy0NcmqhqPhiwMf1+XpSFTZ2xlDaTRrNR7tgCCWOD3nte8qKPvOFfOlVhFxfXdzpWQt7vRgmsQ4drIwwhYVMZ7hvbd2bAPBBOWhOXX3ENsiZm1qSrZzacesu8kbQOP0wZEVpVWhR+v2dhpcD0/xZWaoaHEhTwEBlZTRoajj8SggHyIm+6FFExGIDdzsuu+O43vn7eb34efD7Wbh5z2OFHb1jXsHTtJZf8/pyzN/7ugsSH1WNjvruoJRoOzYbt50FTCEBKJERiA2wg1zcG01NnvOj5pz7iYWgmwZBptgXINXrNQA6ilmHJAAofSTRxjlRUdckkU6BKGsib/iCZH3z4pf/wqy9/5YTp6e6uba2GtRYWagUkGqNGYm+JDKuv6Aa3g7MyC1uybCAiXryIEoGImCwzxxhLH0TgUqRpSgnPdXM33pojU3YmXv6Od6055VTpjEurKdYqmHRpdiIAjUSlqiXrKnCAtZYsIqOoYpaYUAyoHCTwqIprfv7TL3/ooxf+7MczKadlZXMkZKmReF9SFRuOcq/Osffi/j/23jtet6sqF37GmHOu8tZdTz/phRQgCQQkEDooGBAR9aIXRQEbdvTaPrFeFZVrQ8V++eyCoFKkCygthCAlpJNy+tl9v2WVOecY3x/rfffZJwkBvBc/k+znt3/77PPWVZ811hjPeEaeap4PAi676qobb7rtxN3H0zybP7D/c8O1137k/WgnofCf+9yd7PWMs89L9uyqlFTJ8qQwuDPv+0GOHV3HAwsKhClhMdCEmc3/Y4wxxjxJCZDas7FQVBvraWJBCokYDVGUMEDtsb72/j/5o+s/8L61I4edxF4rhQZDKoRRLUqTjmojlpShxhvrdi3cNRo88Rtf8HU//H3Yvbi+sYIss3lXbD+qaW69q7LKs8Q1SWARAKoRQOOLpKqAGiu6tvJPv/p7n3jjm/XQkV1ZmqRYGay0OilBrI+x8FWFKJAUNiH+UsJCBmxkUlbVqNLMQmQGDI/HUQQuQbudWWu99wBs4jYHQ9tpjYw9WkWz/+Bv//0bsXs32u3acCRmBUFYBaQCKGFQ1YlNuyZFxGhYVj4krbTVcnVR5A6ox1hf+ejr/+6Nf/zHcWmpy0jVm8rzGERGM6Mkxgej8BGLC/PFqDy+PrK7Zo7V1W/+xV9sDqtff+X/DEV1YnX1sqde/UN/8FtYnIXh8dqgHJazcwtotcahBqxzjqbe0Dsc/WDGjq7jgQXCxBrJysQ8iKa7UETYmgCNACUuaAwS0rl+IEWSIMuQ5+j2QEbuOHT7u9//gbe858hNd2khVo0EDUHHtVSlWA9Xw0QYBUiII9uYUFj73B372H7kb9/4a9/2XYff9b6Z1kyXnIvowHSFuAyZYDZLtQ6DwRDNJL2tAdlgoPGfA0ZhfOfJj7zjfTQO++d2r64try+v9ZIWeQ1FVY+9BrQc+m3upi5l+3lnBd7XDynYEUyQpknRgVPi3HJmewv5zK48yezmqNzYHEbxIjJcHyaUaEUZt/tZ/9Btd/35H/xx3BwiiJFJI4wRZmUWskJWyJWxywkBo7KApbmFTrtli/WVsLqEEHHXod99xY/94S/+cqsOM1miwRvDxEwGgogolo21pg6wTMcPr8x0ur1OFkI4eP459isun7vqimte9I1DCLXzbKaHNEHwCKE1OzM7O0tJIjFC1bIxOxH0QwY7uo4HFJo5IdOSPaYeQEyw1hJxkBBUE+PI2hh8iIGsgQisgcvu+ugH3/P3b/zcddfx+mbYGOxf3NNvJZurS8WwSFO2TCGoaz5XQawCARGpkMbFXmu4trLYnd24+fbf+rGfeto3feNXvfhbzGIfwxGbzDIQQoA6AlI3roo0TZslZUVjPgdhqFRrm2/5p7cdO36yPS6CD/3+YpRyabQZSTOLPDFGNdQSK/EqSkjyL2XzEILEKAgCJRiFqCAqIphtHYIquXZqjKlDqOvgBRTASTYa+zXWvWef/c/veveTvumFBxfmAWZlwvQqOI3m+71ZiTEy2u1coq9G6xmo005RjT7yZ3/8t7/zGr908vzuXNzcpIlpHysTTONTrSwUBKVgbtdsHVdPrKx6xcCP9+3ZhVihnVz1guf/yf/+88hubt9eOBdiLKuq03GUTITiKVtjGmkNmPElZux38MDDDkc/AHFfp6WIsCHDJsRQ+zpxibNOJXCa1OvrN3zsug+89Z8/+6GPFCdOtH3sROm3WnVdrY8G8PVMkue5G4wHELBlIRiBajSAigoBiL3cqep4bSXLO6Eo/vG3f//YLbe88OXfn511ARKFM2CybGoVkDprJjmYZh7rhOAUhHR2Vrvti5901ZHP3HDs7kNVy4RKk7xnOAzqqgixZVy/n+ZsYtDal178l6RG4MSJRDixMMYYghEREURhUURiYWdsEqzGVCLMyrDu9ntH1tb7B/Y/9lnPevxzrjl44QWU51EIoOYCA2lMqRrzT7A1XmrE2pIaCjh5ojx5/M9+47eO3PCZ3tpmEtUfPgJRJJR20hACqTLBEpuIWPraRElweDjYdcb+u2890sqTbn/25IllQNDvouLYa62dOPmkZ30VSDhxWZYqExkCsVECExQxiPfROma3Y8HxIMcORz+gQKdnZ5uQlwDAGOO9d84lxilFiUFEq3LcaeVJbyZvdVY31o+vriaqNnXlaFgbY51kxnQ6XRUdVuVoqK02yBApQLFptFNAFQKsbW6wcitJFDER6oDufN+//vSHr3v2t3/H477qq7PzzsZ4FK2xeULGEJNAMB2GwmjIWgEgb73gJd9uv7navOOuu274zLXve++n//3aItTVaNMYSgNilNpHEwJFUQlpYjGZIfOFIUAIKkqAUZjG7F6VvKKK3rZa6tw4+HFdi+UIHcd46RMe96RnfuXDLrts5txzvES3Z8/6aJRYZ6YBvE638/TSqFB1MVA9hgYcPvyBf3jje970xpO33dYJcQY29zCgJM2i48G4RApV0ciGyCpqH4Lj1q755XJ008rx+flO4jJJ0uMnluEyVB6k1O9efeWj9jz8UrTyKMFyIoBpdnMUicTEJGqMcW4nV/ngx07N8AGGqKc85s22G10CRYkQtdZOiFw11LVJjPhgoDAW6+uf++xnjt51B8XaWbLwmytLG8eODo+cGB8/efSW25K6RDVKZFIzJIUSInEkruvQ6XQceLgx8F7b3XZkHBqMxt2ZRz39mc/8+uef/dgrMNvzrCGxNu0oiEAciZRZuUlLR8agGiUp5wpiRl1huOGXThy6/dbPfOK6E3fefednb16+6xAVdd9mnSRrWaoG63RfHD2Z/3T6wasEMClN5sCqEsiosWKNJMlm9AXBznTmDuw7/xGXPPbqq8+57Ap05nVcUJqg24MxdVmarAXjgkw/n8BTdQqrSFFaRJDIsaPXv/Pt733DG+769L8n3s/lWYeNjsZcq2uGRjqHxBRaqqoptWUSJl0tvd+V77384sue/KR3v/M9cngtrA6rym/myWuu+zf020jaf/Sbv3PNV3/13oMHMdPxIkFBQikZIiNeRGBdCiblnRaVhwB2dB0PLDS6jq35rpMmQwBAjNEZC0UzjwqAxKiqXiXLEgB+PDIa2VkgQGoYhXhZW2FRRMbtd731j/7s2ne/09WjRKY9xQoio2Blqirf+Mk5aztpTkRVWW7G6Hbvv3s49Kl73Fc942u+7UW9Sy5GYtDuQEnIqoLEAsxKqogEOBRVhVB1Wy0/2HDOoS5RFEgNxkVc39w4vnTs9s/d9unP3vTvnz5+5+fYjxlxi4l4erROEvH3OniDKpgIJqiIEqyjNKE0vfLqqw9ceP75D3/4whkHuJurc+qMmsR0Z6rhKO31QlWCjU2z1fWN/sz8hKMBQBmhMT01EiCCE8dv+uC/ve8f/uGWj3yYBpsziWtbM9rY5IB+2xmhYrO2TMa4UVUnLauqUsTcOWPM0bLkM2YueOrVL/zVXx5+5uZXfecreK1oZe07Buu/9JbXzz38ErTaxcYgz1ve1yE1Sdr2KholtymEEVViZJeAEFVF1dmdUPpBjR2OfmBBT0/Nbj87t27ETwNBNYIE0hjrB5BECkAYD9a7nTZOLKOOuOXON/zBn3zqAx9KY52RH49qBbI2M5u68G2bqQgFoW0yoEajHcgsj6v2/DwSd3I4MDMzT7zmq5/+DS9oXfwwGKtZhqwVwd5HazLHkAhRtZZUodETKVmGKiQg1DAEEERRB0SBKEjK228abKycPH7ixIkTS0tLy0tLy8vLo81BXVZEZEBEZJgnniVkzj73wrzdmZ1f2LNv754DBxf37ksWFtFuAwAbGAYbEEFEVUTEGCsS2FqkDkS1r0XJJbmoMDEDkJo0InqEGhpv/ad//Ld/eNMNH/qYKcs8COqaSQzUWWaaOG+oKgUgqgqY7Vx/Znl5xUfN2olvp3dI+bXf/91P/4HvBaU3//5fvPG1/7soqo3E/Ozfvm720oeh3QYxBKDGzaOxtSMLAz1tn09H2OyE0g9q6E4++gGF5h7+i6qgndYZzEpxEneSKJjA3VYvnFz72Dvf96F/fNuJT91sNkd5TcWg6iy0Z2ayldEm553oQ6m+GpR75ro+jBqdswLxlPllnGsnUg4ctc7t91aL8j1/+Zef+PCHz7ni8he+/OV2zx4A1ibWJSH4MmiWJlKG2LTcGBZtFBeGrRF2ggiAHXNKEA1VFasiu+SSLPrFGC8BQAwiREGMCBFEUw9ATPQNZBABMmCCIRgrTJGNMnOSEgyYiCZJfVIYAUiYIiRGX5MxiXMAVfUosU6quh6PTQi23QLiDe9695v+95/4u2+vjhzp1jqXd8AxNgF28CQgUjAJiRKpU2JwhLHJ8uagiDoz0wkqhQ+t/kyW92BTjKr1sj66vuHSrLdvz+y558MmgBWAuOljp+me5Mk+3eZit9PW/RDBDkc/8HCPm9v7iKDp1AOkRgkRpCQAEdjFiEhv/KPXve8f3nziplsXOO1Fk8QsIezdv/vwsUOd2fbcrvmTaycMNAK7dvVGoxGzAGj89iJDmlmuiiyqVHFtWNtsvTUzj9Rt3H7HJ46e+Mg73v3Ypz3z2S/8xoVHPALWILFJp13HwuSpj15UmVhB3nsGpTa11oqnEIKyptYaQyZz5BIfKnIJJURETNtW3YdTKznNv0MZaQoiTAyKFAQlFnAtMSIiQoVYQUQMNiShGlpH1lq2RlVjrDV4G4VCbXwwVelPHP/ov7z/zX/z15/7zA272qkZDeYSWmz3OMb10TBEYWekjjkbQ6qMaDgaiqTKUY0poWJtyCK1cr8xKEOdz80hEIJBZ+bmo8eKVj5QeeGLvwV5jqylk3uU5lrM0xWehssT7/DPM1V2Bw9G7HD0gxP3OoG5CYENRMAM/pprnrdx19F/PbK8dOwY8u7C7Mzx5aW7lo7P7dtz6RMff9a5Z8V69PY3/+PKkbvWh6N+6qSMILCiMaIzEFawoirR7yY2hpWB1GsrrX5vvpUWvu6zvf4tb7n23e8974rLnv6C51/6tCehlSaGgCpK8L6GSxOXJiaTGGL0CI6ZkySJUFWNEczENgkKY4jYAuKjSvCqBEiadEBCOuEyaHPfj+g9GESkIGIGE4EY0CjEZGHZgaelNoYxJmUWAkGFfI0QoAoQaj+44cYPvO1t73/rW1fvPLS31X7M7n0yHtTWtK2RuhoMymGJLEeWWFFFFQlgQwBHghpEcG25ULriMY8erq4uHToERWpTX2M8KMEJ4ObOOqdo50/5ymdcec2zwSRRYCxO+fQzbQucT2Enu/FQwk4++gGGe6ek7zOOPvW/ic+9KiFCAGGIk8BlBZHi8PG3vOHvPvyv/8bAueef8/gnP/kRT3sqyhIuwWiMW2/929/4jVs/8hEZjPsJm61y5LSpD4BN7DgETQy7ZFhW41KYyNqE4Gya14Y3vN/Q2Du474nPfuYTn3PNwnlnUauDNAFIIqKSEkdFUdZpmqZpbqyFapAoIgDquk6scy5lhiog0mSJg4+AEJnmN5NCWQkwk4SMKhBFRKCqqol1RNQss8QYvA8hxFi1Z3OiiBCp8hDFeDS47Y6jt9zy93/+/y7ffbhaW+0akweR4SiNmljmjMmAQSGEyJrnmQHVw4IrGIAZkal2VBmuWcfWxP7cj/3UTx++7da//v0/nPWm2+0f8f5JL3rRV/3CK8eb68MjJ1/7mt/7vh/9wdkD+9HpVXVMkkSpmbWrrERb2pXtQnMAQIRQw+P/Fw+vHfxXw04++kGC+7rOyjYmNUAEzDSJGZg5Mdam+YHFr/rWF77gu7+dum2Mx4EUrQSdHOubCOGTH/7YkbuOrq2O9/W7GupIsTGvaPoGG2WecclwM0gVZ2bRydKUYyN/s8x1Pc7S1mynPdJQDTY/8jevf8v/ft05lz3ykY+76glPfsrcuedyu8MKGFKifLYLJaioLyMUBGZiZsu5Maa5t9fG8DoqWMk0vRtExBM/bQBA6UUxiT8NyBrDIEAoRqhAFaIMShwlaQak2NxAajEql2656ZMf/PANH/rI4c9+dnTy+EyWdGPsachVZ1u5sMZhkaap58kYAbaGmUUk+FB7JBZoxH+kmFb9Aun8GfsXn/T43q7Z4k//NIvacbYYFr2ZWQQlk+466+wX/9D3zZ51ZiniAOvSyd6cJqJJt+VzTifjHWp+iGCHox94+CLFVk3E3fQ+sBKRKDEAgYK4qsUYok632+lMZBX9GVsWGI2Xbr7t1o9c9/G3v+fQJz5dL62dsfcAxVApACEEgjY03fgRD0fj2dnMGDMej4tCE4MsSwhmc32z32+HMKpWN9PEGVKt6t3K8brPfPBDn/rQa/5033nnP/yqqy696rGL559HMz2UNbIMWUJMVkUBNUygoDWCUbBqBBgkRM14awVUoCDRycIwKZxGBqQZlEUigEBUxShMo3ZGhA8Y1L4qTBnu+ORnbrzu+n//2LXLhw7zuErqIvfVLIxuDpylVjuJoV4+sWKBXsuIhjgOAVSr1ApjkKbM1picAiGSMhFr4+6tjVD7rAsvRLeTnrk/37cwuvMkV0VtzdkXXISgedYC8xkXX1r5QtMkEjNhEkJva/S/L8jWYbDD1A967HD0Axv3c4rS1j8TGcMkdUtQQMhYUBIRJPpQ+zxvD9fX/vrP/uzv//TPZmB5ZbMbec6ks3v3nazGw3KQtjNGtMpJDFbJijZ90sYwmMZlIaK7ZttMNNwYjmvMddPRaGRS023lZVWFse+l3O91B5ujnkmi9xs33vS2G2540+v+LN+92N27+9FPunr/uedceMnFyRlnIE9JlGoBG8cM8RNzJiIwgRWiYAaJUYU2ajcVJYIYIojyJNkh0Di5WgUPAarKLy/fcdOtn7r+E5/+xPUn7zhEG+MsqoN2VV1EItGpWkRlhfpqWANotxE9yiqmiaadLCHDPsDXBsRsNFAdI/Vyz1AmUiRBOSpDmOzBcy5AFdDu9Q+cdej4cOQ19Lp7L74QxiJNAV3f3Mh6LWNSgKNq05e0w7w72MIORz/AcN9n7/YJo9P6/yTc5u2vYgAGDEjmjAqg1oDZJcFrns+97Lt++Pithz79vg8SZd25mdWVlVqrg5eed/F5+y951MM3Vk6mha9OLN/20Y8NjxzvJO2EsbmxIYjGGedcrZEDWXY5vFXTbediWUhsxm1noo+rm6tJxqJB4BimQ5zFGI4f3jhx91s++kExZJNWb35+/1nnnHvxxWdfeH5/765dF54dnGVKmRIYC2tgFawQDwOAQIaISEBKQQMkGgAwCILRaLC8vH7ixHB95dqPfHj58JHVuw6NVlel8iySwCwYY2J0BKMCIQoC1QAB63Ac+n0ebohlaMDevf2V5Y2qFuIi73Q2RmUry8ebxZmLi6tLK9bapVFV91sv/Ykfffijr/zV7/gBHpQ8KOs6pp0FoIXZhQse9YSPfep210q//0f/B84+C86qgZBpzc42wT5v85ud5HF4uk+B7bKOndEcDynscPSDC3TPP+99u9zobbmRdYk2ts5EUFZAfvrnf2Xtrjtv+NDHio2NRz7ikl0XnoOeQzdBLJDnOLq0ee31N3zoY+yStbUNBMzNtazjqOqrqvZqQXmattrt9Y114wyIoSqkROoMuJnfByGoFWPVRChEBLqnnQeVKOqXlo8trR39+L+/PzG11U0TqZV32rO9/vzM7EJ3ptfuZknKvV4XJKoalaKQDxJjFJE777qtLMbjtUGxMaiKMcrK1t4Enzjiuk5C3AN2xhg2GiL54KUiIkOTjhFVFUNCcWF3qqrz8yhLD+XjyxsLix0vYW1Upkye0JqZYevuOHS8nboanO9aeO63ffOlT38aen3pztxx86fPXNjDtV8fVMi7ED370ssOPPzfn33NNY965tORWDAUiNRIOGQifCcomO5xJd4Jqh/a2OHoBzm+wAnOBAWBVJVAIB2rn33EJU+45CL4CqlBqBAL5Bk2KhxZxtLGH7z6t6u10YkTg8dccP6hu+7IsvbS0tKoRJLAWjg2ZRj5EZIEvY5zzo3Ho7oQA5AiCGzezCFXVoE2amsRQlkMQQYMR8bAa6hDFQmx5b0a683Kuj08tNZaa1gsY21tVZjUGDWWmVmZFQzJEleHuvTeh6CQBGhHOFXSkBrODE+GBohIpVUN14USAYabsqRCIUqihHFRa4SzTojHZfQJV4XAI5ah0+vfetexXbPdpJOPvaduu0jN1V//fOzbjbIaS5zZtcsrDet6/5lnQgIMPfzRV3z3TO+iSy7mmVlMWlTQdL38JxwGO3jgYoejH7rYkl02f6iqEFpzs1DxCDZ3ZAGXohBU8pF/etfG5+78wJve7Fc3MphLH3HxZ++80zsel0U2N9cVoK4p1jEGy9yfT0NdrW2WRstOzr3c+cpHRTs3ZYgAWJQgpGBSAiKgpEpKGgiaEhsDsobIum43CrxyCCFUVRgNJHqRcObMTIBGojgxja5JIoma8bhnDKxDnoIVQUxVsw9sjGFoCCEADOeQ5TZLYiE6DVzFNIltVSXaHFZE8BWyjHxVz86my8ub1sIlZmNUmNlWf9+igWnP5CeXjsNZ0+1hZgYQkKmcjXl6ZGl98dyzLrnyCmQJmFODS66+ClHBRqMnk6AJnhWYFgl3sIN7Y4ejH+z4PMmOGCO2UUND06TUuE24LAcU0W+eXH3fm9/6D//7/y2PnFh0qfOhLP3Q6A0337jvwnOe9Iyn9GZnztt/RnX4xMfe/q6jn70xVWZotVEkCWUWFOALCSIREKAM0WSAElGTHBea2EJJYo0SQxQSEaoYogSNCk5IiInIkkkJmSFJGJQOR2vNXG0mIoVhRBZD5GDYe6nqGGOzUhMVceJik1ewMAY2caIUYmCBkkZWEAxgQY3hXzSat7O0bTbWRizIM7M+RHdXFk2yvrq5b+/el7zkZTf824c/+M73Zu3uaul37z8AIWQd5LZ38MAHP33j+Wed++xvexEW5mApqph2W2IgJjIoK5/ZpBnXYBrnjS0vjh2y3sHp2OHohygmsfM9HiRIRIxinVPEuvK9Xbuf+7XPv3D/mf/+vn+r1laP33XXwTMO7D/vzCuufmxnzwL279YTJyjt4I6jd37mlo998No9qTv73LMOHblrIHWeW45aDaNVdHJD1o1DHcHKTScNAEBgVBRU+yAGTMzMzrIxMA5C8KQCImVVIYkRKlGFkCUWohynumjWaECqNkQmJoYSRyVmOENkTaVeoAKIgTD5EKRWX6KVTMzzFKwkESJEAFya1CG6JBkW6GUoB+NeCsdupfK9s8/4mpd+28KzvvJJ3fl3vPVdM1mPtEryLvKOH5eu1dlz/nkXXT167nOf9/hrrqnqyibtSJEZQRGCt0QuTZul3lLXbTWi7FD0Du6BHY5+KGF7RXGbxmu72Sk7C8ZwOGZGq9X1dRnS5MKnP+XCpz4J5RhlgZkeQoAjVAVqodZsfcehf/6TP7/2vR+6YN+5xoRP3nm3T1VbychxGpRiTAqUZUysJsZ65jBRmKgRwxJJSRCN5QCpSbyKj2CBDQCQtKxChZoMBEHB0szf0ghEFoiSKslEnFcV0bloEwdjIKh95StI4Rtpd3AgB8MwitRxy3EoA0eAOBJqAwDMJAZlVUfmIhQXXHHe0p2Hw6hsp25zVA4T96znfc3Dvu5rUXgs7LKtXhUVNl3Ysx+CwgeYcPkTn/DYJz/p0isehVYrjoo0SQi6tLba6XWFDRtnQM1Ql+09ojvpjh3cJ3Y4+sGOLXnHPbrUpgFo80fjwgFAaiWiJEnUcARzkiXWNPaYajOabQ/W1rozMzouqN1H0I++4W/f+rq/PPzxzxzozmyGsD7Y5L275g7OP/Ern/Lohz+85cOH/umfP/aO9+jmOHGurIMSCxGIhVSboa4qCo0xwsBaq5ZYgShURwTUZYgE4ajEVolARiauzgqNUDCxkiEmthY00kIFHiIkQogEx2AgT60n9Y7FcYixrAKiGsMCKJGCI1FtIAwHEPPmZpw/MDv09bNe+I1/+do/iiAIj8eF27X7idc8R6OnXvvY2sq6qFVdD9UlV1453Bz1ztwP4IonPjmMRnXiLCPrdBUofT07O+tjsNYBiBIZ5rRddJ++HDvYwQ5HPxSw5dy5HXT6Q00SFgBZAoHZRiAABmQoIUgItXH5MIw1b4EM5Z07r/3En//W7x768MdnhPfPLPoQCmNaF573+K971jNf/A2wDAJu+dw7P/SRlaOb5y50xNk6xO7cwtLGRtpupWl68tDxfrdVb25280xCZZytpTScjEZVJ3NKYiwQUUdN0nR9vdo72/NlBR87vc7dK+vdGZvn+dKJUTdBu53HgPXN0WLLjn1IO/l6OXTtPIwLZ00YRPaBFUURZ3e3S+iowK7FtlRRnBOBNfnKxpr0kla/u3pidaaT9uZbtx9e+epvef78k68++Yd/oAabZY1utzKutbib5hbAyTuv/WiY7Q02NlYjzrv8cjM7MzmhYrB5LsYChogASl0GIJ2M8wazmUbQE6UzbfMh3cEOtmNHDP8gx5bZw/1YZ20RNLa9ePrGSdDKbAGyNjdpVtRh6ciJOw8fm5tfFGMGEu4ebi6znvOkx730lT/1zO94KTqz6M1t3nz3z//oz4SKs8xlcws3n1hfdrh1tDF38QVPecHz0Gm5mVYg3bVvby0RaXZstajJcJqZ3JYSx14Lr5Xo7NxMWdazvXxtdXO+N7fuw/rmaG6ubVwyHo8X57Nu3lpaHdWlv+jcc8dlyHJ7+MjQZSkz1zXIcN6xYCOE/ky2ORgtLdXzc9loWAyqam1caKd7rBiff+WVL/3hHz6yur5rzx4fIZzuPuvgNd/0IhDfcng5mV2gbndk7AWXX0G9GSRpMRy+/xPX33ziWHZw33O+9ZvNTF/TNIA8oMQgB7DI5P6lOc0YzPfvyr/jb7aDe2Enjn4IYVvD2gT34QixrYNi+5PlqGx1W4Y4BMmSPD+w/zF5+6qrrkr+nx9DUQCKXgdEaGdgHt3+uXxY/eEr/9fyJ+80xfiMM8687diR9oX7e2cd+JYf/oHdj34MDh/5+7f+Q7G+esneM47ccXen1YKx/V29lXJzdX1j7+L8+smVbmZaabayOuKybPe6CNpK07LyfZd152cHOj6xtDHb5XaSrS6v96wxaXbj7befcXDP+mjtjPNmjy8tia86iWnnnbXlNVUER8a6zWG5/4y55aOrC73WcDAunK0lXvi0p37rD3wfqlGnO7t0dDm32clqdNUTr+a5xeHq8vzc/Lioq6ir3r/wmV+JXr+qwgc/dv0jHv/4l37H9zz60VdoatPZudpyEEBhtRkmADRVSwATat4eEjXDuLZNT9/BDu4LOxy9g3tgwuQ0yWFPiLrV6QDkIFFNY5FvW61kLkU5rJ2ItVm/17xyeOzkO9/yjnf86V/JsZU52zad7GRZLTE/6QmP//Zf+Tn0u8Wdd63fcqsziY+QEE2a8tzsgKPk7hXf/xOj4ebP/9gvf8XFZ4yOL60sjxbO2vu5u4+1EHLjmCVWYzg+OVjPd83afrlmrcBEAAEAAElEQVQ0qgbFei/h+d27j5xY1sTeubQkCQ1PLjlrFjpdHZVrJ9fzVmdIcSgh7eRf+4Ln3XL9p8bDMgYkScvtnj8p8cU//ZOYnUVZlLX2bUYuDUzPe9G3oT17/b/8m0s7GxubnGXnXXbxw7/icQhi8/YVV1759KueBO9hqKxLOMfNJW9S3AS0ce2egO65gb/YiTo7eIhjh6Mf5Ph8t9WTfPS9nxYFxICn0/KmegPRuiicc6lJEVEFX/pA1rhW2+XpONRjCcycwd51x93/9A9vXjuxfObsXFlV7GgD1Y+9+ncvfvbTIBU2ytx2XvuaP9m4/cjZ/V1Hjpyghf4dUnzD937XE77mGuTZ63/5Vbt290/cvbTAWW9x9vo7j5776Iuf9OSr57vt1/3m78UQ4Snr9e7cWJ3dt7saDi+7/PKV2+6++eZbWy6bnVu8e3XZ9rrZbOvrvua5488dfetfvfFAu+cDucW5zWLtRd/5ksc861n6x3+xdmJNTqwTsBn0ud/2rdizCyKbtxztzCwkNB4Hv/uih+UHzkCrd/z4aqCkt2d/Pj//sh/50WzPItqtqvLWZoHI5hkMJYlVAYPARIZYtGkIAludbL/TbFOmmNL0Thp6B58fOxy9g9NBQgBUDKAkmBjmgZjZWrYWQB28tbbf7wMIoQ4iYEtsfF0nxlxy+eX/89W/6k+ul+sbhqg90913ycPEKFo9EFY/fv0v/eAr4u13H9x9cHV5ecNQZ8/iL/z6L/UvfzhIcOj4sbuOd7N+12JzY7C5Mfgfv/GrBy45Lzm451NvfH1lNXrf6/aObazqrtk1w7/wu7/TaXd+4Pn/bbbbtmRPrK+0D+zbffG53/kLrwTzv/7Ga8+/8GFLt92teXJ8sPG4r73mMd/8TdB42VVX/fNfvqErFJjGwOO+8lnYtQvD8n3/+qGVzVHf8F1r69//9d+Ag2egLIdKiw87/xtf8PVnXXxR65KLYViitNI2MkQfxnXRSvMy1C5GIhi2AJqJKlFhiO+RY76XvmYHO/gC2OHoBzs+T5/hfT45fUYmxkvbpgWKwKZJBLwEH8UJOFJVllkrZTZqlIFgtRgXqer+i8/HJQaZQ1EgyRCFbYr10Wt/6Zfe8od/dNUFF24Ec3IwXLjgXHH63b/y8/1LLwYZDOrf+qFXrt1yx/LRI6NuX+Z6L/7RHzzneV+FrsPS8be+8+3O0Hgsva4pLF/xhMe/5BU/RAvz/nN35bNzoxMn9+6Z3VjxdSf7zle/CrmFTW49cfzIxvr83t13rSzR4q4XfPdLkTBqfOS6T5BJ18r1bGHusic8oX3h+SiDDMu/e/2b+v1eldnHPvWqJ/y3F4AA5779J38cJoIVLse4RtZZO7nc6cBYa3MXI1VQZI6EVREUTDBNwmNbeXBrO8t9+T7faxTlDnZwCjvFih2chqmFx1SJMC1nMbMPfjAeMnMrz4wxZDnrtMqilhB84cuiYrBzzhOh26kclTH4do40RZKA+PDho3Ozux7zmKvuOHJ0mGXDXu/spzzpF970hn2P/wrkKer4v37iZ+666Q6tsXffuetqHvmsZ19yzbPRacMkN133icO3HqlHdaftBhEHL77kpT/6P2hxFzhZXdlYHQw6CwsnRuXAmR951f9Et4tuD6W/9tOf3WS+dbixnidXfe1zksVdaM+gktf/wz+NjKVdcxvOPvfFL0bWrsr6p37uF+387DEfnvbN/+0Vv/armOlK5jDTi5bRaY1UQwjodqAyv3t32s4BrK2t5Vmnlkiwhh2AGKNiYvoBUcW2OVf/RaCn/+CLEZPItt9fzDeIfuEXy/RnB18AO3H0gx33G5jd+0ki04wPBwh0WhY1YeesA6BQa5vxgsjaGQA7SVsLpdYluQA2aTfvi1GJDCsOXPSwb7jwPLzsW6ARxRjtNmY6IAETrP2+l337+s13pCEMqpC69MrnPe+5L/kOzC6C60Of+MQf/cprdMW309mk09pstX7id34fnS5mFnH7nf/0V39nxa4XMevNLJ61f+ayK5AlGBQfe/u/pb3FE8eXssWFmTP3fd0rfgSsWBn91q/99kathTF5p/OyH/7BxYsu1oAK9LIf+sGWtYuLC6aTAQLbDNki02oLJG/PTOunBgpATGL6aV9VMkqgUBVDbMzkVdbaqQL9NN3zvYR3/IV20f9V3LP3/74evAdI7o+j76FIIShEpmNiJon4+7iTk6mXFO6ZrN/BvbDD0Tu4N/jU7/ucobetd/l09d6kxihAE0I2g1IVEAJbQBgLfZRj7FkAc+mr1OWj4cbbX//3vbnFXZfNzJj0wL6Dlz76yoMXXZidfy5Ss3T7jT/+oz+pR1fO3X9GqMNaDL/xp3+KMw4iz+vb7/joW97x/nd/YPfsPGfJMV/++s/+LDo9FGMU8kd/9v8urW50F+fM4sJP/9ZvIGvD4Nr3fvjDH/24ZK3nfM1zv+KJTzz7yivQalOedVxm0pZxzuQJiJphWwAUzXSxLQ2GMPi0lnrd1i94f9OtTt+A/3+B7iux8nnLylukvO14+EKQSc79879lJ5vzJWJnLvgOvmRsP2a2WKmh7HsZ6Z3+NwmRiIYqRDKWmS2xIZLRiBWwBlWNUYmZWagiRuSt3/zpn7z+X/81rq+vnjj2LS9+8Qtf+u048wzEiHH9x7/66nf91ev3zc1ERsjdS37iRx71rK9EVd9xy23f/7LvzpzLTPKoKx/1gz/1E8gTAEhdeXI1I0anh1DDWeS5RoW1qoaZYYyI9yEYS8Y0MxMtJvQ85WjZzsn3xPbuTbp3c+d/AZyehTiNQ2lr1EDzmkl980sL4yIwuZKdfkXf+g4A2JYMIfBOHH1/0B2O3sGXji94zGy5f9zDohqQqIGtCSoALDmpfUIAGykKsqAkCT4wM5OBVziLzQGcWb71loULz/VrK67f8wRn0+W7jx268ea8CsvHj5994fn7H32Zt+L6XUT2y6u+rokoTVOvks7N+vHIJSlUpayYDZhDqCl1ptUWHwm8naNDjC6ZEPGDkqO3mJgntwgAwKdIE9vSGl8yger2uuh916u3ZgQ3D+xw9P1ih6N38B/AF3PM0FQYvJ2mlcCMOtbGJFGi1LGVZFD1w6FL01KDzTMBRcQEBlXNUSl1CAHWQANiQJooaGNtvZ3mLskxGCIKel2tR9Rp+7p2QWEdYqjqKu13a18PNwfdbtcPx61uD8wIAYZgTQTKukxgnLGiRlWNtc38LSKd5jq2okhpmPqBztGYzCrb6lCa8OOUlU+jS74fv6f7io4x/cD7iKBPvVKmQqGt3ssdjv780J2ts4P/MzQDu7fjlD2TgvSUrx6YiEhVLVkGWbZ5loMJoq7bE7JJ2q6VKgll7atQByjlKQxXKjBms6xDmhUhDsuyOzPjWu0q1Gh3MDczLscrRaUgthbM6v2grJJ+b3U4Upe4ThfGtGbmIFIMx8rkidfGwyJGl6SVr1VEVWOMKgJAREIM97myD/yzRXB6Fnra+c98P1xJMvm5z7Ih3aOieP9SjVML8MDfmP9J2Imjd/AlY/sxo9Ms5hYvT0DUCD90cvM8eUJEYowuSYNGEUmMgyDUUVXT3BZ1TCZ5BimrghUQbbU6a4PNbrdXxlpVM2sRozNJWZZJkjUDCgUQBAsmhXjPSboyWOt0ZzdGa/32TAKtR2WaZwAJYgBFDdYk4iuCJNZBrYiwMSCJMRJvtXOfiqMnvx/YcXSzHwSnKp+Mz6/smOY9TtHuVlS3bd1k28yvbZ4kn8+EhASnheY7XH2/2Ml17OA/gFPpi+lJPKnibw+haJrpmHL05EmRGCNbGyQys6rWdejkeVGEPLVFUbfbSRUiGQGJiGSUEJkIDIsiz3OFRl+m1kmI1ia1j8yWLaq6IiKQWDXW2sKHJHEjXxtLCTmKwQSCRpDxCOQSIhUvaWIbApKgqmqMiTEKonNGoVDW7erDSV76/pj3vzxHY7IfJq0195LufB6d3PSZU2qNbWlr2Xpn4zOOL0jT9/jSHdwPHoIc/aWv733fu32+zyG6vzjri//8z4fTWeO+l+c0pvgSV/dLYpZJHfB+akSYnI26dULrNr83YMKDCiLQttSbUNTJJELQ5DVbWj8hbNnFmcli3ENfAgid8vkzuv0Lp40lzSMcp+yz1RIo27bDvTj6ARX3nXbHM50s3EhWYozNH957AM45KOrKa/RpnoMIUcA0KsY2cYlLprsPTGAgBrGGAUH0opEsEbGohAhnU4kxhJAmCYgQAebmnbEKXqIxhoiIyLgH0sb8/wc7HP1FYIej73d5phx9v2WiU2wuKgwAQlORtYBPOXieupduptQqTd5vt3PoxEgEgFpSUgKgSiJggHi6DDLtaCed1r62LQxN/1ESnFI5T5b0vjga265HDwzcm6Pr4IuiyLIsSzOCxBgNc+29qhrjmLmqQ1X5JM+cNWUdkszGiaJuoq7WCAlop5CA6D3UW8s2MToRxRtRcWQABO+tMcSMEEXAxjQtmGStyuQi8SWfLg817HD0F4EvtV31Sz2HH/AcjXvcud6XCcjWYyrSCKUBkArASlAwiLeyCAwlFZpE3KpEJE1emJrZhQIBlEGsk251JURqHPUN672j6ilN32uZlZqrzEOCo5UQBMYQqUKiaS6t0Ucltm7sPWyqxOMIrxBGiFAgBIzHcTAY1HWdJMlcvzfbZRORMBILS1AgSBREZk5gATCgUO/LxFjLxteltQkxxxBUtembB0DMO0mPLwDd6TPcwf8Z7uMU24p36b5eNuk/ZtApnS5BlGjS5tK8vklJUPP4qfcSSCfO1pP3gwTKp8bm6qkpBVN5GbBF1ltMPUmXPOQowhgmQKAheCYVEesc2WQYhV0+AlbHGNQ4srT68U/dcNMtt955591VHYrxuK5rZu7k2Uy71U2z+W73wrPP/IorLrvwggPtDoIYkEmBGmCg9pGgucsVWobgXAJQcwMkKoYITHTqGNjB/WEnjv6C2Imjv2To/XCfNKErTlMLEJr8cPNGPiXgE2EBwGIx5WZlbUJmTF7MTc/MVraYFKQTBbCSNsTevGb74wICtOndIP0i4uhJ+P9A4pR7x9HDcelcmibGi5R1yLJEgJVxQGaXN/W6z9z83g9fd/2Ntx1f2YTLuzOzg9GmEhjEzIY4YZMAmWL58BH2lYt1KzEzs92zzzrwiEdecv55Z11w5sHdC708hwRMpjdqdIYk+MQykRIRgxQqIsyNS+ADaZP+Z2Mn1/FF4L459PNxmeqXynEPcI4+vcHhvvIcp71y22CSCRveu3vCNATdMO+Eo83Ep40AgrDGbV/B2zIaW+u7Pd2hhO0MjqlkUJsPeyhxNIjroKO6CmqytlPgrpV44x13v+HN77jxc4eXB1XWn4sm3xjXtSgbFyjaxDGzhBirmoK02HXYzKbpeHV1vLok1diwsgYfyhjGs738zDP2PvXJT3zqU64+5+zFdgKJqMZVlljriFSI1RBxMxydmXc4+v6xw9FfBHY4+v4XaOudp/63vZX4tFoiIWwFyzpJSG93kDBg0smYgSkjq9L0kekXNSH1qfz1Fi9P3ra9rXnyUKTp8ui09kXbsstq7rEGp3H0qXV8wHN0gAaYShAMVoZ474c+89b3f+SOIysrw3rsoS4Du9orGZPnbbZUQ2qNEpREHaxTchGZj3v7M+PlpcGJpTAe5onJUiO+HFWbaZeqelgVA8P0iEsv+Lqvfe5TnvS4hZnUGBgCRBTBOdPUJByZHX30F8CDgKNP1X+mf21fH7rniyVO/dhI72G2ftp22M60pEKnv0IJvM3vbHsNKkAbaQHdM6Y87ZOlkR/QVOm1lSfdllq9r3dNwMA9ymLbPYyI6B7rddrYbzr1e9vjU5kDnVqX05fg8+QvtnG0nn7B4VPPnPqEClDAbhEraQSmbpZsJikPOvWZW8u5/YtO52hMxRvTZ+MkU71tE21PpPCkrXtraZv9tX19Pl8cfXqNdNt/7ueo+zJjeocwWRg5/blm77MSomokHgW1lo6s6z+8433/8PYP3LU0kHy2QuY6c0J2XPq6rjOXMFH0NRmuWMGkQohCom1jZlzed+mBmblqbX310LGV40fqoshSl+e5Or9ZLJEL7SxPnBlsrtXjwWWPuOirn/W0b/j6Z2QWDVMnBqqQUBNpYtP/1K31gMMDnaMVKKEAkggjAIkAoZHSilpjY/CJdd5XjZ9vQBzEkTVJCougpNZZo4TowVO6BGsjCNsaHGoAjjKRFaE53Em8JIltWMADAIwiiojl9cFwvtuxQDkubOKYQcwEjMfjNG+BeFCV46Kam+lHiazBMBwsaSNxYmUigxh1UvLSibdnIG2iPwNKIqxSU1JrhGhlWbZaLV8FY4x4sc5J1OajVJWjEhFUy6qCozTLAgIA07BW4/3ccJmAoFaJ6XSyEVFA7hXFk0SQEEwjz9DpcCgzfe+UQqTpZ1n3hXWZ88idHY0KmxjjEr+15jpNdGBCxEPv0zRxAdYAAo2eUoOJkICJjE67JsyWZhqnvloB1i1B9SRwtlNGjoRhHZyzGcE0XyoijbxPgrWJqhLbEARE1pAPUoSSGe0kI0B9AEA22cqlYGqdfGr7fHnjRAlaA+CYcrOqzY0KiSipobKumU1qU4UZSajY3r1Sf+T6m/72DW+9/obb2/MH0tl9G17HkdQkE5tVUUaERkYE4JVskgHiixG83704N+PSY7ferqMijaDSJ2SSJK18rHzNjuCGQCRi0qZ+q5a8QTXTtS/+lm/4uq99Zr8DPyq6ncQySfQ+SJqmUPZ17X1MksS6afQzvd4oTm3Uh1qN90HC0aywMuHoSAgMIYSqdsY6YkMcxRtrq7I0iRmjCqFOkbaTNoTrOrIx1mLaIwuwKBCbe2wghGBUkggQCUtkQBlsnLG+CoiwqRWDEMESmXnsAzvbIqrKylq2iat86ZwTEcMmxGCNraNYM1HU1GFsiQ2IhVUgUCVDpnGLnxyeTRAZJsFgNIJEgBh1cjaqEAyoCtG61Bjn65imJgSEEF1qYgiJsQDq4ZBIbZ6WsUrTBJNIs0k8sIKFAAJBTVRSjaqK2DRkc2oBTFz/t/enRYEKeJoWpmk6QJqQVlWjkJAoqwq0JGVmGkmetxAjrIkEDxEQTcPhrdxFZIxDSKxrA7H0sS5c7shQRDTGKkgVKiTCEOIIVeGUQMIKZRJiAczkWttch1RUTGwmpcQA4iQVgfUCjYaUjGFLzZEQJfo6ZlkrAqIggg/qLJV1mYATl0AVIkoUSYknLez4z+boEmCOKSuBBQhABKH0gRMHsnUM0KSOtOqrj3zqlte/7cMf+9RtQDa/54yTm8XayHOr44nJpAoihaoyFOoZIqTBC1vLDJFoKe6a6aVR77jhM5lQTrZFVqo4KiqXpkmaV7EULQAhstyMwVSxEIMqz8SXK49/7MO/6yXfdPkjz67HpUHsdtuAeO8BNsQTWyugKsu0sZPd5tPUWPQ95OTUD3SOBhAFdKoQJHF6dxxq30rSclxYZptYgKrRKG1nYEGMiAwYkJUYRdVwcxcbAYBF2CiRNrlRgFVABCLwFm8ywkS9K0zKlgFEUS9ME/KNGowlYfIIlnlcjltZiiDVaGxAzlgiImeR2OlnElQVCqZJaKiARqg2yn9sJTQAEgXJaUGFSaAqNlkdjZK85euQJAkELDGzBjGEsnAEtgYUAIVtWkkaezqCalQSqM1SNJ8PESg3cgmJiFs9xI0qrkkY8ESh3CwTk9DEmFRjIGo6BJlYeetmXAkiYbO0eQtkVMKYtCZNshYAAk3jr8lWEUKsQwpxTIg1KMCoaOCGE8VACWongmlApQArKcAUDSkTyABQERA1RhwMIiIFCawKUySWSNzsaa3Fe+9VNXGpcynIlhJ9EJc4KFJCKAI0ptYRc3M9kWmZ4JTC+/QM2JcNEjUAYEkACDxISFUJwhxhK6jAKLBW4Lf/5K/f8+FPrRdZoJaPEgVgG4nImaTVLqqoYFXe6iJqLq7MXBcjGMpbWTkekoS9s7Nt4tHJ5WpjU8YViWpUMqxMVXPUgQmTW0yGMkWjvhqvddoaitVum7//5S/95m98tng4q8SBGZaZQaoaQs0wLj1NEzwtcuxw9AMTsi3/FklOJXPrkLjEl6VLEoQAAFGhMjHVqSO8wKZwBtFjPEKaAqFJFE9+QFCGryYETQTmU3+Px0hTJAbeQwHnQAQhuLzYHJCz2VwXqkEiWQ5SWyYZF86l2NhAkmFjHYmDAtZNzXubi8S0pJW0oIBKw9GTHwCOQU2aMYIUCggDjFGBThdpsl6MuwuLK8NBu911hOhD7hjjAkUBUtQ1QoVeG0XRUOLkNxHIgAkRk4gTmCxPc4A0nLhVbsOUrz1NmgFPkWuTdJhW2E7RLU92mCqGFdIUzkEjel2080o1Egi8FUoLEaChqhOCtQa+xngACxgBA3UJAJEhhMndEwBBxmBpJHmwBjxdu7DNyq7ZkgaABRIoT6aDGQZFaPQSXZYDBDEwJirXojDGACRw08qDxEDMMBCoMdvYWPmemesvG6IGgEkYBIEXElZEYpDzwKagivjsTUuv/r0/Obw2WC3Qmj2TbHswGFRVlbVSNih9Za3zQQRWYbYqBUICUsRArGmaJokdjjbE17OtFtcxEy1W14fLaxbabbcBlL40nPoaULO929aQEMJMJ1tbPgId7VnsrS8dufrxV/7SL/x0K4cxIW9Zo/C+Io2G2Bpic5rYY1shmvnLv0n/a+HBwdHTQwo6TWCRwoFDVVlrNUqsSpu3cXIVoyGWjmGwWQ+KMC5jUfnRII7H8NVoYxUaVSWq+qhRWKOKkkkcmAiGiBxBmcBKRLMzfRDZdpbN9G2nR+2WmZ/F7CzyDnbvg4UPQaBKSJNUQmnYVEcPf/pd71mI2q9q3djMLEUv47FADEmM0JpjII2IJFpvjoyCFRGqpKpqBQEaUiIrlo21hhKbGJepY07mdu+t8lZ69oG4OMszfbt7MYqA2ZCp1zfM5sBUNVZWR4cPby6dqItxMR5570NRFeNxNRpWw3FdjqOXqgyqqipRQ9AgIqQK0Rhjk6bnbclcJe62O1u7gkgBGBUArSQVMIyFsWwss3ViwNTfPaeGxQMuqVNbtfMzH3PFw77isTTbi01fg076wpsJVVqWXBQ0HK/edNPJW29ZvvuO22+5YWN9eTAaiEgIIXgNtUoVtBYRESOgQGTUIBoDJgYx0Uy3lyRJp9Pp9/szM3Odmb7rtSnL95xzftJqdTqdTq9j2znnKVIHwypKZJB3kOeAqeq6iuqynGxGBNsk3KUxsFDcV7/cfw5HS6PXEAYhkG8a4iPMKCIaHke86wM3vea1f7Ey0jGb9sLeYY1x4YltlmUx+hBrbQwIk0zAUKPEjTIdFBlCCsNobkGGwwGCb1lbD8fnHzjYMtZvjtaWTpbjQqMHxNokeAO1RKfu8oiIoXlmx5vrlrxKcWD3fDFeb+XmF3/ux696/Jm+FgNNHRsShkClMcyaBAGni4V4cjP7kMGDgaMVmJb4FdP0KhArDxGXZZsrq71ezy+v/fYv/er6zTf311f6oTZBqAquLNMoLUjKaiEKH6EiImokkorRySBVKBODLFSZtCEhCUElEEuaaZKOE7eW8HqSmoNnfM8rX7nr/PMG1bjTm6ljoCiJsxhslMeO/cR/++/n2pQPHz2r0/bjQcqJF0dCJiokRivBSFSFhFTZ6LTeZVQBK1DS4BBZAAg39hRso4PapN2/dWVZ9u8ZdLOwZ/4nf+NXsz0Lqqo1Pvyu97znb95w/FOfmgsyC7CvUlFfV0bFKJyqhVooKQjGuRyYZD4iYkPQUI7MCo4EgJvyWiQoiYQIasw3hFhZlQAjsNN8USQblVjJRVZCcIjQCKqJ1g37udmr/tsLnviC5yVnHIBpqpencikMCaPh2//6b/7md36fl1cOtjK/stzLXb+bD4ebQiJQiFJgjmQiTTrZjBKREkXb5DzIwFg2qk1dlqJIAIJBRVRAlQ0zs7VJO23P9ruLc3mvc9HDLj3z3PPOueAi6vbRaSNrQwXGRdeqI0iR5RYKxBhjNEkCnMbH20QoX+bjfyI7p2Z/RRIPF8BDxUaFf3rXx37/T/6O3Dy1+qMYVsfjLG950egFxGSMbRhPgpIBGGSnAawwgVVY4X0VQnAJQaKFdtMUdb3Y6TmBiXG4vrZ09HhVDtM0NcawJIBRakxUWAkEA2CwuTnb7+WptUaXjh+SWJx79gEJm6/86R98zJUXzXRIBEa9Naq+ctYS2alZC2M6ggsTjn4oyfUeHL3g9+kWaYxRZhWJMYKImWPtV4+d2OOrfDRG8EmQVgitENsaEyOilVAAkxAIjmBBjpSkqJqbaSW1TYqBRQlkDBlbC8aDIljn8rYmvMHD3Rc9opu3wMY5p9CiqIwKh2CNyfqzl5x7frzldt4cd21rtDJOXJ232hqR1bVTUBBlDQ2JVcEowKpM0bESGVJWmCJGwtiqJyVBUBfhAuyemT1DZOOVYrg22HP2wYwUCBoj2/S8gwdvWVw8UQX2vpVlJmgmYkWdqAvCEllFmzsHIBQjgiEKTOoIBsTgSJay3DMCWUx7TJrUhZJnRGFQI2SGGJAhlMWAiAA2ZJyyAScRDDLiKgkmcQWb2sdRkXSdSfodUGguA6RKahQSGUbFdtP+Qj9JWQxlWUbWSeVHEqpqBAaRWkMpbGI4MZYI47KYXjHAatDIRQAFR6LacGSjxEpGCA6aQ6kO6msRCauxOMpFYuHMp9/8Npe32Lma7dmXXPLEpz3zssc8JrvgYabyed6BM4gBTEEkxBirKknT7eK8bQfml5dQCKpQmhpJCUwEl8BahT/6q3/++7d+IOkdOLlW1aNBf6HX7bQ0ltFXxuZ5nhEsSDhqiFKHoOAIBjEmeiBApSjqLG0nJiaOeq2k30pjURy743NHl5ZkVKZMubUdy3krS5LEsilGJaYVIiGosrJR2F275sbjcT0oNdY2y2Z68yuD0dKJI6969Wt+4Pte8vSnPcqXAbGY6brUORUhkkm+8YHmkfJ/HQ+SOHry91QQywCDxsNRq90GUKyt5d3++Pa7P/sv//LZN75xdPvtfjDY1+7sFkMrK0kx7KU2ogJHIQFg4AgOMFDW4JUQWITgZOJx6Yl9agdRRkqVTWLait2e2bMYdu/67z/78zjzwKgsXKfjsrQRclDwqCsMh4ff+4F//O3flzsO7WLjB5uOkOep86FXhjzGnEAcBVElcgiswszRcnQsRpnZgszYK3Fh2TNHMjW7oUsLk22MQ2WTfPee/Nwzv/5nfxjn7qukSlptigbkNq7/5Ftf9+c3fvCDyXjUjoGHo573rSjdEFMFGyKD2qkwxXpaFqQAiQBDbWXMOjC2SWVMJJ6o3EgImlgmFVAjKBaj6pSMSkrGSNMiqCSaRCRejWJYDD1itzcrnc7RqBtz/ad818su/+/fiF4rmGYUCJMYAJGYEcgXxofPvOf9v/vzv7hy6+fOXZwfr65kBoklQmAJJkYXo/FqBaQwbtJYGJtIHzDKAi5Va8dja4NNhBlsCCaR0Akhid4JWSbDiCKVhCqGVqc7LMtAXKrW1mnqurNzBy96+H//3h/qnXc+ZvuxKKKzSd5qCmOTQ/AUR29ziPryHv6iUFIWUFSt2RTACPip3/jL628++rljw4DWnt1neV+trRxKXbAybrdSUTscV6rq2CTGOOeqECIlNaWRs4gU5Lb6QB0rSeVQazWkepSIr4YbLkhmWEuvvk6bnVGXMWrqUgAgElIlisRAEsmur43mF/damwwGA0MxSVghxvjjh259+MXnfu/LX/L851xJitFgrd9OEyamLa/qRu8j23IdDyXKfjDkOppSwqTlVwWg6fUXOq0ONU0SUREjVk/q8aPvfv2bPv2uf5lfHi2My9mizKLvdN1ovM4GUnlTa68zMxgWaZqVUtvcRaOj0ciqznX7y6vLlLc3OvlK5tbbrdmHXfCoZ3zluU96EnbtAgj9ORgrhCZv1tytEQQaUQdsbJaHjy3ffNudN9106I67xisn45FDbmXloNhFYV1bne+0j60cmkk6UpfOGlWNFCVzFYU6BkdORuHMfWeeWFntLOw6PirvKsbFrvn5Sx5+4ROuPvvyy+2552Omg5kWrIp4MkZE2QcShQg2Vo9/8lMff8+77rj2Ornz8GIdO5ujHlBWRYHQ3jsbiav1wigYERDDUGOjTUdpdsu4eP7Lv+twVa9XZStttdv56sZqCEGD+rquxqNiOCoGm3E4oqKwZW1rb6qqnyeJaBKCjodzJgvjYlgX3W5PVTcUo97MkSz5tl/7lb1XX4WFfmQGmJRYLBrFFQUgEgKqqOsbx+/43I3XXf/Jaz96+LZbZLS5ceTIhfv3bhw5MutMRso++Co0Y/KUAUKkRutOHpBu59Krn/D8H/ieu1eWP3PjLTfeePNNN9w0PnpsthrvSRMtChlVCaCCTichos3NKslgstwzV6RjEWGjNlut9Btf9p3P/Pb/jvk5tLKxqHMZwVh2zb25koLoy8rR285ZKf3YWnacBlGvXJJZ8fjF3/3LTx1a+cStR11nX6e/a2N53WnBfumHvuObRifvtPDW5apaFIUjMcC4LMilkvY+efuJ248OlgdxdtcZS8eX0zyvy9Guha6Um2vHbr/swrOO3n7DcOlIP3Nn7dl9cO+eiy648KILzj944MDMTC9zltkwI0kwHOqxkyc+e8ut73jX+679+KfqYIcj2Xfg3OHQE5nVjfU9e/YsL5/MEqgWiOOZrv31X/2Zqx5zQTdB8EXuDENUtSp9mqbExsfQyFUfcvMPHzwcrVM1LkHvxdHatBJHBaQuBlx7GxTHlo++7b0f/avXjz9789ndTiuJ5Av2lQmKOmYuH/vAziE1VaxnZnvLy8uxrs88eOaxpdWTEof75w888fGPfN5z8ksuiWkS83bSm4XhaJyAdXIoNRk1EUAgFkpRUQWUHnWAKqLHXbfd8e73fvZt7xx85uZz0qwbJSMJo5EEnxhW1cgSnalNrGKwbLigTtYlly/VYbPdyi+64NKve17n4ofZAwexuIjEiYJbGSxqXzIzAInRqBhjwYJiEI4fi8eOf+DP/vzYv107s7TRjzFhrI42dKY19lUarVFAg4EyqbCpXb6R59mll37DL/8idi3CWoDhLIohjIFJIQQJ8AG+RlGiqFDX8fjxm6677pMf/dDR22/ZlWU9pbU77lzIWqEaz3Z65bgYGRrPLR5K0+/+nd+afeyj0e96w6xEytscR4WMiEaEKMGzKMcIHxDDJ9/5juve8673v/FNM4SD/fZoadVFpAbctEoyhFnAqkSCkozvdS6/5plf+z9/Dq02ah9Gfri6aUfjN732tz/1/n/hUbGn19fRuJekg/WVlC1LjDGO6xgItp2Ic1E1ELd7i77dsgf3/OAv/5Ldvxezc2trm7OzixqZqFlmVcJ0gBfwZeCUbXNwROCDhJSzCHtybej6ndf85bve/MFP3D2SlRGQzhibJxJaWvbN6h+++qfPm0Wi2BhWbM1sz9YRflzMdPPD69713W/82Qf++q0foPaejbUClM7t3TVaP6l+M9VqNo/n7em96PnPvvLh5/cdJYSE4QwMQATiU11Lg03JW0wWRQVh3Hzb8lv/+X3vfPeHPnPDHbPz+6JaY7NOr7+ytgqKvZYbDpf7HdfN9XV/8psLc+lc2wRfZNaQqqoycxBRVWvtJAnzkMIDPR99qkOYpl0K027m5sEtSW6jaSZw0pmFCnxE2t73LS98+uzse//4T9aOHpeqmDWpGYfcmJqqOnhNbEGR69hK0vHxtQ4S2HR5ZVjmnbqbP+Pl39V/9GW48ELM9I1NTSD1StYKEGm7G/EpNWBQY1TJEGUJMgIbsCDB2WfsP/uRj3zLq3/rxmuvf1hvthgOnJKFCU0/CZjVcASBSMhKDHWUllk3MOedcfELrpm55unYtw9sYJwWXkWarzfGGGMgBFhoBEENUadrz8vswf1Xjcd3zM5+4LWvs0kO1SKEhbRd1l4tx0Y0rQpVgUaoV65DQKuDhd0wBuMSqtJiznMIT0R7TCBBEIggeHNg7yWPfdQl3/Gt5S03vvP1f3frxz7uZ7pJ3nJrni35UBGngKhqp9eHdduLazIx9pfJBiRWS2CeZClTBekjv+ZrH/noR/uyuuP6646dOLZ/fiaONjJjw8izQImp6YpRhsAaMyxDUQqEYRxaLdtyM60FBP+iH/tJfN/3vvl1f/Hmv/mb3a3OseUVLn0ifle3pUESJkekdRSFY+s1xvXllWPl8MTdP/Xy73rVn/9FWZWzuw5I7cmkk7kD09b8/5xRWaJonJsH3ie9ztve9+k3vO19dw1i3ZrPZhfqaKP37VZqi4GN43N3YUFgvZ+fSwOwsrHZarVmu/koYM+M+8jNg/d/4INJ1l5eH84eOGttaU1VMsuDjcFMh1sUf/0XfvBAGw7Im+ZbQfASQhARANYYY4jZJiaUQ59kaSezEbjkvIULv/cF55191u//0V8cOrpGlEnU1bVaQTGIV2ttZ2lpeZTjFT/6yt/73VelFt00jzFAvbMWgITIzBCKEq19iHH0g+CuITIiT/owsN34YootjmyYenMwHJdlqXFQjeG0++ynPfFF31jO9TahSkajJEzMqFDHTIKLasUyscSOSebndi2X9XKSXP51X9t/9lfh3LNi4qpGr8aGnK0KTwQDNdBGRkAKncQZ1hKTMcRWiGMUX1V1UZZJgjzHpQ+75ge+1++eXxI/BilbgRU1CiZYCDgap8YIJyYRQxvwZT+/5JpnzD318ejlhZYFSyG+YtWEVMWHqg5V1KAxggjGgW0VdcP7zbqK4ttXPPzS5z+vfcHZG45PlqM8zetxkZElavo+SKhppYaqCrQmgrEQBNXaupgklOVwCYxTJg8aaxyAxpbrxKCV6eIcZjrotbNLL37uz7zym37ix/Xsc24bF0PAQ2IEARolhOCyVtOB0jiVNNlkJWl+i5ICRNaYjE0aYKog1biGKA4c+K4f+zHNW5VQ7UMr7w02PQBlKIsyEdGkeV652+3m7Q6SFGyEmlKogU1w4AzsO+M53/PyV/7+H7gzz/Dz83bPYrZrroDWIqnLuknGQUwdLZRCXW4M5nLb9vXw7rtf8eJvyZRQlhxi0/YTgXiqhVnx5b9HLYpSmcci4tytRwZ/+rdvPjHUitvkOjbJASB4lVrD2JfrUiGJMUWIVUmKxX5PoKuDsQInB3jVb/7OyqAclb4zPzscbrrcrJ24e2P50MXn7Osn+NPX/Py+NsabFcraiA/FSH2ZWLTypNPOOu0sy5xzhrRqtWy3m6ROWbUalqGo2gme8sRHv+Rbv8Gg6HXcuNhgA2ttrz+ztr4ZyZJrpfnM9Z+85dd/7bWiGNWAsbVv3BCIiIxxAKxJvuwb9L8eHtgcLZOz4p7nAk1D7O3+RI2wN0ksmGyWdXbPD9ljNpt93OX9C88cGqlZA8XIGlyIqfepj2kodBi4dqAsy8aCFdGZKx558DlfhflenOvTbN922nAmqEQGHIyqVWGVhqBJwQJSIoF4Ue8RPGswibqck3aSzcwOjR1IxCMvfuQzn5buXTStVtRG4NAox4hEnSAVMlGzPB9pGFuWXb291zwVB3ehk5rEOWuMJU2iOlUOEV6NgCgiqkSEiChWKTGOkzzkOfIMZ+5/xstevNpL1w36e/esr62bKtjYbKiGoKX5EZKyLhCDRq9sjHMmTcRxCBUSptRwwmyJDMSoMAVGKYJ2B3PzmJlDb3bvk5767T/1s+0zzh6DKu8nnRIiiIBNEBQ6MYIihUxs6hqa5gqoQpQYGXDOpnkr7XbVGQSPxYWveu5zyNpOqz9eH7GHEoKBZwSWwBKMRBYhqXxd+RJ1LdHXEgMDLkHeKqP6Vg9nnHXgGU975Z/96SXPeOqyM1UnD0nKSeq9r6vKsrGGJHpD2u1ZXxQH5mZpY7B5+6G/+JVXY30TRCRRSCIhNppl2l4O+fKBjXGOW6PAaxX+8K/+6aaja2gtJN3FADMcDKQaIZTleKAI7XbLOATxMATrKkUB1MZVnG9G/Pyv/e5tdx+lLO/Mzw7HGyDvy7W5vltoY/XQja/51f9n3yykxnzHdpxYilnmjDMTYThBFeI11DUMaj+MUop4lbrbSWZ6aTkK3Ra+5pqrnvLEx0KqxHKeuLry3sdWq72ytj47vytEk+Xzb37bv7z17R8rxs3BZ0W4WcfT3BAfYnhgc/QWhE6jY0xvOu/tJpxmrSRrj6MfqaSLu5A79NuXPvlqn2YVGMbGGFXVGEIMFjEUwqFmgjp798aaX5h91PO/BmefgdmZ2OoNRTerio0hA4FPU0ekdLqHvYBUwQzjGr/cxjJXBMFLvTEat2dn2wf3wOmjr3nmpqPNWMGZSAhQzxNTTZoo3qgMWltbddoPf9pTcM6ZyG1pjJrUwJCoqoohJXIudTYFiKwTsGozpIoZRpTLSIVN0Ovse+bTsWux6rSWRuOZ/kz0gYQgzdQ5CEFIm1Y9Ng5pSmlqjWXmJkSMqhpEAgnYsDNkmFhVvcS80y3rWsnGLFeXYW73wvkXnfPor0CnPSal1ERChLV5DtdIaCyJpW2WSBN/PhAiRBCixih17TeHmydWT1IrQ+rgzFVPeVp/dnFpeT21rSxLqNnVjMiIBGESEjFE1lhr4SzbhK1Rw2JJmbnd535fswzOhW7rxT/xo1c//5pjxXDNF5wnxlpVzVOnGosiWMvVKOyb66/deXyB3Bl5792vf9Oxz3wWwyGggqgQ3dKdTUoSX95TLE/bEZCE/+pN73nfxz6N1uIIaXSZBI+yTA23MgeJqsqpCYBLM69GjVFGCXjBh6697ru/90duueWWpzzlKVdc/siN1ZVeNw/laPdcn+uiY/wrvvvFuzvGeMwmaLMM1pYgNTRCJIYQAppbNbJknGPDbIwPofZlCF4lEKTbtplFr42XvuRblpeO9LrZ2tqKc64sy7SV5Xk+Lqo0axdFzPPZ3/zNPzx0ZD1EGJM0jelMLFEl/Gfcl/wXxAOboxkwONW+qlPXSmASSE/HeWD6GK9sbo5DVE5qtRuqIyX0Z/dc9QTu9Dys4Sx6haccmR2FVqW7DUxVuywdEA3zfPejrmhf+WjMzw+iLo+LErBZ5iHMcNZAGvsdKKkwhBFJI0kkKWIsQ/SkYhmpVWfLhEqmVqujoA0JOLCbLj3/8GitZDXtLBqqDYLV2lFNGglEpErjSkxnfjNtnfeMr4RNh5xuCo/qSMo2WquO1UU4hRNP1biGsjKrMeSccZbZCbgWIG2h00PWOvvyy2KrtVqN837fOCcCFSNKChImZQYZMqyqMFOTZRESGMtpmks02ngXgQzYwFljEmOHoyEbGwCY1BuLCLN772Of/sx0157S2dhqlcySpvnMHLIukmzKaJbUThR4yizMEU6cY2cSGx1r5tBr28WZ4/VI89RHyc6/YG1cFJE2xrXCAmwBKBERmBQMNkooqrKsA0JoLmMRpjJcG14vihoYQ5d9aRdn3Nn7v+l7XnbhYy4vSWoEsuQceamDqFoEDTaiWt3sgRZsPjx0dNa497/9HeXG+tYhRlMDv0k/zpcZCow9/v2Gw3/3j+8Yalq5dm2TWhQMm5rcUsqUJYkAw7Ia1VgpsDrE6hgB+NydG6961e/8we/+7lye/cyPv+LnfuTrHnnBWQfme5vHjy92u8XqYLC8mml41pMvOziDGYeNtSVC6PXbYAIzrCFryYEthBBV6lAJAOPyvJPlbWZDRIo4HA19qFfXxpdcvHjloy+1DjFUM7PdVisbjQa7du0qimplfbD/wNknT6ytbxR//hd/e/RYBYAZGoGJe84Dm6z+w3jAr/Z2FtZ7RdNbr8E09dHt9RqDsFDUiOyStuQtzM2GTqu0LGkShEU0YceVuBqscM4dK1bXQm32LuquOczP1Cp51um1uo4YKom1zFANU4cdCDhOMwXNUllj2Bolam6HQ2NKQah8XYu22n1AEcJFj3qUMHvvQaLwkSVyFNbIkxwCtzshb2+4BHO7wEkl3El7jlwMEBFVijGKiAKJS/LGs40mt92TaDqChKJyXXpAH/aYx2qaLezbf+ToURCpRqjHRDzADDJQ0+h/TnN3EijVdcWWQNAIrdFsAIYF2STNrU2iaASTTZFnAt3/uMfQ4sI4a0naicjJpFm7hzSB461ywqldCQbYCDWWG1EQBLVqY8bRTnprZekWF+DMeZc8vDc7l6a5bUpMU0nttDghAJiZDWAMMROMQOoYilBmrdRLsOxavZm1oqh9Ld3Oj7z61fnuXRs+DKqqFh2No7XU7aZ1ACxi1MVud3RyeV+rM8f2zX/9d9nMDEQIwo0XHzA5s/4juQ7ZugO7B071Lk6P+QgqgLUar3/zewcx3fC2QGqztgrYmMxSNR4Nh0MlA9OqkEuGkCOddUdObv7iq/7wB17+PavHj/2PH3j5a/7XLz7+sjPqMfbOtJeO3j3f74w31zuJ293Pf/0XX6llbQAHLPR7QOMByKISJHqRIPCKqIjgSInCAeSjFxHrrIiEOnZa7dzZmV7L1/iRV3x/NRq02tnyyaXl5eVWlt156M4sb7c7M4eOLi/uOTPvzP/5X77pyNHV2kOAIAqNZLB1kDzU8MDmaAIswdLEOacxQ6Kt52iSzp3EZWQaD+XEcA6ay/IZcjYot3LM9x/xtMevmlrbaTA0Rk2wKSchCHfyTQnd7u4x0+FycPlTnoh2mrRSI1VLQh/UVmIJItKEEmFyBjUmPY1ZQbNUjRUFN45rBE7gUpMhYbBhMJDBtvafcR48yciT90YjQimxghVypM5QOxtQPXR07hWPRqsPNf2kSz5kALNQAkqJLbnmIqTwU8ckJiEEkuCgLZP00o5F4rIuejPnPPYrKtWNjY0ktc5pmsbExYQkAVuQEeGy1tGwUYaAjYiQ40hoiDgYwMAyEkIKGEEEPKSOgUApJxZERLUvqWvRMs992UuK1sy4yozpk2mdfd75MB5WQJ4QQZFownPaiNaMMmvjOZgCLaWWcFtMAmq5VhRFu9NZXCijNxa+HovRyFBVRLFBTIwEJYpkCURwVkWagyRzNk9T4sikUQOD83yWXcf2d2Nh73f8+E8WeR5abbQ6XpG3uptrlQ/QDnHPVuWwb8iMBrK6spjY6970RvjKChJwQkTNdC5qVGlf0hEtUxOQbbeDEECCRAFEREIMde2rqq7rGhgDb7/2lre8/5M+XWzPn+k1qYcFRFBVYVwaY0JUMfl6RRXPL9V48wdv+96f/l/f9h3ffuOnP/HD3/sdf/hbv/RVV1/ZtWgBCwn29NrdzPlQZ7ktRicsxued35mZSaBghSFGhDE2BD8cjyKUDQPYWK+VceTEqAymrKFqRKSshuIrCZ6ECCRBjSIxuPiiA0wxIePgZjpzVVXNzMwElcJ722oPA2+OMbNwxo/++P8sKozHsM4UoRT2pR8I6nsMMXgo4IGtvcN/xA9BjIDQNMtt5XrhZjo+4dGw7jAYTkQMWSUNosJGotQW3lm0U1hu5uYRZGu+tU6noPLWry8WDIAbd02TmqQNTshEBhM3M7AFahUaRQQmGBqL3z03A2vAKSsSMGkEjNKkF8sqU2P0QRNn6IlanISUDRCnVjVkEyQ2a3V0OCQi1UgKqJAyKxonT1KyaqaU2XwSK23JfqezWxWYCB8FQJpmAHxdExnnDBKCRvQ7uy66cNfFl2yu3OCIK7K7ztiPVgIrEGomfE/nUak2M2Ib/zw1W3GxEQgxGEE5Ehtmm6VC0gzemk6mncTRBhIJoKabRaFESqZJbDaiGQERDKBklBChhpVY5w6c2dmzrzh0pOsspWZtc5BktqbgoQ7qINYwJKbMzvuN4yehOh17i8m2BTczbv+vQIkAlRgdMScpoHWUSnFkhH9+/7XdXQfXSlNYNVkm40gkWeLEl9badrc/GI6zvFvUg+9++a8Nj9x09nzrpS958XOe8ZTd/bYTJY2IPjFZQuQIlhhMXqIztGfvgiGwbs0c5InVOGnebivYh2CtZZv8zM/93r996LrHXXXlj/3wdznLKmi3uhRFgljnUE9CJkNwhNl+f3mdYmDRJhslAihYwCCJmhBkNAp//8b3fcdLnhwEbIwgJqn7v7QtH2B4YMfR/wFs3U3LJNJBI5yenZ8D23FZqKqDizEaYxw7VGKFtRYJ6ozL2x0Yg8YwaGJhylAmYSPMyiy8pej4gj9GyUa2IIoMsmCXdHriEknS6FJ1KZnUwHFkqkE1waslW5blrj27wQzmGMUY3q7F3b5Ht42Aug8oAYbh3MzcbOP/q0oaibyhYFQswRFSUApyCgsYwE5G/ylPS+9gmRpZAQCMioGoBlFxzjnDjf4m1hGimJt52OMfe1cYHWN/UutHXv14qBSj4akFVabp1eA+hzFOLKqbBk4yADq9rk5zSjIVWd7bxYX1VJWCFUZhBamYLBoXrRUmZagVcmAzd+ZZF15ycaUxWrZZujpW0zIuNRQArwogsdGQMImEu++6Y7spgU40l7jfzX8/2LYPt7oViUUmGpuJIM0ZInz047d87BOfBDkAIqJRyDoGOeeIqKgrJYhIDLXGMpbD73vpi37rV37uW15wza5e3jJwLDGUzCwEGLB1SkYoqQLItS648JKm5DON6psZc0bRNM5S5WV9U7/vB37iA//68RNL47/+2zf/24evqyOUUiATTQkOzFF8c5F3DMc4sPdMpmTbejJA209MJfZB/u71b1xaRhA4lxInDLNTM3woYhpM8+z8AltT1lWEsrMxRma2bDREVgohAGDj8k4PhkHNaCjo9sJ9UynCF0vQTVTY8IU2FvXG5d2usFHnyFomx2wYliNpYI2QIMaY0teLe/c0veYiAj41KPaLJ+gJiJCkc4sLAoBIhSSyKqsYUqvCUDNxjrvHobIVOG+/lWlcAhsGBMUYiQhE4iOITZIBVqy59ClP6F98/lpuw2LfnnsmWnne6U3cuomxzaPh1Hd93sUnALMz8wCkGbg11cjzdJ7LtsQ0Ti3z9MNZCcIQZiVuCn6NErzfv+IrHhuYylBz6ghQVQfmAIoSAHUmMMRojHHpxHFohETVqFu9VF9w498HTjUlyj0fBxEZY0RERYJoBIqIt7/nA0Lp8upGmubO2jAeN10eIqhD9EHq4PMs8dXYUvjxH/m+b33+k8/aPZdobFu2QF0WzlnVONksxoAN2AYRY935511oCQagSTGClFjByobgKh9Axlj60Ic/vjGMQfKordf9xd/6CLJuWNQCYuMQtelZJahlGIOzzz7bTkb1TFoxeTv5KglYjfvcXYf/5X0fFIXCMExV1/SQHCL+kFthmqApFWPSIMeU9bpsjQCiSqQiQkRKUKYAjVA2hgwjzwBt/Hoaf3nZ6qD5PBXLL7Q8kK3aElOr01aajAlhUKJklRoHaYHGZqyFSn9+DmZbCYVpy3b//hdhWlYVIlJt7vbN3OKCYDJCRTHJnEOZlIGmP5toGoROP0em8kYGsU76dkBETGRAJESKEIIKODEAEBBiQKfLu+Yv/sqn9C+/+MrnPVssAQhRti/6ZKoh9J5x9L22cNPsPrM4L6aRkwukmSA2yeuSKFRJdHLFmkpsT5WXJwcAoiI2nqjkQA6KCx5xWTRmVNVENNuBjmsTxHpQRCSuDdeGhFg1QgQaJ5WzSR5s8tn/x5DmEsVohkVOdmIkKiNuun3t45+6Oe8tRGEhNiCEoKGRREZRSvOWMYYZacIZy8POXWgBXQurtfjxaLhOEMOGGE2yzhijGhuvDAbt3rWruZk6NZ2yWTkyAVJ7n2TmrjtPHDzjvKKgjYHs2nPuxz7x6SIoMcrgRRWmUWaAjKhGYjDh3HMPEEVFIFKaHmbNj4Cby0AdJMna//jmt6s04xmsCn3JQ+sfFHjA56P/w5io9AiqIDBsomSSJNFxLSqsElVUBdYIJBK7LBUSWIKKYjJ5GpM0dBMHbE2TAul0ZNT9/wYmHpAUQAyStJWIBNUoMTBFbuI3hUCUiY3RhphaOcwktmo6XRq5xfZ4hLamHpHgVKq0yZ5DQVHVqoKoPzfrJU7KXDxpYJ+YQkhUBWLcoj0QCUhImvFQEyk4TbcCTTLWGqJLkxCksSyp65gA1uWRa7Ow8Jzv/PZnftMo3b8PhEqbWVu8fcmVTtN40L2ksdO+fwLRzNysAhGaNJtDJp81zcJPP2f6zsmI2Im9qkKE1Wy7/DAMI0TT7aqxdSh8Hbrt3mBtk2M0AmZEQoixmcSonMwtLGxTvMRJjZZp+7d/sZhk5HFvdUcze4EMiBnA5hDv+eC1m0V0eezPLw6L2nM0iQu+YiNBkLXaSkZURCRhJEaLjTK0MscAGYIknY7EOBgOslberL1lqASKtYWS+MxO/KRpOh0IxEIsEAanaa6KoydO3nbb7d2Zs/bs27M53DBJd2NY9TuZzRIYiHpjyHvv2EECyArhjDP3grwqiJVlup7TSkazd0ZFvdDrfOLTN952x9o5Z8wykKb5Q7GD5SEYR08hwGRctDbFdwKY0jwDiY8BgKp68SY13ogadblTbYa+Rmjgyay/SbQWm662xoiT8EX+YBL+qqoHe1C0CXmtRX2USmJN6o0G1kgsxihbBol1DtbAMEKwxoR42sm8dSN/6o/mwKZ7vkxVowoIrU47iihYdJKpVRNhoiBGDaK1aA0EUARHpaAIyiFCIuBJa9Km1VMmc3oBsDEGAFgjSVGXlgWGBseOwWuMGvI0PefMaE1JVHhvjGsudacWrykVbht+dipfvfWXaPOWdrcrW7OpppPFWU7L9myVH5QQp9O1PCSyRquR9dQXNbO1jFlaWYlsySZV5RNySQT5SSoYQAgB4Kgkxuw7eEazJKccymgyleY/xCp8GkFPDpFmwYSIBKiB9UI/eN2nKGuPaiWX1MH7UKWZgwQAMUa2SemjD0JEGkNKMtPKVLCytiYgYru+vlnUvt3p+CjQ2JRYyddGQkZi/DiWQzutNZySESoTDMFkxpXj8MhLL3nYwy70oRqNBiur6zaZOXxiwwNqOHKIHGE0UABrpNDExAu7ZgK8aH2vPQolaS7/bJNauKj0jW96qwBBQI026qGHh9w6b62wTgLDaYoyAkCapqoqKkQkEkTEWquqzOScmWiEJUyK+KduxT+vpvULgiC2MfVvWqAt1VR7il581ABRqBpSo8TMxpKqsrNNOUpjJEKMkZoo915Cr/tPfUjTTahkkzTq5KLV5Bub1LSQCmlEDBSV5FQoje0r3mxVbU6gbZtEETw0Ooqoxiw1jh39q9977cm7jnqvptNWYm9IE5vnbTo9jt7CF8VvhCRLt2xb7jPdtMXgONW+OPktaEa5iCBO5vg0V+I0OXHsOABrbV2HGKNhUp2sp1HAKytHcGTu79ndZPN1+hVyalv9xzDVzGwtp4AZxhglLb0fVRgFvev4itic03xzMCJrRMQQjCESEaVxUTVSOSKyzAgxczCE+dlZZ91oNEqyvJW3vIhzrvk2hkAC1DtDKsFXIwKIIyiAgmozCQUMXl1fUahhmV/g7/rOb2MKrZx271nYHFTHTwzGHnVQgY1EEUbIKEwAB5Aosm5bVBtRJIBpZ+ipLaZA3p2pvLq08+Z/frdpxjI/5ER3EzwkOJq2AGJiJp78D8BErqVglqirq6utVstZ45zz3scYg/fiQwxhdWkZIihKWGfZisjEYF20id3MNIL74n9YYKJQiKlLmnzCONZiORrS1HpIHbyqGmJDxD5KPbUZa+UgImsApKnTSV8dbQ293hqN2zx66i+avEZFEuvYWjhXFEWn0zPGGFJjCCRRo4/ex9ojRIIa5tRBIxDAJFBLdjJbSWMCsSocJFY+lFUsg/qwubkJJoRaRhu54+FnP/sL3/09H37jP2J9kIolsrVEJVIlw06luaM4jV/1tB04qf0TKZlm0jeSJGuezjrtKMKWyFkxFA2aYqfwxB8KTE2BAZOMQVOdbASKyqIaAwHMZBi2SZGMR5/91CfzJPF11e2203ZrJOotyIGZ1Usva3da/cGooHbvqqc/HS5j65itkiEGMSuiiFdsdf+c9iNy3z8qkKYsCBEV0SAiIohRY9SqKkUE1lCCN/zjPwvnSNIAigxmTtOkrIpmP1trlQjWCVhCtGykrjKLzEAVotJut9M0FZWJga0IM5jZWq7rOsa4uGvX29/5zkoxKoZCEnzlq8IwaRQJsZO3CMgSo4qnPfXyl73shaPR8WK85mxmTLcskSU2wARJxzVT0qnUFj4pghHG6kbp8nbW7o3LYiq/adJo0tRehDAuK3KpyzrLK+vXfuw2BYigGvDQw0MrH93k1BgyFeFugZsx1M0AeVIoIpESyBCnxpJaAzO9kyWFkrFNloQnt2BbCdQJ10xChPv9DUzGAsIAYLAhtmQcqRIMwUEFYFVSUeUmkGeBQvUeEfL9Zj5lMkp8ErU0U4gATKa7NrRFCiirqkz1T2AwMdgS2yk1sgIqQmxj07ijaAI1YjCgzAQCU292RmJtDaEKWNt43a+9enT77XOCzIPiJEDkaTXsftCsJ3+eiFoAw0SGp+4iEqf5a8LEIfb/Y++9421Lqjrx71pVtcM556aXX0cauhtoQSQJCmJOiJgFDAgyhtEZdXTGMOPojDPj6DjBMTFjBFEBHVGCkkSCZOimSd10N/06vX75xhN2qKq1fn/UPueee999TTeiP+nX63M++557zj57V+1de9Wqtb7ru3aB8Ixse0UUcGRH42FR9Jwzm+MJIAuFo9Em6urGd72DJ+N9vSIMh6cmY2TIFnrD4aQ3KAbWjpt4dn09GywcuPpqWlgEGyVOnvAApY4BaBakpF3bxMl33ufoWHSn5uQ8RMQYEuGoqsxNwF2nV1t2EVaJIdtXiGZLQ51mb5IBYEhNqn6uyckHqCh1NTONMZHAzMblxvQ8+NxwZH0whLwcjEfD0pqy14PGGNq87PvQ+LYpskKAmOFZz3xGFP+7L3ll0Tv86r/8a0tfNpmcKzNF8Af2HTx16gyIvERlF01+191r41oVcHmm08gwKRgq08eI2MBwVVfG9W/6xCef/oVXyzREfLHJxaWjAWA7RCHpKe4QCzGywhChC4aBiCzICGecqcmgLKCESm6SM5nYgEXBkmq/CKjDEVEKF32qLZAgG6bLIVdnKHNwuWqumVXlFOIDR2gLDtpZgl1pA6IEzZit3Hdp6h0L/+0pZBYFmjrDVTVKmqUErNSVyGMCkyW2lp2BgTKEiUgEEaqBVMhSyoxJhq5RaFJOrQaLyEzh3Np7/+wvTn3k44NRXRP3YuK1g6VUW7u7/p+Gp1E1lXNgdlaYkucmpMydqQruTOXt3wA6zZhRqDJZLst+mqEG/d5kvNEMR2WZ3/Kud61/8vbDrRxYHKwH7wWub6NRMQgia6ubvYXllmIoymc97zlYXoaxCqsEQeK80wTq4G2mf5nfEpk9dbR22pNn7d2OjgLGmAhVYKvGbXceD5QJdYAc0qlyVgLtnpl4tnQDWMN07ks7KSspWQGiGpgctmyDrlW1n2xNgM2t4WWLyxwajZ6Yi7xQjb28F1VFYlSo6jVXLTzn275qPN76jRe/whn85Lv/dv++xdUzZ53J2jbkRRlUrMvV2CzvjRtZXRsXJfK8hEwbTjoXMmFmC9WmjYu93vvef/0/e+HXtKLFRaiuLhJfx3kyK3wIAKQMYbQRQgZWQ2RNHO3EzIhiwI5Nl1POFkwRHKBBJcjUVZJoNlMQXSXq/dpGFWFWpq4qMyjVt7aRnbITY2AILpINZDxRAClBoyDGC3FBbPNl7/o7J8kB2MEyVBFFo1BXs4YVSdsykWFmQ9bwlIMSZIgdG8fGWZMZa62lhA8kVaNqNIHvDNQRsDm0zDe8+S3x7NkFiUUIrjtpYvboPDOyVyOBaUb1tB8J7zF7jrtlDZHN3IzFSOYyWbBzluIZdlCnQWKgabyx1lpTVxVJWOiVJSnuPf4XL/mDQ3lWhjZsbVjBvuUyspwbVuVCMWwa6tnT9dgPysc+4ws/75nPRJ6BDZhSBqTOudd1B0jjU27vw+ea7hkbNg1wdn14/PRqYCPUVQbTpHN1OwydVB53i8TOA8Zz8zlNeT9Sc0Xgg7ZCQW3kwnPWcP6qv/lQa3oeiMQhhFjX6VRMbNh470PbUPQQXH6k94Mveu4v/eK/PX7XrSxxvFm3dVY1BdlDdVjysuJlZTjJ14a8ttnu339J3iuzMtOdoezplAbDLoohUwrMJ249tjUE8TZS6aKSi7HPM2HIdkazD0ZgmDUKANWoTDCcKIowNWQAgDvnbnJ0YkqKP+/wvf8iBjBp0a8QpSDGi22ji0hFVHUKQkgvpLwVkZkpPq+r58e70t6ar4PuzhLEJRnRAtEOQrhTrxmFEYRqgqpCU0vbiG9JPUuARhhSg0gSEAUyJU7RTAVRUFe3vvHNW3fec9BlK3nWyxxNk3Y6TguiKVpZZ3DobdzLHiKYOQKkUzDWuYR2w9Rg7BYKHc76PC/Q1KmQop9N4wngEJr1dYSg51Zf/du/fct73zOQUCJWm41RWGtjlKgIMFT2/cIg7F+gS/Z//7/9SbjE0EcMY0AQZVFOVj2RqmqHvd7xmlFz7HztHJ5z95GZQuhS/Vsv9548M/YSySa1JQC0AxcDLDTvHRLqUk/m4MVzV4Q6jGWqScACE8hGttH2NFt88Uv/H5XZuVrrSGRdCAHeAwrVtqmzLCuKosjYcbSCIwey537z03/h53+syLRtmqXF/Zdccm1eHlQskF0RsxC0VCp8UCEMh+vr62eRJovdY5UEJgYUxWKIZmtY333vOWIElb0n8we1XLw6mubsCChHH6FqQFF8yu9IdnRQiTGICBOxCiTAR4IYsAMbApMSAcypshuhI1G6vy+FIioCVKBRNCAGCZ5UtKuSG5UkTuEWlCjg4szXsbsg5fx/Ha32eTQ0XcdnlmeERqFU8gQCEkUEiWoUDaItpHWpOBKL4ego6bfIiIiBgqfQsG9tbCjU1Naox1jfwLk13PLJt7zkj/LNjYNs/Oo5mYy7ZJLpAh4QwYwaf67lu1qs85CR7pOZHU15lpwMAIxKFmEjrHSzi5G5PEPqDi5ddmHnitWqyUXLotC77nr9H/3xX/3JHx/tF/XaZmGNYZgMw/EY1i6sDIaT6sxotEaRjuz717/yX+jgMhZ7w6YiMslVZIRMJNPFKy/4fO0VR1Sdq+Ky42KQECGKT/8HxV3H77V5GbZJVFgoadhOX6NzIslUMUtKi4oEhVWakhlMOVigytP6amQsiAOoVbM2lte/9eZ+QZFsiDEvMwkdxSvYJr7QzDoGtfXIiDrg+c/7ql/8Tz+LUEfxq6urd99zQijfGDXj2tchsrNlv++cKcsyz/Mplp1nxkOqZinCIGdMHgOzyT95+x2awvsXn1yUDp45SYEyKCTEjtlLlJhIlWyCf2jUCESTDJq2Rdta12GqO9oKmC4uB8yn894fEfFTWiIP9UCIHKJR71tLQUiYSCAGiMSGITqzo6c6S7oSJrvCoPPmsG5jnPZyfqhS9wgTNBqNCoIKVAUMiFdiBziFEUBIFD6gEcTOjLccwQxSBEHjUbeTW2678wPve/srX7l28y2PPnw0q2shhrNECiNgSY7iNFkICcCfolDdeVFR1W59D5s0jpDCRZhpB43MJdegw0cLI3D3uRLYkiOO1Rhbw+r0qbe95i/e/Ko/xcbW/v3L1agWBNdHYJ4E0dwpu62JHH7EZWdJvu/nfvrg467DvhUPMUUGUhIynXHOUIEhqE4r2px/3/d0axBj/jJsA6WJYK1VDQGBObvn5ClblC0bdCnvM03XHb67ZCmXhpSmbusAeGaTMmWUp8FDMDMLVIQkgiLUQDUoDhy56g9f/pqnfM4jH3mIg1dY5gy+9dbmmbM+qIi6zPi2RhTDijDJ+/2nPfVz/vzPX/qjP/5fbrt79ZpHXTuqgzfGlU7q2mTRgSbN0FrrfSQiSXqZUrxeAGJl38Yyzyf1hGHJZPfcfSLFh+97jDwo5eLU0awUaTay0xMkERBQUEQQBGzJgIw4Fh8ZMdOw0jTYGmF907W9GTkEkLItZm7gB8L8QuISkJYZMWA8KurJgm9zVaG25ZZgQIHIJYowTkaxThX0NrgjgUu2H+lZvB5ASl3bNkO7EI1McwqjqEZVgaZVOgWxYhBViZQ0UktR23OncfoEfAsltC1AaIKGOBmNfVNNtkYba+fOnjl1+viJcydOVavrZjR2w7Hd2HjcwSP9KDEGzczWeFNYDQFQAZspCIZE94rYd6iTed1M05an8tuUese7hzHpBaAg3ZzcWdIkahm6umpa9Xfd8+o/+L2/e+NfZ7G+Yml5vLp6YKEYb9Z5aYZRGmOaQMP1Uf/o4a28959/63+XV1wqRdlqFHAUFcOpX1BFjJBoBKCYqFj2aMlMR++I6xJsRmQiZenGzb5kBpGJEpoYxWJtc4ttNmNNmR0UOwHg3ae0y/uRCBF5tjO6oQBpK982mnmYFHexp4aVDjd/86Wv/KWffF7Z6w+Hmwu90uW29R5kQtsWRaEaiagc9L1ve6VhxsoiFhf7/+2Xf+rf/PR//cCN7z106VUSPGkIYVTVTRC0Vb28tOTbdmeocDtYGsVn2cJwY3WQk6pubGx4D3dRMt9dXDpagQgCYGFAzDTLHZHxeMhGJ9Xm/n5RjyeWXQjMhhvSbFD4Tb8Y/NVVfePP/PuzaKVXBHVCBKTyfwGAgomcoftrR6uqkkR41VjYXMaTsDU6BLrKN5lv7IIbN15UGCzeQw0zYuttlrPoZG2jFwQKUiXH3ZofnAxE7fQvzRLtSDGvDJTAie5DBYZPnjulGYVIbQDEM6hARuCa1XMUjg5tvP22//nN3z4Z+8Xlla2tjYWFfj2sVNpeP1cf0AhFYWbLlBOXitj6gmiBTDmZODIgXY8h9kuz3IdRhZIxgFERQ6mYYbcKUULyw8vsKiXYR5cProRUpyCz1koExEB0cXlFV8+KBBhSUSNMzMycEk+IaH1zrewXqEet+JzIGgXYr6+CQn3s2Lv/8FXv+IvX1qOtQ45C24a6chXIctWidYT+fs6yiQ+nffWN3/Qdz/zO5/Qe9zlaT1rD1g0C2BrjRUTYEcGonDulmxsff/977/zEx3IyjtgY45zLXeacM8YwMxn23m9tbVVNzcxFUew7cGD5ysuOPPXzsbBknFWwtjDMEETx4kiNClMr1BLWx5WvPZU6HdWJPEZJSdUqEYKazBkYSKwbX1hkWdbReFGwnfWcskNTiqlYy86qY5/ou8jmQXlTRll/4Z03HfvdV73ve77+KSu9pa2mGuSWAWu48mPNFWTI8biJRV5ImPhm07hMxT7hsUv/+1f+1a/95h++5a0fuPqKR9xz6pyvJlc+7LpP3HZnv7fUtpznPR8mrAAsaXIaJgeNFEU2mWwVRZY5771fX1/PHUSi+VTLrQefXFw6eia6N93NXKYTEcFGIslctMSWnMSyif7sOaIoxok6AEIBFImjEqAOym3bgoSElXdsDezsPSLN3gtClmUUSVofm3rJudx7hjQIngWRDMEmVkY1jihCeJZxNfNdTK1m2sH2MDeadQoNTJCKDqbV7SmEyBIJwmKZTVBtY4wUrSirdabv6NEr+9pGGpElzyN1vcAhc0QOGpiMY2O6ZbqIBFFpSQa9fg5XTybjpvV51pRl6EppbKMAZ7IzbR1T7BwYQgql2f2a2Vyi1LFeJVhk+i07KyEG1SCBlFjZwFjDgyzLgkdVF1mB1qMax7WNzZPH3/zaP/v42/4ufvLkPlPkUI6xNxiQ9+PRZgZ7aP/yPeubR684eOPdd/OBw9/zAz/0hd/wzStXXQm44HpZmQsoxEBkfBuJRSTwaHjLB9/7R7/2q+u3feLo4uLG6VULssSJRpGZU50HH0MIIUJDCFElL8ssy9YgX/J9L3z29/2A3VfEaHPLUCCKcVamPhNjswg0vp1dOiEAIlDT4R53P9TJT00dWTNYpasbSUgFGGdjhRFTlbdZhQW4XhS+5+zp3/vT1/YcPf/Zn7/UK4fVeKHsbW1uFEVhjKnbmrM8wAgQQiyLIsbAHIPXz73u8L/98Rc5tK/+q7+97KpH9xfLj733/Zc/+nEbW5WCjXFRpz3AFLdDAkA0GuYokYhUtGqbnQP6IpKLVEd3iqJ77i8IeFJiYTNqPLcamonNwrAd2axgMhKMATEEFLpyfuIA7hVGSFl119bXIyFlpfktoI1vev3FNggTKEJECc5H7xVBkMiGjBArZ2IF2sR/QJccKYe6NbB5XhJnbGKlVTupfISsN77RJbdwYKkoXGxbLyHAUN14o6CWiJQs1GigGKFLSwvS+o3RaFwHYnBeSmZgEqMeUUevjJCUNgnSBNPlMiRQ8xTwu+etSeiyXYqeKARRBRNZZ4hIo0TfxhoHV5YWmoBjx9fHn/j4jR/76A0fueu22ybra2G0kbXt0WwhJ0zaCVhLS07igUERhiPXHyz3BufOnD16+WVf/j3Pf8a3Pwf7j6LoAVFZUlw5Mxkp1HDGijagGr/sN3/j1E0fw7ktiu1iUBIGompgQBMNKMBETjW3ltkKlFoJk9G4Gv3NX776C7/6mUeWj+SOQwvrgJxjaI0xImoYJTFF1JNKEk7901FcPLVRdoUwpolcagBmYVF26qRpF/tH2jD+09e8RWL1TV/5BYcH/bHGbGlJQtu0Mcv7o6rKy6Kq2r4t20nDbG2WheCVcM3D9/3Hn/uJ06dP3nFis/X5I6551LnVLYO83+tX4yFnDIAkBTC3n8QQI1ujqswcPOq6lTiHrbqY5CLV0TvkAsiHlDlohNtWLGcLvQVblO0mDRb7xrjR+oQ6PLFFKmQHR0TtZCIkuzMWAGeMAClGwwQlYgKUI4S8ZajJsgaVAkS2FVVJxbzTAhCU8ulm0NckmggkPh2RC2AODFuNaNpWII0VcSiKoiRaXD506t5zvmmHW5PIWkc/adrKt0v7DohXoeglIoqSBgNVGY8r+GiIVxZ7jeFNCcN60pgesK14p55a6RItEzRjWlp0zlP6qZga5rCMCesSVaIXMmRBlig36qvqTa/88zf8zVvGbei7smczu77eG40HuVlaXEHtQzNZXuwFjfVkpKK9vFeNRUMoyv69mxs/8JP/6pHf+RwMlhC8tB4ZE9s2CBuxnDGBnMCHONmqzpz4xPXvf+TyUv9AyOvW2gwyJZeNUERVUo3WZmogEpjJurz2NaJcsX//zXff88mPfPTI1Y/FIPPBk8kIMS1QWKEKQ6CA0NagT/vh5TmweMJQTx102qFEEuGzjRybtjC9pcUFhK2b7775D/7fX3tpnv2Vz7h8sRCgsMXa5uqhvJe7ggULZdZWrYGzlFWjuuwXXuLZs+uHDx/47d/+ted8xw806zIZjRDMkUsOH7/7+MrSYhtjSkSaMX8xRMCKOBvnIhJjjApzUeLQLkYdff/vsxFw7U3lM2QZuVHrJxLHw4kPcXmwD2BwikQLwxjJSMXYgkg5lSyc2/rWp7iNMs22AC3tP9p49W1TM9oY89xlJhrHQ2lZkk5mIRCJJHcECWYcEEk6KqAHfB3Sr+ZNk4TEM9YZKiyZyKGRuq5CkNjQmAeLkbVxpWYsyFx/ERrH4shAneYkZIVyJeutqla1Fc6EQKZSCVHJUG6zORUAmk9I2S6uMr1Nmhbjn0rSHJZ4SIhCFBXR2CVTKnV4kyzrsmdC07abI1bOoxzqF+1k7KsmxgiGmpZIHQkbWh+Ni4XMlsW58XDp8MGHP+pa5AbSYHmferQa2BpSjSGyCWwYELCohpMn7nnUw69a//jNS6XROqoVUU0+KBUCiVFWxJTaKl4kqIpX8UYZHAaw546fABGC5Lkji0lV9cseoiRmGO7K4RCAT8+Oph05ndtTIAOAnZZmJ6NgIQ3u4P7DTTM5eWZt/8HLN3Xjf7/0VR+55Y6f+eEfWLK4pI99+/dvbG71isyQrceNtdY6CwV7BN+w46Wl/mg8ObCv94qX/84XfcV3VOPq8oddd8/xuw/sW6qqCRtATBrkoLjdLOYYI1vjJQDIskw7LtmLTi46Hf2A7jIBg6xwSqHSLV8fb6veJYcmxkxCODVpE2N55GSJkJHACi9+ms6LHVvXrdR2snZw247a1ud5XmbZwr79BZt77rh70ZFYwzAEGFIgletmJah+WrdsLk94mm68t0RRkuBjw5SpETaWMxJLn9zauPTqR584udq2Y2OcLTMxWjcoi4GJVLZqoo/Uehk31RhNfcCUfUEvCkVB5hbKQW7thHYoFAKzQhPlM2iK09huMvS+Gzs70LaalhCR0OqW2BgCEIIINtfGmwIXYytyoL+cEeLWaDLaQiODfuF6ZdVMJnXgxB0tSg6NxnEzMWXeqP+jl//xC5/0GCwVYOHCGmQigRS5cyyI3iuJs7BlNhwOT58+nQP9/hJo3PgWBCVhEBklIhglsGqIIuwgQJDG5RmA1Y11t+/geGsENjAQRILN81wRiQ1UUqa9MSiKnMgn9OIDGAV7leZJ15kVkdL4YCARhEVWyfKinoxXRxsAxSxXGownW++/+c7nvejHfudX/3tRWOthrcvyItZN2StiCKpCRHk/ayajtvULi0suN1WLPMPLXvrb3/jtL2gmq0wVcV8RoFmy61llihZNmOu8mtT9PA+hNsDi0tL0Pl90ctHp6KnVMKcsEmx2N8WBAsJqhpsbWZE3pFqwP3rgGT/8/bjyUmQZFhbBDEORGICBgRgAWBzsfdpeb1s909R8JcBYiMDl2NjAcIz3ffAd/+PXmmqkUlFgE5PpFKFE4kFg5Lt8Hffd23my/E8ppMgzZyOcQGKAqqi0Tjete/hXf9k3/+zPY+UA+gM4QD3qKoxqG23YnExOr547ce+9J+68995Pnjl1T7O2Vo1aOwl5I31BacirjkatNDW0gyx2zEcpT27eZKb5dJX7y0c581QWpdMQoxd4lRisIWbDzP3CAsiOHj157tyEpPKhLDKHrL/P+KpqJuMic5esDHz0W1sTH7C0rzi7XoPj4vLSqdX1N/7V65727c++9su/YrKx6hZXYDiENgMbziACHyNFBwO2S/v2VwGHj1xSt1I3w0GvTJ400h3t9N4DiRMFEqMkLH6eDfbtWzqwD9UExjQSDTS3RjWkWUhFU5nXfr8P3qRU4eEzK0oszCl7FJEtKj/KSlm59PCwWqvGo/37j45GmysHLv+W7/7+7/ymr/mpH/n22JTjNpSOVNvKN4Y4zwwT8r7L1LS+NbYYD8fL+wePunbwnG/76je99f15IePxalEsey9z1eZmCynKXD72EzfIJpV3Fvv2L6es3otQLi4dfV9BqJk66IiS00MlrOgV5cRPNiSckYDHPBLXXIXcYrCQ+JUMAYlvKGEM2CT+t+0q12mdG1RJGEZJCAasgAGLckqHNkCLhaJZKG9fX73KmEVnlA1JJE2Fi1JScdeD3cmFnyFRQkAi+leCWCCSgox37rb1dRw5iJV9YowndcUiM9soqNR6Wbw2LsbwcLTQGqFG09z6V39z7F0fuOfdH9DVjQOuKJ0l5YxNutBESBkzHXYDHWPczOHImLL73/8JpjOklUDGMinIkKrGGIPqBPjCZz3ry1/4gjNnz934/g++9Y1vPnH2dKnot2JCWM4LSzpc2xSCtVDGpKqdw+LS0uZw6+DSUmbp//zyr/zPJz25d+jyNvim9sw2y3O0IbatMhW9MsTG9opHPPFJP/zvfvZ1L33Z7Xfc2SsGGwjUsaFOC6kolGBdRsbEGAUKY9sYnHPFkcOP/MIv+JKveSYGA1jXc8WomkBC7jIgWRLcBmGLPM8x5fP89PAOO4gypulX3TUnIRUrntCEZlQulJv18MzJk8UgX+iV1XBEXs9tjhcOXfbGd3/o+KkzP/cT//zy/dYBKpGMK3IbQ91Wk6IsSTn6YA32rwwmEwXRv/up73vDm98UGllcvnxzc5i5/s52deATY0yM0VobY6RMF/p94k/Dn/dgkItLR89JF9Gezdy9Qd/7xjqWKJaNYWONa5pqoez5qu5nvQ1jQ1ZgMMD+Q8gtsgxE25EWxdRbtreOpsxQyuya/5wlxJaZyUfefwhnzuWHj0TrjLO+HWcwzCoxpWVrFA1sUpJ6URTJEjfWnp+At7fQ7v+iRMuMEA4fPnyriLVWpCFnlJSCUJSUxsKGSdn2BuAMvSWyFjEos4gQCRWKjNEDAHAEB8AjyrXf9dxrv+Vbwrvf/99/9N8sZmUJ8qNRZIFhSBAAxKPJuKqqxcXFzNkm+lS6JYQgEgTKnHiQs2nq+MyhAQJERJUQUz4+TyaTgTGhrUmUUxY4k28kRgAIuW2LMqws49HXHnri47/qmc/8qn/5L/721a9+/ctfce6OO0qvywslgpdQGYegIEJssNAzk9VNAkjIh3ZxafCa3/6dZ//kz2XO2Cxjl0GgUNMbBD8ZtS0c50T2wL6nfe8Lnva85/kTp4/fcez9737n2traiRMnzp05OxqNQgikGlWJaGFh4cCBAwcOHTx66SUPv/rqRz/60UsPuxzLSzAZmIRYovaKglQIEB/JmCgoimytwYEDB/rl1rhuXZbPMqS7Wukdd64REWedTtPBQwjJNU872fITur6pJtQbEFFZls2wgRGVaqFnJuOhH2/mhMwpmgk1kilgMNwcrazsE5u/6yPHX/Aj//G//fy/efzVC+TZwGjtB4ULtUrTSKTMlqSoJpOFhV4dVFr5jm/7hpf/vzdNxlVm+9OHETvXTHzu3Fq/v9BUVWY4z7Orr3m4Y8Tonbno8lguRh2dMBu7KKQT1nbm8epSaxWIAkXi6ojsYHpwBawLZLWLf3VaY6qaueMGJrNr2yEW5rYKlo63Q5kZ3NVBYWEEJVFONQTIKCup8jyA/34753a5O84HgswW4F3GJYFJ2ACSYqHTbGMikBVA5/J0hBQGrCxAMIhEDDYsbel7RWm/4PN/4L/+wv/7z/+t2tg0EFP24TI2buJbgMpePyt6s9pO1lgRMcRsrTFGp9RJ96OHu1lLEqeSSSwYDDCjKGOvj7yHLIcaXHrZl33ndz/9K77y//78f8CpMyeu//CBzA0GC007trmR4Is+MbHhKIqC6WBRjM6deddrXvfV3/Kd+RVX8tLyeGszc2UQlBnyotf4OvoghjLmPCvI5S7vH92//zlf8VVoG3gP1S5EG6PGSEWRKlIihhgDiEyeI8+ACGKw66CHHaGfMDshRBUFnMNllx71zU0ZL6js4WMmujCf1hyKaQ7uJs65RuC9B6k1mvXcdVccftoTHlk4EKJARYJIgEYSimooG3z0E5/80Edue9ijP++mG97387/y4n/9g9/1jCdeksH5WkMU53I2ViHGZmDul3kMbWYNW/OUJ3/eH/7Jq4tipWp5D5IkJYCLIouthzNFaSfj4ZEjBwAYc3/H/INJLjodfZ5zVpRkF7YBQFqcAhxD7JbjQkoZXAktAMPoaDSgM4eqgBL2FzgvZEh7bWdtoFmqIGBBrLBerIBAiWImTR9Mhoi30RAPMIYyr511ppqn7H0RagEWqGiK4DGrpqKPokY6rJwBRNOSIaBbjDCUONWlBgEUCHawH9FDaeVLnzH441cMb/1kZrjSCJeBbZkXnrrJqgnBWdPLCkBCDEQkoszc+sY6R3TBhTxNLzVERCSp6Wn6e+dJUYUoQbltw2BpPwb74Ew9rHJiWjyYDVb+2c/+7Fte/orVs2ujM2dJMRzLSi+P5MmYNgbOCFGbtrJZnlXhntuP/fGLf+t7/8N/jJtr/cWVWpQzFxVM6LsihT8ZqhIkiil6RTlAbAGCy8HcUZLESCEk5sTOXZaYX0WURC0rWJUFSHy43U0zUEVUjaoZ03XXXhOqV2eDRQ+JD8TXMc0Ilx3Ud4B1XEXEGGL0pJw7POKKQ9/8zMcc7cHsTDpSRSBseTz76z7nR//t7915/K7+wcMfvOWuX/qNl6x91zd+05dfZ5ibgJ4bgBixVbCqsKV6NHFFzrb8vMc/xvsGNgQBODtvjDKAXl5sVhMSZBlXVXvFZZcAU4DmRYbAu7h6uy3naeoOXTu/SyczwhlSNrAuGWakYHCnoIm1M4enh+78ep96m9buiuQH16R4jIIiKAhHdOnRqYyVYZk6Xv/+F4CmocuU+ZaKvHAkjiQEZU1bpJwHdGm6ULAKJSBYYiKbOSM04boY4BZaM3tRqDz5a7+qWeyvA5WxMKaqGmI2ME1dQ2NmDRAJQgoLNSDnXFSB8k6U2A7ZBkSrJh2dWIoSPWBiqBDqMmaIaNJ6JUbwUUnzMrpSIoW6Ka+59lk/9ENP/5ZvGfbKU3V1xSOvuvNsFYkalUnQ1sAWmZIgNguGj/bKt7321Tf8zZtNui2kajgymggFWBKbHCsZYeNhhAkuR+bEmAh03HZEYky0JjBHIjUMZ+As8kxcFtlEZmEC7wpig1jJABIM8IjLj2RQp4q92Do/1fCYLVCmwORUEIfBzG3bhthoaEjGh3soIwYeCzEuSrso7VJsl6MsCfYZHM3wjKd+3vrGGTX2ms95/B2nhy/509ddf9Nxz6zGRTWi7JW9aEhkOAzROPHtQh8HDx+qQhM5RR128LImS8JYAhQUfaguueTw0iLMHDneRSUXq46eE0EHM9j7W4IahuUADYZgkx0pUzAIABZQTLXYSIgiUyAKu7ZAIOyx7RK755AMrLBKnBLZVCM0GoqGgqVoutj2TL1+2qJAl6ZHXQkiVWVNBCYEsDCJUTDtdWUEUFAQlrCz2CsDpMzg8dbYUeb6C+gNrn7GF00WF28bDocuw8ICW6teog8QNYaMih+PdDTCeExC8BEA2GZ5qXRfOnquORJjTAE5Ugu1mthJmSJDDQVDea9Uw+14KARXluqMZJldXqkUuOySr/ih79/3eY/BoX3Ht7YuvWzFK4ux3qAixMzkZekMcww91cPWvPI3fx2b67qxDhUvQQAxqGufqMZFhJlt5oRQBe8h0RgqCi5LFDmKHGWBsjBlaXo9LktxWSDjQTXgu+zK7V4rSQruSYwEctamkNqhRVx+cL+RQKl08NTV84Dnb2XAkrJEWILJnPcNIIZkPF71HksGSwaLRvoIPfVl9GWoet6XARur/gue+DmH9y2MRqOxJ81Xjp3c+Ms3/M0kBFhqotQhCpsAUmMjuOz1nMuZUUc8+jGP9NrAkpLM6MKnI01A4n3NRghtXY2e+pQnIoIBvSjrzj6kozshIuUd47uzoUlBooyoQYwis4BolwKIlCOnnSXAn15teQaoy/omUEd2Rh1tiG6TmZEK7f0Q3n+Yx8z3MiNfBtG2HQ1wJCPMnZVFs+2UwCH9vLOqBQCMkAEzGDQt5kugQTlgcOsFyyu49LInfv0ziyuvPD6uoJSXfTIOorlzUEBDBpCxqFvUnlS8jwqKQBvua2FL02pjMmdHQw3UKEiJA3MwCIbUkJDsO7CSLSxYYxVofWx82/jgBgs+KpYXfupXfzmuLJ2ohpzlTFYE1hrHJrQ++oaimBBt2+wTWbvlE3/x4t+k6HOg5KwVYcBlzjpDZFS7iU0kxBgBFiDR3wk0SDpv2wYvKinTMkC8xqBRFATDMLt4uZQkaiCIJTiCExTAEz7nOiOeHrgd3dEQTLE0HXBeFYAxhojKPLcOvhoZgDzQBvVeYtQQIcpRTPQ9Ey/d7w4s5Ihtv7dwZnV97Cly9p7rb5i0IQJeoiCSQeL4bUOrIECs4aaJRy85kurozmjLpz1Nf6X1NbEQKzOe8UVP8y04cVNefCrrouvw/ZRUg5UgEr2KB0JEIKOwCganmNTUYzLHMsmAfUAvA7ONKSKATWSJnGpaJ1J/VUSW6Wumpv+edvScWp/pfQI44SKErBoLskrJgzGNbXIkROIIimDtnvOUFY9IiF2uGOXO+UY01UJYWPiK5z//K7/rux751KfWlQ8BCJqZjIwBhAwhtMfe9e4PvP6Nn3jnuyBkiAWoATh7IcNpVvZm5pNSVXTlWCzAktw2hoRFjNa+GvRLOEdQjr50pl9meea0iQTGYt9ccek//7mfLg8fOLe14euGmmAbzQPbJspItBFDnDPC2fWrB4PXv+QPj7/n/ahqq0JtBEAEEQSJ6FjHYYkXisIojDJFaBD1SpEcudzkjpwBG8AADmzVWDVG4EA25filPEySztJEWuCoI+QEB3zBE55gRDqLYf7KfDoDgxlGAFVkWeGcqavxcHMtc9DETmic2EKzRcoGlJc2y3w7gcCRWoKq5r3y4KVHW6DxCnIhqMuMM8xIVLSI0YdQhzghhMxiUm0ldB3QVaKY2tEKqJDG2AIBiMuL/Sc87rrUSvPpp79/FstFp6OnBWWBNBy6MLcQAkPmPYBKiECL4DUoImkkIhgDUjB3A4ogs1Ta+YzsPeOGe221o0pIKb4zDL+AIhhCIpyYwAQSSTWSxORcQUdaPU8BTOgyfVM3ZQqzUpKdMwqlsh1QQHn6wwiIagSUlE2KfpIoRyGfIoQp0yRh3iJMTG2eXi7WrlfpaY8x5v2ibtuxb3H46Fd853f82H/8D8XBfbDGq8BRQIjR63h0+tZbX/zz/+HF//ZnX/3bv4vTp434iLi6vh67e8TdCRK5hGqKkVHXGzESbZSOsYhFSKbE9YouiKtFlpEo6gYKww4KH7zEqGTU2E3vUfYe/vmf/yXf+A1rlS8Gi8a4WAka32dXEDKAwbHW/aXJ6mohhFf85m/izKqsbyxklhWGEFRUY8IeeO+Db6jz0YshskyG2TCYlBmkUWKMMUaJUDFMjtkamx7JWawwsTxDEjerQDSlnxrFdQ8/UqAyaBjpOlFXJHJ6h2d3mxVmboU3j4yeDQ8l8QGqkmWWmTc3N9fW1mJaxjgDdkrOE1qFVw3qjUHJOLqfVwqjzSa348nwHLQ9dGA5d1ZEcjbOGNKI4NW3pcvEh1AHjuhnZu3MWq8orTGs0hVc7iQt3YQgrK1Bs7ycHz4EZwGALsq64BdXnxWIql4h05qtTCBKi9HQL8sgIYqQKER9jGJ5+ei+4GKMcdDrWxDGk5T4rZRK/aULqF3qnHZaSmnv7a6YoQK1hFqE4BANhJGXxovxLakXbSPFQFGYlIyDNUQxQ8gRUCO2aNt0HFFKQb5UckqTnUtQICJERCCCJELj1EcxqX0AtPGImllL6qFBEUxBbIlhKEAEQjFkHmWACQh1ckQHhYIbhZ/BBKizhEw34agSbOlEkZdlf2kFpNi3uHLtwyqpPftgdRLaGt5YbTfXz37k4+7Oex9Nlo/defZD1wNetdm3PAAkgqKSdNzRClFEpSgExBiJI+pJde5cFnxpM0DExZh5mCDaskRLYFW0LaIUWQ4yiKwCBdi5KkayJjDDFZsx8iWXP+v533vJYx+zFaEwywv90lqEmDHXFSxloYV1rL5aZpy66aZX/sav8cYaNMR67NsmRp9lNsRWJDiX0pSmtQShXUyYk4UsMCCDVGpYKQq8wAtiBGKCSxIsYMAMy2yssdGHRDrVhJYJVx3KnvToQ83oRJlFisGxY5v72qsqqahEYqnbOkSf1kPGGEq1sqbledPNi9wt06yFamzaCtDBYFA14hUVsOXbRtFGnVTeQ+E4GmSZjVItAk/7vKsXsbWAzcKvPuqK/dJsBj/uZ2YybqVFphkHyRVGqLS9XHOWQgNuu+mO0frYqEAj1EMUohADMRCmSAuF4Vhtrt35z17wrQRIqADEi7Gc4UWmo5Mk1q/ZCpEh0AjvKQZSoq5Ed7f4ChKJQKIIalOIrTNcDcBTGAPdTyDv+ULUVTWEAmShCRCbCqtocpGrdq+onV8ixpBAIJh6leM0uLR95J1BcN3+PPGMUWe3K3zTdutziJAICStIuTNwKChFaatk7BiFARFgYExn8M1w1UIqu046RbwwmMSQMAmTGhajqoKmyYXvuv5DB4IcDNqfVLq+Ce8t1BJMauq06dvRy86FmtACUo+3HDR5ZklTG8SoGFUzLchiDYlIh3ibHYNJBBpBsFAnEXz0kq997neO2JSLKyJaVd5aG6DGIUBBGI18bnjANqurW977ng++5c0YD0tHRWZCbOu2IbYAmqbJ81xnRmvXONGp0ToNBnaQxulaf/sOziBC02VWqpUFSVUHgeUCX/a0J/SyqHFCMTBoUA7gckNUZJYgBgqSNCmoqhLH6a06rzbkdIWkSkQiEqIK5ZMGEWiaxlEKypA1NKyqIJHBOZv1c+e+5umfb6pzZdwoZevmD77jX/zACwtLAHJnSSIMQSIsS1XBR0ulIdQVqkkwag3Z5FHsJo6ubcwKDaFwevTQ0pOe+OjWx4VB2bbhIXz0g18Iqdi2JNJPhWCKAA61pwCrhgHSSICQkArqWCDTAG2j6XydZoouJRA4AcdAkc7HBu185HYLE5BvZ+AqWCFBDEEJxrBYAJIcmgqvKkI22oyCVB4KUEcqKR1OQ51JkDVYpGwzUrIC0a6YFgHMogrKia2mJmA0GSsZkAOpEqeHhrrFMxuBFYRJDSXESGRyphhDLsQurUxnqDt0mW6YmxMU09UGQ8UYl7w6rMRC8ILT52579/sXyPSsrULYPHfuUBSrRMQ6Ra7vuHoEgLSrZqggrJ09Z4gQI0W1ECiTYGorQgERMcaEENIdSrzLDBiyJqiBKZTrYDw0X1h60rOe9Z43v+XcB25YtrmxvvKhhS7sH3jv8x7aEZYGWdu2Bdkzt976ppe/4uDll1/55V+6fu7swoHDTdOykgXleV4Hbwy5Dvq+w8u1YxDs6N+FGSlEYA0zgkZDeQAYeMbTnn7kT998fFRZtWhbyT2iRBUxsF3vCVOHb5yhItQAMh/fTugeAKpEcKoc1UbY4yfWL33Ycm5Ym6pgm1uWEHPOOAZlMOnlBw7k5cLv/sZ/36rDox9z+dq5+uoDRWgrX1V5lkEDoGwJrOxsGyTrufEEowrDke/3DtfKQNiGc0xNRlKMR1v7lotrrrn0yisPWMD7ltRI/JRVLx+EctHZ0dSxHCSnhHYVSYSaqo1RyNiUcKgmkV0itt4Sq6qX2K3IziObnl1EofOfwe6bXSDQ7fak6F+HslDE2GEtyBIZIsNijRhEAzES2YhhsW1dQzllFyQwlqaCjDtBhKQgkEn2vhpMDShSOAJ1FV/j5sYWs2VYmrGOTHuWjBojCHWD4JHKODLH6A3DAiYFVxVTnpPpfLHTsz99sWFHXVK8ZsxQHHvne5oTZ3pAxqyxGY23oJFJGWR4ahdPK7hryhIiCDoSDJDee/zudI+MwEZOFcF31jNkYlu3DYigMYWDCXBMqmqZWeHIZVmJIFhZ+aYXvmCSuZAV+dLKuIlNhMvzGCNbm2UIIm3TFox91tz2/ve+6RWvCMfu7EE5+l5eIKa1jklwwOkl6NJKkSz/6WtqIzOlcpWz2mY7hoiAEFSAhJwxAEQlRD20aL7gcZ9bauwzU4zjza3kwYh1yypGphetGxiz4Tc7I+aWKUzosOUEZ13ZBPO+6z9KhvpFKdGzeGfUivSyzNqsqVoDY4Cc4sOO7H/qYy7PBUeXi2R0QyI0kqUQPRyHGFBkcKZqYwPccOOtEg0hi54VdlokSNMDxAqQMKOpR1//7K9NvIEhhCIzMT7k67gIRNOKmyCc3LSddpxUdSskbKIxnhFJ1KiSePFiNDiJRqJ6xKYbT5SKCQIAAQZqNJq98ZtzJMnnaWplk3B1QATEhwaYJsWwSagJUrbCNpBJ+kppMq6RVu4CaLLtd3s20ssIrBCpofmUEFIyEA0gQGVtbY2VjbAJZhpynPONKJOyb9qU0CyMAAQopVRmCCiAAiBKCdbhCYERiSKx0hSdAlYlFeJIrFBEgW+xvvn+N7zpUJZlEIWPJGrTRDoHO5nv1Mwp0K09BILjd93JpNDIBKOw0vmguqLjyW2gOplM5qk8DWAS7MABgHOW2I5DjIzLvvRLnvQ1X30mxJGoKwtLqIYjaUNoAzKMfOtyIrQDo8uIn3jHO978spfl5LbuOUkCw0xEddv08zLliE6TnrY1cpcvsq2at1/z/Z13VxHbacIrVCUj1hAheNaXP+PwoCg0lKTq215eZFnGzFbJdhlIqagNYle5BrMJD90sLinU2JncICJ2rl8H9473fbgNELJlr+dDG5vKOoreO2OzLFPVth7v6/X29TITYVv0gFhVlriwDgAblxCHkTDxDTI7qqtygJf/2f8jdtUkqLhkQoAJ03FCpAba69nllf43PPuLIe1otFWWJQBnL0Zfx8WloxWIlHLaKAHFOu9l5KoOIaowizHSpXjpNpGFYeOsxIjJBDFMTZJuqKfo4R4W0O6TJ+G5LVRIoKJdkdVYt5hapslyEnByKTIRa4J6SDWZJB0tGqEJpQ0zB7HAVJ3RTFt30MDtOx5FAEGM6+vr6LhBZNc+018RQsRkDIaSCKBsCLQT9dV5pVPMkCGM3fQ929atqG0DVte33veBe2740ME8d9H7poahhZXllG+OrizL/KzGCoamUk+GQVCFxnvuvouZE6/K/I2ed42HNoyHI4ggVTmc1kgggwh4H0CIMbo8M4tLyN3Xfe8L2pWFk8OtfGEhd1m91WSWAARF3aAc9CSEemt42cpibzx+00tf9sGXvXx5cQVta1QCVJgiOKjMLNYuow9TaEKHNSOgY4rp/t3LUlSAjUkJTQQSEQIyZ1jw5Mccue7hR2ysjLSW2Tk3g/mTMnfIH4mImK2iuvsqU7+UkCK5jlSVQVEpwHjKb7nj5PU3ndxqEWDVGK8ioYWJ3jcATp48mWVFFGSM8UbzP3/5f+cWB5eXWTnLciYLZZDxUYhdG0IkWwwGd9wz+eD1N/pIbAvnsjTC5/qa8kMDpPmO531TYVAWzkCr8XB7iF1kcnHpaAARHMCRO0uNOxgat7WXSMJGTFLAYqBWxAgjGkMud4Vpo05qqE4Xhd2AIkVX/HSv6zk1hmiq74BZ7oBCFRIRVaACkRhajsJRKYDjdgQuWaNsICKI0kwqSACihwhBQDO7LO1/Hh12dxhFZ6sIWEQgihBGw62UrdalgUyP0OEC1SSvdDupoGkhoGwcARIUMInnI+kaUu6Qy8osTGrS+dJKg0BBlIBMAjdVuPPud/75X+STCY+GHH3dTMiZ/QcPgF1Kl0iB3Xn8rIATo5MBGQWCIMq5U6cJImAyrOlloZbJMAxZZkMcvJ+Mx4gBCiKSDsiHQD5wIxxAEqN3zinbBjr4vMd84Tc/G/tXhm1wMJkiN9YQjIGxaNvaGiKBa+N+5nj89F/++m+G62/A2irUT3ylNpsgROJt5TsLOKelBeZM2i5iPL30uq2MFEgg9JiiyUoEY4gYMKDcokf46md8/iUHehzHhnwbWx9DkAg2RGbe6cw6s6A7zxulfNcpGa9XMMQY9r6p2obyotb8j1/3tzffuzUEQj7whiahNsa0sV7bWL3kksvGddgahY0R/vkP/9Rfve6d//7fv3htM7aehCDKCvUS2bgAZduLkSYNXvKyVwZ1JivYOmIrO20ChjI8kX/YlUe//Vu+bmtcE8LCYFDXNQjxYkwzvMh0tEz9t7KdZMWkgCi8IorpWAGABNJUWBgWskolTNYGqlqESCpApz9nK+r7UVlwt4JOzx8pp8RFaETwLMqiJqqJarq1qkQSMaoMRYSoNg1iAGlADNAATWYa7XzSZV4/YIbI7jQAiyJGxBgmE4aoxmRtzSl3SqQZRtgKpG6hwqIAGWIoSBKtEc2hvGlu2b7HBZHonQQXBFUbT5z86FvfeqgsxxurIqENbcxMsbIInq+5kNwlqYLMrjBXCv/FydZmZ58bqwRhiYyuhpl2gDf1PtbttAp2d3EAqASIOGfatiWiGHRUV1l/odX47Bc+/+ijH3l8Y2OiMLmtW+8jiiIrCqyvxyzLlpeLrbUtU7fXHDpoz6392s/+HFbPQgKHOqImqL0Qkeb5vFs7Ptl5xQAAIQhNuZNS7n7wjQV8VX3F0x977ZUHM9sSeY2tagyAJJQfJTs9uUi2K1FNgULd+dLiMiqElA18aOumNlnPLe5/47tuuP7YqTFQEdXkqOy1iGTsgYNH1raGtsgj2R/58f/0sU8cr33xxje/9x3vubEK7BVNFK9ofSRyVRWZMyLc+smzf/KK15Lt95f2bY5HPgbMPYYJSEXkGc2XfvGTl5ew0i+atmJg//79EqK5+AKGeNDr6JR+lhKFY4waNQRpW6/gum0UsIaqzSEa/96/fftSVjabozCuW+8lBAm+X5QaJbYxJ8feh3PrG7ffhckkjIZtW6cqouI1ObZD6733XQa5br/S0zCZ1CFIwpBAoRFNHauhZ4U23hAjszraPH3v8ZTc6JQKZQtlA3KEnCUDMmxtbRQGN7773ZiMEZs8z8a+iUy+jdFPERwQkRiib3wTU2SNESKiJp0GImSGnWEMx1u33BYmw7YZN6EW1rZtRSSEloiYWQQA+7Yt2N7y4Y+iaYyqH4+kakId2VJK+BNB66MPknwRwQuYfIyiUEYbYysiBO/bvrM2Bgy3mptufu3v/UE2HJGvssyAVXI7Zuk97AoojLMxhpQ8rcQdyJARRBsfCXzuzCpEoajuvOv4ncetmtiGpm0jNEJVo2pMDpyO+aQNx++8A02F6IMGYaSEx9xYp4QoxhhmS8bmxSCAG6Hyqqu//Qd/oF0ahEG57oNbGCyu9MdbbfQoCgzHjTXZ4UMr9XBrcubMMvTkh2/4/f/48/XNH+u1bR+qoZHQqIpCosYgwWvsAgrbOHWdhjdkTwU9E2sdgZhmKa6au8wiLpa0YvH8536Do8mBlaJaO9lf6pE1VdXAWGsy3wQiir6NUWOEoMsFTcGYQBpAATT20QNRpPG1alzZt299uHl2ONnQ8qV/+ebfe+0Hj4/hs3K9lkZYTVb7uLS0tLaFH//p//qGt37ADS7Zatwd927+9//9e29/9w1VgMl51IxrX3uVflmo4FWv+sDP/PQvNU3uiuU2MltjMjMcbh44cCDGqFEyi6Ya7lsqrn74kR/8gW9RQUQoXOajj9GDwnzBw4tHHuQ6epcQ0LecwbJoaTMDgUrJhM3NrKoHkRbJLJhsgV2R5UTkIS0Aa+q2ilW1wvaO6z+IprHattUWKBgDIo2Nj61nx1mWXciY7vUGAEKYovYM8tyYzIDhCkccw/oqAR99//stS1FkLnF9EgEaNHppQ2yDr/ctDnLRevUciKE6qYZtaNros8xa46CMEGOMIDGMzBk2EkOYTCYpGikpwRaoJhMwYVLdduOH2o2NQ4sLS/0eQwBxbKy1lgGVGGMUIRgN8dypkxCBxkGvzMvcFgYMH2KIqoAk9iUyYDKZU6CVOG7roGDHrZ807aRwNN46h9Bgff31f/yHG8eOLTsjoVESZFYLVx7ch7LAlJ+amTsCbkFywxhjVLWumwMH9qFtEMPHbrjhc6+9ev3smeXBQlsrKxJNVUcfOI2Jlc5++AMfRF1BJPq6asYut1GjIctqSIgT7JhYwE3ULO8h6lVPesLTvuFZZyC6srheNeNJVRQmy9gYNC02N4fjUVVm2SWH9m9trV6y1L/tve/+pR/7Ud7cNKPJIpmesTMwnTXWsAkSg8TO/TMvOl0r3KfseFxJCCK+Fm0fdtny8775a8frJ7OB3Vw9kzu2eeZDlESTqCqCIh8UGVpCY0zLectZy86Ta8h5OHE5GOqsV6nrSQhtVgyKheUjV1370TtP/dbL/vyPX/PWE1swRS9wFuCU3d33jl7won/59nd/aN+hK+49O7z2usc/8rFP/OANN/2v3/r9d7znpvWJUtZbWT7YBD5+qv7F//oHv/rrv3/nnWeL3v6TZzbGVb24f2l9a/XwkYO333rb/n3LloWlHfTM5tq9L3j+t1mCJTUiyatDmHHXXHT+jotMRyuoQU9MnEwotFYjJmMMh7jjbqxvDOqwEJH7WBpnQG2Iw6aSXqaDYuIr9c2Bsrj9+utx7z26ubGw3K/95NzW2Uk7Mj1jSg4IMbn5dIa0mwvli0LUWgYgUSWqqAZtG0ZFPvraSouzp8988jYDEYRa20p8i+hVAGFIBikAbZoFY8zGsPnIxzDccjEs9xc4BecksXR26TmW2IFj6y2hV5RlmRvT5dk1El1GaCZw9oa3vf2yheXhydM9Zqk0Y0NRpG00BmOMyzO17CGtrz55602YbCDWkeIoTIaxGfm6Dk2gCENsKZJUoZ6EetxO6tiYnAP5zWpNEMrCFZmKHy0WjNXTr/0/v/XJ9793fPKekiIzomPJTMPm0kc+Ev0BiCFgGGgXCGUGA15EIFnuyjJX38IyVs9+5P3vbTY2l/IeCw6uLCaXq0mlWLQD4bHKeH3Sc/bMJ24Nq6tlZvu5nbSjUTUJkVQ6l/rMLezYWc7hCvQXXvQffh5HDlX93sRH35L3MXglGGsAcIyxjaFqqgOLvcn62cW2wT33ft/Tv7i68aNoPbY2MZ6oDxrFt60PnpktW8BMiZJn23mf1K5IKXCe2yh5sUnVGXYkmbTf/a1f+rhHXtaj1qGuq6HCR40wnOe5YReiqi1OjTAUbEZsCjaFhhFbgqFgExgC64JNH4rBQtbrhaitD02kE8dPhqzHgwN/9tdvf96LfuJ//e5rzwxRR6xt4uf/86+ePD3My32cLwxWDrznrW+9695TBy+59uSp8X/6xV9/8e+84uyq32zwyle+/tu+9QV/+udv2BpGmJ6a4pLLLndltjlaP3BgWUMonF0/e6pfmOhHlqvnfPszv/xLnmAJRgMTTbMHZqU0Ljq5uHJYWKE+UGH7xiXQL1jRjG9573t4OM6y3IoqAhmORpqoQdUZznu5DEe5cab19fGTp977gSP7vxJbm+VgUCwuqrISi0oUCCS7wDgKIVhrwaoSFWKsBcQRKyNOJrYaQunOd7y7P5kYX8HaBsFAQ3IcQllhFUQS6/H+lX3rE/++1/z1M572hWWWtbG2cL6pLRKGiZmZQCnmJo1HzjAIrSobWDYs6msTWkxGOHZs4467rrG2arwtYs4IdZNnnDlriX30bQiU9crM9SxvnD4ha2d5ZYE9OQIbstaJqiEGOpZ9ZiYyxsB7XxjuF7lvIe2IVCwIBjh15k3/93dueM2re6ubR4oib+ookhlu2Yx8e+3jPg+9AsxCxGREp6U3CMQwBIUkPDmiBynK4kPvelc+qdC2MaIKws4AAM2D90DA1Q878sm11T9+8Yv/1W/8GiZ1XU9M2V/oLdiUAgQgovIxQJnZGiYykICFBSyUL/jX//pP/sf/qM6uL7isrcesYOcyQ5mxOdtJM1rdHPV6dNmR/cfPrtabm5ddeulPftd3vOAn/s0Tv+6ZuORKJmYRGII1KVe/9W3uclYCYdui1mlbseODXZISd0CScD855w3AUu1z2Qu//es3X/JnH7vznBhk+cAYG3yw1uZ5Vvtyq23+/PUf7CMYNJqYdBN5M7EQ6uAnXs+uh2Hre4OlwBmFqODe0csnW6unTp9dXiiW8/1/9lfvfOc7rr/q8KFmOHzPuz4k3OfcrJ9a6y0sX/WEJzjme+4+1uvtv+v4iZe/8g2vf8Pb1lfP+lYlOkbfFoM8c5VgYzSsfCsiTUtSt/tXFsab623lJ6MzT3nGk//FP//uXgELZVKDrtznP1ABz88Kubh0tJBwz06N3DZORgYGq2c++t6/KzlYtmzVs6IwDRBNruDReNIrF/K8zLP+6tZkuSw+/sa3H3n0o7B/xRofHUtWRqCOCILcsEAN7UjiSmIdg1QliERjrWrTNBUAi3zgGBstbvrEB1/+Z/beU5ctr8RJxZkBpMtMj0IiEGGRhawwoiteTt/wUdx9HA+7bDycLF16CQx19KpMSqQxpUlqZjLxQSPBMiFVRWy1HjEzQvuOP/+zFdW1u+79nEsuG5261yrKwUCFpKnVGMcmALVvJ+2kv++AUalPHO8t9mNZ9g4ekthM6uGgWAJbAATLdooqEVjDJGphMjUIEULYXA/3HH/rH73slr95277T6/tEj5blZDTs9web3gdBQ+bhj3scsjwwBWLH3PqQsY0+kqjLLBEUGnzTtm0WBTHe9o63xcmohB5eWgobo9r70lohVggIEdO6aJCCuQz6sbe9/aa/euN1X/NVS/tWYmjDaFNNwWwNG5DJs+2glPc+z6yQbpw89fhv/uaN02df9b9+sxGgDTlrZnJtY2wbFMgLY2M8eumRm287ubhoBnmvqiY0Gv3uL/3CLTd/9Lk/+K8wWOSlBfRyeG1iC+eKot8xRStt4yO3wZszC1q2sXrpy3n4R1f0hyB6cLCw0TZf/LjLTnzNF93zu382RssUrTVt22pQKvtqi80m/Mlr38JtTYhKQiqJaCmmLERoALlsaX0c3YILYpSMRp2M6l5vyZUlh+bcxno7DNZmK6H3tnf83TUPe9S9J86VbnDwQLG2seULf8cnb3/8k55y913H9i1fEfzkzrvXer195Ex/od/UMm5FTWwRR3Vlc1uUWduMF8teNdw8uH/h1Mlj11x15EUv/LZ9y2ABIbrE1QujCNOKNBejIX1x+ToAgAShidUYIsY5ue22G//mTeN77uppQFNF8VFDK7GNQjAGrEH8pCWYMssXXXZJuXDP+65/7ytehdvvwbmhacV5sQJHnLGxuCApvaqqBECNZaBp/ZhNLArLa+s4dQ43fvyVP/Pz5o57Pv/o5e25c001UkRVtQIrsGosWSZL1jljJmsb+zkz61s3//lrsbq1sv8gR2I2xDYVPEzkDAIomLI8WmvywjpnDTuJPdZ+r8Da6uh9773hTW+WtdVCIkWJSh7Y3Bo2viVKbuBoVMrMHRws6ubG6M47b3z9G/0dd7nhBKvrvDUcuAxe0HjUDVU1xg3GTdisqnMbxiuPa4wqVA3GLU6cOvamt/7pr/yvd/7hK/aPJ5e7/GhZbp0687BLjoTxpOfKKFg5eHDxyitgTCSjbBSIKqIgVucMIKoBMWbOZmWGIg8njr/kt35zMXMymfRdVtqsn2dIPgBK9WQQDIShhHOnTh3olYdc+dv/6Rc/8ro34ORpM2lyZufYUERsoq8kNhS9huB9G6CRzLmt4b6jl2rdfulzv/NZL3rRsN/fhBkHIrZ5nmvUpqoRo3N8550nD+/PYh1La0Zn1g8VdsnXH/ir1/zL5z33vX/9OoQGvpXhRi83vSKTUBMCUgAPO0DL58me7td5Y9tZdhahb5EDX/clT/qWr/vSvg2TrbW2HpV5rqqjcTXxiNng3i2/FnrrobcWeuuxtx5666G3EXprobfqs5EUFXK4XhPQ+sTzxaRmsjmOwcAMkC0//LonDw4//K/f8r7Fw4+45e6zNl/e3KyGW9XKysGN9dGRy6667dY7Q2vqxgCD/uCosUuVz4+fHk2CobwfmCPxysEDSyvLRJpblxu2Rs+cuvfaR1z+Iz/8vU//gkfHEK1RC+UpZ+5OBvOLTmXRg3sRkXo32xJE69ogoBlhaz2ePP2xN7/tY699Y+/k2sEzG/tjhEw4M9Fx5QMZRwK04pyLhvK8LEwxUXPbcLi+2H/EV3/pdV/9Ffuf8gVYXBRruddD6cDSpRpjGy49S3TxbWUdE8WmnWQZEwjDCU4Or//9l1//ylcunD77KDZZPRnsXwgaJuujTMmoocRGNE3fjcCkjeXCgWNNvXXF0Sd/1/Ou+cZvkMLxyn5kRh2LsWISIROpkjBENGM1KtxW8A0mWzh77uNv+dub3/K3d7ztncuNf/i+/dX6xmDQmzQTlxUE8LjKwETUxNZbmkiEKzaZz2aODh46eO21j3/a0z/3C5+Cg4dQDJD3EjgbbGAyCCF6mUyYCBqa4/fc8OY3v/uNbzh18yfscPOSzD18eXFBsXH21MEDS2tnNy87cuDUuDnT7+97xtO/7jf+F1aWasmIswzU+qCszjqjgPoYWmOBGHU4Gp869dZXveqvXvJ72cbGgRCzsV9g69vAZRkY0USlkHIMUy2o2God9OgVD7/p+L3FZZcefNS1X/u853zO078Igx6KDFkOJiiUWI0Rsg10NB71i3JgWEcjqhsQ/9F/+cW7/u7vRsduL6MsO6Zmor5VwDlaWVk5dWpNCVWLA0eWT5zbWD7YPzGqepc/4q6N4ZWPuuY5L3rhdV/15chtENjFRXSp4TahFaf/YlajcU6mYM3tYY1ETqvgKMYYDDfPLS4trNaei8FNa/i/r/yr1/3tBzYbt//QlZvD2osqYjEoJLSmq16faKeADjUqja+jBGccNBMqhIqyXPLe+3ZrZbkf6mo03DRkFwe9nLhaX63OnjHeL+U9Dbq5vpnl5crKiohsrm4eOXTozOnjLqOyzFe31pdXDuzbf/Dee09WbZP1cpOhbmuXUcE8Xl8fGLfUc6Hd/JEfecH3Pv9LochJNLaWLCW2aBKVIEpEhvjCfCYPVtEHo47e0Z/kzU0526KQgI0NNBMM1z/+pjde/4Y3N3efKFZHV9m8v7rZa1vikBXOWxo3rXOZRKgPxphIaOpqqbe0NhkevPzaj5452R5c8cvLh6579OO/6isOPPUpWFoAAYMCvXLK2k/aJVwkijptx8Mss1CPegRr6mO3v+fVf/2xv3zzkTosbQyvcpnce2KQ8ZbUgdSJSahkgCLUM4RVCAsL/RB0OGzPqYwO7b/DN4/9mq99wpd/2f7HPha9Hhb7WOihl4O5gXgFyGVAFiImFUKLzc073/HWd77udeN77qruuac/bo6WPRmNmvGk3y/r4I0xmbG29TkZgoQYKXdBIUo8GNw9mejSUsyLExtrC/tXaueOXPeYlSNHLr3iyoMHDw76C8zsa982zcbqubs/efutH/9Ife7cQNRWVR7CoV6et021utoj7veKs+treW4RDZaXj2fuSd/z3Y//8R/G8krjYUzOwkriKTJzFgWthyOoyImTt334xr9+xStv+Nu3LCFeMeiN7z79iMP77z2+enjf0lZTe07FX5Oh2vE95TZndmsbI8lLXl5c8w0tLTaGr3r0o65+zGOe+PSnXXHdo7G8DGNgLVy+6Vsh7tlSoy/IYTIJIVjfvvbXf/W9r3n1mVtvPZwV+3JbQCkGZ3g4HFprvQ9s2XvpD7LRVkt9swrLK8tD344kPurzn/xNz//Oa7/oGRj0Uu3dKTmp1a6mDJ/H0D81G3fo6ITVE4Udjds8zyyq9a3Vxf2HW2SnA85W+MlfePH7PvbJlYMPX594Wy6qy6vQauNhc1IoS4qsErqKgnlhx+urEC36y2SKppY8G/gQbBZVAzSyom1D5kzf5qP11Xbt7IFef7y6YZWOHj48GVbn1lYtm8WFBV83i0v9qhpPmslgcaFpw3BSuawsiqyNvvET68iwOMJSbrdO3ZVzeNH3Pvd7vufb9q/AIBZGAYEwMScXkIqI6kM6+rNWtnkluv/S4rBu6iIvDNCMJxzVWRcntWnr42943W3v/rub3//BsLa+gmw56rLHknLZ+EwCoBHRk0aWRGVHhiHRBmEv5GOE1OCaEAdlzTQmjIl5/9Ilj7r2kU9+/PLnXFc8+Uk4fFTawEUOl23Vtc0LR6yTceYDNofYWJt84tZ3v+Y1H33HO91o6zLn+qEdBOkJclEmBKvKhAgHNswAvIpAmZkZ7WQIIBI3xk6sq41rjGtMtl7Vm+Av+45vu/wpTzz8lU8fU6DD+1sweSxqRpOgtx/72Jvfcv0b37R+xx0DRGqrLLZZ8C4GEyPFwFGFYPIM6BARpDCglOvRNE1g9syeKTAHBpg8IzqOhARcSRURocoKC2bAihgVK7CiRsUI8sxSFHgNIsHaYDO12bgo7rH07176u3jSY7GwEGGbOljYrHAVvMaQ+2hC8Pee/q1f/K/Xv/VtRVsvsSlim8XWxWDVW4mkUAJMHrnLT8RcrIlBqiRKUaWBBpWQqlBG5L1+LLMJ9LqnPOWZz3vOlY//PBw9qtZ5tp4NwWRqjBoSQAMwueWNr3vV7/zBLe95/0HODmSZjqvYTKwhQ7DOOIIVMSFQ4xtgVKK1FKLWShMA/cGVn/uYRz31KV/znOeao0dQ9oOSzXpN60VQFr37s5yfdkpSBhVIoT5pbk9ZCzTArSeq//F/XvLOG27hwaGRFi0KFAveR3LWZi76ICE4ttDomwris9xwYpRWy2qgFuqEUKMVCoYIooY0M6ZvssJgX1GeO35vszWsh1saZaHXVxFjaDwe53keQmuMUUKMarOibYOzxXC4ubS8WDg7Gq9DveNoZeTaUz/6g89/7vO+df++rPWSWRgShc5hOPi8NxeTPKh0NDq2g1kygPetBUnrc1cg6rve+KbXv+wPy+PH7No528aVvFhUi1FlRlUWYs9ao0IqQhIJghgZAJxzViQPyJPdwZqSsuAMjBXmCnEksbbse/nmwuKlX/Osr/yeF7hLjoYY7fJiA26hBugpYWPz7Ac+9K5X/tntf/t3Bxp/WVHyeIvarUy8icqiUVWgkSGq1hhDnJBnU84zsgxoC0jy6ymMwCqMwg29Zgf2b/XLMyU/40XPe+xzv2lcgIuFEvno3rPX/9Ub3/OXr9u66bb9iiXVMNooMjIaDKJVtVBWNUSkaHwLgMGsapQNASAGrHUCiYqEVE7F94TUFjYiJobiGUkyAYl0rav5lG6OQgnGZiLQqF5RG9dmubhiqyxw5SU/+Gu/gmsfplkWhFRg1BnH66P1hbKwbRjffe9v/edfvvU978snk2VmrieZBKORxU8X/hFAVDM/pGfvrbUdi8WcsKI0WRBsBX9qNLIHlnXf8pVPfMK/+Lmfd1dcAWs8ZQFMYjgyCwwi8QjwuPf0u173hte+5I9uvf7GlTy78tDhth5paFliSTpgW4hy5ZvoQy9hfSyz8TBbTTs2TlYWBw972Hf/6I9e8wVPxfLK5nC4tHIoAgxDD0BHA5iy2aUcQmUlBJiWaORxehx/+dd//1Vveue+K6+r0dvYaJcuvXyrqtR72KwsSkRp2oolWssSW07038qAZSWoFcIk1JRxZqxhQvDwkbx3MV524KA0zcbp0+1oFJsWEossDxqImZkTqUATAjM7Ww43hoPBQpnnp0+d6JdueaknocpztKMTP/vjz//KZzzx8iuOeh+ZuyKQRZHthLnsoLi5uEQfpLiOLvM5xgjKnAtNfcctt918yydC5sLhQ7TYC40/2fiT44o5W9i3sFyUG95DIiSIxBijFy9QE+MSXB7FxmhESTVp8PRgMKshIpMplET9pB5X4UNveNszvuzr3IHDvprQQl8NV7F1xAXbu26//S9e+od3/917n3LgkqVxtXrXvYeWBtlCrvAiGlWjSCANUFVtoqRKHQBUZ9VmYTIoiRCgVpHY8i3B9NREW26uj06dGr/l1W9+7Fd+Tf+yg4imGY4+/MEb3//+G06cXc3yvBZlFbu4WPnKgg0pQ6c4J2IFmcwoIMQqLASJpAwSJw4kAgaJgpUFYEH0sEod/z1rilaCFZw7IMH/5mhCSSvvExOHZ1MbbpiFzRD6iKsejl4fbKIKs2M2Uaip6wWTW09+bfLh991w523H6kmzb7BSD4eFK4LEoJHhAFV4IbCiMHmXW5/GwMyO3llmaZqKL5uTOsuysre4kpVmaenY2trwIx+/9+7jDztyVCPUBDaOQNamXpimlrzMcOllT3vutz/p6V/0wXe+6+1vfP3dn/xko+yyvIjqCRLR+mgNCbT2ITM254yjyWCZ8kzQDHHsPR/5Q3nx93B59Rc+dWnpQNs2ap2Pvu96D2g5vzOHXKBMFAxsZmkxM//mR77vqmse+fuveK1ocfTA4XNn74YtYBx8Uzcts3XWkjGqMQpiN5emPCdOJB7GgiEGniKAaEmto9y6fpk1vm4mQ0ggxLoe93NHzBGk3BXCTZ4J0lgW2YGVpXvvvuvA8sCaoGEY/LCX5//8B57/7d/6rEHe3SAiJSJr7dxD/JA86OxoTN0dVV1l1lo2FIXYoG6r4ajMLSZbkIigaBs0DYJPiiRubpBEiV5CjNHHGEWjCzo8fsaFxGsRREQ0lTeUM2fO+Kb1VR1bLyH6GOrWbyqdRP7dP/4TR578eFjCykINTNq2yFwJoq0RNsajD3/sY2/82+beU/ucE197DUKBlUEibCKJEgNCkRhi1XSaEQI2QsKsygJlJSJJZBoGYGuKoW8xGLQL5YlYffMPfm959ZXIHNhN1jawNe4poWlx+sz6yXsnm+uj9fXgq3oyHg6Ho+HmcLhZjWtpm2ZrBFGNojHCRwlRRVQ18dl0EfapV7Aj86BOFZKoTkn72rblqQaZ09FS9ItkcbdEjbHeGTHZJHPf9kPf/6Rv+UbsW2gsZ5QTOATxTVs6ixhkVN18/Yf8+ubGiXtzLxunT99z7Daj2nGiahRKNUc41H6+yM7szblz53SugngSFqm3tlTV9fsVqfaKibUPe9xjfuinf2r/NY+EsYFZYFQMK7OAIeRE/DjWrbMGxqKuq9Onzp289/r3vffE7cdu/+jHVu+6x1TtsnELNndGlXxsaqmDVdMzpRJPBGNgXGSj0h189CN/5U9eps40xtaQvCwL5J9SOe18ZpMdPXPyITE314LNWlC6CviTv3jPS//4VaNGe/uPbHgbTNY2osq9ckDWNU3TthVlBtBtjmlKFTvVqFgoRKHRgArmns1KYx925MjNH/7wmePHF/OCFbFtyqIw1o7aVplSUQXHJoRgyJRFfs+xOz/3MY8abp6rJmveb37uY6756Z/+sc9//CU9hq+CaMiyLE2jqlERrZk3Hy9qO/rBpaOxraYBtL41xG3dQKQ3WIAAvoEEaFcdD9bCMiTAN2CDjigf3RaJXiODchelUZkS/EcQIdXFS9xmIvAeAiyvIC90NGpizAYDznIwiQiCpxgICmPgPSYjhIjMoa7B6PxvxqQSJUAqGpjQsArugpBAl7kwdR8whDqG4qCICgWKHJaRW89qe+WkbfMst4k0zLfwHk2DpkGWbSezTfMiAIVzUIUIQuxeiR5vYxPATh4SVdWtyUS2YTPQVHIQSCWf54n2SQFStpQoKwTsDQuzkgnGLRw+uHzNIzzBqxZsWU0y5k5vrQ4WFno2I4loAwQYjRE8cgfotAtTGgcCsnJubOiO9+nf+a1EOIPgESOqGnmO/fsSIjkQR7IRBDCLIWUjMCq+rUm9zTOyBlE6V05ounT12mNt7eTtd37ixg9//IYPH7/rdj/Z8uMxN8GJsUKqRLagsrhrfe2S6x45uOySf/vff8ntW0ZRrE9Gg97Awj0gHd1VV5hxRZGqqhCtbo6WVvY3wLkarsB7bzj+f3/vpbfcc+Z0cNnikf7CighNJjFBNsd1xZYk8aHvYNQCfDAMS0wAS7QiTikDHv2wh7/77W8zMfbzTNtQ5Flb1REKw7VvnSHnXCqSpiEGX+1fWlw7e2L/cjmZnPuSZzzxZ37mR6+8tJ+CLZmjtg0xxjx3IYQQ2l6vN51y5vXyQzr6s1QumIyFqqktG+ccFBARHzjLEBWqknQuM1gVMcRI3NE2EmBUVJVVAAvKAZs4IChVEU1RmsST2BUVTIz1EmI0g3I0mRTGubyHGKEEZ2PbGmfAKjG0oSnyDIjtZJKVJRIA9LzVAKJsq5hEO6o6bRXQVb5jADzlau66FoV75Wg8hLGwJlFnJG5QaGQRKwApfNxW0Ilhlaa+Z8xUmELnNF3nIJjquC5vutyt++b/3X2zBNJOiaEBNpGIYVKpPS6KWiKTLdkiCCIhM2fG6/2FBQYsKFRNWeRoPaxDDFAFy5THLs26M58GA6JdKXEGhMjM3mP7fYCkypAWBPGhDsHlhbBpRRUENoasVTJkrM4PNhUNbIwgeu+tYcMUJ+N6NDEqhcthHUJA22CyhbpG42Xcbp1brydV1usXi4u9I4cDwx7YV8WQ71uuJURFrxjwfZTLmp17l47GjNNOQCqqSoiiwbgIE4kCQMCt92y+7M9f9+aPfPLu1Yk1ZZ4vjCeBTNYfLFc+RJVIszo+BGIogwQxOmZrDCsQA3tvolqRyw4cuOnGDy2VfWmbari1b3mlrZu2bXsLveFoVBaZMRSbuiyy6Ov1tTOPuPLyerxxYF/5Ez/2A1/xZY8bDYeLC06iz22euwxAjBqjT4PdWvuQju7kwa2jAYQQQgh5nkO0qqper9+FFKmzEqKKIEbVLFWHhxBUIV1ZLJAxTsHTzGKBqEnPqjWAeokRCiLmRMgZvUTvfZnlzljftiTKzF4igAjNMgumNnoRIdbgpSjKdEydwgQBqKq1NmVXg4k5pRMLVFNcCHMMoqwA2BBXVdXv9dumSa3Ni6JTmIwINKEhotwYQKXx1hhAEu98Kr4hYAOwRCOYur9pxpofvE/OcUocdNNytyC3t3Y2ez1RKjOtCiYQK4GUlOC9r0M01jqTZWQgihgDlDITIQRMhqM8z4ssh25z+Xc3pQNXpjvfUVl0sAeN8zpalVRj2gIMjZnjtGyKQYwxMA6AlwA2aQYgJUhKyWdSWGPqpnF5BqLhZOycKfKCEFkBlW4Jsj0IBSRo244ito1ghmGIIs9hua6qot+v2zbPcug0Y+NTDvnzdLR2FP7p1CJAG4Ia62xRSxRla6gRrLV4/Qc/8fq/+8CNN94U1RrTrzxEmWwR2MQu8ccorExJeg0FpmhAUKEgFpoBGSGOxuONNQ5C3reTcS8vyiKfTCa9vNe0de6IKJI0hOg4rCz3m8nmc7/t2S/87m8tS5C0g54VqSzbpo5EJs9zAG3bWmuZu9K35/X7IR39WSoXaj51FBmq2ratMYaImA0pNEBEyRIzJJUjoe7BmhYIldkinZSTlQdANabxmwpqeBVPCucMMYAgMTRt7hJlHAFqQBKjxGidU1VJXoupHolARMzgZrQ6u3CEqWQfTblkaGrF7tTOnXgRaxhAVTdZljHTeGu0sDDQIFFFUoUwRBHJyDi2ScUmJElIKwglVpSpoknnS+nONF8ILyHNE+MqlGFSVHOXMa1mr7JGBITgTSomyRQS/lUZJIZciMEYqypWQMxQ9TGAdWtra2VlBcBoNOn1eoZNlDj/DHfecIVAjTEyp8Vmb8xcsdKZpwaAZcTonbExhLYNWZaFEPKiaINHmsSIUm3LlGkSgwq0KwIDYWYGSGCgSHMqBGzJQEARUWNrmZgsooS2tVkGw23TBNW8KBWofV26QmJgZZNYrz6V7PHMzuzolPNBYGMArtsGbK11bQhtkJBlE8Z6wBve9IHXvPYNp85uRSqayJT1WuXAxpMTZJGcqIsphSROLAuDSIVEjYpTzSFn7jm+kOej1dV9S4uOaLw1XF5aGG0Nc2OdUdXA3JYFD7fOLA3yr/3qL/nBf/Y9B/bnmQEDliJzZMRJVRdFnylL5ogxBkCMPvHintfvh3T0Z6ncd/Np9y57pHFtP+wy/XJHDRMFx+na2CSdMl32CnfRcOqgqp02nWb4yi67iABM62enfXhH1XoA25Wu5hd720VXmQQwwKwN262fs691lumYqmfTzGXbVT7slGlXb6XrHSUCF90NNt85d8zNKHuXdrov2Z6KptnQPHfkWQsx9TGbuWTometnfs7Y/u2n1Z6uVRee6Xd9M7sgs7vTTbG7bgdt/5xSL7cvoui0dzJ/hB19egAyPcX5ktIXt5sdAQ/UgACbNd7z/ptf84a3fPSWO1qyo6DRFnCDQHkVmGwfpvDeDzKGBqiSRIqBVViiFQ2jkR8PnWIhz6JvR8Mhqwz6ZTsZORKXgag5cnjp2V//Vd/49V91yZHCdA63dCGEZvWZ95iTLnQFzvd+XARyUenodIcvbKXIBd5DAT8dGQZITDTJEzKtfAjWqeYw8xhtwa4ncLvq6+xUmtBs297S5EslFgindbrOthfQ0QQolFRB0vHuKLp0iE4rzrpk5rqYKnGk4o3J78BzXoP5X0316nxfuiM8IJntPrtK8wjYud50Wszu1NGyc9/zdTQ+1Yi4YKsurGRnsutqJDH34+e7LtLcCEmXdN4meGA6aGZV4MJafn6fNooxtgoYN5Ey0ypuOnbu+o/d/DfvfNc9Z9eHDZlisY7ZpNZIzprManApDhk9YsitWekNBmU+WVvz4/F4fb0Zjy2jyByAGCsr7YH9C097+lO/4dlf/XmPuyRzUEHGadgoTzM/MV0/4QH4NB7S0Z+lsudDsNcu51sauy3cPfYSdLXDE56DefuZVAVrZyMJKUgFxC0Z3a7SLbuVCHj7uZ1FexLimKbRrTR2lZW2308/SZ5WMJiw/TlIVGnHEWaPPSVqdJ51eVtnp250ZruQMkEAs31GsO6cOSi9V4amfTBdO3w6MrPfcd4xpguRbR09r9dmQgBPuzZ3hD2Nyj0bMP2tbn+04+g0P3gS+pDpAvvuPPLOdu46Ps1tZ5/Tp6OD7v/TS9soPaOEViCMCNTA0CMa3HVi+KGbb73+xps/+olj61tN2Vvu9wYI0ZHJsizLnWNC8L5u0NTV1ub66dP1aFg65wyXvfzqq6/+nOuu+ZIvfvLDH35kYQHRo8jmE3J0tiLcblFi7tvhS7uPK/CQjv4slZ06WvfSGLMeyl777NTU5z/eCdaUfL+MmR06JQHpDtLh8NCSm+noZG/vbM8cVjg1RSMo1Y3tTkA7nR7zXRAgFbXj85XCNG62Xfs5rf1pe3bB/Gyhs3eYoWJTTHDXSbHrXLPf6nYnH4Ccv5I4X+9OdTTN+Tp2OV52Wt889+v7q6DTD+Pc7+dWINvnmBs8CT3D51+XPW3tHaeZHR+7R+y2/KPoaN+2zhUAlDCsfOUDucwUphbAIABNgBAmFT520/EP3/jRe+8+efr06eP33LW6ukqKhX65POiXLlvq9w4uL1165PBVVz7skddcfdXDLl9YgCiU/OLApYitBUQ0tr4oMu18eruHLe0of/WQjj5PHkw6eqbF9rTL5jopstezjflf7bgk05Gx84iybQynZaaQQsEgszvC1oHDsOu8nWtEoyYO36nGY93dKZ0dZVo3aI/L0N3Hndl0BO2sd96lenX+nw6EJwoQzPlaZo9LqlMlOm35/dkCczr6QjLnCk/I87kf75zhsFufbvflfkucHpMuMHLuS+Zm+11rtT0fKtprcv17ygWXkbuvTAenCUIpfNoGMIMYAajbEDW63BG4jjFGJbbOwgBNhDJ4CotPN8YQHIEinIFzIEWIIAIbBK0UXkIEuLAlkyWwiNAca+92BAbC2yMJD+noPeRBoqN3ek6T7PkwKDDz3M3s4nnZw4ac/3f2QFKHr2DtrKMp7AkEM2vP1Pe9W0en33UjThVAoO7wPBcR21XgTjoNTvO+iG3/9U7HcGKOn+loswuCvaN/nSk9Xc7v9tjPPz17Tw9dj9L12NGqXdvtgOF5v58/3071sj3lXVBB727ZDi/QfW63nSqzA+w9Ic2dRXc1YLe5fb7I/Ejby8O29/rv/sj909EycxRFpTYIYK1hVrQRgFpHTAjSet845yxnAPsQQaZDTOscM7qApCMPICLDCXAYlQAEw6GNjTHGwrZeQhvzvDDWTkPh22Ub03izD+no+5YHA1/HzuG5x5Cd+2je6bznUnvbQTlvlylA2zou2bwzkDJ3tJI7lvB0nqrb+YlM/+WuEsesmbq9h8xFnFgTpCR92Q3luWnmPL9q9147+3neOzGdY+YbLN0swuY+cyj2vLyUnmHibru7VdvtVOxcBcwfbs58ThvT7cLnY0u6CXj7qu3S/KmO+P3Zdv6oC/U0fb9r5UJ7vNut4uc+T+mpex9/XrPP5u/zG3MfsnO2SB5n7j6h+c+7S9v61rA1xhLAhJIRI8VWjSNDhskgRJNFIGpogyLPComSQKPEBgCxBgkaBMYAFKMwW2MoqrStN7ljRI2AISbNLBgeEqfcMwyiafmndCcvUOr7gpPPRSef/XY0gDln5Ux2wQ+AXYu+uZXX7t9ty/Z8PwXMzU4kU9XZaRjttDbN68TtM16o3QzdBld06L1k7dOU4wYwc593h93pEJid8TwMmYKpsyuxwzux67YLzXwl6UnaYQPvkr1tzL2/3r3L7PLMz5fpzFN3QfeJmX4o5x1hx83dacY+UEm3+MI/5517PgDRmYKeHWfHeOB5LTp3lk/LTjxfR28fUban4O68JFEBm5BAlgFGXTWZBSBERAzvPSnZoohtMJlLwzqEoKrOOQAiklJuFFO0OQGAZYoSZIpxpoRkNwaqKXdRiaCsiQ+ElBG6fl/QiL4PD//8Pg9S+/rB4OuYynlj/T7kAd3RmU6fV4b35ywPTM5f7e7haz1vv/u7Rt77WHt8P9trtx/5MyN7hGSnb7hz2my3YY8LPmveZ4lItw5DV1JkJ+huJjzdme/fyNwFhNhLdKqy9/7JXj/c/RMAmPnQ9gRZzkyTuU+2j0BEO+84n/dGpnY07ey7XGBqnnXkQjqad354/p6fbfJg8HVM5YE8tw9IQc+/57kT7Tn4Pn05v/179+hC6+4HevTzv9+pj+c8EJ8pue9oXveMnQ/M+OzRyBcQ2h0M+MzJha2N3dr2fP14/k/2+FbPW5rt/Pb8z3b52z/lo5EsgV269QEFfmcr289ydXwBefDo6Ifk7yv/eLpw3ly68LcPBo/kTq2hu9FBnWaZetIuPjnfHN7zc9ynIr4f089nszykox+SJOeP/s/wQmGvQ82v8R+kcj7WcBsh/sBggg9SmfdgnP/5/fnwU3712S0P6ej/3+WBPqh//7G45xnvw1N8ofPef9163sEfDDbyAxeSTw0P/ycm51v+ST7TcaxPeVke0HV7UHiip/Ig6cZD8vcQmVtIytyHOn3hwi7Cz5QluNdx6MGlxEm2tzvkIWv6MygP1Jf9WSAP2dEPSD7bb/+egf4kF7Js5QL5QA/oXHvJ9iFn08CuwNGDz4DYZd9NoQsPOT0+MzKvoB88g+chHf2Q4D7hxfeR9PXprShl5wJ6Hm6OOe/k/PEfHM/bnho5vb9QBsxFJffHt4YLwOywE673YFp//VPU0Z8Z/+yF/GV71XcAgFRtfv7b9F5Vp4z8s0Pz3J678ZgP9LwP1K9HF7C5dtTm2HEu3nMfVYiIteyDJyLvfVmU48m4Vw68903TLCz0AYjK1tbW4uLiXC5kd4Ap89OF+nv+GRUAM7z3IYiq9npF6xtAnHMEVFVVln1ARAMRvG+stUz5/b0091tmxW6A3fXCP/OyfQUUFIGuBFqMYqxRDSLClKmCmUPwxmYzFXOBIXM/T3tf9+WByN7j7f4d51OO7VkFd5lLhmSAugJC2zlPsx0ACNTMYUXT3RRFNMwKJdBkMsmyHvM/+O39R5DP/h78w8qD+frEGJvGE4w1GZTbNvR7C6pkjFtYWAJ4PK6Cl8WF5Z2Rrr+Hy49ENBpLZVnmedk0DUCZy1KZMOqS2QSASCCOouGz3r9EMmc4CyCgABJjGKoxKpEJIYQgvlWCm+ddedDkl11Y5m/ufA6v7t5tW4PP3PqzkAkAiTGGEBRo6iDCWdZLXE7/KL34h5V/gnb0Pz3pCIywc606T8jxT3ko7I1qYqY8t7NyYnleigjA3nsiMgYA93oLIXhmq3OlmO7TeTr/1XmGFgmA4XDY7w/YwBgCMmYASqQEx0zBw4dYWFZFqpz0DycXWtn8A0uniWJUY6xhR8w+xrwooocxXZnfWdMuUNnvwSSzQbUz4YVkLyucLwRAcs7FyARH4KaOABmDLH8w2FgP6eiLVNLDz2zbNgyHw6XFFWu5qaO1jhkxoGk8oFmWAdh+Nv7e0a2lxWWA6yqEEPv9HEDTNFmWA8ichaJtuW2EDRyz/NOb/B6oL2tOdugLEVENMWqe5XUV8gy+gQrYpbNcTGp6e1DRBT6f/zYxz8yhqinNedw2EAmZs3lh0u4ieBD4Oh7S0Z9B2RX/uX/ygLXeA7Uu5yN7U241oK6bPMtFkDm7tLjStl7VJdO1rmVra+vAgeXgYQxiQIghy+z0UHtyf9xnLzquHwAIQay1RcHra+IdvJe8yNsGTR2NMSroD5wCghrQpmnKvHyA/d1bLuyv/9T7//1OvCtnJ/lb2WUWaglQwd13rg76K5sbw0svO7zvEKUKf8lnPguKPGjV9Gz86yw4PNO8O0ea2u1vd/w8Oaa5LMwdx9bruhWRouRHXH34H771/xjykI6+//Kgcg6WRQ6gqlqibDye/N7v/cHx48cXFxd/+qd/uiz4/7z4FceO3XnZZZf9y3/5L4oS1jhAdz42APZKortvUbbWxoDxKP72//2922+/fWlp6UX/7IVXXHHZy/7wFcduv6soihd+7/OvvvaAihNti7z4zPX4/0dJKoYBRiq2rTyZVGWxoIKPfPTun/mpXzh3dnj40JHv/8HnP/MbnjivkR/M2nleOgU9s5Rn5o7O7YDpt1OhMPUdGcN0+tT4P/3Cf/vIhz+W5fYZX/yUf/fvf3xhMftH68E/nDykox+ofNp44X9aEkKqrEFEyLLihutvPHbs2PLyskTEiLe//Z3Hjh279tprmdE22KkqHwjkbg8eNYigV5pjx+6+/oMf6Q96z3/+C3olPnzjTe9/3w29Xu9bv/VbJYLIqBLTZwx4l+ziv7++e+C+DtrxflrVj2CIYBiW+75x0Wcb6825s8PZWS4WNa17AunOZ6Ga7TC7FDNYPQDEiOWlvm8w6O8HwtbmuCyzeZfRZ6/8E9fR6QbMr3HmMerJMzW3+477StPNeUTqe9KY6txXHeebAHGKMDM7VPM8Rdz2ILB7HHy7DXHuO7Pdkt31/XbBjGiniXFes3dckPPLUJ+fGAKA0xqaCHnmVGANW5sfPXppWZZlSW2Lsiwe85jHxECZgyokgg0BDIrTE3FXAn27R3sKgUL3Vv+/9v47XrPrqu/HP2u3U55261RpRhpVJKvakiz3IssVy5VmGztAKMaUFEJJCEmAHwkh1JgkQAwO/tJjsLExBhvbWJbkIktWb6Myfeb2p5yy2/r9ce6duTNzZzR2vkkur+/zful19cx59nPKLmuvvc5aawtA1RV7R+0W2q3OZZdf6lyZZso59Cba33TFZa28NzW5rb+CLIdO0rouk2SjJju1ErBBGQLQ6FnyRN6i9b9abcR43P5w0s/PuOE5ASeHdK+z5Jy+a8G64wTSq52BYpqmAOoKIfDU1MzOHRdaW0l5oq8yYyMxfbwFN6yEjexaX9fy7xQBCJyU72lD52M6k5/PaYlZ1nkhnrjc6Zaz1Zo/Q486UZ4ADSiQBwtmCIGqqrZu2S6lFJQqiRCDlPKc46E2aQT5ZssfHbnxRGYFRDT7UbIBUJdlmuvRcNBq5856bZL5Y4OZmcm6ADNCwMJ8NehX3kVjzMWXtIsKWQskojYCDOesTkTwUQozf6xs5ZkQKIqY5SJtwdbOWZGnshyhKDA5BWkwGC22u81ubBDUHvVha2QZmJHlWJhDkkJppDliDEJRVQRJGgxbQ0pUFQBoA6E4axHIMVc+1ESyKtBuT0aPqkSi4RycBwm0O6sB0N6XjEprAxgOkkgcPtjvdNrGCGtZKTIGIQLkdSKK0UAIlSat5aWq20mtxWhUTs9kzjfPvranOSI3uX1ZA8I75kjlCJ02RgWERN6CdSCgtk5KmediZQlKQgqECGmQ5gDFuqqSJF9ZKpVMpSQp0ajexkAZACiGUAr9fpiekZEhTQTC0uLi5OTWYujLEaZn1HCAuo7dCaETcIzMLITkgKOHMDuDqoZ10Am3u7QwP5ia6hChrjAauakpPTdX9Hq5SVAUMW8JjvCeARoNXaetlxYxMQEfQdLrpBqOlrqdLceO9BM12+1gMIQP1dRsuuq1RQg8EjIEp4iUlGbQH3a67aqsY9BaihDADJMgBHgHqSA1x+iN1jHAWmgDqcAMAkYDCKAsMTkF61D7UauVVQUpSUpAEJZXMDkDCPhQh+AImVHywDNxckIoheEIrQ5WBiszs53INgSXpilIBhet9VmeAsG5SpukLqR3ZKumyWKSCqWZRCDBAIIjQDBTDFzXLksTQahKZBlGI0SwVqQklILzcB5agRlaw0f4MMrbWXSCI4ajWgpNEFm2+qTOoqqgDVodNPlEY+DIVkgnhHQ2amNirIWQ3rHSBkDw0TmXpq3+cpUmKUcszvPCfF8pw3CTU/nUjDIp6gplWfQmUiIxGrDR1Azq2ro01c0dWoc0wcoSul0ERgiIEdpgOBpOzbSB0F+pomu1M6wso9sDNHQaN5q31s8o649sypipzZfjf0MZrQHi6Ly3JlHeuxgQvMrSdLCMr92z/6Mf+cThw0efePzJiYlp7yIzp5m8/IrdVzznwre+7Q29CUCgKOrhaH7L7M6HHzjyX9//oUHfDwdViHWnayq7UFbDnTsuOHZkQNxptya6vey65178gpdcdukVO6xdDoFsmX/sI5/9u0/dWZUhhMBMnU5Pa7rqmou+413fnGQuyxNE+eD9h3/1P/328lKZJu00TUmE+YWDV1+358d+/L1p7oVyzE5QAuSI4pEHj/76r/72oQP9NGm32rKs+oGLD/6P34iweUcwlwCCV6NB3Pf08u/81ocWF5diwPT09MLi3Pvf/4vKoN3D8srCxGTPVvzL//F37rv3cbD2wfYmzY/+kx+47PILlME6GQ1mv1qfLL5272O//mu/NVgWWuVVPbjs8l0//W9+JEkhBIgwGtUczb//+fcfPLA4GlYmobzjv+t7vv2qqy8H02OP7v8v7/+ArVFXMTH5cFjkeQri5eX5drudJK2qrM8777znPu/qa5978Xm7ppMU2qhg8cM/9NOLc3WWdep6tHX7xHe86w3XXHexNuSc1aqzcKz6rd/88KMP77PW+lB0J8zLXnHTO7/z9dZicWH00//y55aXirq2s7MzRPyu7/z2l778yroGw6e5Gg3cX33scx/9i09J6pblKGvx6775ha99w/NNgkHf/tuf+eVjBznPemW9vGVb+3u//9v3XHShlolOAFFW9VCrnKNS0oBoOBj94Hv/6ahPHHJnRZIk2vgk1cETyI/K+SxLg1NS6rxlfuRHv3/PxduNkf/pF3/v3rufyJKZ0ai0bpTmbLKirAZa5bbUxZCmJreed/6OV9zy/Gueu3N2WwrytuLlhfiz//ZX5o4OhVDOF92e/ol/+UO7L9hiUvbeEhEzKZUAAIM5+lA6i//6m3/wwNeeKUcM4rxFr37tS175qhdMz3aYLTMLkQCoS378sad/9Vd+c3lpFH2SpRPexSzLfCicc85KKUknVipIYUIIK/3583ZN//tf/Jk8T//dz/z6Iw8/JYSIAd3OVu/9YDjPHAQlWZ5u2ToxMzP5Xd/zzolJnSQoqjprRe8rJfMYo/MjpZSSOTOYOQahpC4L3HXn17545317Hz905PBiMXKdTq+qqonJfGVw+I23vfo1r71l1+6WNCgL94mP3/7hP/sERxUjhPQEydGUZS1kyNKckCmlhsWcVCFLW3kr2b5z4vu+/z0kqw/9/p9++c4nXZUXwzi7tfOq1z3nDbe9qNPLNwpNPD1kfLPGtW6+HP8Cx+OOgLX1TxMByFIKQICl0ZkHDu6r//zPPvPxj95ZDNl7K2j2ga89fe2118zPz2vV+uuP3X3v3U8/cO+hl778ea97w/WJTszkNgDBdr50xzOCJzjqNJ10hV5a6UvVmT90hJBOT07sP2aPHHn6/nsPfO4z9/zQj37Hc66bzVIhoiyLsPfxI84KsKzrkGextn1j8l5Xk9bD4Uqqe3/1sc888ej8yrLNM0+EVjs5Orewddtsq6UhNCCsZQJFH7UUt//9V7905yOp3saxmJhs94dlf3jkwL7hhZe0gUhkmCMha7fA3j9436GVJZuY1iP3P53m9Iu/8D/+9b/9zqXFstXJ68oi5I89fOyZvRUhhlgtzK+kSVcIeF8r1Swzm4aWq5UMzEyf99jDRxEmBEVt+InH9ysFqWHr2iRJmibB4aEHn9r/9JKtqdvLIBcneztaeQpA0dT99xxWomtrqWXqvepNqNoOB0NrzFAr9p4fe/DRu77wxPU37HnL2259wYsuQURVYO+jC9UodW7RusHiwiDRXa1ToPbea4VuNz12dOmhB55O03aMdnnR4aVdAMZgZrq1uDDqL/HSYlWPyrn5I+XoTy684Ke3n6eY2NpR3mqNhvbxR44Y5Qf9kcnql75Ud9rTIDjjjxwcHt4fOBa1W15Y0ErmaZIQEDwgyOhMCMPEVV2kaauVt2wlVxZJou1qM2Su6iVQ6ZyTkqTS3W5rZXkYYunjSjvflqXSOzz4tcOPPrjcbbeq0gkplamZ7KgcZHmdZ71gW/POPv7o3ffec9/Lb73iO//RG6dmMiH0zIxemF85uH/A0TgbhFrWsivEqimDiJgRvCcia32aGq1TRPn4I4fu/9pTAi1AgMqXvFi38w4YwcsmjpEjaUW97pannzwWnEFUZeIPH1rsdqZCYKNzpVLrBjasRC6llK125oOfO1rkSVoNMXdYzB9O6sqnSfupRxaUprydELEQau5I+dTj/SQ9cOzw7z3vxiu//TtekKdJ8F4KIlLMVkpNJJoNWkIIBFOO8ImPf/GDv/tnB/cvt7JpKdJ+3waLcpgNFmlU6t//wKcfuv/wG9/0ihe99FIhdVXwgX3zxN2y8HmeE6RAWpaUt3SfvLOFVHEwWpIqJsnI+bKqds9M59bmTz8xePrxWiKPIVmcn3/Bi2Sn3V6zaJ1m795Q+Jzxq/+bbLobOnFLx183UQSFqh5JJcqilCLjiGDx/l///Q998KPFQJVDaXRv65bzL7744vmFY1VdOItOa4fkLX/98S//3n//2Ef/4ivOQinZXw5Z0kJsKZpWNFMMzKH9Q+JcqzzLWsxh7969deW3bbmQfeeRBxY/8uHbg9ON4TUx3dRMdVrbE72lnZ1nxBaJ6flj5dICABBkDPjaPU8kejYz27utXYNlWY3SVrpNUrsqAMDZIEWqVasqadjHZz79pXa+LTUz7DuDJW2L1taZS37vA38iBQYrBSAIiiOkgFa5q0072zXZuTTT5+3YctUnPnbnX33sgU47MzojaCURfZ6Z7cH2BE/VpQBLa71z9bqKJbBYFdYMWwkOmaIpir0smY4+8R7MVml2rpAKWiNPJ9ut2a2ze1rZTF1BUl4MgQglOpJ6irak6jwjt1DseieZMbtlRmvtXOx1tk10z6uL5M7bH/zYR/9usIJgUVfYtnVXcKngiVa21dbS6JZ37Kx3NpSFrSsEZ4zuKdER6BQjHg2ttWCgrtFpTweXTfX2BNuTmBksyw/9j78YDVAW1iQKwI7tuwXaRs5smbm4k+/QciI6uArBqTydVaIr0J6ZOl/JljGZcwGEECBlKkQKFiSUVklVurrCZG/n0UNlf0nYMh0sU2ImYhBKqXa7DZZ15aVMer1enqdVVXJEMYLRnamJ8xM1LbinxVRwabs1Oz21XclkOKgOHVywlZY0cWDf8p233//AfU+Vo8hRCIkYSMmW4G63tTvV24JT3sE5H2MUQiglGodIrTWInGUpoES3k2/vtc+f7J4vqWd0p/ECDiEopYQkIUGEqqqWl/r9/rCqKudCrzsdfaJpS6gn61Hm67Sdz05OzjgXFhcXhRDD4UhJtFtwVearbivZ3V80Rm4RPCHQWlkeDfrFYMWWI5EnOz7yP2//0O994nN/d9jXaFofgBBCq0zKNHgmklplStJDDxz+0Ac/cvjAiH0rODM5seXCPbuvvPKyq695jnfxyiuurwp5x+33fPXuB71HkoKZ67o2JrV1dFZEn9RF6qqWK7uLxzBYiUZ3Ou2pxORSaubgnEOErTBcIcETCBOKpoNLjW4Ff7JIeRaxs8k06DU2mx4NMIHEqlcNN77oAYjMIQSfJJmtGZE+/bcP//XH7yLudvO2kqGoFvN268prrpzdMnlg/7FPfuLzMxOXR5+1kt3zR9zffvLOG2+6Zmab6HTkQYs0TW3htTJ5S3Unp9/zPa/duiPTJj704KO3f+6BQweGtgwrK8WOned/+m/veN+Pvj3LydWoS4yGkT3VVey0u9EbRTh6qH/XHY/d+oZLW+324hEMVqIt87qI2y66rL/EwXqhvBSZs0iBuq7b7UlEJFp+/rNP7n3sqEAvQHXbW+raU6RWmt1z92PLi+hNtatypLWOMVZWOOeKUa1JVMNqNGJB5daZyz78p5+6+NItVz93i3csFFrZlKR+NRpOTnWha0FpmimIfC3vQcOJLjjoV0p0BLd9EPPHli64uG0MiIgEFUWtVd7vg0gOB3WnRYOlvslbRFIpBA9B2qheNTRp0hmNlqdmku/+vtuue94FJKt77/3aZz9972c+dffs1MUTvZn+qL73nof7fTs1a6SGMqEobKIzk2QxjIzOBZHQmTZZtMIFxJBwSB1DKiOES1NjEoDgHNs6LC3YiU7L+5jI7aM+/uYTX9pz0fnf9q6byqKvpF6YGzqrgkiDQ+2DswgeTEgMCGZxYS7R3STRrnYxwLo6SXOTAkBVhBhj3sqkFIJgKxw6sDQ7vUvRDEeVpuqt3/oyofs6sUmqBOmVlQGAdictqvm8Y0mh1UZZDRYXFxWJNGknKT/n8qu+971vSnP/xBOPHdi39Jd//qXBMhk5NdHRRw8d+uynv/LyV1wvBKoRnCXvaLDsuh3lIgUviKC1CtHFGEKIgkgIJSQA1JVXQg0HoRwJJQQzjwpb1y4EaIJUIcITiCGqOkYubn7htcOBq8o47Lvzz9/5wNf2eW4npp0kunJ1q42pmU63p/OW7g8Wdp639dhczFOhZBq8ItPOEuODnZ7tXHH17MJyiyD3PXP4wL6lpcXiogueuzS3+IHf/sj0zHsuvTJ1LgjJAFVlTLMMrIOHVGCPz/7dXY89cmDrlgtiUErqt779da96zYVT03jowfjzP/vE3qfuyVuqP6y++MUvPPeLu6697gqh/LYdE9OTE+2OnpzYItB97KF5jkqgYxTbMLdj5xYf1cLSASHdebt3Xn7FBbWFEFAqMRrlUBKi1rrVTp8lgOUfSP6lzSejgbXXvXFtbR6BmOdZjCRIxwD2+N3f+dPUzHbbW5eXl7sTyYtuvOZ9P/KtE9PoTWJxDq961av+3U//TqJaWdoNof/AfY8/+siTO3dd5iy8j+12d7lCCKGsRlDu6msu3HMZQLj2+h2vefXLf+B7/8ORQ8ecFaNRQdBzx5Zmt03lOTqdCUm5NN1Ei127LixHcWnlSO0G997z8MtuuVQp3P3lvUsLBXymVX7ezov27zsyKhab1/FKAUDzKt/W0BJ/+sd/xSFvd6eH/XrPRbsW5ofHjoWDBw7PbEv/7E8++Z7veXXwlBglE+GBdjtvtVqGuqMBGZ2VReh0eocPHvq1X/nvv/qff7LVNlUBQFjrjM7zvDuojizM93dfNOtslZyIAeF1L8VFMarTpM0uR6RRsSSFqUrAFe1OqwmiIEBrEwOUMlnWEboOIRLBOTgXAeEsEk1CIMTysssv2HMR+qP0tjffdMNzbxqsvP+JR5ZHwwpSFqOyKMrgjVKI0Scm16JTjBZY2Lp21iYKLsaIkAmCs4hBEqSS5BlJktgaJPzEhBJCddpTxLlRKsZw7PChzkT60T//7I03XbPnsm5dohj5mekd1VBXVSUTlWdtZggJkyF4yrOJdj5j6+XSOSIphKjryhhDQqTZatom74MUEowYBJHurxRlWU1MJS97+bU7dyPJwYwkwfwc0gxJisgsVQSBBC644IL5w7ocZFKq0WjZ2mLPRTLN5SWXXlWXGC63PvnxL8/PDSYmOyGYuWP9soDWKEu0Wp0sdS4VYMGRjFFSovH9IFJEFEKw1kspiWSWpZIANoKyNOkxB51sybN2I4yEoBBsjJFI5u38ksvO/+mf+WfdTjvNMH8UiPj//bs/+eoXjxXDsqr7U7P6LW+95c1vu4YBqQFCqw0A80fQ7eV5ng6Hw8TkoHD1Nd/0T3/ixSpBjPj7zz74h7//tw/df6i7fZeR4qkn5r/6lUcvu/KaJDXWDYxJvI/BQWrUlZdCVRUefeSpTnu6rthoY0y2ZXbHzAxI4YI94nXf/KIv3nlfnueHDj/T7pKQvttL3vq2V7/pTa8moCygFQ4fwL//uT98/JE555wQYuf2bd/13e+45npjMjBAAkVZJzlWFhFjiDEYo733IF/bQW1jqp7N4rzpJfXmlNEbIsDkfdBKPvxw+cRjh3dsuWo0KgMXuy7Y8mM/8a2Ts5CGSzua2tp+/gt3vvp1N9/+2YdGI9vqZAsr9Sc+8bc33nyZUogBzjkhEq211ElVLdXWOatJoCrRKF8TE1NapcfmDs9uS6SUVQFvIUjVteMwUkqA6p3nb33mwINMdu/jB8oRshR//7kvl4Wfnkwntu/YuXNnWdY+2CxDZJtmcK7SRnnLBHr6STx4/1Na9YwxWXv4vOfvfvD+Z+YWiuGoP8PTn/m7O9717lfnWU6N15OAkHC+krImoSYmW2U1WlkZuhAfefDI7/72X/3A+15HAAkfUSRZZzBYstEHDyGR6Byx8UFcn7lRACLLshhjXRUcVafTSdM8zyFMDxFp0gWgFIJnKWVVWtHk7xDEQJrDJFA6ttqG2TtfSZmmKUhASXCE9wBEp9MZohQK0rAUSVFACPT7Q62mU92phytpnguBJBGkwIxgIQBjTJ61nQsxhtGoXFpaEQTvwUCMsa4Yip2ttNbbt+0Wqjy0f/7n/s1/++8f/BFfo9OeWV5a0SIl4ixLstyA4Bw4oKqs0a0s7ZXLfaPbUpgsTUlEktG7SmkJphhZaQHIyOh0W/VIppmRUmaZIoHGwaOZa7s9EGEwqCYmU6FkMSpHfd1fKUajMjhDFLXWUrEPqCyMQr+PlZUVEpxlidKQkBMT3SSB0ihHiNH3B8smmUZ0WS5rW1W1MORIeKmVEhpSec/MQQghJWINrQyRrEqrNIQWMcaqgtCBBGutAQHI4D1z2LK9DQCM6S0IDjYcSzInRBKZa7dQ2TlpIBWSDDEGH70UempGHJ3bK1To9Nq2Lgb9uWE51ZmASiEkbnrhJc88dWzuaL1//6FW3puanX7wgUeEuAaEyDUgklSDPKCEQG09oM4///y9jy0XQ6ocrywOfuu//NFffMRcdMnWl77s+bfd9oLXvvYF7Q58gFBo9RBjVVV13ul5F7qplBJbGaPqYOAS7EwiK9ufmFI6A4voQj9Psl4mvQ2tjpTalvViZpK6HEoxqOqVNFsvoMPJgbKb0ayxIZtZRq9foccQYlnURne0lLd//ktbt+zqr1RVVW/ZMaWTKDWkRuBhlitX27xtbnnVCz/20U9n+XRtC6VUWfgQ0O1iclKMRqNhv263ZN7KyoHYv++YD53eROvOL9z7sb+4y9ukKsqj/QPbdrTb3XDZFb0QMFiBUkopEyG14ayFq6698Itf+bTWmXeSGK7Ck08c1loPi/ldu3ZNTadEIcuM0lVVDUkCIQISTFrhc5+5J9FTrlYH9h+64QW7XvfGS7ShRx67d5YnFxeXh0X15BPDS69ou9oniQIgJby3Ac4kpqiW0jStStftTHkW/+N3P3rFFc954Qt3EUkpmZlX+iszW7sTE1McEEKtZBN/clKSMBB6vW6/36dgEtMxxhRF4R0EsVIkJeoSgsBM3e5UNSKppPektWnyiHlvB8PlTE1VhVXSOMePPrzAPG1S3H77l+664/4H738i0dPMrGSYme1dcmkqNMBIk2w4qJIJxVEbnTsX6tprssysVFbXcM4xU1nWE2ne7fY6nZ6S8EEtL9cTExNEy9PT0wcPHO72WnPzh7dv3zJYDk/vXf7N3/jMe9/78uHAOcetdhpCGAwGy8uLREgSRA9jkrKoUu1jkK12dzSsqypNM6rL0iQyxgAW3ntmThKdpkgSPRoNEpUTxWee2fc7v/2ny/19EKNWO+33B71eTykq65WXvuz5b7zttVma5SkSk3vHShiC9q6wNRYXYH3/0MG5xx459NnP/j37lk70ymA+77qbbr42MkBot6GNtLaanM1GgyiVsLaSsmMSAij46JzVqqW0aIaqtxgN0Gp1smywsliTULUviWSWQWkJUkDkCBIk1WpUwcryspKm1WpTBERd1n1f5yArMCJp0wxMAAXnC21kCIGQM7w28vCBo1onJmUXBpCo7Eqaplu3JRdcuEvKu2dnZqyNdV0+9dRTQqAsiiQzzlttWrYKzCylsLZstdoveNFz/+avbw8xbeftid7M0SNLTz41f/99D3/0I5+cmW1Nz0y8+S2vv+VVF8kEZVGY1OUtE8JIKhWCr63UiYpc5m1ZjQLDj4rlJINJABW5Di4MjciLsuq2J41RJELeUnVNQlFkv24rn/UZybGav+mUbLub1T9608loWq0iARwPEiJAErjVajuL4RD79x+oSgfk3W5vbm7fhRdPdjqN05gBECOChzJOmVpqNxj1ldYcpZIIDj4AiJNTvWIYRkPqtHb8m3/16z6O2q1uK585+ExRVzh/1zYf6nYP/+i738oMqZDnWOkvxQAicmEQMXjFq3b8yZ/R4vzo8CHxhc8f2LXrvMOH5qWi4ejIO77zmx954JBJAygOhgtptjMEECGGGAMWl/HZz3yRQ+ot9Xq9JPPTW/CCF1/8x38yzHLt+kZJ/ef/829+8oq3pKkaDm27ZRbmh628x07Wddnu6Mmp9mClCiG4Qkx0dr3/1/7wvO0/nuguIPqD5enp6dHwmLMcAlSiADqeQqzxn22cpEmg1WpVQxFjLIpiq2pJBUYEyRiQpKhHAEtbs6BUgGMgjlIQAGijqqpqddFqtepaLM2v/Oy/+Q2ludU28/OL05M78mRbUVR5m5O8/v4ffFcEEDEaYtCvpyYvcDWytDscrBidaa2kVMweDO/R604uLj4xO7NtOFjyXCiZxgit0eslg8GK1nphYT5vy+nZ/Lzdl+7duzdNuoL1H33o0xddcH23vb3TmnIucqS0lQNCytUeFAOMTglaikSKyCxiEEIKQynHEGMkIkAKsTpknXPee4rWmNaW2fPu+Px9QlVKR8aClJqoShK9uHT4xue9RAkioCpw7NhilvY09Y4eXdi6fXLv44e/693/XKg6BlWXpMVMmprKDvrDAze++Jve/PYbnbcQxgUAMU3TunLORUibpqnWEhxBJKWUKuFwImJFKbTasLaq6zLL2zH6bncqBioLqASRvVREJDjG4bBIkkSrpNebAlCXIQYZAxOE0amPXimtVXI8DC9JEkaQUoSATqf79BNH2+0uM7tQdLu5UlAyD57rApdfdpEgBYgYLbNLMy0VfBEyyiNc9EJrTQL9/lK31xmsFDe/aM/bvu2VH/rgRyOp0vraxk5ri60r6Vv7nxodOTj/8AO/9Qcfmr3trS9+09tvErJ0zmqdxiCk0FIjRBjdqsoBs2SOUpH3NoRUSk6SJMYIULc7GR2qykqhh4MCEFKkRrW9BSm2tkhSRcSRQwjBqBRg5nAiCIsaF9WAVZG+ucT05rqbk6G1qW81aWEIwSRaCCilTKKBaK2d6E1rvWpyZQQgJqlJc2zZOiEVpCQiyrK8LCtjIDWqKgrphYgmEcxhMBgolUrKjx4eLM27C3ZduX3rBTPTW55/8/U/9a9+5GWvvHpYFMyQGkpJrbXWUmuZZpRkaHeV1pqjuu+eJx99+GAMQgg2adi+UyVZHBXLztedTieydw5KG+/gLR5+cHnv44e0agMyzzqvvvW1QmB6Btt3bDlw4Ih3ZCv91a88duQQgkOaGjCEEErpqqqM0T7YNJPv/aH3HJ3bNzExcezoSjEwf/bHX1qc995qKTJBikgqpaQCczwpGptPhNutrPSFEJ1OJ0ZmZu99jFBa1vUwct0kDLM25llHCM1MSpmyLJWCs2DmTqdlbTUaDfr9/uzszMz0FrAqi6hlOzjTyqe2bt16w43X/sRP/dBzb7xYJRiOiizDzp07AQyHwyzLAOE9vIdzvmlt57C0vNBqtfI8b7fbeZ6f7JciVrsEOYjyne9+48SUqerhoF92Wlv/+A/++uD+hao8w6auFL33ZVlaa4ui0iqRUoDBkaRMtc6VSrTKCJoDRiPUtWu1WlmWheAGg5VWO5dyLcuzc6PRaDQqATjniwIhIs3Q7baZg3W1lBRCcDZwVIcPLhs1IalNMEVRzM5O/vwv/Ktf+uUfGwyHEDXgiND4b8QYCRJ8PD9nozjTGR7nxHuF0bCIEUqBRLNfQQDFGEOn2zZJ4oN11jprQawUpNSAAiRYnWSWZQEiIkVCEoGj5KhiEMFTbauiHFgLZlbK5DnqGgcOHIhspeLKDojY1ui0p5wljiZGkEBZDru9DhA6PZN38K73vPEP/vi/3vLq501vkWmryjvBh5HSYtuW3eVQp3r7Iw8d+uM//NjffvL2YmS1zpcWl9durPmfAiuwBssT6jArsAHUqmfBiU6iwHrtL6SgLMsESYKUlAhSPkQf1m3ZsXqZePILm03EptOjAaz5nJ+UjJhIhOCVgjbodlvOlUmaj4Y1kawr7xxE4LJeMYlWQgZrvEMM5BhKZmBeWFhYDeELhdKBQ6W0ERKCw03PvzrN1F133iORPPDgvUbnN7/oite+8fprr4fO0R+UJDKlSGmQ4OACsw/RZzl2X7DtyMGng8cX73xgcWEUI5jszvOmtu+EMi7Eyjndm2o764JnsHKWtcBdd9zfX66TaaRpmiTmOVdcMupDADc+7yX7n3RZ2iuK4cH9B77y5Udf+4bLIBAistzUtpDKkOAQQq/XetFLWm96yy2f/tTtxpi6lJ/4yy/2ej1fd4wSYAWWUoIEuAlnPyk2eZWqqvr9vhG9qoxpLjudSa0BiiRrpSk4HYLo9/ved4hUnqeDKjpfCpVXNZhDu91ePGptHfZcvP3QkaduvOmqwVDu3bt3Yqo7WFk+cvTQd3/Pu268+dKbX6RUjsh1dzIdLMWjxw4Phu1We8vCwpxK0eumSQLIFPDFsE7TJM/TPE+PzR32YWSyUki/LkR4bYIhCzF6+S3m8b3P/8Bv/4UkVVfha/c8hpgJSk6NU1hd7YZ2N2WPREggtDt5kgBAjDFaTySVJCIorZvtZdKkE4O1djQYDHTi3/zW1/cmtY8DHwohhHOu3W77UN5089WdHopyRNxiKqxfoWCUBsjtOG/2yquvfuzxh/c+NtfOtyzOF1ddfeVb3vaqq6/pgpDnacSImYVo8meqGFiIJmZdggGWIMaGb7Jo7dFYgJs81NIkAOBDJRAABGaJJookaq0AEKnoEEIAS7AG6zUxvVZZq2JLcgRBCdJCtqKwk+3JJNFCIMbYvBoxBtMznWK4kiRaawrBxdXs4ilzVAYgpw0AV5aFIGNrmppp97rqB973ZiXxuc/d/4Xb73py73Jdlk8/9WCqd9gya2c77rvn/r/5xJ0vfdmLwGh0/xN9lgVYI7bAdl0SBQLMqhvYCZoKFGtiGt6DBAshrfVaazA1CdOb855I50HHUwJsOjahjF4fAnTCQkTEQPTOCaEv3LNzVC5OdLpShbr2de1ihFKUiQ4RS2Ei456vPmxrElrkWcf6RWO095AekUOM0VaFIDZGdzv5m9/6ml27Wwz72U9/9cI9u7O0+4U7Pj+/8lDWfc11N820221GCFEB3nvrHUP6qipabdxw4zV33v5IjPmxoysxPK2UqZy9/nnP0wY+lO1OWhY+ePYeUlJV+MTo5WN44L7Hp6dnh8N+u5MvLw9/6Rf/KMuVtW7Qd8tLw9KIsiwDy8995ksvf+VlKoExSFM1GC7OTm4rhrVQkYQXCt/33tc8s+/xvY/PC6iycBwskCmlYqxD9Kd5Ha3/N4OxZctMmplQca83WVQLdeUX5jA5K0bDykzmUgoHdDsTh1eqqnC1r1RWZbkCgZmllCEEIiME5hcOg8of/OG3b92GX/u1P/7Lj36qk+2amGj/xZ9/9NHHd85sf/s3XZ2OhvXc3IHd51/U7bb3cZFk6I/qVpLNHat37E5GA1eUy7MzO+BR1/Vw2Pc+tjoJ5ChJ1+982AjoCDCoDoy3fsvNTz+978t3PV4Mh3ne2r/viDFZDE06NA+K63KqeGsrb1lpZoQmlhIMKZpNqsAMjhACHOEcvIveeynzNFMTU/nLX3HDebuhDLRp/I5hDMoq5i0BAZMoJWBtybBCOKmk88X5uy5553e+Yf/+G37llz5w9PBKqzVx8ODBn//5//C2b33ZP/+pV1WVT1LNTI3sE0IER01U4bpWWosxP95qoJOFNgEUI2KgRgsUlEgBIHL0gCLAJPL4r4kQTtUfj58pCtGoqJIJzBQ8h+gju36xuLCQ1jUSoUHwDk88sU9parUVUSTFQkIIeAutABJlZZ0fdXud4F2W9RBlcAERy0tYWBilaXLLrVe94lVXtXLcd8/gA7/90Sceqo4dXZnd2pnobj14YFFLDPu23TUxVqAIaAixmvKF5ap3/4l0pmsByRuIjlVCdBxYK6mlDg7MKgAxktI46Q3NJmYTzhti3X/HicxRax1CCAFXX3vZxGSy0j8idVxcnD96ZO6xR5aWF4DYir5VDXH0MO64/T7ETMnUe8+wNz3/mma6bbc6iekY3eGonfODwcr0bL5tB979XbddcNFEaY8dOfZMlmVfvPPe3/5vf/T0E0MhkuDZOY4cItckglKGSEqJG2681iQUok1Mtrw8lEIJgZtuep4PCMG1290YUNdBCkOAIKUk/vKjX9m79xkpJUkbuO9C/557vva1ex/64l33PPXkPiFUlrZaeW9yYsv99z320INzSjXLsFLpEKON0YPCcLTYnURvCj/4w+/I2662yxMTbe+jFAkROVfH6NfsbHIDXYwFgMkpsWPnrBBxZmaKIz3z1LEv/P0TtkS3tTP6bHEufOpvH1xZ6VtXtrtKyLrb09OzeYzQhpRSg8HAJKrVytudvNU2g9HyxAze+Z7XfdNztncnVKebz88Nnnly6d/9699YmUcr6+7Zc1ExwrZt26amW2W13O3mHPHnH/5YMYAS+ezMjsX5+onHnPdxcnJydsuk85VU8aqrr1hVGxvVElhNOEVhOML0VrzrPa/fvacjVCWkA7AqoCmA/Ak9GgBiiJY5aK3TJO+v2OVFjAaoShBAAnWF4dAyQBJag0STwyg2hkvm1XR9kSEUTAqVIG+Jshp557yLziJNc6WUEKQN1XYQ4vDiS3HjzVtveP5leZtJ1onJEj3xN39956/+p4/kSQrOmsX42gZaJ4uJE5mD1hxyTkoMe0IY5VnHO7myBFvCVamr0rrIbZmXQ7GyXIMFB9Sljat+KWI1vddqFTVWJgZCCFUINgbHAImoNGktjDFbt+zotKcEYLQUhCOH+e67767tKM3Z+pW5+UMXXLAbDKWwMIeVRRhltJyYP1rYSrlKfeHz9//ID/7s61/9Qz/+z3/lB/7xv/zX/+oXg0e7A1K48urOz/zbd0xvSfKWPHRoX6/Xa0JO2h3jfQ2yEDXECAKg+kRmrtVqOf43rmvouJpFa/XRfKNSpImpCtQlgkU1wvwxpxU4glcJax82W2KMVTahHr0ecbwNInspjTGyKvnCPfkLX3z9Jz9+N8jt2LFj3zPP/PZ/+5Obbn7Om978wm6XntzLn/rkXU8+fkyJlvd2YenQnkt7L3/lC0zj7VSWg8EwkR0plBCyrvsxehf0xZea17/xpb/2S38cubO8vLhly4677rj3nrsf27r9epNqKRGjFTKSIBBJocoK27aLLVt7Tz6+3E4nRqORUKHV6Vz+TZcSQQgRQgBE8NHonCOSFMUQf/7nHwFEVQ/bnWQwOtLpmRACU5Xm7OwwxMTHUQhU2ZLL6tN/e/vzbnqzc6jqYaebjVZKJVsh1EIGEALj+humXnnrDR/+4ztJdpWWQvrILsYKZJkDB0l6rX2Z1uUnkyCYBBdcuP2+r35FikxrvbS4/JEPf1YI8Ybb9qws4p6vPPOJj31+eWnovfShtH7l6muvnJzU1lktDRHVtcvbCSAPHtynkmG31xqOcNGlnVfcctNv/PKHc3P+nj0XHzt2RGj+gw995gd+5OUH9i2dd97kzGxnWNwnaKIo0e5kX/j8vX/6R7uve+5Fu3ZP3nfPgc986p4vfP5eie72nZNlObj48m2XXNY5yURIzYatEYitNqzFpd/UfsvbX/Fr/+kP6sLVtoxeK00npchZ/bVIjGHRloKdk5/59F0mDUop5sCwWZaU1UgbXHv9ZVdeeVGSQikCBeYgpVQq3ffMgg+TSjsSXkgwRyEglLducMGFO7M8CxZSKmbyzrVaSSQxKpZqC5Pive97y9NPHXjsoaXBYADWwcXHHjl8YF/Ydp4kgRAhBGIMRBSZBNGqrflE5kUGYaMMm6vmjhjwyMNPMo1IFpFrIm7Ucx9HvYn0m2+7RSqhVApGCM07Yb82gZ1kS5Sy6SqCPZiDD5aijTH6vitH8olHapOG/sro7q889MW77vMuCskm5a7Sr7r1xT4gFPgv7//g3NzSk08+uXXr9LBY+Nc/85NXPGf79m275o8NF+aqlcX9tubDB0ZP7417LhZZC2WBQ4d8XY+EDNqIshp0Cc5BGghSgDtxh9RMKu74U5/UJVZ7+PojEbSapsZo0+/X7//13//aPY8JypIk6fb0D7zv3ZdevgUQq7NUc87/5T2G/jexyWX0cZr3sJGk1JoQ8c1vvHVxzt9z9xPDYSGp88TDw6ce+9IXPrN/bm5uNCr37zu0Zfa8JMlG5fzUrHn1a19w6eU7pAQYSap8KFuJ8lZCqBgkRxkjSOIlL73pL/7nZ+cO+8WFlUsuu2owOvjRj3zy/At6115/kTGobSlkEMy19UqpEGASXHf9Fc889TnrCilVjOGC3Rdu3ZpICSl1f2WYpilzQSSlAIC7v/L4wsLclpmL54+NdBK39rLrnnvpwuIxIRADSdH2dfbow/utl0opkvmXv3Tf0SNvnNkiYxAcpUAGKBIhb1NRojOBYT9+2ztec+jA0t1fekolE4D3vha61CqSiCFIdcKBn0/kqCICiyzHy15x01e/vPfwgYMz0ztS3/3aV58Z9v/2v71/Jcv1wsLSaFhr1UkTxah27d7+ylteQhKxJqEbN4lWjPDOn3/e7vmlvUpqWyMGfNu3vf6zn3rwsYcWl5YWiDAcFvff/+CRwy/cvmMSAu9895vuvOPexWNuNIqJnh2Owgd/568/93dbGf7okcW5o8VU9xIljZK+28tvfuHVRRnzTgBrMNZcfVblSxPuHEJ81Wuuuv++5935+ScOPVNo1VmX01KsrmUZ4KSuWELZ2hWV/8iHP1vZZSlMjOivDHsTrbLqT0zJqqq2bzsvNUlRLjNYiAygg/sXfuk//I5zdeTAHKz1xhgfCmVspxff9Z4333LrC/M0LYsAlswcYzBGKE0hQklMb8ENN13+1BN3Ls0NJye2WxsO7Jv78P/8+Ht/+I1EiAwhQ2SvhPBBCCUBD1IAgwKOm01P3+tv7bXh8nL/i3d99e6vuqoaMUKMkaMyRiaZv+Ty7bfeekurDSEBRuDmpWJce0V2vGsQWMUY0GQIiSAiKSURBKVSth+5f/En/tmvD4aLSdJyNnpHiZl44onHZrfmL3vlDddcd36TF/CZZw7ed8/TWkyPluKgKIycAWPbtokbb7rh7z9z79zR4fZte8rC/+p//MM9F287f9dWIfnzf/+lo0cXRkM7MdUmtfLilz7PBwz6rtPTMSaIDFaIzRKqhqgh4tpbweO9+nguGgDHLezNcQGGEMiz5NCBhX1PLyjZYmaSha3oRHphWnvPzGvxCJuMTXdDZyHG4KxVBsv94fU3zr70FdexWJHKdrq5s+HY0ZWnn5x75ql5Se2L9ly2tDSXZF7o4fU37PnW73httweSqGt0ugnIRZRlvWzdAOTSTLQ6GA555y56+7e8Qahqx3lTd33x861W66kn93/tnkfKEcCoqpKEVcYxyjSTWQapcN1zL+9OKBaDdg9pxpdceoFJm2SVblQsZS0BUftQaIPREJ/7+88wfH+wYN0gYnDLrTf/i59896/8+o/95//yYz/3C//8N37z+3/hF9994cWTkMs6LXUS5+aO7nvmkFZQMu33+73J3IeRMp5RWRcA5G1x/gXZW97+ClJLwiyynPeYJznSSRAynLRuO2lb5QhCVeNlr7js1tc+vzsZdVJNz+ZZbg7sP9ZfDsM+LS+WnU5PG9Ru0cWFa66/8Lk3XHH8ZCEgy5WPg8ouHz76dKebDob9qRksLVVZC7e+5kUkh3nHdidJSHfXXXd86Pf/wHt4hy1bzXvf906VDGe2mMouCQln6dGHDz38wOFRXyS6u23b1qWVQ4PRkefdePl7vuv1SUonVrLkIEqIEUQJcqMRpIbJoFN81/e8NWuFrG0j+scLnBBDDICFDFkuW+1kampm0K+Hfe+sCM4YNaHEpLe6GDFi0mkneQvaIGIIMcpbcmq6VxbOOym4S7EnuNtKZ1O9xVXZ8qKLPul10lYLVTVMUijjK7vs40hpbl7bQuDb33Hrc66+sNNT2sSV/kJd13fccUfzUMwQkiOXQlmmIUQFcqBwsk532gKcHKiEHEKMOt00hDDo17YmgRZxO/okhmQ4cGXh0wQxAgxrASDPU4gaooQoQfXqnA0AEFIJIalR2sgxiohR4CGobrfbVRmDN+WIi5HLW6nzwwv2zF593QXf+/3v7E0iMgaD0G6nQqDXm15ZLlqtDpEsCuRt3Pam18xubWct0R/Mj0aje+95+I7PP/TL//EDH/zAX9771UcBbzJfVMcuuWznD/3wW/IW2h299txrMrdperkM2QfVjRGjscafWkmihBhArkAMmm+bUeB9TJJESUNE3vt2JwcAeFBYsyNhzZdm07EJZXRc9x/WDNPE0ESJVMaHOLOlDYlb33D1B//oFy67ajLvFUlnsH23GlT7ZNqv45EyHGhNjWZ38j/+wdt+4qe/b2YbSKOsB6T8oFjK2lTYo0m7kGlfmNGwXIyMCA+BF7zkOS971XP7xYHOhFzqH+0PB1+5+759++eFRHeiB1ktDw9uP797ZG5f7UAC19942fl7ug6Hat5fxUO3vv7m5RWA4HiYdX3lj0KtJK3YH2FlMLz7nq9kLeNjJUzl4sotr3lh3gEpBEJ3ClBwhCuv25r2Fh3tH9lnVFr/5/f/5nCITrc1PTuxONib9Yo6LESy3Z4sqygkInDN8y748Z/+7sXhQ5TMmdaKw3zWCfNLB3XSZHgPTI4pMMXVVyQiQgSTQSZ453e95p/+xDu27aZjyw84HGlPuqXBwWE5P7UlK93R0u+77KreP/vJd7zvn3x7e4KKapi2tI/oF0eFGVme83TMtEYeS1t3dEFod1PrcdvbXnDTiy9cKh6COVb5+ayVHjm6tHfvEaVBCs9/8RX/5MffeeHlWcXPFO6gSm2aa2VkWRe96eTI4n29LcVbvv2FP/ZT707biOQhiCRcRKTCYZ7VMst+oJGLFoQQXWWr7jT+xb/8xypfDPJwkEd1PvRYYeHKCiRgA2wYiGSliofqOO8xFEpmrZwpsKghI8kYyfnoQUYbrPQRyaq0ZLVk+VigZZk4aZiFI+VUWtdxXqUuoGp3e5FVVaMokbTY8kLaqdO2i1QIHQIgFAIjbePW199k2oNjy49MbaU6LO0/uO9nf+6XrUVVQycqUuF43rT7rI7pzLlQgYiZfODa2sAMotpa610Tu2jDssMxSuZYz5EZRCqUYZ2IiEAy6jSwGAWMOhOJ9bAeTDApXMDC8lGha9MqHeZaE1GaOBg5ocHCRw4MERmVRe1XLC/IdEhmseb9I/8kqxWWgyocVdkwaQ8v+qbeq15/3a++/0e37lSkYFLoFFEUplXuO3y3yhfbE2F5eCDNAYHz93R/5ufed+W1O3RW5N2Yd9SwGM3O7l5a9ELki/39z7lu6xvf9sJf+KUfhEIErC8jah+rAEsSoyqM6qPQi6zn6nggaRdVWGCBwOyCDewhCITaodUxpPtOHFCtuTLs91w0r2CFgHWjwXDe+iWdWJN6IR1EM9X5tQ23cOrOCZuGzThvbOQEI7TKADDHGH0gKaVM8jA1K379N//Zwf3uji98+aEHHyuKLUVRpGm656Lzrrn2isu+adeOnYokyqpOM5EkJnjevmPy+37gnRxTjlII+Di48OIpoaA0g7D9PLz6dTdduGcbkRJCEMWyXhYCwxGec9XF3/2931qVodVqaU1NioMkxeu++UUvfdkLEpNW9WjPxd0kBQjPveFSqd5qTBpiMbstyXJETt7+LW9SNFVXUUoNqndfOCkUvA3U5MsRmJ7Fy155/bYdHaVMURRSSqVUZEiFt779dVKkaZqQtN0JLSTYMoDIsd0VN958+Y//1Pd1u1N1FUajgdLxwj3nNS9M1tfhiY8UinKYJt1tO+i2t1537XMveeShvU888fSRwwuDlR0hhCRV27Zfdelluy+/Yvf5u2c7XeFjpY0gikqLiy7Z+s53vzE1E8yUpOgP5/I2ygrawHvf7qrv/K5vvuI5F7XyyRioqiqhnPe2LCFVzHLxmtfd+Pybb3j80QP33P3I3icOjQZeCD0YDHbt2nnxpS++6uqLr7y6ZxIUZZAqeseAz1vqtrfcYiuZZ+3aDlod1emYEAIJpKkKHpdfOfu+H31HXUIIWddVxOj5L3iOSVHVUBpv/7ZXu9ogZuAELFdtJuQAENI8T8tqWSh35VV7IiNv4bY33+psBBtwgpiAE7ACJBAjlyF6oxNGlMpdceUl2gDAa1//4uffPJTUDoFBdnI6kQrOB62llHjeTZf+wPu+xTutlCiqhcDD6ZmcCFLilbe8+AUveJGSWfCQyk9O50KsuqsLIYxOmTlGGGMAEQPaHbzprbe+7OWlkq3gRfCneFVHUABZqWx3QpsUPkSGiBF5C2/7ltcf3l+18l5lV3Tin3PVRSYRITjnK6UURx8DtVritrfc8vJXOKM6zDQYLhljlDRCKK31xGR7x87Zmdm0dqVJsbRcJFkOIEnkq179oufffEOiJmOMtetfsGerkACh1RE7dPf9//UnvvbV+ScePfa5z3z1sUcOHTty7Jqrb7zyqgu37Kxf/LKrLrwoJYGyilku2DPDNdn+QuBuV77tW19HnII1c3BxZc9F25UCIxJEjNF7D1ZK4RW3vPDa665hZhKhtsMbbnoOEcrSpal+422vedGLXhI82u28skudrln1/Fn1TI2gr3PDmv+DbM5XmRs6Kq6ua0IIzCyEEEKBUZUwGtYCjDRFVaGq0GpDG0SG805rGaMXQoTAUugYEANiRON0FSKnOUWuYgxGJ2AVPOoKTYialChK1ppMgibErnn1LAQCQylYy2VZZ2lKQG3RpKsF4OombQVCgFA+y1QIqCqbaDMcxlZLxIisBQA+RKVE42EqBJxFXbMxpDWY4RyMATOqElqDCN7DutDtyqoOaSqtjVIKKTEchFZLWouqCkTo9qS1bJLj7dtYDE74BhBEjMLWkUgZDTDqGs4iz09EdzIjMkJkhjVGCNk4fykw5ueqqYkUAANLy6OJyRYjaE11XWudCCFGg9hqCTBiXBXfUrGUtLi4nGftNFUcUZXwHlohSVCWyDKQQGyc4YDITMQx+hhhtFlZ9mmikgRFARLIclhnlRJCCOdYK+nsqoGxKCMJ3+6YyPVoNGy3elXltVzbe+U4BADeQSkI2ZiGV5NyDAZVnqXABmYGZlgLraENnAMJKIUQ4GyT4QQhIElQu5C3JOCZJYHAKEZIEoQAayEVymo0NdVyFisr5UQvA1DXEBJZDh98jKv7YxmjADhXa60BWBuNVivLXivVdMX0lF1519ouRB4VyxMT7cheKeVsMCYtRsFWMkubBRZM0jy4ZY5SSIDqmhOj+ytBq9UYPCEh5WpP9h5ESFM0SkWIqGub58Z7CAHngiCpJILHSt9Nz2gQfIhCCEHor4ROSz79JP/Cz73/ySeOjfpi965Llal/83f+UdZZPWFtbZKSdVWWGkAw4CwbbRbmizzLmwEyGJYzs9naRmghcmQWghJiDPowBkRghg9otcEM573RatBnJakq0W7DByQZICzIAate4asxMo3FflOx+fZhOSONi8zxKmzid4kkR1gblRJSoq6hFKREiAx4ho/RC6GIhBSaWXCEtT5N1ZpLlge4qPpCiMQkzgUiqWTSJH7jCCKAEEKIMUopiQRHMIMEQnBayyaCAID3USnhfRRCeO+boXWcEBiAlATA2tVvi6JK09X0icxwNjKz1lJIxAgibtIBj0YjIplnqbUMkFnb6Tiuui2wtbZJ1ZEkuvESawaMtbbVytaa93imjtW0MgTR3HYIbHQCwHsohaqujTFSEDOsjTFGY5RUCMFJST74ZkNYKbRY2zAvhEDEQgBgH3wIUQhB0EqppkIaL+CyHBljtE42alxunte51VZWSjUVC8QQ2PuYJM2bQ8TYCFNupFiTZ4NICpJSAquRezGyI+LInpmVSBlNdM/xmIW1OmFqQuMa85r3XgihlAmB1xXmNbULUiQAGM3wjpGjd/A+pElbStnIVikJ8FVdSqkEmbp2UkqttRAIgYlICNS1AxBCEEKkqQHgXIgxJomOMRIEM6z1SimlwcwxeikpBHYuSJEoRd5BCAixzrty9aEiKBIQ2ccYlVSRm+qCFAlHigFSHe/bjhGEEIIEAGZyjqVQUqJJyxXWCjMjhNiMBQBpmji32pMbtanp/zEooxEZzJASPsQYvZTau5gYGRzuv7dvVPeOzz/yF3/+ydotvuzW3T/0T94BxHY788E2TZYYw2CAvIshsDHpcd+WGBFiALzWIsIysyBFMDGsxrasbZAARmwmeu+jFLrRsZptPH2AkH7V1tG8PTwhozeZkN58+7CcjaZzrL6bphhiFCKEGIQChPCRQRzRZLazSsMoBSkiYgyI7AUZktCGfLCRHYnA7IWE1kJJTZBSNjl9QjPYgGZscJIkUhIjWFvE6AEwkxBSI2WOIXrnrLW21WoJKQhSa2JUPljvPUFqlQuhQghVZUl4qaisndFpjJ5ZA3JVY0qE9z6yjSE6Z7M0aTpQlmsAoEjCSamti8aoEFhKcs4pJbQmHyqtdYirkdMMYW21Lp7qdCjEIIVkBBJMIjIzwwVGlpGPw9oRQEobIRTAIXhGABTAUglnnRDwgUJwShMJZiZAllWZJImSqqzKJBFFOTLGgERtHTPSzATP3vvVZWysra0jW6WUUkKRcqESShmVRkZdF+QauaakFFVVJImsbLE6hIKIEUmSNP1BSlmWIxJsPTcOCc5HIjbaSKLa1yGUSpo1pWm1jpp6IGgfQrBOCGitjTHWWu/9mi8ar01sofHTCuyI2LrS+ypJEi0SYZQxCRCdt0Agotq5RGdpkgKSWQgRIlckgnXOey+lllIZo4kYUEDwYdT0QwYxixhZSUEEISQzA0REjZYgJSJHQd4HCKU4UgSDsfb+JqzeLaORzkQSECF4rbQPPgQnhKltTZ7TNAXHZs4gIauqTpJGO7GM6DwLkiF6oSIjMCKDSUIKlqwAAgIRM6OqyixLmoAX5x1Y+phIqaqyyvNUKRGCBLxJxOLS8tHDy9t27uy28MwzrYWlpyq7FHi61W7m4ygYQhIF3WSLF0JoLUOsGHVV10opKTQgnKuTVAIxRs/MJAnwIBGjqCsnJIiYRABCYrIYQUQgb703Ji3KUgpNgqUSILn2Il2g8XrclArrPxgZTURNLl2iEzvZM0dmT4JjFMyNb6kAoJSQQi6tLDOHPG+nJrfOE1gpJSWNirKVpwwOURCxkiYwCyIpNCsSAoIEURQkQgxGryquBBgjeG27qeBFCBxj5Bi0lkqbJntkDFFKYsQYndYSEMGH1fmZotJCEjxFKajdbgOCmb23zUpWqeY6JKT00TI4xACAGYFrUDNeRYxgbs4JH6xWSipDq0oUEyhy1EYarRvH25Oz5a5G60mhgaiUCiH4YAGWiqSgCE9ESglmYoTIUQghJcVVp1SSQsKAQERQSjM8mjsDlNJEAiBjjCCRpkaQiBwTk0SOgrQ08D7GGKUkJZXKJGBccM5ZaDQrFecdINI0JRAgnPONEgpEY3RzQud8MzMwN8ozjDGN/AIiETnHWhsGMRNBE5EQ6wPSGkSjnCWJDmF1ldN49SkpQnBrPS+Cw5qYhiTFgBAwxmihAREZkiRAzNFoCaC2XpDTKvceUsIY5YMjsJAxUVKS5tVcNI3qECPXxJJJEBRTZEaILAVpTceFRoyxcaBWSgHMiFLKAC/E2hA+OSCTVx2lJSDBsbFuNyuMNFONLg8IIqa1EcXMRNEYAyCEICWcC8yBBIvV8xOIIiJH6XytVRaCr+s6yxKAiShLdWQAgUgcD6FszkaCP/7xj/zuB/6fqcntthKDleiif903v+TH/sX3hVgwhLVWa02kYgzMdPxtStPoWjdnclKkSaKlgA9WSEVggogcIkeljGyZJiAZoKaSiYzWIGIhBZHLcilINQ8LxjqPjk2mPq/jH4yto+mlTasTrd42CSbEEIMQiiBCIGYSQoTgtNaRrSDywSupnHdEIoSgVVJVVZ6njTgL0QMIIRKklDKEJgNvDNFJISMHQdL76H2UkqRq4oxBkGAdQnMbUWkmcIjMDLCSkohiiI6IBClm5X1cM4tziA4AkRTUrPqjD7ZZbDa7dIrViCkBiBBAJImIOTSDqq6DUkoIAJEo1rY2xqytZGVdO601EYHicam9zuf/RC6LunZSUjM3eO8jN1nkqenBTXU3+/8KIQgUYhBC1HWdJnmMCCGs/dZG9kY3ZlERYwzBNQ62AKy1VVX1er0QghCaCHVdM4cmy76UUgix2pTEzrkYo9ZaCNG4+kohVw1QSllbGWPWzAtRSt2EpDd/gdjoYmjiuiNLkZRlbYwRgmh1wK+ZevjEznUhNGYWZ4yq1/R0IlJSnTZu1yKAgMgBiIJEZAQPIZQQEmCiAMAH713UKo+RiUgqrqoizYy1tdZakATgvNNK1bYkwUDUSjsfpNBS6BBjDBBCNR2yMSNUVaHN6pKLiJxzicmstbpx31lNbHjcRQHeB0FKSh0CAzEEJxVijGAphCJiNEEH0TFzM1WvqSMCgHNO66SuS1CUUijZzHCCV7dx5+C52T6iqiogSkXUzNuIzCSligHOBWNM0zpaa+fcoUNHYhBl4aenttmaO91UJy5vC2bu9/utvKO19t43c4aUzWpPrA0TL4QS1ARxh7Is8iyNHAUJQJRVpZRqvOuAyIghOGYiGEBErqQU3vvEpI1iyow1d/u1mFwWgDihAG4S/gHZOhpDh1rNiovVmR8hMqSQPkQwlDLewdlojA6BpTRADMEKwVIqQaJ5faGUYpaNtRdM1lZZljVhgXE1PYyIQUgBQdK5AAiljFIC8AyK0UdmJSWhMX4J76xUJIVkJkCEEJSSRFyWZWKkUiSE8D4qpWL01oYszYqiyjLjnDNGSCliZB89cwQQiRuR7UNwlolW1++g2LxDk1IDKMpRnuVCCIJo3nEDyvtaSqwmdWvE9AYCGo2iaq1t7MUAtGriyENAEwbdLFwkKDJz5BAjA/AuRr1q41MKMTKR9NYaLZxzWosYm+Rq8D5477MsS5IsxliWLk00KCaJBjSARbDMmwAACAxJREFUxsrcZPWTEnUdkyQ5bv1nROYIiEZOAWgcXWKMjUxpVItmtiaiGLksyyxLmq4iSMcYnQ1ZSkDzQo9Acl1tEAAwlCIAjQkLgFYaiD76RiVfjYhZXXwQWITASimO5JwD2BijtYgBtg5JKmMU3ltjUo5BCEhJzvHqUoYFWIYQmUSMMXIESEodo5dKEqQgYmYGSyGYA3OIkbxvTL1GCCGFttZKKYhE8BZGrCnRa2rWuuQVHCG0BqOuXJabGL0UaCpKSmJuVmBaSsOIBCrLQinpnNdai0ab0bGZhkkws4gcuZlypAIgCKPRiIjyPG9mSiBaa7VRMUYCSSnKshRCNEqDcy5G7N6921pPRFrJ0ci2WiKyJsQjR45t374dwHA4bLfbANbCdL00q3Y5712aSO+9UgbMIcQQUdeuWUJlaWat9+yZWRtZ11YpqZUGVIywVTTaNNLOWmuMAa826Lppe5Oq0ptQjz5TROaGNbg+pHWtlo/nhD31hGJd+fWT0/pUAKcfX994J3ttn0hMtkEyl3NjLRvDuvRuax/otEuHdcfXc5aL/i+m8opneLRv6Jy8PsHC+sSY68qse5e7VkyenARxw8Xp+m+b08kN3F3PpiEdf9L19Y8NjIGnfI/T+xvO7pi00cHTx+B6d7pTfr6u3lbdG7CuQk4uxuurNKwd/7pUxXVefadmo8RJj3lq+25Upim4bk1DJ/aqP72JN/75ycU2vJkNO+2ZBuk3PHj/j/APSI8+Axt1hVN74ClljidzWN+nj4+zeFpf3/BU4uQLiTMMy3OENvq84cS+fpycy8x/ljOc4z2fPizXn+pMVzkX4mnVeMppz3JLOKebp5OlH59dOjVPevb6X3fmDT+v3hhOEzRnuej6X53Os8g7ACecdk6X1Cfd21kl7ElX3LDMmZ+F1304Syz1ep3mpBRRZ6ml0+eAjb46y6/OqfzaXW1KMb05ZfTXNfjPrkWuK7Aqh9W6UXT8QrSu5Poxdu6X+8Za99nOtrFYOfnoBnrchsQz/D3lousPrq/ADVWS04udcvzsoyVuJKSaD7RO4ztdTp2pjU67HG2kpG7w29Mf4Qwq2ImV0/GD6uSSx71H5MlT/oaVueFJzn6rp7CxLWujYnFtUXKWYji5G5yucp51bNL6a+HUwnS854p17+uwUZc7O+cy05xS+BwfZDOK6c0po/9XOJc2Fhs10pkWSv+72+wM2s2pkqW5jY1EztcnnU/5Z9zoMc+0Kj9dHKwXPXGjk6wbIScN4NP/Hue4jFi/v8Yp355FJTxtTJ5Nz93wDGfSLtc/YzyD7D8X8bGhAAXwrJr+yZyoz7NzShOfY38+RbM5a5lz6oGnLT03thGdmW/8Zd7a1U81TvJpS+dNJ6CxWWX0/0pNndv69KSS62QinXb8/wTneM9i3exy9l61oWKyoXV2w3F7pvs5k7J5LvPi6ed/VsFxmiA+yRa8fqJdPwI3nIDX3/966xY2svye/YbPcG+nHhdnqNVnsxhsIKY3XN6dvY3OfuSUb8VpM9BZONPoOFNtnHKh0789Q7qCk057dgXi7GaldeU3XpI+6xn+L7MJZfSGutXpbFjmWQUW1g2VdZ3mJIPa+tvAmfvH2fvNOf7q3GeU9Wd4Vp33FE7kYTztVNho5JzFBLH+h6f/81kH6rmz/pxnatZ4Dprnmepn/YrkXCTUmZTfcz9+ptN+vTNcw1rnPKFbbLgWObnwxjeAs93zOWmvz2YDOek2cLJclqd9e3a+sRo7k7ls/esHfKN99X8vm1BG4wz1taEB8SyttWE3xcar+6alTlI042md/vgRPlkJPaUkn+Z6Eddp6accj6vJcE8dY6f89hQfj1Me8/RnPOUBzySDztQjz6QwHlc64rr7ERu1xVla8ByHQTztr3gWeXFyEMc5X6UxH5/Spmda459JEn1dusIZWuFZpOG5L1nOpOquP8O6rzbwq8FZr3UuGvTpJdf1Ulq/ZJCnFX5WTVmc4fiGrXaWlWI8c11tIjanjP6GeVZVKJ6t2P8t7/UzXjeerOvhXG0CJ5U8y4dz5HSl4+u6gW+As+jsp3D8ls7dmHhKMd5YypytUc5+V9/YV98Y8Rxy0p9V8eTT/nmu09Ip324Ovo4hLM7wedOxCf2jz5Ez2UPOYpYSz9bhTv/JKSX/X7d1nP1OzkU727D8/44+96w3c+6C9eu66LnUz/+Z+WB9+W/giuc+i2wqzrGSz70tvt5e/f9t/gHlvRszZsyY/8/B40lszJgxYzYxYxk9ZsyYMZuXsYweM2bMmM3LWEaPGTNmzOZlLKPHjBkzZvMyltFjxowZs3kZy+gxY8aM2byMZfSYMWPGbF7GMnrMmDFjNi9jGT1mzJgxm5exjB4zZsyYzctYRo8ZM2bM5mUso8eMGTNm8zKW0WPGjBmzeRnL6DFjxozZvIxl9JgxY8ZsXsYyesyYMWM2L2MZPWbMmDGbl7GMHjNmzJjNy1hGjxkzZszmZSyjx4wZM2bzMpbRY8aMGbN5GcvoMWPGjNm8jGX0mDFjxmxexjJ6zJgxYzYvYxk9ZsyYMZuXsYweM2bMmM3LWEaPGTNmzOZlLKPHjBkzZvMyltFjxowZs3kZy+gxY8aM2byMZfSYMWPGbF7GMnrMmDFjNi9jGT1mzJgxm5exjB4zZsyYzctYRo8ZM2bM5mUso8eMGTNm8zKW0WPGjBmzeVHM/H/7HsaMGTNmzAYQiMYyesyYMWM2KTy2dYwZM2bMJub/DxFDddKvWnw8AAAAAElFTkSuQmCC"
UYGULAMA_ADI = "HER-İŞ ORMAN ÜRÜNLERİ STOK TAKİP SİSTEMİ"

UST_BAR = f"""
<div class="topbar">
  <a href="/" class="topbar-btn" title="Ana Sayfa">🏠</a>
  <span class="topbar-title"><img src="{LOGO_URL}" class="topbar-logo" alt="logo">HER-İŞ STOK TAKİP</span>
  <a href="/kullanici_degistir" class="topbar-btn" title="Kullanıcı Değiştir">🔁</a>
</div>
"""

HOME_BTN = UST_BAR


def sayfa(icerik, baslik="Stok Takip"):
    return (
        "<!DOCTYPE html><html lang=\"tr\"><head><meta charset=\"utf-8\">"
        + "<title>" + baslik + "</title>"
        + BASE_HEAD_EXTRA
        + "<style>" + ORTAK_CSS + "</style>"
        + "</head><body>"
        + UST_BAR
        + "<div class=\"sayfa\">" + icerik + "</div>"
        + "</body></html>"
    )

@app.route("/palet_giris", methods=["GET", "POST"])
@rol_gerekli("depocu", "muhasebeci", "patron")
def palet_giris():
    if request.method == "POST":
        barkod = request.form.get("barkod", "").strip()
        ad = request.form.get("ad", "").strip()
        depo = request.form.get("depo", "")
        adetler = request.form.getlist("palet_adet")

        if not barkod or not adetler:
            return sayfa('<p class="hata">❌ Barkod ve en az bir palet adedi girilmeli.</p><a class="btn gri" href="/palet_giris">⬅ Geri Dön</a>', "Hata")

        toplam = 0
        for a in adetler:
            try:
                toplam += int(a)
            except (TypeError, ValueError):
                pass

        if toplam <= 0:
            return sayfa('<p class="hata">❌ Toplam adet 0 olamaz.</p><a class="btn gri" href="/palet_giris">⬅ Geri Dön</a>', "Hata")

        con = db()
        try:
            with con:
                with con.cursor() as cur:
                    cur.execute("SELECT id, ad, adet FROM urun WHERE barkod=%s", (barkod,))
                    row = cur.fetchone()
                    if row:
                        uid, urun_ad, mevcut = row
                        cur.execute("UPDATE urun SET adet=%s WHERE id=%s", (mevcut + toplam, uid))
                        ad = urun_ad
                    else:
                        cur.execute("""
                            INSERT INTO urun(ad,cins,ebat,kalinlik,yuzey,sinif,renk,adet,depo,barkod)
                            VALUES(%s,'','','','','','',%s,%s,%s)
                        """, (ad or "İsimsiz Ürün", toplam, depo, barkod))

                    kullanici = session.get("kullanici", "Bilinmiyor")
                    simdi = tr_simdi()
                    for a in adetler:
                        try:
                            palet_adet = int(a)
                        except (TypeError, ValueError):
                            continue
                        if palet_adet <= 0:
                            continue
                        cur.execute("""
                            INSERT INTO hareket (barkod, ad, tip, adet, kullanici, tarih)
                            VALUES (%s, %s, 'giris', %s, %s, %s)
                        """, (barkod, ad, palet_adet, kullanici, simdi))
        finally:
            con.close()

        icerik = (
            '<div style="text-align:center;font-size:52px;margin-bottom:4px;">✅</div>'
            + f'<h2 style="margin-top:0;">{len(adetler)} Palet Girişi Yapıldı</h2>'
            + f'<p style="text-align:center;color:var(--muted);"><b style="color:var(--text);">{ad}</b><br>Toplam: {toplam} adet</p>'
            + '<a href="/liste" class="okut-kart okut-mor"><div class="okut-ikon">📦</div><div class="okut-metin"><div class="okut-baslik">Stok Listesine Git</div></div><div class="okut-ok">›</div></a>'
            + '<a href="/palet_giris" class="okut-kart okut-yesil"><div class="okut-ikon">🧱</div><div class="okut-metin"><div class="okut-baslik">Yeni Palet Girişi</div></div><div class="okut-ok">›</div></a>'
        )
        return sayfa(icerik, "Palet Girişi Tamamlandı")

    icerik = """
    <h2 style="margin-bottom:2px;">🧱 Palet Girişi</h2>
    <p style="text-align:center;color:var(--muted);margin-top:0;">
    Her paletin adedi farklıysa, aşağıya her palet için ayrı satır ekle.
    </p>

    <div class="kart">
    <label>Barkod Ara (mevcut ürünse aynı barkodu seç)</label>
    <div class="urun-arama-kutu">
      <input type="text" id="urun-arama" class="arama" placeholder="🔍 Ürün adı veya barkod ara..." autocomplete="off" oninput="urunAra()">
      <div class="urun-arama-sonuc" id="urun-arama-sonuc"></div>
    </div>
    </div>

    <form method="post" id="palet-form">
    <div class="kart">
    <label>Barkod</label>
    <input name="barkod" id="form-barkod" placeholder="Barkod" required>
    <label>Ürün Adı (yeni ürünse)</label>
    <input name="ad" id="form-ad" placeholder="Ürün Adı">
    <label>Depo</label>
    <select name="depo">
    """ + "".join(f'<option>{d}</option>' for d in DEPOLAR) + """
    </select>
    </div>

    <div class="kart">
    <label>Paletler</label>
    <div id="palet-satirlari"></div>
    <button type="button" class="btn-kucuk mavi" onclick="paletEkle()">➕ Palet Ekle</button>
    <div id="palet-toplam" style="margin-top:12px;color:var(--muted);font-size:13.5px;">Toplam: 0 adet</div>
    </div>

    <button type="submit" class="btn yesil">✅ Girişi Kaydet</button>
    </form>

    <script>
    let paletSayisi = 0;

    function paletEkle(){
      paletSayisi++;
      const kapsayici = document.getElementById('palet-satirlari');
      const satir = document.createElement('div');
      satir.className = 'sepet-satir';
      satir.id = 'palet-' + paletSayisi;
      satir.innerHTML = `
        <div class="sepet-ad">Palet ${paletSayisi}</div>
        <input type="number" name="palet_adet" class="sepet-adet-input" min="1" placeholder="Adet" oninput="toplamGuncelle()">
        <button type="button" class="sepet-sil" onclick="paletSil(${paletSayisi})">✕</button>
      `;
      kapsayici.appendChild(satir);
    }

    function paletSil(id){
      const el = document.getElementById('palet-' + id);
      if(el) el.remove();
      toplamGuncelle();
    }

    function toplamGuncelle(){
      const girisler = document.querySelectorAll('input[name="palet_adet"]');
      let toplam = 0;
      girisler.forEach(g => { toplam += parseInt(g.value) || 0; });
      document.getElementById('palet-toplam').textContent = 'Toplam: ' + toplam + ' adet';
    }

    document.getElementById('palet-satirlari').addEventListener('input', toplamGuncelle);

    paletEkle();
    paletEkle();

    let aramaZamanlayici = null;
    function urunAra(){
      clearTimeout(aramaZamanlayici);
      const q = document.getElementById('urun-arama').value.trim();
      const kutu = document.getElementById('urun-arama-sonuc');
      if(q.length < 1){ kutu.style.display='none'; kutu.innerHTML=''; return; }
      aramaZamanlayici = setTimeout(() => {
        fetch('/urun_ara?q=' + encodeURIComponent(q))
          .then(r => r.json())
          .then(sonuclar => {
            if(sonuclar.length === 0){
              kutu.innerHTML = '<div class="urun-arama-oge" style="color:var(--muted);">Ürün bulunamadı, yeni ürün olarak girebilirsin</div>';
              kutu.style.display = 'block';
              return;
            }
            kutu.innerHTML = sonuclar.map(u => `
              <div class="urun-arama-oge" onclick='urunSec(${JSON.stringify(JSON.stringify(u))})'>
                <div class="ad">${u.ad}</div>
                <div class="detay">🔢 ${u.barkod} • 📦 Stokta: ${u.adet} • 🏭 ${u.depo}</div>
              </div>
            `).join('');
            kutu.style.display = 'block';
          });
      }, 250);
    }
    function urunSec(uJson){
      const u = JSON.parse(uJson);
      document.getElementById('urun-arama-sonuc').style.display = 'none';
      document.getElementById('urun-arama').value = u.ad;
      document.getElementById('form-barkod').value = u.barkod;
      document.getElementById('form-ad').value = u.ad;
    }
    document.addEventListener('click', function(e){
      const kutu = document.getElementById('urun-arama-sonuc');
      if(!kutu.contains(e.target) && e.target.id !== 'urun-arama'){
        kutu.style.display = 'none';
      }
    });
    </script>
    """
    return sayfa(icerik, "Palet Girişi")
@app.route("/manifest.json")
def manifest():
    return jsonify({
        "name": UYGULAMA_ADI,
        "short_name": "HER-İŞ Stok",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#0f1115",
        "theme_color": "#111111",
        "icons": [
            {"src": "/static/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
        ],
    })

@app.route("/transfer_gecmisi")
@rol_gerekli("depocu", "muhasebeci", "patron")
def transfer_gecmisi():
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("""
                SELECT ad, barkod, adet, kullanici, tarih
                FROM hareket
                WHERE tip IN ('giris','cikis')
                ORDER BY id DESC LIMIT 200
            """)
            kayitlar = cur.fetchall()
    finally:
        con.close()

    kartlar = ""
    for ad, barkod, adet, kullanici, tarih in kayitlar:
        kartlar += f"""
        <div class="hareket-kart hareket-giris">
          <div class="hareket-ikon">🔁</div>
          <div class="hareket-govde">
            <div class="hareket-ust">
              <div class="hareket-ad">{ad}</div>
              <div class="hareket-adet giris">{adet}</div>
            </div>
            <div class="hareket-alt">🔢 {barkod} • 👤 {kullanici} • 🕒 {tarih}</div>
          </div>
        </div>
        """

    icerik = "<h2>🔁 TRANSFER GEÇMİŞİ</h2>" + kartlar
    return sayfa(icerik, "Transfer Geçmişi")
@app.route("/qr_baglan")
def qr_baglan():
    url = request.host_url
    img = qrcode.make(url)
    bio = io.BytesIO()
    img.save(bio, format="PNG")
    bio.seek(0)
    return send_file(bio, mimetype="image/png")

@app.route("/barkod/<kod>.png")
def barkod_resim_endpoint(kod):
    with _CACHE_KILIT:
        veri = _BARKOD_CACHE.get(kod)
    if veri is None:
        veri = barkod_png_bytes(kod).getvalue()
        with _CACHE_KILIT:
            _BARKOD_CACHE[kod] = veri
    return send_file(io.BytesIO(veri), mimetype="image/png")



def rol_gerekli(*izinli_roller):
    TAM_YETKILI = ("patron", "muhasebeci")

    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            rol = session.get("rol")
            if not rol:
                return redirect("/")
            if rol not in TAM_YETKILI and rol not in izinli_roller:
                return sayfa("""
                <h2>⛔ Erişim Yetkiniz Yok</h2>
                <p style="text-align:center;color:var(--muted);">
                Bu işlem senin rolüne kapalı. Yanlış kişi olarak girdiysen
                sağ üstten kullanıcı değiştir.
                </p>
                """, "Yetkisiz Erişim")
            return f(*args, **kwargs)
        return wrapper
    return decorator


@app.route("/pin_gir/<isim>", methods=["GET", "POST"])
def pin_gir(isim):
    if isim not in ROLLER:
        return redirect("/")

    hata = None
    if request.method == "POST":
        girilen = request.form.get("pin", "")
        if girilen == PIN_KODLARI.get(isim):
            session["kullanici"] = isim
            session["rol"] = ROLLER[isim]
            return redirect("/")
        hata = "❌ Yanlış PIN, tekrar dene"

    icerik = """
    <style>
    .pin-avatar {
      width:72px; height:72px; border-radius:50%; margin: 6px auto 4px;
      background: linear-gradient(135deg, #2196F3, #00BCD4);
      display:flex; align-items:center; justify-content:center;
      font-size:28px; font-weight:800; color:white;
      box-shadow: 0 6px 20px rgba(33,150,243,.35);
    }
    .pin-noktalar { display:flex; justify-content:center; gap:14px; margin: 22px 0; }
    .pin-nokta {
      width:16px; height:16px; border-radius:50%; border:2px solid var(--border);
      background:transparent; transition: all .12s ease;
    }
    .pin-nokta.dolu { background: var(--accent); border-color: var(--accent); }
    .tuş-takimi { display:grid; grid-template-columns: repeat(3, 1fr); gap:12px; max-width:280px; margin:0 auto; }
    .tuş {
      aspect-ratio:1; border-radius:50%; border:1px solid var(--border);
      background:#12151c; color:white; font-size:22px; font-weight:700;
      display:flex; align-items:center; justify-content:center; cursor:pointer;
      transition: transform .08s ease, filter .12s ease;
    }
    .tuş:active { transform: scale(0.92); filter:brightness(1.3); background:var(--accent); }
    .tuş.bos { background:transparent; border:none; cursor:default; }
    .tuş.bos:active { transform:none; filter:none; background:transparent; }
    </style>

    <div class="pin-avatar">{{isim[0]|upper}}</div>
    <h2 style="margin-bottom:2px;">{{isim}}</h2>
    <p style="text-align:center;color:var(--muted);margin-top:0;">PIN gir</p>

    {% if hata %}<p class="hata">{{hata}}</p>{% endif %}

    <div class="pin-noktalar" id="noktalar">
      <div class="pin-nokta"></div><div class="pin-nokta"></div>
      <div class="pin-nokta"></div><div class="pin-nokta"></div>
    </div>

    <form method="post" id="pin-form">
      <input type="hidden" name="pin" id="pin-gizli">
    </form>

    <div class="tuş-takimi">
      <div class="tuş" onclick="tuslaEkle('1')">1</div>
      <div class="tuş" onclick="tuslaEkle('2')">2</div>
      <div class="tuş" onclick="tuslaEkle('3')">3</div>
      <div class="tuş" onclick="tuslaEkle('4')">4</div>
      <div class="tuş" onclick="tuslaEkle('5')">5</div>
      <div class="tuş" onclick="tuslaEkle('6')">6</div>
      <div class="tuş" onclick="tuslaEkle('7')">7</div>
      <div class="tuş" onclick="tuslaEkle('8')">8</div>
      <div class="tuş" onclick="tuslaEkle('9')">9</div>
      <div class="tuş bos"></div>
      <div class="tuş" onclick="tuslaEkle('0')">0</div>
      <div class="tuş" onclick="silTusu()">⌫</div>
    </div>

    <p style="text-align:center;margin-top:24px;">
    <a href="/" style="color:var(--muted);text-decoration:none;">⬅ Geri</a>
    </p>

    <script>
    let girilenPin = "";
    function noktalariGuncelle(){
      document.querySelectorAll('.pin-nokta').forEach((n, i) => {
        n.classList.toggle('dolu', i < girilenPin.length);
      });
    }
    function tuslaEkle(rakam){
      if(girilenPin.length >= 4) return;
      girilenPin += rakam;
      if(navigator.vibrate) navigator.vibrate(15);
      noktalariGuncelle();
      if(girilenPin.length === 4){
        document.getElementById('pin-gizli').value = girilenPin;
        setTimeout(() => document.getElementById('pin-form').submit(), 120);
      }
    }
    function silTusu(){
      girilenPin = girilenPin.slice(0, -1);
      noktalariGuncelle();
    }
    </script>
    """
    return render_template_string(sayfa(icerik, "Giriş"), isim=isim, hata=hata)


@app.route("/kullanici_degistir")
def kullanici_degistir():
    session.clear()
    return redirect("/")


@app.route("/")
def index():
    kullanici = session.get("kullanici")
    rol = session.get("rol")

    if not kullanici:
        ROL_ETIKET = {"depocu": "Depocu", "muhasebeci": "Muhasebeci", "patron": "Patron"}
        ROL_RENK = {"depocu": "kisi-yesil", "muhasebeci": "kisi-mavi", "patron": "kisi-mor"}
        secim_html = ""
        for isim in ROLLER:
            bas_harf = isim[0].upper()
            rol_adi = ROL_ETIKET.get(ROLLER[isim], "")
            renk = ROL_RENK.get(ROLLER[isim], "kisi-mavi")
            secim_html += f"""
            <a href="/pin_gir/{isim}" class="kisi-kart">
             <div class="kisi-avatar {renk}">{bas_harf}</div>
             <div class="kisi-metin">
              <div class="kisi-ad">{isim}</div>
              <div class="kisi-rol">{rol_adi}</div>
             </div>
             <div class="okut-ok">›</div>
            </a>
            """

           secim_html += """
    <a href="/transfer_gecmisi" class="okut-kart okut-mor">
      <div class="okut-ikon">🔁</div>
      <div class="okut-metin"><div class="okut-baslik">Transfer Geçmişi</div></div>
      <div class="okut-ok">›</div>
    </a>
    <a href="/palet_giris" class="okut-kart okut-mor">
      <div class="okut-ikon">🧱</div>
      <div class="okut-metin"><div class="okut-baslik">Palet Girişi</div><div class="okut-alt">Farklı adetli paletleri tek seferde gir</div></div>
      <div class="okut-ok">›</div>
    </a>
    """

    icerik = (
        f'<div style="text-align:center;margin-bottom:6px;"><img src="{LOGO_URL}" style="width:110px;height:auto;border-radius:14px;background:white;padding:8px;box-shadow:0 6px 20px rgba(0,0,0,.35);"></div>'
        + f'<h1 style="font-size:19px;line-height:1.35;margin-bottom:2px;">{UYGULAMA_ADI}</h1>'
        + '<h3 class="alt">Devam etmek için ismini seç</h3>'
        + secim_html
    )
    return sayfa(icerik, "Giriş - " + UYGULAMA_ADI)
icerik = (
    f'<div style="text-align:center;margin-bottom:6px;"><img src="{LOGO_URL}" style="width:110px;height:auto;border-radius:14px;background:white;padding:8px;box-shadow:0 6px 20px rgba(0,0,0,.35);"></div>'
    + f'<h1 style="font-size:19px;line-height:1.35;margin-bottom:2px;">{UYGULAMA_ADI}</h1>'
    + '<h3 class="alt">Devam etmek için ismini seç</h3>'
    + secim_html
)
return sayfa(icerik, "Giriş - " + UYGULAMA_ADI)

    butonlar = ""
    if rol in ("depocu", "muhasebeci", "patron"):
        giris_sayisi, cikis_sayisi = bugunku_ozet(kullanici)
        butonlar += f"""
        <div class="ozet-satir">
          <div class="ozet-kutu ozet-yesil"><div class="ozet-sayi">{giris_sayisi}</div><div class="ozet-etiket">Bugün Giriş</div></div>
          <div class="ozet-kutu ozet-turuncu"><div class="ozet-sayi">{cikis_sayisi}</div><div class="ozet-etiket">Bugün Çıkış</div></div>
        </div>
        <a href="/kamera/giris" class="okut-kart okut-yesil">
          <div class="okut-ikon">⬆️</div>
          <div class="okut-metin"><div class="okut-baslik">Giriş Okut</div><div class="okut-alt">Depoya gelen ürünü tara</div></div>
          <div class="okut-ok">›</div>
        </a>
        <a href="/kamera/cikis" class="okut-kart okut-turuncu">
          <div class="okut-ikon">⬇️</div>
          <div class="okut-metin"><div class="okut-baslik">Çıkış Okut</div><div class="okut-alt">Depodan çıkan ürünü tara</div></div>
          <div class="okut-ok">›</div>
        </a>
        """
    if rol in ("depocu",):
        butonlar += """
        <a href="/depo_stok" class="okut-kart okut-turkuaz">
          <div class="okut-ikon">🏭</div>
          <div class="okut-metin"><div class="okut-baslik">Depo Stok Durumu</div><div class="okut-alt">Hangi depoda ne kadar var</div></div>
          <div class="okut-ok">›</div>
        </a>
        <a href="/transfer" class="okut-kart okut-mavi">
          <div class="okut-ikon">🔁</div>
          <div class="okut-metin"><div class="okut-baslik">Depolar Arası Transfer</div><div class="okut-alt">Ürünü bir depodan diğerine taşı</div></div>
          <div class="okut-ok">›</div>
        </a>
        """
    if rol in ("muhasebeci", "patron"):
        butonlar += """
        <div class="bolum-baslik">Sipariş</div>
        <a href="/siparis_olustur" class="okut-kart okut-yesil">
          <div class="okut-ikon">🧾</div>
          <div class="okut-metin"><div class="okut-baslik">Yeni Sipariş Oluştur</div><div class="okut-alt">Depocuya hazırlatılacak ürünleri seç</div></div>
          <div class="okut-ok">›</div>
        </a>
        <a href="/siparisler" class="okut-kart okut-turuncu">
          <div class="okut-ikon">📋</div>
          <div class="okut-metin"><div class="okut-baslik">Siparişler</div><div class="okut-alt">Açık ve tamamlanan siparişler</div></div>
          <div class="okut-ok">›</div>
        </a>

        <div class="bolum-baslik">İşlemler</div>
        <a href="/ekle" class="okut-kart okut-mavi">
          <div class="okut-ikon">➕</div>
          <div class="okut-metin"><div class="okut-baslik">Ürün Ekle</div><div class="okut-alt">Yeni ürün kaydı oluştur</div></div>
          <div class="okut-ok">›</div>
        </a>
        <a href="/toplu_yukle" class="okut-kart okut-turkuaz">
          <div class="okut-ikon">📥</div>
          <div class="okut-metin"><div class="okut-baslik">Excel'den Toplu Yükle</div><div class="okut-alt">Excel yükle → otomatik barkod → toplu etiket</div></div>
          <div class="okut-ok">›</div>
        </a>
        <a href="/liste" class="okut-kart okut-mor">
          <div class="okut-ikon">📦</div>
          <div class="okut-metin"><div class="okut-baslik">Stok Listesi</div><div class="okut-alt">Ürünleri görüntüle, etiket bas</div></div>
          <div class="okut-ok">›</div>
        </a>
        <a href="/depo_stok" class="okut-kart okut-turkuaz">
          <div class="okut-ikon">🏭</div>
          <div class="okut-metin"><div class="okut-baslik">Depo Stok Durumu</div><div class="okut-alt">Hangi depoda ne kadar ürün var</div></div>
          <div class="okut-ok">›</div>
        </a>
        <a href="/hareketler" class="okut-kart okut-kirmizi">
          <div class="okut-ikon">📊</div>
          <div class="okut-metin"><div class="okut-baslik">Hareketler</div><div class="okut-alt">Tüm giriş/çıkış geçmişi</div></div>
          <div class="okut-ok">›</div>
        </a>

        <div class="bolum-baslik">Raporlar</div>
        <div class="rapor-satir">
          <a href="/rapor/excel" class="rapor-pil">📥 XLSX</a>
          <a href="/rapor/xls" class="rapor-pil">📥 XLS</a>
          <a href="/rapor/csv" class="rapor-pil">📥 CSV</a>
        </div>
        """


    icerik = (
        "<h1>📦 STOK PANEL</h1>"
        + '<h3 class="alt">👤 ' + kullanici + ' (' + rol + ')</h3>'
        + butonlar
    )
    return sayfa(icerik, "Stok Panel")

@app.route("/ekle", methods=["GET", "POST"])
@rol_gerekli("depocu", "patron")
def ekle2():
    on_dolu_barkod = request.args.get("barkod", "")

    if request.method == "POST":
        ad = request.form.get("ad", "").strip()
        if not ad:
            return sayfa('<p class="hata">❌ Ürün adı boş olamaz.</p><a class="btn gri" href="/ekle">⬅ Geri Dön</a>', "Hata")

        try:
            adet = int(request.form.get("adet", "").strip())
        except (TypeError, ValueError):
            return sayfa('<p class="hata">❌ Adet sayısal olmalı.</p><a class="btn gri" href="/ekle">⬅ Geri Dön</a>', "Hata")

        barkod = request.form.get("barkod")
        if not barkod:
            barkod = barkod_uret()

        con = db()
        try:
            with con:
                with con.cursor() as cur:
                    cur.execute("""
                    INSERT INTO urun(ad,cins,ebat,kalinlik,yuzey,sinif,renk,adet,depo,barkod)
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        ad,
                        request.form.get("cins", ""),
                        request.form.get("ebat", ""),
                        request.form.get("kalinlik", ""),
                        request.form.get("yuzey", ""),
                        request.form.get("sinif", ""),
                        request.form.get("renk", ""),
                        adet,
                        request.form.get("depo", ""),
                        barkod,
                    ))
                    cur.execute("""
                    INSERT INTO hareket (barkod, ad, tip, adet, kullanici, tarih)
                    VALUES (%s, %s, 'giris', %s, %s, %s)
                    """, (barkod, ad, adet, session.get("kullanici", "Bilinmiyor"), tr_simdi()))
        finally:
            con.close()

        icerik = (
            '<div style="text-align:center;font-size:52px;margin-bottom:4px;">✅</div>'
            + '<h2 style="margin-top:0;">Ürün Kaydedildi</h2>'
            + '<p style="text-align:center;color:var(--muted);margin-top:-8px;"><b style="color:var(--text);">' + ad + '</b></p>'
            + '<div class="kart" style="text-align:center;">'
            + '<span class="rozet">Barkod</span><br><br>'
            + '<b style="font-size:18px;letter-spacing:1px;">' + barkod + '</b><br><br>'
            + '<img src="/barkod/' + barkod + '.png" width="260"><br><br>'
            + '<img src="/qr/' + barkod + '.png" width="140">'
            + '</div>'
            + '<a href="/etiket/' + barkod + '" target="_blank" class="okut-kart okut-mavi">'
            + '<div class="okut-ikon">🖨️</div><div class="okut-metin"><div class="okut-baslik">Etiket Yazdır</div></div><div class="okut-ok">›</div></a>'
            + '<a href="/liste" class="okut-kart okut-mor">'
            + '<div class="okut-ikon">📦</div><div class="okut-metin"><div class="okut-baslik">Stok Listesine Git</div></div><div class="okut-ok">›</div></a>'
            + '<a href="/ekle" class="okut-kart okut-yesil">'
            + '<div class="okut-ikon">➕</div><div class="okut-metin"><div class="okut-baslik">Yeni Ürün Ekle</div></div><div class="okut-ok">›</div></a>'
        )
        return sayfa(icerik, "Ürün Kaydedildi")

    icerik = """
    <h2 style="margin-bottom:2px;">➕ Ürün Ekle</h2>
    <p style="text-align:center;color:var(--muted);margin-top:0;">Yeni ürün bilgilerini gir</p>

    {% if on_dolu_barkod %}
    <div class="sonuc-kart sonuc-yeni" style="margin-bottom:4px;">
      📷 Okutulan barkod: <b>{{on_dolu_barkod}}</b> — bu ürün stokta yok, yeni ürün olarak ekleyin.
    </div>
    {% endif %}

    <form method="post">
    <div class="kart">
    <label>Barkod</label>
    <input name="barkod" value="{{on_dolu_barkod}}" placeholder="Boş bırak = otomatik barkod">
    <label>Ürün Adı</label>
    <input name="ad" placeholder="Ürün Adı" required>
    </div>

    <div class="kart">
    <label>Cins</label>
    <input name="cins" placeholder="Cins">
    <label>Ebat</label>
    <input name="ebat" placeholder="Ebat">
    <label>Kalınlık</label>
    <input name="kalinlik" placeholder="Kalınlık">
    <label>Yüzey</label>
    <select name="yuzey" required>
        <option value="">Seçiniz</option>
        <option value="HG">HG</option>
        <option value="MAT">MAT</option>
    </select>
    <label>Sınıf</label>
    <input name="sinif" placeholder="Sınıf">
    <label>Renk</label>
    <input name="renk" placeholder="Renk">
    </div>

    <div class="kart">
    <label>Adet</label>
    <input name="adet" type="number" placeholder="Adet" required>
    <label>Depo</label>
    <select name="depo">
    {% for d in depolar %}
    <option>{{d}}</option>
    {% endfor %}
    </select>
    <button class="btn mavi">Kaydet</button>
    </form>
    """
    return render_template_string(sayfa(icerik, "Ürün Ekle"), depolar=DEPOLAR, on_dolu_barkod=on_dolu_barkod)




# ============ TOPLU EXCEL YÜKLEME ============
# Excel şablonu: Ad | Cins | Ebat | Adet | Depo  (Barkod otomatik üretilir)
TOPLU_SABLON_BASLIKLAR = ["Ad", "Cins", "Ebat", "Adet", "Depo"]


@app.route("/toplu_sablon.xlsx")
@rol_gerekli("muhasebeci")
def toplu_sablon():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"
    ws.append(TOPLU_SABLON_BASLIKLAR)
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2196F3")
    ws.append(["Örnek MDF Levha", "MDF", "210x280", 10, DEPOLAR[0]])
    for col_cells in ws.columns:
        uzunluk = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(uzunluk + 2, 10), 40)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="toplu_urun_sablonu.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/toplu_yukle", methods=["GET", "POST"])
@rol_gerekli("muhasebeci")
def toplu_yukle():
    if request.method == "POST":
        dosya = request.files.get("dosya")
        if not dosya or dosya.filename == "":
            return sayfa('<p class="hata">❌ Dosya seçilmedi.</p><a class="btn gri" href="/toplu_yukle">⬅ Geri Dön</a>', "Hata")

        from openpyxl import load_workbook
        try:
            wb = load_workbook(dosya, data_only=True)
        except Exception:
            return sayfa('<p class="hata">❌ Dosya okunamadı. Geçerli bir .xlsx dosyası yükleyin.</p><a class="btn gri" href="/toplu_yukle">⬅ Geri Dön</a>', "Hata")

        ws = wb.active
        satirlar = list(ws.iter_rows(values_only=True))
        if not satirlar:
            return sayfa('<p class="hata">❌ Excel dosyası boş.</p><a class="btn gri" href="/toplu_yukle">⬅ Geri Dön</a>', "Hata")

        baslik_satiri = [str(h).strip().lower() if h is not None else "" for h in satirlar[0]]

        def sutun_bul(*adaylar):
            for aday in adaylar:
                if aday in baslik_satiri:
                    return baslik_satiri.index(aday)
            return None

        idx_ad = sutun_bul("ad", "ürün adı", "urun adi", "ürün", "urun")
        idx_cins = sutun_bul("cins")
        idx_ebat = sutun_bul("ebat")
        idx_adet = sutun_bul("adet")
        idx_depo = sutun_bul("depo")

        if idx_ad is None or idx_adet is None:
            return sayfa(
                '<p class="hata">❌ Excel\'de en az "Ad" ve "Adet" sütunları olmalı.</p>'
                '<p style="text-align:center;color:var(--muted);font-size:13px;">Beklenen sütunlar: Ad, Cins, Ebat, Adet, Depo</p>'
                '<a class="btn gri" href="/toplu_yukle">⬅ Geri Dön</a>', "Hata")

        eklenenler = []
        hatalar = []

        con = db()
        try:
            with con:
                with con.cursor() as cur:
                    for satir_no, satir in enumerate(satirlar[1:], start=2):
                        if satir is None or all(h is None or str(h).strip() == "" for h in satir):
                            continue

                        def deger_al(idx):
                            if idx is None or idx >= len(satir):
                                return ""
                            v = satir[idx]
                            return str(v).strip() if v is not None else ""

                        ad = deger_al(idx_ad)
                        cins = deger_al(idx_cins)
                        ebat = deger_al(idx_ebat)
                        depo = deger_al(idx_depo)
                        adet_ham = deger_al(idx_adet)

                        if not ad:
                            hatalar.append(f"Satır {satir_no}: Ürün adı boş, atlandı.")
                            continue

                        try:
                            adet = int(float(adet_ham)) if adet_ham != "" else 0
                        except (TypeError, ValueError):
                            hatalar.append(f"Satır {satir_no}: '{ad}' — Adet sayısal değil, atlandı.")
                            continue

                        barkod = barkod_uret()

                        cur.execute("""
                        INSERT INTO urun(ad,cins,ebat,kalinlik,yuzey,sinif,renk,adet,depo,barkod)
                        VALUES(%s,%s,%s,'', '', '', '', %s,%s,%s)
                        """, (ad, cins, ebat, adet, depo, barkod))
                        cur.execute("""
                        INSERT INTO hareket (barkod, ad, tip, adet, kullanici, tarih)
                        VALUES (%s, %s, 'giris', %s, %s, %s)
                        """, (barkod, ad, adet, session.get("kullanici", "Bilinmiyor"), tr_simdi()))

                        eklenenler.append((ad, barkod, adet))
        finally:
            con.close()

        if not eklenenler:
            hata_html = "".join(f'<div class="on-izleme-hata">{h}</div>' for h in hatalar) or "<p>Eklenecek geçerli satır bulunamadı.</p>"
            return sayfa(
                '<p class="hata">❌ Hiçbir ürün eklenemedi.</p>'
                + f'<div class="kart">{hata_html}</div>'
                + '<a class="btn gri" href="/toplu_yukle">⬅ Geri Dön</a>', "Hata")

        satir_html = "".join(
            f'<div class="on-izleme-satir"><span>{ad}</span><span style="color:var(--muted);">{barkod} • {adet} adet</span></div>'
            for ad, barkod, adet in eklenenler
        )
        hata_html = ""
        if hatalar:
            hata_html = '<div class="kart"><b style="color:#FFB74D;">⚠️ Atlanan satırlar</b>' \
                        + "".join(f'<div class="on-izleme-hata">{h}</div>' for h in hatalar) + '</div>'

        barkod_inputlari = "".join(f'<input type="hidden" name="barkod" value="{b}">' for _, b, _ in eklenenler)

        icerik = (
            '<div style="text-align:center;font-size:52px;margin-bottom:4px;">✅</div>'
            + f'<h2 style="margin-top:0;">{len(eklenenler)} Ürün Eklendi</h2>'
            + f'<div class="kart" style="padding:6px 16px;">{satir_html}</div>'
            + hata_html
            + f"""
            <form method="post" action="/etiketler" target="_blank">
              {barkod_inputlari}
              <button class="btn turkuaz" type="submit">🖨️ {len(eklenenler)} Ürünün Etiketini Tek Sayfada Yazdır</button>
            </form>
            """
            + '<a href="/liste" class="okut-kart okut-mor">'
            + '<div class="okut-ikon">📦</div><div class="okut-metin"><div class="okut-baslik">Stok Listesine Git</div></div><div class="okut-ok">›</div></a>'
            + '<a href="/toplu_yukle" class="okut-kart okut-turkuaz">'
            + '<div class="okut-ikon">📥</div><div class="okut-metin"><div class="okut-baslik">Yeni Dosya Yükle</div></div><div class="okut-ok">›</div></a>'
        )
        return sayfa(icerik, "Toplu Yükleme Tamamlandı")

    icerik = """
    <h2 style="margin-bottom:2px;">📥 Excel'den Toplu Yükle</h2>
    <p style="text-align:center;color:var(--muted);margin-top:0;">
    Ad, Cins, Ebat, Adet, Depo sütunlu bir Excel (.xlsx) dosyası yükle —
    her satır için otomatik barkod üretilir ve ürünler tek seferde eklenir.
    </p>

    <div class="kart" style="text-align:center;">
      <a class="sablon-link" href="/toplu_sablon.xlsx">📄 Örnek Excel Şablonunu İndir</a>
    </div>

    <form method="post" enctype="multipart/form-data">
    <div class="yukleme-alan">
      <div class="yukleme-ikon">📊</div>
      <p style="margin:6px 0;color:var(--muted);font-size:14px;">.xlsx dosyanı seç</p>
      <input type="file" name="dosya" accept=".xlsx" required>
    </div>
    <button class="btn turkuaz">Yükle ve Ürünleri Ekle</button>
    </form>

    <div class="kart">
      <b style="font-size:13.5px;">📋 Beklenen sütunlar:</b>
      <div style="color:var(--muted);font-size:13px;margin-top:6px;line-height:1.8;">
        <b style="color:var(--text);">Ad</b> (zorunlu) • Cins • Ebat •
        <b style="color:var(--text);">Adet</b> (zorunlu) • Depo<br>
        Barkod otomatik üretilir, Excel'e barkod yazmana gerek yok.
      </div>
    </div>
    """
    return sayfa(icerik, "Toplu Yükleme")


@app.route("/duzenle/<eski_barkod>", methods=["GET", "POST"])
@rol_gerekli("muhasebeci")
def duzenle(eski_barkod):
    if request.method == "POST":
        ad = request.form.get("ad", "").strip()
        if not ad:
            return sayfa('<p class="hata">❌ Ürün adı boş olamaz.</p><a class="btn gri" href="/duzenle/' + eski_barkod + '">⬅ Geri Dön</a>', "Hata")

        try:
            adet = int(request.form.get("adet", "").strip())
        except (TypeError, ValueError):
            return sayfa('<p class="hata">❌ Adet sayısal olmalı.</p><a class="btn gri" href="/duzenle/' + eski_barkod + '">⬅ Geri Dön</a>', "Hata")

        yeni_barkod = request.form.get("barkod", "").strip()
        if not yeni_barkod:
            yeni_barkod = eski_barkod

        con = db()
        try:
            with con:
                with con.cursor() as cur:
                    if yeni_barkod != eski_barkod:
                        cur.execute("SELECT id FROM urun WHERE barkod=%s", (yeni_barkod,))
                        if cur.fetchone():
                            con.close()
                            return sayfa(
                                '<p class="hata">❌ Bu barkod zaten başka bir üründe kullanılıyor.</p>'
                                '<a class="btn gri" href="/duzenle/' + eski_barkod + '">⬅ Geri Dön</a>',
                                "Hata"
                            )
                    cur.execute("""
                        UPDATE urun
                        SET ad=%s, cins=%s, ebat=%s, kalinlik=%s, yuzey=%s, sinif=%s, renk=%s, adet=%s, depo=%s, barkod=%s
                        WHERE barkod=%s
                    """, (
                        ad,
                        request.form.get("cins", ""),
                        request.form.get("ebat", ""),
                        request.form.get("kalinlik", ""),
                        request.form.get("yuzey", ""),
                        request.form.get("sinif", ""),
                        request.form.get("renk", ""),
                        adet,
                        request.form.get("depo", ""),
                        yeni_barkod,
                        eski_barkod,
                    ))
        finally:
            con.close()

        icerik = (
            '<div style="text-align:center;font-size:52px;margin-bottom:4px;">✅</div>'
            + '<h2 style="margin-top:0;">Ürün Güncellendi</h2>'
            + '<p style="text-align:center;color:var(--muted);margin-top:-8px;"><b style="color:var(--text);">' + ad + '</b></p>'
            + '<div class="kart" style="text-align:center;">'
            + '<span class="rozet">Barkod</span><br><br>'
            + '<b style="font-size:18px;letter-spacing:1px;">' + yeni_barkod + '</b><br><br>'
            + '<img src="/barkod/' + yeni_barkod + '.png" width="260"><br><br>'
            + '</div>'
            + (
                '<div class="sonuc-kart sonuc-yeni">📌 Barkod değişti. Üründe önceden yapıştırılmış eski etiket varsa, üzerinde artık eski barkod yazıyor olacak — yeni etiket basmayı unutma.</div>'
                if yeni_barkod != eski_barkod else ''
            )
            + '<a href="/etiket/' + yeni_barkod + '" target="_blank" class="okut-kart okut-mavi">'
            + '<div class="okut-ikon">🖨️</div><div class="okut-metin"><div class="okut-baslik">Etiket Yazdır</div></div><div class="okut-ok">›</div></a>'
            + '<a href="/liste" class="okut-kart okut-mor">'
            + '<div class="okut-ikon">📦</div><div class="okut-metin"><div class="okut-baslik">Stok Listesine Git</div></div><div class="okut-ok">›</div></a>'
        )
        return sayfa(icerik, "Ürün Güncellendi")

    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT ad,cins,ebat,kalinlik,yuzey,sinif,renk,adet,depo,barkod FROM urun WHERE barkod=%s", (eski_barkod,))
            row = cur.fetchone()
    finally:
        con.close()

    if not row:
        return sayfa('<p class="hata">❌ Ürün bulunamadı.</p><a class="btn gri" href="/liste">⬅ Stok Listesine Dön</a>', "Hata")

    ad, cins, ebat, kalinlik, yuzey, sinif, renk, adet, depo, barkod = row

    icerik = """
    <h2 style="margin-bottom:2px;">✏️ Ürün Düzenle</h2>
    <p style="text-align:center;color:var(--muted);margin-top:0;">{{ad}}</p>

    <div class="sonuc-kart sonuc-yeni">
      📌 Ürün gerçek hayatta zaten üstünde barkod ile mi geldi? Aşağıdaki
      <b>Barkod</b> alanına o barkodu yazıp kaydedersen, sistemdeki barkod
      onunla değişir ve bundan sonra taradığında direkt bu ürünü bulur.
    </div>

    <form method="post">
    <div class="kart">
    <label>Barkod</label>
    <input name="barkod" value="{{barkod}}">
    <label>Ürün Adı</label>
    <input name="ad" value="{{ad}}" placeholder="Ürün Adı" required>
    </div>

    <div class="kart">
    <label>Cins</label>
    <input name="cins" value="{{cins or ''}}" placeholder="Cins">
    <label>Ebat</label>
    <input name="ebat" value="{{ebat or ''}}" placeholder="Ebat">
    <label>Kalınlık</label>
    <input name="kalinlik" value="{{kalinlik or ''}}" placeholder="Kalınlık">
    <label>Yüzey</label>
    <select name="yuzey" required>
        <option value="">Seçiniz</option>
        <option value="HG" {{ 'selected' if yuzey=='HG' else '' }}>HG</option>
        <option value="MAT" {{ 'selected' if yuzey=='MAT' else '' }}>MAT</option>
    </select>
    <label>Sınıf</label>
    <input name="sinif" value="{{sinif or ''}}" placeholder="Sınıf">
    <label>Renk</label>
    <input name="renk" value="{{renk or ''}}" placeholder="Renk">
    </div>

    <div class="kart">
    <label>Adet</label>
    <input name="adet" type="number" value="{{adet}}" placeholder="Adet" required>
    <label>Depo</label>
    <select name="depo">
    {% for d in depolar %}
    <option {{ 'selected' if d==depo else '' }}>{{d}}</option>
    {% endfor %}
    </select>
    <button class="btn mavi">Kaydet</button>
    </div>
    </form>

    <p style="text-align:center;margin-top:10px;">
    <a href="/liste" style="color:var(--muted);text-decoration:none;">⬅ Kaydetmeden Geri Dön</a>
    </p>
    """
    return render_template_string(
        sayfa(icerik, "Ürün Düzenle"),
        ad=ad, cins=cins, ebat=ebat, kalinlik=kalinlik, yuzey=yuzey,
        sinif=sinif, renk=renk, adet=adet, depo=depo, barkod=barkod, depolar=DEPOLAR
    )


@app.route("/liste")
@rol_gerekli("depocu","muhasebeci")
def liste():
    depo_filtre = request.args.get("depo", "")

    con = db()
    try:
        with con.cursor() as cur:
            if depo_filtre:
                cur.execute("SELECT * FROM urun WHERE depo=%s ORDER BY ad", (depo_filtre,))
            else:
                cur.execute("SELECT * FROM urun ORDER BY ad")
            urunler = cur.fetchall()
    finally:
        con.close()

    KRITIK_ESIK = 5
    toplam_urun = len(urunler)
    toplam_adet = sum(u[8] for u in urunler)
    kritik_sayisi = sum(1 for u in urunler if u[8] <= KRITIK_ESIK)

    filtre_html = '<div class="filtre-satir">'
    filtre_html += f'<a href="/liste" class="filtre-cip {"aktif" if not depo_filtre else ""}" style="text-decoration:none;display:block;">Tümü</a>'
    for d in DEPOLAR:
        aktif = "aktif" if depo_filtre == d else ""
        filtre_html += f'<a href="/liste?depo={d}" class="filtre-cip {aktif}" style="text-decoration:none;display:block;">{d.split(" ")[0]}</a>'
    filtre_html += '</div>'

    kartlar = ""
    for u in urunler:
        kritik_mi = u[8] <= KRITIK_ESIK
        ozellikler = " • ".join([x for x in [u[2], u[3], u[4], u[5], u[6], u[7]] if x])
        kartlar += f"""
        <div class='urun-kart {"urun-kritik" if kritik_mi else ""}' id="urun-{u[10]}">
        <label class="etiket-sec-satir">
        <input type="checkbox" class="etiket-sec" value="{u[10]}">
        <span>Etikete ekle</span>
        </label>
        <div class="urun-ust">
          <div class="urun-ad">{u[1]}</div>
          <div class="urun-adet-rozet {'kritik' if kritik_mi else ''}">{u[8]} adet</div>
        </div>
        <div class="urun-ozellik">{ozellikler or '-'}</div>
        <div class="urun-depo">🏭 {u[9]}</div>
        <div class="urun-barkod">🔢 {u[10]}</div>
        <div class="urun-gorseller">
          <img src="/barkod/{u[10]}.png">
        </div>
       <div class="urun-aksiyonlar">
        <button class="btn-kucuk mavi" onclick="window.location.href='/duzenle/{u[10]}'">✏️ Düzenle</button>
        <button class="btn-kucuk turkuaz" onclick="window.open('/etiket/{u[10]}','_blank')">🖨️ Etiket</button>
        <button class="btn-kucuk kirmizi" onclick="urunSil('{u[10]}')">🗑️ Sil</button>
        </div>
        </div>
        """
    icerik = (
        "<h2>📦 STOK</h2>"
        + f"""
        <div class="ozet-satir" style="margin-bottom:14px;">
          <div class="ozet-kutu"><div class="ozet-sayi">{toplam_urun}</div><div class="ozet-etiket">Ürün Çeşidi</div></div>
          <div class="ozet-kutu"><div class="ozet-sayi">{toplam_adet}</div><div class="ozet-etiket">Toplam Adet</div></div>
          <div class="ozet-kutu ozet-turuncu"><div class="ozet-sayi">{kritik_sayisi}</div><div class="ozet-etiket">Kritik Stok</div></div>
        </div>
        """
        + '<div style="text-align:right;margin:-6px 0 10px;"><a href="/depo_stok" style="color:var(--accent);text-decoration:none;font-size:12.5px;font-weight:600;">🏭 Depo Bazlı Özet →</a></div>'
        + filtre_html
        + '<input class="arama" id="arama" placeholder="🔍 Ürün, barkod veya depo ara..." oninput="ara()">'
        + '<div style="display:flex;gap:8px;margin-bottom:14px;">'
        + '<button class="btn-kucuk mavi" style="flex:1;text-align:center;" onclick="tumunuSec()">☑️ Tümünü Seç</button>'
        + '<button class="btn-kucuk gri" style="flex:1;text-align:center;" onclick="secimiTemizle()">✖️ Temizle</button>'
        + '</div>'
        + '<div id="liste">' + kartlar + '</div>'
        + """
        <div id="alt-yazdir-bar" style="display:none;position:fixed;left:0;right:0;bottom:0;z-index:9999;
             padding:12px 16px calc(12px + env(safe-area-inset-bottom));
             background:rgba(18,21,28,.92);backdrop-filter:blur(14px);border-top:1px solid rgba(255,255,255,.08);">
          <button class="btn turkuaz" style="margin:0;max-width:520px;margin:0 auto;" onclick="secilenleriYazdir()">
            🖨️ Seçilenleri Tek Sayfada Yazdır (<span id="secili-sayisi">0</span>)
          </button>
        </div>

        <script>
        function ara(){
          var q = document.getElementById('arama').value.toLocaleLowerCase('tr');
          document.querySelectorAll('#liste .urun-kart').forEach(function(k){
            var metin = k.textContent.toLocaleLowerCase('tr');
            k.style.display = metin.indexOf(q) !== -1 ? '' : 'none';
          });
        }
        function urunSil(barkod){
          if(!confirm('Bu ürünü silmek istediğine emin misin? Bu işlem geri alınamaz.')) return;
          fetch('/sil/' + encodeURIComponent(barkod), { method: 'POST' })
            .then(r => r.json())
            .then(d => {
              if(d.ok){
                var el = document.getElementById('urun-' + barkod);
                if(el) el.remove();
              } else {
                alert('Silinemedi, tekrar dene.');
              }
            });
        }
        function secimSayaciGuncelle(){
          var secililer = document.querySelectorAll('.etiket-sec:checked');
          var bar = document.getElementById('alt-yazdir-bar');
          document.getElementById('secili-sayisi').textContent = secililer.length;
          bar.style.display = secililer.length > 0 ? 'block' : 'none';
        }
        document.getElementById('liste').addEventListener('change', function(e){
          if(e.target.classList.contains('etiket-sec')) secimSayaciGuncelle();
        });
        function tumunuSec(){
          document.querySelectorAll('#liste .urun-kart').forEach(function(k){
            if(k.style.display !== 'none'){
              var kutu = k.querySelector('.etiket-sec');
              if(kutu) kutu.checked = true;
            }
          });
          secimSayaciGuncelle();
        }
        function secimiTemizle(){
          document.querySelectorAll('.etiket-sec').forEach(k => k.checked = false);
          secimSayaciGuncelle();
        }
        function secilenleriYazdir(){
          var barkodlar = Array.from(document.querySelectorAll('.etiket-sec:checked')).map(k => k.value);
          if(barkodlar.length === 0) return;
          var form = document.createElement('form');
          form.method = 'POST';
          form.action = '/etiketler';
          form.target = '_blank';
          barkodlar.forEach(function(b){
            var girdi = document.createElement('input');
            girdi.type = 'hidden';
            girdi.name = 'barkod';
            girdi.value = b;
            form.appendChild(girdi);
          });
          document.body.appendChild(form);
          form.submit();
          document.body.removeChild(form);
        }
        </script>
        """
    )
    return sayfa(icerik, "Stok Listesi")


@app.route("/depo_stok")
@rol_gerekli("depocu", "muhasebeci", "patron")
def depo_stok():
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("""
                SELECT depo, COUNT(*) AS cesit, COALESCE(SUM(adet),0) AS toplam
                FROM urun GROUP BY depo
            """)
            satirlar = dict((r[0] or "Depo Belirtilmemiş", (r[1], r[2])) for r in cur.fetchall())
    finally:
        con.close()

    genel_toplam = sum(v[1] for v in satirlar.values())
    genel_cesit = sum(v[0] for v in satirlar.values())

    kartlar = ""
    sirali_depolar = list(DEPOLAR) + [d for d in satirlar if d not in DEPOLAR]
    for d in sirali_depolar:
        cesit, toplam = satirlar.get(d, (0, 0))
        kartlar += f"""
        <a href="/liste?depo={d}" class="depo-ozet-kart" style="text-decoration:none;color:white;">
          <div class="depo-ozet-ad">🏭 {d}</div>
          <div class="depo-ozet-sayilar">
            <div class="depo-ozet-tekli"><div class="sayi">{cesit}</div><div class="etiket">Çeşit</div></div>
            <div class="depo-ozet-tekli"><div class="sayi">{toplam}</div><div class="etiket">Adet</div></div>
          </div>
        </a>
        """

    icerik = (
        "<h2>🏭 DEPO STOK DURUMU</h2>"
        + f"""
        <div class="ozet-satir" style="margin-bottom:6px;">
          <div class="ozet-kutu"><div class="ozet-sayi">{len(sirali_depolar)}</div><div class="ozet-etiket">Depo</div></div>
          <div class="ozet-kutu"><div class="ozet-sayi">{genel_cesit}</div><div class="ozet-etiket">Ürün Çeşidi</div></div>
          <div class="ozet-kutu"><div class="ozet-sayi">{genel_toplam}</div><div class="ozet-etiket">Toplam Adet</div></div>
        </div>
        """
        + '<p style="text-align:center;color:var(--muted);font-size:12.5px;margin-top:0;">Bir depoya dokunarak o depodaki ürünleri görebilirsin.</p>'
        + kartlar
    )
    return sayfa(icerik, "Depo Stok Durumu")

@app.route("/transfer", methods=["GET", "POST"])
@rol_gerekli("depocu", "muhasebeci", "patron")
def transfer():
    if request.method == "POST":
        barkod = request.form.get("barkod", "").strip()
        hedef_depo = request.form.get("hedef_depo", "")
        try:
            miktar = int(request.form.get("miktar", "0"))
        except (TypeError, ValueError):
            miktar = 0

        if not barkod or not hedef_depo or miktar <= 0:
            return sayfa('<p class="hata">❌ Barkod, hedef depo ve miktar girilmeli.</p><a class="btn gri" href="/transfer">⬅ Geri Dön</a>', "Hata")

        con = db()
        try:
            with con:
                with con.cursor() as cur:
                    cur.execute("SELECT id, ad, adet, depo FROM urun WHERE barkod=%s", (barkod,))
                    row = cur.fetchone()
                    if not row:
                        con.close()
                        return sayfa('<p class="hata">❌ Ürün bulunamadı.</p><a class="btn gri" href="/transfer">⬅ Geri Dön</a>', "Hata")

                    uid, ad, mevcut_adet, kaynak_depo = row

                    if kaynak_depo == hedef_depo:
                        con.close()
                        return sayfa('<p class="hata">❌ Kaynak ve hedef depo aynı olamaz.</p><a class="btn gri" href="/transfer">⬅ Geri Dön</a>', "Hata")

                    if miktar > mevcut_adet:
                        con.close()
                        return sayfa(f'<p class="hata">❌ Yetersiz stok. {kaynak_depo} deposunda sadece {mevcut_adet} adet var.</p><a class="btn gri" href="/transfer">⬅ Geri Dön</a>', "Hata")

                    kalan = mevcut_adet - miktar
                    kullanici = session.get("kullanici", "Bilinmiyor")
                    simdi = tr_simdi()

                    if kalan == 0:
                        # Kaynaktaki kayıt tamamen boşaldı, direkt depoyu güncelle
                        cur.execute("UPDATE urun SET depo=%s, adet=%s WHERE id=%s", (hedef_depo, miktar, uid))
                    else:
                        # Kaynakta miktarı düş
                        cur.execute("UPDATE urun SET adet=%s WHERE id=%s", (kalan, uid))
                        # Hedef depoda aynı barkodlu kayıt var mı bak
                        cur.execute("SELECT id, adet FROM urun WHERE barkod=%s AND depo=%s", (barkod, hedef_depo))
                        hedef_satir = cur.fetchone()
                        if hedef_satir:
                            hedef_id, hedef_adet = hedef_satir
                            cur.execute("UPDATE urun SET adet=%s WHERE id=%s", (hedef_adet + miktar, hedef_id))
                        else:
                            cur.execute("""
                                SELECT ad,cins,ebat,kalinlik,yuzey,sinif,renk,barkod FROM urun WHERE id=%s
                            """, (uid,))
                            u = cur.fetchone()
                            yeni_barkod = barkod_uret()
                            cur.execute("""
                                INSERT INTO urun(ad,cins,ebat,kalinlik,yuzey,sinif,renk,adet,depo,barkod)
                                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            """, (u[0], u[1], u[2], u[3], u[4], u[5], u[6], miktar, hedef_depo, yeni_barkod))

                    cur.execute("""
                        INSERT INTO hareket (barkod, ad, tip, adet, kullanici, tarih)
                        VALUES (%s, %s, 'cikis', %s, %s, %s)
                    """, (barkod, ad, miktar, kullanici, simdi))
                    cur.execute("""
                        INSERT INTO hareket (barkod, ad, tip, adet, kullanici, tarih)
                        VALUES (%s, %s, 'giris', %s, %s, %s)
                    """, (barkod, ad, miktar, kullanici, simdi))
        finally:
            con.close()

        icerik = (
            '<div style="text-align:center;font-size:52px;margin-bottom:4px;">✅</div>'
            + '<h2 style="margin-top:0;">Transfer Tamamlandı</h2>'
            + f'<p style="text-align:center;color:var(--muted);"><b style="color:var(--text);">{ad}</b><br>{kaynak_depo} → {hedef_depo}<br>{miktar} adet</p>'
            + '<a href="/depo_stok" class="okut-kart okut-turkuaz"><div class="okut-ikon">🏭</div><div class="okut-metin"><div class="okut-baslik">Depo Stok Durumu</div></div><div class="okut-ok">›</div></a>'
            + '<a href="/transfer" class="okut-kart okut-mavi"><div class="okut-ikon">🔁</div><div class="okut-metin"><div class="okut-baslik">Yeni Transfer</div></div><div class="okut-ok">›</div></a>'
        )
        return sayfa(icerik, "Transfer Tamamlandı")

    icerik = """
    <h2 style="margin-bottom:2px;">🔁 Depolar Arası Transfer</h2>
    <p style="text-align:center;color:var(--muted);margin-top:0;">Bir ürünü bir depodan diğerine taşı</p>

    <div class="kart">
    <label>Ürün Ara (barkod veya isim)</label>
    <div class="urun-arama-kutu">
      <input type="text" id="urun-arama" class="arama" placeholder="🔍 Ürün adı veya barkod yaz..." autocomplete="off" oninput="urunAra()">
      <div class="urun-arama-sonuc" id="urun-arama-sonuc"></div>
    </div>
    <div id="secili-urun-alan"></div>
    </div>

    <form method="post" id="transfer-form" style="display:none;" >
    <input type="hidden" name="barkod" id="form-barkod">
    <div class="kart">
    <label>Hedef Depo</label>
    <select name="hedef_depo" required>
    """ + "".join(f'<option>{d}</option>' for d in DEPOLAR) + """
    </select>
    <label>Miktar</label>
    <input type="number" name="miktar" min="1" required>
    <button type="submit" class="btn turuncu">Transferi Onayla</button>
    </div>
    </form>

    <script>
    let aramaZamanlayici = null;
    function urunAra(){
      clearTimeout(aramaZamanlayici);
      const q = document.getElementById('urun-arama').value.trim();
      const kutu = document.getElementById('urun-arama-sonuc');
      if(q.length < 1){ kutu.style.display='none'; kutu.innerHTML=''; return; }
      aramaZamanlayici = setTimeout(() => {
        fetch('/urun_ara?q=' + encodeURIComponent(q))
          .then(r => r.json())
          .then(sonuclar => {
            if(sonuclar.length === 0){
              kutu.innerHTML = '<div class="urun-arama-oge" style="color:var(--muted);">Ürün bulunamadı</div>';
              kutu.style.display = 'block';
              return;
            }
            kutu.innerHTML = sonuclar.map(u => `
              <div class="urun-arama-oge" onclick='urunSec(${JSON.stringify(JSON.stringify(u))})'>
                <div class="ad">${u.ad}</div>
                <div class="detay">🔢 ${u.barkod} • 📦 Stokta: ${u.adet} • 🏭 ${u.depo}</div>
              </div>
            `).join('');
            kutu.style.display = 'block';
          });
      }, 250);
    }
    function urunSec(uJson){
      const u = JSON.parse(uJson);
      document.getElementById('urun-arama-sonuc').style.display = 'none';
      document.getElementById('urun-arama').value = u.ad;
      document.getElementById('form-barkod').value = u.barkod;
      document.getElementById('secili-urun-alan').innerHTML =
        `<div class="sonuc-kart sonuc-yeni" style="margin-top:8px;">📦 <b>${u.ad}</b><br>🏭 Kaynak Depo: ${u.depo} • Stok: ${u.adet}</div>`;
      document.getElementById('transfer-form').style.display = 'block';
    }
    document.addEventListener('click', function(e){
      const kutu = document.getElementById('urun-arama-sonuc');
      if(!kutu.contains(e.target) && e.target.id !== 'urun-arama'){
        kutu.style.display = 'none';
      }
    });
    </script>
    """
    return sayfa(icerik, "Depolar Arası Transfer")
    
@app.route("/hareketler")
@rol_gerekli("muhasebeci")
def hareket_listesi():
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("""
                SELECT h.ad, h.barkod, h.tip, h.adet, h.kullanici, h.tarih,
                       u.cins, u.sinif, u.yuzey, u.renk, u.ebat, u.depo
                FROM hareket h
                LEFT JOIN urun u ON u.barkod = h.barkod
                ORDER BY h.id DESC
            """)
            kayitlar = cur.fetchall()
    finally:
        con.close()

    kartlar = ""
    for ad, barkod, tip, adet, kullanici, tarih, cins, sinif, yuzey, renk, ebat, depo in kayitlar:
        ozellikler = " • ".join([x for x in [cins, sinif, yuzey, renk, ebat] if x])
        giris_mi = tip == "giris"
        iade_mi = tip == "iade"
        css_sinif = 'hareket-giris' if giris_mi else ('hareket-iade' if iade_mi else 'hareket-cikis')
        ikon = '⬆️' if giris_mi else ('↩️' if iade_mi else '⬇️')
        kartlar += f"""
        <div class="hareket-kart {css_sinif}" data-tip="{tip}">
          <div class="hareket-ikon">{ikon}</div>
          <div class="hareket-govde">
            <div class="hareket-ust">
              <div class="hareket-ad">{ad}</div>
              <div class="hareket-adet {'giris' if giris_mi else ('iade' if iade_mi else 'cikis')}">{'+' if (giris_mi or iade_mi) else '-'}{adet}</div>
            </div>
            <div class="hareket-detay">{ozellikler or '-'}{(' • 🏭 ' + depo) if depo else ''}</div>
            <div class="hareket-alt">🔢 {barkod} &nbsp;•&nbsp; 👤 {kullanici} &nbsp;•&nbsp; 🕒 {tarih}</div>
          </div>
        </div>
        """

    icerik = (
        "<h2>📊 TÜM HAREKETLER</h2>"
        + '<input class="arama" id="arama" placeholder="🔍 Ürün, barkod veya kullanıcı ara..." oninput="ara()">'
        + """
        <div class="filtre-satir">
          <button class="filtre-cip aktif" data-filtre="hepsi" onclick="filtrele('hepsi', this)">Tümü</button>
          <button class="filtre-cip" data-filtre="giris" onclick="filtrele('giris', this)">⬆️ Giriş</button>
          <button class="filtre-cip" data-filtre="cikis" onclick="filtrele('cikis', this)">⬇️ Çıkış</button>
          <button class="filtre-cip" data-filtre="iade" onclick="filtrele('iade', this)">↩️ İade</button>
        </div>
        """
        + '<div id="hareketler">' + kartlar + '</div>'
        + """
        <script>
        let aktifFiltre = 'hepsi';
        function ara(){
          var q = document.getElementById('arama').value.toLocaleLowerCase('tr');
          uygula(q);
        }
        function filtrele(tip, btn){
          aktifFiltre = tip;
          document.querySelectorAll('.filtre-cip').forEach(c => c.classList.remove('aktif'));
          btn.classList.add('aktif');
          uygula(document.getElementById('arama').value.toLocaleLowerCase('tr'));
        }
        function uygula(q){
          document.querySelectorAll('#hareketler .hareket-kart').forEach(function(k){
            var metinUyar = k.textContent.toLocaleLowerCase('tr').indexOf(q) !== -1;
            var tipUyar = aktifFiltre === 'hepsi' || k.dataset.tip === aktifFiltre;
            k.style.display = (metinUyar && tipUyar) ? '' : 'none';
          });
        }
        </script>
        """
    )
    return sayfa(icerik, "Hareketler")
   

@app.route("/rapor/excel")
@rol_gerekli("muhasebeci")
def rapor_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Stok"
    basliklar1 = ["Ürün Adı", "Cins", "Ebat", "Kalınlık", "Yüzey", "Sınıf", "Renk", "Adet", "Depo", "Barkod"]
    ws1.append(basliklar1)
    for c in ws1[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2196F3")

    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT ad,cins,ebat,kalinlik,yuzey,sinif,renk,adet,depo,barkod FROM urun")
            urunler = cur.fetchall()
    finally:
        con.close()
    for satir in urunler:
        ws1.append(list(satir))
    for row_cells in ws1.iter_rows(min_row=2):
        for cell in row_cells:
            cell.font = Font(name="Arial")
    for col_cells in ws1.columns:
        uzunluk = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws1.column_dimensions[col_cells[0].column_letter].width = min(max(uzunluk + 2, 10), 40)

    ws2 = wb.create_sheet("Hareketler")
    basliklar2 = ["Ürün Adı", "Barkod", "İşlem", "Adet", "Kullanıcı", "Tarih"]
    ws2.append(basliklar2)
    for c in ws2[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2196F3")

    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT ad,barkod,tip,adet,kullanici,tarih FROM hareket ORDER BY id DESC")
            hareketler = cur.fetchall()
    finally:
        con.close()
    for satir in hareketler:
        ws2.append(list(satir))
    for row_cells in ws2.iter_rows(min_row=2):
        for cell in row_cells:
            cell.font = Font(name="Arial")
    for col_cells in ws2.columns:
        uzunluk = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws2.column_dimensions[col_cells[0].column_letter].width = min(max(uzunluk + 2, 10), 30)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    dosya_adi = f"stok_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=dosya_adi,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/rapor/xls")
@rol_gerekli("muhasebeci")
def rapor_xls():
    import xlwt

    wb = xlwt.Workbook()
    baslik_stili = xlwt.easyxf(
        "font: bold on, color white; pattern: pattern solid, fore_colour blue;"
    )

    ws1 = wb.add_sheet("Stok")
    basliklar1 = ["Ürün Adı", "Cins", "Ebat", "Kalınlık", "Yüzey", "Sınıf", "Renk", "Adet", "Depo", "Barkod"]
    for col, b in enumerate(basliklar1):
        ws1.write(0, col, b, baslik_stili)

    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT ad,cins,ebat,kalinlik,yuzey,sinif,renk,adet,depo,barkod FROM urun")
            urunler = cur.fetchall()
    finally:
        con.close()
    for row_idx, satir in enumerate(urunler, start=1):
        for col_idx, deger in enumerate(satir):
            ws1.write(row_idx, col_idx, deger)
    for col_idx in range(len(basliklar1)):
        ws1.col(col_idx).width = 256 * 18

    ws2 = wb.add_sheet("Hareketler")
    basliklar2 = ["Ürün Adı", "Barkod", "İşlem", "Adet", "Kullanıcı", "Tarih"]
    for col, b in enumerate(basliklar2):
        ws2.write(0, col, b, baslik_stili)

    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT ad,barkod,tip,adet,kullanici,tarih FROM hareket ORDER BY id DESC")
            hareketler = cur.fetchall()
    finally:
        con.close()
    for row_idx, satir in enumerate(hareketler, start=1):
        for col_idx, deger in enumerate(satir):
            ws2.write(row_idx, col_idx, str(deger) if deger is not None else "")
    for col_idx in range(len(basliklar2)):
        ws2.col(col_idx).width = 256 * 18

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    dosya_adi = f"stok_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.xls"
    return send_file(
        bio,
        as_attachment=True,
        download_name=dosya_adi,
        mimetype="application/vnd.ms-excel",
    )


@app.route("/rapor/csv")
@rol_gerekli("muhasebeci")
def rapor_csv():
    import csv

    si = io.StringIO()
    writer = csv.writer(si, delimiter=";")
    writer.writerow(["Ürün Adı", "Cins", "Ebat", "Kalınlık", "Yüzey", "Sınıf", "Renk", "Adet", "Depo", "Barkod"])

    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT ad,cins,ebat,kalinlik,yuzey,sinif,renk,adet,depo,barkod FROM urun")
            urunler = cur.fetchall()
    finally:
        con.close()

    # Barkod hücresini ="..." formülü olarak yazıyoruz; aksi halde Excel
    # 12 haneli barkodu sayı sanıp bilimsel gösterime çeviriyor
    # (örn. 1,23457E+11) ve barkod okunmaz hale geliyor.
    for satir in urunler:
        satir = list(satir)
        if satir[9]:
            satir[9] = f'="{satir[9]}"'
        writer.writerow(satir)

    output = io.BytesIO(si.getvalue().encode("utf-8-sig"))
    output.seek(0)

    dosya_adi = f"stok_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return send_file(
        output,
        as_attachment=True,
        download_name=dosya_adi,
        mimetype="text/csv",
    )


@app.route("/hizli_islem", methods=["POST"])
@rol_gerekli("depocu")
def hizli_islem():
    data = request.get_json()
    barkod = data.get("barkod")
    tip = data.get("tip")
    siparis_id = data.get("siparis_id")
    kullanici = session.get("kullanici", "Bilinmiyor")
    rol = session.get("rol")

    con = db()
    try:
        with con:
            with con.cursor() as cur:
                cur.execute("SELECT id, ad, adet, cins, ebat, yuzey, sinif, renk, depo FROM urun WHERE barkod=%s", (barkod,))
                row = cur.fetchone()

                if not row:
                    if tip == "giris":
                        return jsonify({"ok": False, "yeni": True, "barkod": barkod})
                    return jsonify({"ok": False, "msg": "Bu ürün sizin sattığınız ürünler arasında değil, çıkış yapılamaz"})

                uid, ad, adet, cins, ebat, yuzey, sinif, renk, depo = row

                # ---- Sipariş kontrolü: depocu, açık siparişte olmayan ürünü çıkış yapamaz ----
                kalem_id = None
                if tip == "cikis" and rol == "depocu":
                    if not siparis_id:
                        return jsonify({"ok": False, "msg": "Önce bir sipariş seçmelisin. Sipariş olmadan çıkış yapılamaz."})
                    cur.execute("SELECT durum FROM siparis WHERE id=%s", (siparis_id,))
                    sip = cur.fetchone()
                    if not sip or sip[0] != 'acik':
                        return jsonify({"ok": False, "msg": "Bu sipariş artık açık değil."})
                    cur.execute("SELECT id, istenen, verilen FROM siparis_kalem WHERE siparis_id=%s AND barkod=%s", (siparis_id, barkod))
                    kalem = cur.fetchone()
                    if not kalem:
                        return jsonify({"ok": False, "msg": "⚠️ YANLIŞ ÜRÜN! Bu ürün bu siparişte yok.", "yanlis_urun": True, "urun_adi": ad})
                    kalem_id, istenen, verilen = kalem
                    if verilen >= istenen:
                        return jsonify({"ok": False, "msg": f"'{ad}' için istenen {istenen} adet zaten tamamlandı.", "tamamlandi_uyari": True})

                if tip == "cikis" and adet <= 0:
                    return jsonify({"ok": False, "msg": "Stok yok"})

                if tip == "giris":
                    adet += 1
                else:
                    adet -= 1
                if adet < 0:
                    adet = 0

                cur.execute("UPDATE urun SET adet=%s WHERE id=%s", (adet, uid))
                if tip == "cikis":
                    kritik_stok_kontrol(barkod, ad, adet, depo)
                cur.execute("""
                INSERT INTO hareket (barkod, ad, tip, adet, kullanici, tarih)
                VALUES (%s, %s, %s, %s, %s, %s)
                """, (barkod, ad, tip, 1, kullanici, tr_simdi()))

                siparis_ilerleme = None
                if kalem_id:
                    cur.execute("UPDATE siparis_kalem SET verilen = verilen + 1 WHERE id=%s", (kalem_id,))
                    cur.execute("""
                        INSERT INTO siparis_hareket (siparis_id, barkod, ad, adet, kullanici, tarih)
                        VALUES (%s,%s,%s,1,%s,%s)
                    """, (siparis_id, barkod, ad, kullanici, tr_simdi()))
                    cur.execute("SELECT SUM(istenen), SUM(verilen) FROM siparis_kalem WHERE siparis_id=%s", (siparis_id,))
                    top_ist, top_ver = cur.fetchone()
                    siparis_ilerleme = {"istenen": top_ist, "verilen": top_ver}
                    if top_ver >= top_ist:
                        cur.execute("UPDATE siparis SET durum='tamamlandi' WHERE id=%s", (siparis_id,))
                        siparis_ilerleme["tamamlandi"] = True

                cur.execute("""
                SELECT SUM(adet) FROM hareket WHERE kullanici=%s AND tip='cikis'
                """, (kullanici,))
                toplam = cur.fetchone()[0]
                if not toplam:
                    toplam = 0

                return jsonify({
                    "ok": True, "ad": ad, "adet": adet, "toplam": toplam,
                    "cins": cins, "ebat": ebat, "yuzey": yuzey, "sinif": sinif,
                    "renk": renk, "depo": depo, "siparis_ilerleme": siparis_ilerleme,
                })
    finally:
        con.close()


@app.route("/sil/<barkod>", methods=["POST"])
@rol_gerekli("muhasebeci")
def urun_sil(barkod):
    con = db()
    try:
        with con:
            with con.cursor() as cur:
                cur.execute("DELETE FROM urun WHERE barkod=%s", (barkod,))
                silindi = cur.rowcount > 0
    finally:
        con.close()
    return jsonify({"ok": silindi})


@app.route("/etiket/<barkod>")
@rol_gerekli("muhasebeci")
def etiket(barkod):
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT ad,cins,ebat,kalinlik,yuzey,sinif,renk,depo FROM urun WHERE barkod=%s", (barkod,))
            row = cur.fetchone()
    finally:
        con.close()

    if not row:
        return sayfa('<p class="hata">❌ Ürün bulunamadı.</p><a class="btn gri" href="/liste">⬅ Geri Dön</a>', "Hata")

    ad, cins, ebat, kalinlik, yuzey, sinif, renk, depo = row

    ozellikler = " • ".join([x for x in [cins, ebat, kalinlik, yuzey, sinif, renk] if x])

    html = f"""
    <!DOCTYPE html>
    <html lang="tr">
    <head>
    <meta charset="utf-8">
    <title>Etiket - {ad}</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        @page {{ size: 58mm 40mm; margin: 0mm; }}
        * {{ box-sizing: border-box; }}
        body {{
            font-family: Arial, sans-serif; margin:0; padding:0;
            background:#fff; color:#000;
        }}
        .etiket {{
            width: 36mm; padding: 1.5mm; text-align:center; margin:0 auto;
        }}
        .ad {{ font-size: 11px; font-weight:800; margin-bottom:1mm; line-height:1.15; }}
        .ozellik {{ font-size: 7.5px; color:#333; margin-bottom:1mm; line-height:1.25; }}
        .depo {{ font-size: 7.5px; color:#555; margin-bottom:1mm; }}
        img.barkod {{ width: 100%; max-width: 34mm; }}
        .yazdir-btn {{
            display:block; margin: 6mm auto 0; padding:14px 24px;
            background:#2196F3; color:white; border:none; border-radius:10px;
            font-size:16px; font-weight:700; cursor:pointer;
        }}
        @media print {{
            .yazdir-btn {{ display:none; }}
            body {{ margin:0; }}
        }}
        
    </style>
    </head>
    <body>
        <div class="etiket">
            <div class="ad">{ad}</div>
            <div class="ozellik">{ozellikler}</div>
            {'<div class="depo">🏭 ' + depo + '</div>' if depo else ''}
            <img class="barkod" src="/barkod/{barkod}.png">
        </div>
        <button class="yazdir-btn" onclick="window.print()">🖨️ Yazdır</button>
        <div class="label" style="border:1px solid red;"></div>
    </body>
    </html>
    """
    return html


@app.route("/etiketler", methods=["GET", "POST"])
@rol_gerekli("muhasebeci")
def etiketler():
    if request.method == "POST":
        barkodlar = [b.strip() for b in request.form.getlist("barkod") if b.strip()]
    else:
        barkod_param = request.args.get("barkodlar", "")
        barkodlar = [b.strip() for b in barkod_param.split(",") if b.strip()]

    if not barkodlar:
        return sayfa('<p class="hata">❌ Etiket için ürün seçilmedi.</p><a class="btn gri" href="/liste">⬅ Stok Listesine Dön</a>', "Hata")

    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("""
                SELECT ad,cins,ebat,kalinlik,yuzey,sinif,renk,depo,barkod
                FROM urun WHERE barkod = ANY(%s)
            """, (barkodlar,))
            satirlar = cur.fetchall()
    finally:
        con.close()

    sirali = {s[8]: s for s in satirlar}
    satirlar = [sirali[b] for b in barkodlar if b in sirali]

    if not satirlar:
        return sayfa('<p class="hata">❌ Seçilen ürünler bulunamadı.</p><a class="btn gri" href="/liste">⬅ Stok Listesine Dön</a>', "Hata")

    etiket_html = ""
    for ad, cins, ebat, kalinlik, yuzey, sinif, renk, depo, barkod in satirlar:
        ozellikler = " • ".join([x for x in [cins, ebat, kalinlik, yuzey, sinif, renk] if x])
        etiket_html += f"""
        <div class="etiket">
            <div class="ad">{ad}</div>
            <div class="ozellik">{ozellikler}</div>
            {'<div class="depo">🏭 ' + depo + '</div>' if depo else ''}
            <img class="barkod" src="/barkod/{barkod}.png">
        </div>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="tr">
    <head>
    <meta charset="utf-8">
    <title>Etiketler ({len(satirlar)} adet)</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        @page {{ size: 58mm 40mm; margin: 0mm; }}
        * {{ box-sizing: border-box; }}
        body {{
            font-family: Arial, sans-serif; margin:0; padding:0;
            background:#e9e9e9; color:#000;
        }}
        .sayfa-ic {{
            max-width: 500px; margin: 0 auto; padding: 16px;
        }}
        .ust-arac {{
            display:flex; justify-content:center; gap:10px; margin-bottom:16px;
        }}
        .yazdir-btn, .geri-btn {{
            padding:14px 24px; border:none; border-radius:10px;
            font-size:16px; font-weight:700; cursor:pointer; text-decoration:none;
            display:inline-block;
        }}
        .yazdir-btn {{ background:#2196F3; color:white; }}
        .geri-btn {{ background:#555; color:white; }}
        .izgara {{
            display: flex; flex-direction: column; gap: 4mm;
        }}
        .etiket {{
            width: 36mm; margin: 0 auto;
            border: 1px dashed #bbb;
            padding: 1.5mm; text-align:center;
            background: white;
            page-break-after: always;
            break-after: page;
        }}
        .etiket:last-child {{ page-break-after: auto; break-after: auto; }}
        .ad {{ font-size: 11px; font-weight:800; margin-bottom:1mm; line-height:1.15; }}
        .ozellik {{ font-size: 7.5px; color:#333; margin-bottom:1mm; line-height:1.25; }}
        .depo {{ font-size: 7.5px; color:#555; margin-bottom:1mm; }}
        img.barkod {{ width: 100%; max-width: 34mm; }}
        @media print {{
            .ust-arac {{ display:none; }}
            body {{ background:white; margin:0; }}
            .sayfa-ic {{ max-width:none; padding:0; }}
            .etiket {{ border: none; }}
        }}
    </style>
    </head>
    <body>
        <div class="sayfa-ic">
            <div class="ust-arac">
                <button class="yazdir-btn" onclick="window.print()">🖨️ Sayfayı Yazdır ({len(satirlar)} etiket)</button>
                <a class="geri-btn" href="/liste">⬅ Listeye Dön</a>
            </div>
            <div class="izgara">
                {etiket_html}
            </div>
        </div>
    </body>
    </html>
    """
    return html


@app.route("/geri_al", methods=["POST"])
@rol_gerekli("depocu")
def geri_al():
    data = request.get_json()
    barkod = data.get("barkod")
    tip = data.get("tip")
    siparis_id = data.get("siparis_id")

    con = db()
    try:
        with con:
            with con.cursor() as cur:
                cur.execute("SELECT id, adet FROM urun WHERE barkod=%s", (barkod,))
                row = cur.fetchone()
                if not row:
                    return jsonify({"ok": False})

                uid, adet = row
                if tip == "giris":
                    adet -= 1
                else:
                    adet += 1
                if adet < 0:
                    adet = 0

                cur.execute("UPDATE urun SET adet=%s WHERE id=%s", (adet, uid))

                if tip == "cikis" and siparis_id:
                    cur.execute("""
                        UPDATE siparis_kalem SET verilen = GREATEST(verilen - 1, 0)
                        WHERE siparis_id=%s AND barkod=%s
                    """, (siparis_id, barkod))
                    cur.execute("UPDATE siparis SET durum='acik' WHERE id=%s AND durum='tamamlandi'", (siparis_id,))
    finally:
        con.close()

    return jsonify({"ok": True})


# ============ SİPARİŞ SİSTEMİ ============

def siparis_tablolarini_olustur():
    con = db()
    try:
        with con:
            with con.cursor() as cur:
                cur.execute("""
                CREATE TABLE IF NOT EXISTS siparis(
                    id SERIAL PRIMARY KEY,
                    musteri TEXT,
                    depo TEXT,
                    durum TEXT DEFAULT 'acik',
                    olusturan TEXT,
                    tarih TIMESTAMP,
                    aciklama TEXT
                )
                """)
                cur.execute("""
                CREATE TABLE IF NOT EXISTS siparis_kalem(
                    id SERIAL PRIMARY KEY,
                    siparis_id INTEGER REFERENCES siparis(id) ON DELETE CASCADE,
                    barkod TEXT,
                    ad TEXT,
                    istenen INTEGER,
                    verilen INTEGER DEFAULT 0
                )
                """)
    finally:
        con.close()


if DATABASE_URL:
    siparis_tablolarini_olustur()
def siparis_hareket_tablosu_olustur():
    con = db()
    try:
        with con:
            with con.cursor() as cur:
                cur.execute("""
                CREATE TABLE IF NOT EXISTS siparis_hareket(
                    id SERIAL PRIMARY KEY,
                    siparis_id INTEGER REFERENCES siparis(id) ON DELETE CASCADE,
                    barkod TEXT,
                    ad TEXT,
                    adet INTEGER,
                    kullanici TEXT,
                    tarih TIMESTAMP
                )
                """)
    finally:
        con.close()


if DATABASE_URL:
    siparis_hareket_tablosu_olustur()

def push_tablosu_olustur():
    con = db()
    try:
        with con:
            with con.cursor() as cur:
                cur.execute("""
                CREATE TABLE IF NOT EXISTS push_abone(
                    id SERIAL PRIMARY KEY,
                    kullanici TEXT,
                    endpoint TEXT UNIQUE,
                    p256dh TEXT,
                    auth TEXT,
                    olusturulma TIMESTAMP
                )
                """)
    finally:
        con.close()


if DATABASE_URL:
    push_tablosu_olustur()

def kritik_stok_kontrol(barkod, ad, yeni_adet, depo):
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT min_stok FROM urun WHERE barkod=%s", (barkod,))
            row = cur.fetchone()
    finally:
        con.close()
    esik = row[0] if row and row[0] is not None else 5
    if yeni_adet <= esik:
        push_bildirim_gonder(
            "⚠️ Kritik Stok",
            f"{ad} — {depo}: sadece {yeni_adet} adet kaldı (eşik: {esik}).",
            url="/liste"
        )    


def push_bildirim_gonder(baslik, govde, url="/"):
    """Depocu rolündeki, bildirim izni vermiş tüm cihazlara push bildirimi yollar."""
    if not PUSH_AKTIF or not DATABASE_URL:
        return
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT id, endpoint, p256dh, auth FROM push_abone")
            aboneler = cur.fetchall()
    finally:
        con.close()

    silinecekler = []
    for abone_id, endpoint, p256dh, auth in aboneler:
        try:
            webpush(
                subscription_info={
                    "endpoint": endpoint,
                    "keys": {"p256dh": p256dh, "auth": auth},
                },
                data=json.dumps({"title": baslik, "body": govde, "url": url}),
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims={"sub": VAPID_CLAIMS_EMAIL},
            )
        except WebPushException:
            silinecekler.append(abone_id)
        except Exception:
            pass

    if silinecekler:
        con = db()
        try:
            with con:
                with con.cursor() as cur:
                    cur.execute("DELETE FROM push_abone WHERE id = ANY(%s)", (silinecekler,))
        finally:
            con.close()


@app.route("/vapid_public_key")
def vapid_public_key():
    return jsonify({"publicKey": VAPID_PUBLIC_KEY, "aktif": PUSH_AKTIF})


@app.route("/push_abone_ol", methods=["POST"])
@rol_gerekli("depocu")
def push_abone_ol():
    data = request.get_json(silent=True) or {}
    endpoint = data.get("endpoint")
    keys = data.get("keys", {})
    if not endpoint or not keys.get("p256dh") or not keys.get("auth"):
        return jsonify({"ok": False}), 400
    con = db()
    try:
        with con:
            with con.cursor() as cur:
                cur.execute("""
                    INSERT INTO push_abone (kullanici, endpoint, p256dh, auth, olusturulma)
                    VALUES (%s,%s,%s,%s,%s)
                    ON CONFLICT (endpoint) DO UPDATE
                    SET kullanici=EXCLUDED.kullanici, p256dh=EXCLUDED.p256dh, auth=EXCLUDED.auth, olusturulma=EXCLUDED.olusturulma
                """, (session.get("kullanici", "Bilinmiyor"), endpoint, keys.get("p256dh"), keys.get("auth"), tr_simdi()))
    finally:
        con.close()
    return jsonify({"ok": True})


@app.route("/acik_siparis_sayisi")
@rol_gerekli("depocu")
def acik_siparis_sayisi():
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM siparis WHERE durum='acik'")
            sayi = cur.fetchone()[0]
    finally:
        con.close()
    return jsonify({"acik": sayi})


@app.route("/service-worker.js")
def service_worker():
    js = """
self.addEventListener('push', function(event) {
  let data = {};
  try { data = event.data.json(); } catch(e) { data = { title: 'HER-İŞ Stok Takip', body: event.data ? event.data.text() : '' }; }
  const baslik = data.title || 'HER-İŞ Stok Takip';
  const secenekler = {
    body: data.body || '',
    icon: '/static/icon-192.png',
    badge: '/static/icon-192.png',
    vibrate: [120, 60, 120],
    data: { url: data.url || '/' }
  };
  event.waitUntil(self.registration.showNotification(baslik, secenekler));
});

self.addEventListener('notificationclick', function(event) {
  event.notification.close();
  const url = (event.notification.data && event.notification.data.url) || '/';
  event.waitUntil(clients.openWindow(url));
});
"""
    return Response(js, mimetype="application/javascript")


@app.route("/urun_ara")
@rol_gerekli("depocu","muhasebeci")
def urun_ara():
    q = request.args.get("q", "").strip()
    if len(q) < 1:
        return jsonify([])
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("""
                SELECT ad, barkod, adet, depo, cins, ebat FROM urun
                WHERE ad ILIKE %s OR barkod ILIKE %s
                ORDER BY ad LIMIT 20
            """, (f"%{q}%", f"%{q}%"))
            satirlar = cur.fetchall()
    finally:
        con.close()
    return jsonify([
        {"ad": s[0], "barkod": s[1], "adet": s[2], "depo": s[3] or "", "cins": s[4] or "", "ebat": s[5] or ""}
        for s in satirlar
    ])


@app.route("/siparis_olustur", methods=["GET", "POST"])
@rol_gerekli("muhasebeci")
def siparis_olustur():
    if request.method == "POST":
        musteri = request.form.get("musteri", "").strip()
        depo = request.form.get("depo", "")
        aciklama = request.form.get("aciklama", "").strip()
        barkodlar = request.form.getlist("kalem_barkod")
        adlar = request.form.getlist("kalem_ad")
        adetler = request.form.getlist("kalem_adet")

        kalemler = []
        for b, a, adet_str in zip(barkodlar, adlar, adetler):
            b = b.strip()
            if not b:
                continue
            try:
                adet = int(adet_str)
            except (TypeError, ValueError):
                adet = 0
            if adet <= 0:
                continue
            kalemler.append((b, a, adet))

        if not kalemler:
            return sayfa('<p class="hata">❌ En az bir ürün ve adet girmelisin.</p><a class="btn gri" href="/siparis_olustur">⬅ Geri Dön</a>', "Hata")

        con = db()
        try:
            with con:
                with con.cursor() as cur:
                    cur.execute("""
                        INSERT INTO siparis (musteri, depo, durum, olusturan, tarih, aciklama)
                        VALUES (%s,%s,'acik',%s,%s,%s) RETURNING id
                    """, (musteri, depo, session.get("kullanici", "Bilinmiyor"), tr_simdi(), aciklama))
                    siparis_id = cur.fetchone()[0]
                    for barkod, ad, adet in kalemler:
                        cur.execute("""
                            INSERT INTO siparis_kalem (siparis_id, barkod, ad, istenen, verilen)
                            VALUES (%s,%s,%s,%s,0)
                        """, (siparis_id, barkod, ad, adet))
        finally:
            con.close()

        toplam_kalem_adet = sum(adet for _, _, adet in kalemler)
        push_bildirim_gonder(
            "🧾 Yeni Sipariş" + (f" — {musteri}" if musteri else ""),
            f"{depo} için {len(kalemler)} çeşit ürün, {toplam_kalem_adet} adet hazırlanacak.",
            url=f"/siparis/{siparis_id}",
        )

        return redirect(f"/siparis/{siparis_id}")

    icerik = """
    <h2 style="margin-bottom:2px;">🧾 Yeni Sipariş Oluştur</h2>
    <p style="text-align:center;color:var(--muted);margin-top:0;">
    Depocuya hazırlatılacak ürünleri seç, depocu sadece bu listedeki ürünleri çıkış yapabilir.
    </p>

    <form method="post" id="siparis-form">
    <div class="kart">
    <label>Müşteri / Sipariş Adı (opsiyonel)</label>
    <input name="musteri" placeholder="Örn: Ahmet Bey - Mutfak Dolabı">
    <label>Depo</label>
    <select name="depo" required>
    """ + "".join(f'<option>{d}</option>' for d in DEPOLAR) + """
    </select>
    <label>Not (opsiyonel)</label>
    <input name="aciklama" placeholder="Örn: Bugün teslim edilecek">
    </div>

    <div class="kart">
    <label>Ürün Ara ve Ekle</label>
    <div class="urun-arama-kutu">
      <input type="text" id="urun-arama" class="arama" placeholder="🔍 Ürün adı veya barkod yaz..." autocomplete="off" oninput="urunAra()">
      <div class="urun-arama-sonuc" id="urun-arama-sonuc"></div>
    </div>
    <div id="sepet"></div>
    <p id="sepet-bos" style="color:var(--muted);font-size:13px;text-align:center;margin:10px 0;">Henüz ürün eklenmedi.</p>
    </div>

    <div id="gizli-alanlar"></div>
    <button type="submit" class="btn yesil">✅ Siparişi Oluştur</button>
    </form>

    <script>
    let sepet = {};
    let aramaZamanlayici = null;

    function urunAra(){
      clearTimeout(aramaZamanlayici);
      const q = document.getElementById('urun-arama').value.trim();
      const kutu = document.getElementById('urun-arama-sonuc');
      if(q.length < 1){ kutu.style.display='none'; kutu.innerHTML=''; return; }
      aramaZamanlayici = setTimeout(() => {
        fetch('/urun_ara?q=' + encodeURIComponent(q))
          .then(r => r.json())
          .then(sonuclar => {
            if(sonuclar.length === 0){
              kutu.innerHTML = '<div class="urun-arama-oge" style="color:var(--muted);">Ürün bulunamadı</div>';
              kutu.style.display = 'block';
              return;
            }
            kutu.innerHTML = sonuclar.map(u => `
              <div class="urun-arama-oge" onclick='sepeteEkle(${JSON.stringify(JSON.stringify(u))})'>
                <div class="ad">${u.ad}</div>
                <div class="detay">🔢 ${u.barkod} • 📦 Stokta: ${u.adet} • 🏭 ${u.depo}</div>
              </div>
            `).join('');
            kutu.style.display = 'block';
          });
      }, 250);
    }

    function sepeteEkle(uJson){
      const u = JSON.parse(uJson);
      if(!sepet[u.barkod]){
        sepet[u.barkod] = { ad: u.ad, adet: 1, stok: u.adet };
      } else {
        sepet[u.barkod].adet += 1;
      }
      document.getElementById('urun-arama').value = '';
      document.getElementById('urun-arama-sonuc').style.display = 'none';
      sepetiCiz();
    }

    function sepettenSil(barkod){
      delete sepet[barkod];
      sepetiCiz();
    }

    function adetGuncelle(barkod, deger){
      const adet = parseInt(deger);
      if(sepet[barkod]) sepet[barkod].adet = isNaN(adet) ? 1 : Math.max(1, adet);
    }

    function sepetiCiz(){
      const kapsayici = document.getElementById('sepet');
      const bosMesaj = document.getElementById('sepet-bos');
      const barkodlar = Object.keys(sepet);
      bosMesaj.style.display = barkodlar.length === 0 ? 'block' : 'none';
      kapsayici.innerHTML = barkodlar.map(b => `
        <div class="sepet-satir">
          <div class="sepet-ad">${sepet[b].ad}<br><span style="color:var(--muted);font-size:11px;">🔢 ${b} • Stokta ${sepet[b].stok}</span></div>
          <input type="number" class="sepet-adet-input" min="1" value="${sepet[b].adet}" onchange="adetGuncelle('${b}', this.value)">
          <button type="button" class="sepet-sil" onclick="sepettenSil('${b}')">✕</button>
        </div>
      `).join('');

      const gizli = document.getElementById('gizli-alanlar');
      gizli.innerHTML = barkodlar.map(b => `
        <input type="hidden" name="kalem_barkod" value="${b}">
        <input type="hidden" name="kalem_ad" value="${sepet[b].ad}">
        <input type="hidden" name="kalem_adet" value="${sepet[b].adet}">
      `).join('');
    }

    document.getElementById('siparis-form').addEventListener('submit', function(e){
      if(Object.keys(sepet).length === 0){
        e.preventDefault();
        alert('Lütfen en az bir ürün ekle.');
      }
    });

    document.addEventListener('click', function(e){
      const kutu = document.getElementById('urun-arama-sonuc');
      if(!kutu.contains(e.target) && e.target.id !== 'urun-arama'){
        kutu.style.display = 'none';
      }
    });
    </script>
    """
    return sayfa(icerik, "Yeni Sipariş")


@app.route("/siparisler")
@rol_gerekli("muhasebeci")
def siparisler():
    durum_filtre = request.args.get("durum", "acik")

    con = db()
    try:
        with con.cursor() as cur:
            if durum_filtre == "hepsi":
                cur.execute("""
                    SELECT s.id, s.musteri, s.depo, s.durum, s.olusturan, s.tarih,
                           COALESCE(SUM(k.istenen),0), COALESCE(SUM(k.verilen),0)
                    FROM siparis s LEFT JOIN siparis_kalem k ON k.siparis_id = s.id
                    GROUP BY s.id ORDER BY s.id DESC
                """)
            else:
                cur.execute("""
                    SELECT s.id, s.musteri, s.depo, s.durum, s.olusturan, s.tarih,
                           COALESCE(SUM(k.istenen),0), COALESCE(SUM(k.verilen),0)
                    FROM siparis s LEFT JOIN siparis_kalem k ON k.siparis_id = s.id
                    WHERE s.durum=%s GROUP BY s.id ORDER BY s.id DESC
                """, (durum_filtre,))
            satirlar = cur.fetchall()
    finally:
        con.close()

    kartlar = ""
    for sid, musteri, depo, durum, olusturan, tarih, istenen, verilen in satirlar:
        yuzde = int((verilen / istenen) * 100) if istenen else 0
        kartlar += f"""
        <a href="/siparis/{sid}" class="siparis-kart">
          <div class="siparis-ust">
            <div class="siparis-no">🧾 Sipariş #{sid}{' — ' + musteri if musteri else ''}</div>
            <div class="siparis-durum {durum}">{durum}</div>
          </div>
          <div class="siparis-detay">
            🏭 {depo} &nbsp;•&nbsp; 👤 {olusturan} &nbsp;•&nbsp; 🕒 {tarih}<br>
            📦 {verilen} / {istenen} adet hazırlandı
          </div>
          <div class="siparis-ilerleme-bar"><div class="siparis-ilerleme-dolu" style="width:{yuzde}%;"></div></div>
        </a>
        """

    if not kartlar:
        kartlar = '<p style="text-align:center;color:var(--muted);margin-top:24px;">Bu filtrede sipariş bulunamadı.</p>'

    icerik = (
        "<h2>📋 SİPARİŞLER</h2>"
        + f"""
        <div class="filtre-satir">
          <a href="/siparisler?durum=acik" class="filtre-cip {'aktif' if durum_filtre=='acik' else ''}" style="text-decoration:none;display:block;">Açık</a>
          <a href="/siparisler?durum=tamamlandi" class="filtre-cip {'aktif' if durum_filtre=='tamamlandi' else ''}" style="text-decoration:none;display:block;">Tamamlandı</a>
          <a href="/siparisler?durum=hepsi" class="filtre-cip {'aktif' if durum_filtre=='hepsi' else ''}" style="text-decoration:none;display:block;">Tümü</a>
        </div>
        """
        + '<a href="/siparis_olustur" class="okut-kart okut-yesil"><div class="okut-ikon">➕</div><div class="okut-metin"><div class="okut-baslik">Yeni Sipariş</div></div><div class="okut-ok">›</div></a>'
        + kartlar
    )
    return sayfa(icerik, "Siparişler")


@app.route("/siparis/<int:siparis_id>")
@rol_gerekli("depocu", "muhasebeci", "patron")
def siparis_detay(siparis_id):
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT id, musteri, depo, durum, olusturan, tarih, aciklama FROM siparis WHERE id=%s", (siparis_id,))
            sip = cur.fetchone()
            if not sip:
                return sayfa('<p class="hata">❌ Sipariş bulunamadı.</p><a class="btn gri" href="/siparisler">⬅ Geri Dön</a>', "Hata")
            cur.execute("""
                SELECT k.id, k.ad, k.barkod, k.istenen, k.verilen,
                       u.cins, u.ebat, u.kalinlik, u.yuzey, u.sinif, u.renk
                FROM siparis_kalem k
                LEFT JOIN urun u ON u.barkod = k.barkod
                WHERE k.siparis_id=%s ORDER BY k.id
            """, (siparis_id,))
            kalemler = cur.fetchall()

            cur.execute("""
                SELECT barkod, kullanici, SUM(adet)
                FROM siparis_hareket
                WHERE siparis_id=%s
                GROUP BY barkod, kullanici
            """, (siparis_id,))
            kullanici_dagilimi = {}
            for barkod, kullanici, toplam in cur.fetchall():
                kullanici_dagilimi.setdefault(barkod, []).append((kullanici, toplam))
    finally:
        con.close()

    _, musteri, depo, durum, olusturan, tarih, aciklama = sip

    kalem_html = ""
    for kid, ad, barkod, istenen, verilen, cins, ebat, kalinlik, yuzey, sinif, renk in kalemler:
        tam = verilen >= istenen
        ozellikler = " • ".join([x for x in [cins, ebat, kalinlik, yuzey, sinif, renk] if x])
        kimler_html = ""
        for kullanici_ad, toplam_adet in kullanici_dagilimi.get(barkod, []):
            kimler_html += f'<span class="rozet" style="margin-right:6px;margin-top:4px;display:inline-block;">👤 {kullanici_ad}: {toplam_adet}</span>'
        iade_form = ""
        if verilen > 0:
            iade_form = f"""
            <form method="post" action="/kalem_iade/{kid}" style="display:flex;gap:6px;margin-top:6px;"
                  onsubmit="return confirm('{verilen} adet verilmişti, iade alınsın mı?');">
              <input type="number" name="miktar" min="1" max="{verilen}" value="1"
                     style="margin:0;padding:8px;flex:1;">
              <button type="submit" class="btn-kucuk kirmizi" style="margin:0;">↩️ İade Al</button>
            </form>
            """
        kalem_html += f"""
        <div class="kalem-satir" style="flex-direction:column;align-items:stretch;">
          <div style="display:flex;justify-content:space-between;">
            <div class="kalem-ad">{ad}<br><span style="color:var(--muted);font-size:11.5px;">🔢 {barkod}{(' • ' + ozellikler) if ozellikler else ''}</span></div>
            <div class="kalem-miktar {'tam' if tam else 'eksik'}">{verilen} / {istenen}</div>
          </div>
          {'<div style="margin-top:6px;">' + kimler_html + '</div>' if kimler_html else ''}
          {iade_form}
        </div>
        """

    rol = session.get("rol")
    aksiyon_html = f'<a href="/siparis_fis/{siparis_id}" target="_blank" class="btn turkuaz">🧾 Fiş Olarak Yazdır</a>'
    if rol == "depocu" and durum == "acik":
        aksiyon_html += f'<a href="/kamera/cikis?siparis_id={siparis_id}" class="btn turuncu">⬇️ Bu Siparişi Okutarak Hazırla</a>'
    if rol in ("muhasebeci", "patron") and durum == "acik":
        aksiyon_html += f"""
        <form method="post" action="/siparis_iptal/{siparis_id}" onsubmit="return confirm('Bu siparişi iptal etmek istediğine emin misin?');">
          <button class="btn kirmizi" type="submit">🗑️ Siparişi İptal Et</button>
        </form>
        """

    icerik = (
        f"<h2>🧾 Sipariş #{siparis_id}</h2>"
        + f'<h3 class="alt">{musteri or "—"}</h3>'
        + f"""
        <div class="kart">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;">
            <span class="siparis-durum {durum}">{durum}</span>
            <span style="color:var(--muted);font-size:12.5px;">🏭 {depo}</span>
          </div>
          <div style="color:var(--muted);font-size:12.5px;">👤 {olusturan} &nbsp;•&nbsp; 🕒 {tarih}</div>
          {'<div style="color:var(--muted);font-size:13px;margin-top:8px;">📝 ' + aciklama + '</div>' if aciklama else ''}
        </div>
        <div class="kart" style="padding:6px 16px;">
          <div class="bolum-baslik" style="margin:10px 0 2px;">Ürünler</div>
          {kalem_html}
        </div>
        """
        + aksiyon_html
        + '<a href="/siparisler" class="okut-kart okut-mor"><div class="okut-ikon">📋</div><div class="okut-metin"><div class="okut-baslik">Tüm Siparişlere Dön</div></div><div class="okut-ok">›</div></a>'
    )
    return sayfa(icerik, f"Sipariş #{siparis_id}")

@app.route("/kalem_iade/<int:kalem_id>", methods=["POST"])
@rol_gerekli("depocu", "muhasebeci", "patron")
def kalem_iade(kalem_id):
    try:
        miktar = int(request.form.get("miktar", "1"))
    except (TypeError, ValueError):
        miktar = 1
    if miktar < 1:
        miktar = 1

    con = db()
    try:
        with con:
            with con.cursor() as cur:
                cur.execute("SELECT siparis_id, barkod, ad, verilen FROM siparis_kalem WHERE id=%s", (kalem_id,))
                row = cur.fetchone()
                if not row:
                    return redirect("/siparisler")
                siparis_id, barkod, ad, verilen = row
                miktar = min(miktar, verilen)  # verilenden fazlası iade edilemez

                cur.execute("UPDATE siparis_kalem SET verilen = verilen - %s WHERE id=%s", (miktar, kalem_id))
                cur.execute("UPDATE urun SET adet = adet + %s WHERE barkod=%s", (miktar, barkod))
                cur.execute("""
                    INSERT INTO hareket (barkod, ad, tip, adet, kullanici, tarih)
                    VALUES (%s, %s, 'iade', %s, %s, %s)
                """, (barkod, ad, miktar, session.get("kullanici", "Bilinmiyor"), tr_simdi()))
                # iade sonrası sipariş artık eksik hale geldiyse yeniden açık yap
                cur.execute("UPDATE siparis SET durum='acik' WHERE id=%s AND durum='tamamlandi'", (siparis_id,))
    finally:
        con.close()

    return redirect(f"/siparis/{siparis_id}")


@app.route("/siparis_fis/<int:siparis_id>")
@rol_gerekli("depocu", "muhasebeci", "patron")
def siparis_fis(siparis_id):
    con = db()
    try:
        with con.cursor() as cur:
            cur.execute("SELECT id, musteri, depo, durum, olusturan, tarih, aciklama FROM siparis WHERE id=%s", (siparis_id,))
            sip = cur.fetchone()
            if not sip:
                return sayfa('<p class="hata">❌ Sipariş bulunamadı.</p><a class="btn gri" href="/siparisler">⬅ Geri Dön</a>', "Hata")
            cur.execute("""
                SELECT k.ad, k.barkod, k.istenen, k.verilen,
                       u.cins, u.ebat, u.kalinlik, u.yuzey, u.sinif, u.renk
                FROM siparis_kalem k
                LEFT JOIN urun u ON u.barkod = k.barkod
                WHERE k.siparis_id=%s ORDER BY k.id
            """, (siparis_id,))
            kalemler = cur.fetchall()
    finally:
        con.close()

    _, musteri, depo, durum, olusturan, tarih, aciklama = sip

    satirlar = ""
    toplam_istenen = 0
    toplam_verilen = 0
    for ad, barkod, istenen, verilen, cins, ebat, kalinlik, yuzey, sinif, renk in kalemler:
        toplam_istenen += istenen
        toplam_verilen += verilen
        ozellikler = " / ".join([x for x in [cins, ebat, kalinlik, yuzey, sinif, renk] if x]) or "-"
        satirlar += f"""
        <tr>
          <td>{ad}</td>
          <td>{barkod}</td>
          <td>{ozellikler}</td>
          <td style="text-align:center;">{istenen}</td>
          <td style="text-align:center;">{verilen}</td>
        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="tr">
    <head>
    <meta charset="utf-8">
    <title>Sipariş Fişi #{siparis_id}</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        @page {{ size: A4; margin: 12mm; }}
        body {{ font-family: Arial, sans-serif; color:#111; margin:0; }}
        .fis {{ max-width: 760px; margin: 0 auto; padding: 12px; }}
        .fis-baslik {{ display:flex; align-items:center; gap:12px; border-bottom:2px solid #111; padding-bottom:10px; margin-bottom:14px; }}
        .fis-baslik img {{ width:52px; height:52px; border-radius:8px; background:#fff; }}
        .fis-baslik h1 {{ font-size:18px; margin:0; }}
        .bilgi-satir {{ display:flex; justify-content:space-between; font-size:13px; color:#333; margin-bottom:14px; flex-wrap:wrap; gap:6px; }}
        table {{ width:100%; border-collapse:collapse; font-size:12.5px; }}
        th, td {{ border:1px solid #ccc; padding:7px 8px; text-align:left; }}
        th {{ background:#f0f0f0; }}
        .toplam-satir td {{ font-weight:800; background:#f7f7f7; }}
        .yazdir-btn {{
            display:block; margin: 18px auto 0; padding:14px 24px;
            background:#2196F3; color:white; border:none; border-radius:10px;
            font-size:16px; font-weight:700; cursor:pointer;
        }}
        @media print {{ .yazdir-btn {{ display:none; }} }}
    </style>
    </head>
    <body>
        <div class="fis">
            <div class="fis-baslik">
                <img src="{LOGO_URL}">
                <h1>{UYGULAMA_ADI}<br><span style="font-weight:400;font-size:13px;">Sipariş Fişi #{siparis_id}</span></h1>
            </div>
            <div class="bilgi-satir">
                <div><b>Müşteri:</b> {musteri or '-'}</div>
                <div><b>Depo:</b> {depo}</div>
                <div><b>Durum:</b> {durum}</div>
                <div><b>Oluşturan:</b> {olusturan}</div>
                <div><b>Tarih:</b> {tarih}</div>
            </div>
            {'<p style="font-size:13px;"><b>Not:</b> ' + aciklama + '</p>' if aciklama else ''}
            <table>
                <tr><th>Ürün</th><th>Barkod</th><th>Özellikler</th><th>İstenen</th><th>Verilen</th></tr>
                {satirlar}
                <tr class="toplam-satir"><td colspan="3">TOPLAM</td><td style="text-align:center;">{toplam_istenen}</td><td style="text-align:center;">{toplam_verilen}</td></tr>
            </table>
        </div>
        <button class="yazdir-btn" onclick="window.print()">🖨️ Yazdır</button>
    </body>
    </html>
    """
    return html

@app.route("/siparis_iptal/<int:siparis_id>", methods=["POST"])
@rol_gerekli("muhasebeci")
def siparis_iptal(siparis_id):
    con = db()
    try:
        with con:
            with con.cursor() as cur:
                cur.execute("UPDATE siparis SET durum='iptal' WHERE id=%s", (siparis_id,))
    finally:
        con.close()
    return redirect(f"/siparis/{siparis_id}")


@app.route("/kamera/<tip>")
@rol_gerekli("depocu")
def kamera(tip):
    if tip not in ("giris", "cikis"):
        tip = "giris"
    kullanici = session.get("kullanici", "Bilinmiyor")
    rol = session.get("rol")
    siparis_id = request.args.get("siparis_id", "")

    # Çıkış modunda depocu önce açık bir sipariş seçmeli (yanlış ürün çıkışını engellemek için)
    if tip == "cikis" and rol == "depocu" and not siparis_id:
        con = db()
        try:
            with con.cursor() as cur:
                cur.execute("""
                    SELECT s.id, s.musteri, s.depo, COALESCE(SUM(k.istenen),0), COALESCE(SUM(k.verilen),0)
                    FROM siparis s LEFT JOIN siparis_kalem k ON k.siparis_id = s.id
                    WHERE s.durum='acik' GROUP BY s.id ORDER BY s.id DESC
                """)
                acik_siparisler = cur.fetchall()
        finally:
            con.close()

        kartlar = ""
        for sid, musteri, depo, istenen, verilen in acik_siparisler:
            yuzde = int((verilen / istenen) * 100) if istenen else 0
            kartlar += f"""
            <a href="/kamera/cikis?siparis_id={sid}" class="siparis-kart">
              <div class="siparis-ust">
                <div class="siparis-no">🧾 #{sid}{' — ' + musteri if musteri else ''}</div>
                <div class="siparis-durum acik">açık</div>
              </div>
              <div class="siparis-detay">🏭 {depo} &nbsp;•&nbsp; 📦 {verilen} / {istenen} adet</div>
              <div class="siparis-ilerleme-bar"><div class="siparis-ilerleme-dolu" style="width:{yuzde}%;"></div></div>
            </a>
            """

        if not acik_siparisler:
            kartlar = '<div class="uyari-kutu">⚠️ Şu anda size atanmış açık bir sipariş yok. Çıkış yapmadan önce muhasebeciden bir sipariş oluşturmasını iste.</div>'

        icerik = (
            "<h2>⬇️ Çıkış İçin Sipariş Seç</h2>"
            + '<p style="text-align:center;color:var(--muted);margin-top:0;">Yanlış ürün verilmesini önlemek için önce hangi siparişi hazırladığını seç.</p>'
            + kartlar
        )
        return sayfa(icerik, "Sipariş Seç")

    siparis_bilgi_html = ""
    if tip == "cikis" and siparis_id:
        con = db()
        try:
            with con.cursor() as cur:
                cur.execute("SELECT musteri, depo, durum FROM siparis WHERE id=%s", (siparis_id,))
                sip = cur.fetchone()
                cur.execute("SELECT ad, barkod, istenen, verilen FROM siparis_kalem WHERE siparis_id=%s ORDER BY id", (siparis_id,))
                kalemler = cur.fetchall()
        finally:
            con.close()

        if sip:
            musteri, depo, durum = sip
            kalem_satirlari = "".join(
                f'<div class="kalem-satir"><div class="kalem-ad">{ad}</div><div class="kalem-miktar {"tam" if verilen>=istenen else "eksik"}">{verilen}/{istenen}</div></div>'
                for ad, barkod, istenen, verilen in kalemler
            )
            siparis_bilgi_html = f"""
            <div class="kart" style="padding:12px 16px;">
              <div style="display:flex;justify-content:space-between;align-items:center;">
                <b>🧾 Sipariş #{siparis_id}{' — ' + musteri if musteri else ''}</b>
                <span class="siparis-durum {durum}">{durum}</span>
              </div>
              <div style="color:var(--muted);font-size:12px;margin:4px 0 8px;">🏭 {depo} — sadece bu listedeki ürünler okutulabilir</div>
              {kalem_satirlari}
            </div>
            """

    icerik = """
    <style>
    .mod-gecis {
      display:flex; background:#12151c; border:1px solid var(--border);
      border-radius:999px; padding:4px; margin-bottom:16px;
    }
    .mod-btn {
      flex:1; text-align:center; padding:12px 8px; border-radius:999px;
      font-weight:700; font-size:14.5px; cursor:pointer; border:none;
      background:transparent; color:var(--muted); transition: all .18s ease;
    }
    .mod-btn.aktif-giris { background: linear-gradient(135deg,#00C853,#64DD17); color:#fff; }
    .mod-btn.aktif-cikis { background: linear-gradient(135deg,#FF6F00,#FF9800); color:#fff; }

    .kamera-cerceve {
      position:relative; border-radius:20px; overflow:hidden; margin-bottom:14px;
      border:2px solid var(--border); background:#000;
      box-shadow: 0 8px 28px rgba(0,0,0,.5);
      aspect-ratio: 4/3;
    }
    .kamera-cerceve video { width:100%; height:100%; object-fit:cover; display:block; }
    .kamera-ustkatman {
      position:absolute; inset:0; pointer-events:none;
      display:flex; align-items:center; justify-content:center;
    }
    .tarama-kutu {
      width:72%; height:42%; border:3px solid rgba(255,255,255,.85);
      border-radius:14px; box-shadow: 0 0 0 999px rgba(0,0,0,.28);
    }
    .tarama-kutu.basarili {
      border-color: #64DD17; box-shadow: 0 0 0 999px rgba(0,0,0,.28), 0 0 24px 4px rgba(100,221,23,.6);
      transition: all .15s ease;
    }
    .kamera-araclar {
      position:absolute; top:10px; right:10px; display:flex; flex-direction:column; gap:8px;
      pointer-events:auto;
    }
    .araç-btn {
      width:42px; height:42px; border-radius:50%; border:none; cursor:pointer;
      background:rgba(0,0,0,.5); color:white; font-size:19px;
      display:flex; align-items:center; justify-content:center;
      backdrop-filter: blur(6px);
    }
    .araç-btn.aktif { background: var(--accent); }
    .kamera-baslat-alan { text-align:center; padding: 40px 10px; }

    .sonuc-kart {
      border-radius: var(--radius); padding:18px; margin: 4px 0 14px;
      border:1px solid var(--border); background: linear-gradient(180deg, var(--card2), var(--card));
      animation: belir .2s ease;
    }
    .sonuc-basari { border-color: rgba(100,221,23,.4); }
    .sonuc-hata { border-color: rgba(255,23,68,.4); }
    .sonuc-yeni { border-color: rgba(33,150,243,.4); }
    .sonuc-yanlis { border-color: rgba(255,23,68,.6); background: linear-gradient(180deg, rgba(255,23,68,.14), var(--card)); }
    .sonuc-ust { display:flex; align-items:center; gap:10px; margin-bottom:10px; }
    .sonuc-ikon { font-size:26px; }
    .sonuc-ad { font-size:17px; font-weight:800; }
    .sonuc-satirlar { font-size:13.5px; color:var(--muted); line-height:1.9; }
    .sonuc-satirlar b { color: var(--text); font-weight:600; }

    .geri-al-btn {
      display:block; width:100%; text-align:center; padding:13px;
      border-radius: var(--radius-sm); border:1px dashed var(--border);
      background:transparent; color:var(--muted); font-weight:600; font-size:13.5px;
      cursor:pointer; margin-top:2px;
    }
    .geri-al-btn:active { background: rgba(255,255,255,.05); }

    .elle-giris-alan { margin-top:8px; }
    .elle-giris-satir { display:flex; gap:8px; }
    .elle-giris-satir input { margin:0; flex:1; }
    .elle-giris-satir button { width:auto; margin:0; padding:0 18px; white-space:nowrap; }
    </style>

    <h3 class="alt" style="margin-bottom:2px;">👤 {{kullanici}}</h3>

    <div class="mod-gecis">
      <button type="button" id="btn-giris" class="mod-btn" onclick="modDegistir('giris')">⬆️ GİRİŞ</button>
      <button type="button" id="btn-cikis" class="mod-btn" onclick="modDegistir('cikis')">⬇️ ÇIKIŞ</button>
    </div>

    {{siparis_bilgi_html|safe}}

    <div class="kamera-cerceve" id="kamera-cerceve" style="display:none;">
      <video id="video"></video>
      <div class="kamera-ustkatman"><div class="tarama-kutu"></div></div>
      <div class="kamera-araclar">
        <button type="button" class="araç-btn" id="flash-btn" onclick="flashDegistir()" style="display:none;">💡</button>
        <button type="button" class="araç-btn" id="kamera-degistir-btn" onclick="kameraDegistir()" style="display:none;">🔄</button>
      </div>
    </div>

    <div class="kamera-baslat-alan" id="baslat-alan">
      <button class="btn yesil" style="max-width:320px;margin:0 auto;" onclick="baslat()">📷 Kamerayı Başlat</button>
    </div>

    <div id="sonuc-alan"></div>

    <div class="kart elle-giris-alan">
      <label>Barkod okunmuyorsa elle gir</label>
      <div class="elle-giris-satir">
        <input type="text" id="elle-barkod" inputmode="numeric" placeholder="Barkod numarası">
        <button class="btn mavi" style="width:auto;" onclick="elleGonder()">Ekle</button>
      </div>
    </div>

    <script src="https://unpkg.com/@zxing/library@latest"></script>
<script>
let codeReader;
let kilit = false;
let bipSes;
let aktifTip = "{{tip}}";
let flashAcik = false;
let sonIslem = null;
let siparisId = {{ (siparis_id or 'null') }};
let kameraListesi = [];
let aktifKameraIndex = 0;

function arayuzuGuncelle(t){
    document.getElementById('btn-giris').className = 'mod-btn' + (t === 'giris' ? ' aktif-giris' : '');
    document.getElementById('btn-cikis').className = 'mod-btn' + (t === 'cikis' ? ' aktif-cikis' : '');
}
function modDegistir(yeniTip){
    if(yeniTip === aktifTip){ arayuzuGuncelle(yeniTip); return; }
    window.location.href = '/kamera/' + yeniTip;
}
arayuzuGuncelle(aktifTip);

function hintOlustur(){
    const hints = new Map();
    // Sadece üretilen barkod tiplerini arıyoruz (CODE_128) 
    // Format daraltmak taramayı belirgin şekilde hızlandırır.
    hints.set(ZXing.DecodeHintType.POSSIBLE_FORMATS, [
        ZXing.BarcodeFormat.CODE_128,
        ZXing.BarcodeFormat.EAN_13,
    ]);
    hints.set(ZXing.DecodeHintType.TRY_HARDER, true);
    return hints;
}

async function baslat(){
    bipSes = new Audio("https://actions.google.com/sounds/v1/alarms/beep_short.ogg");
    document.getElementById('baslat-alan').style.display = 'none';
    document.getElementById('kamera-cerceve').style.display = 'block';

    codeReader = new ZXing.BrowserMultiFormatReader(hintOlustur());

    try {
        kameraListesi = await ZXing.BrowserMultiFormatReader.listVideoInputDevices();
    } catch(e) { kameraListesi = []; }

    // Arka kamerayı tercih et (isminde "back"/"arka"/"rear" geçen varsa onu seç)
    aktifKameraIndex = kameraListesi.findIndex(k =>
        /back|arka|rear|environment/i.test(k.label)
    );
    if(aktifKameraIndex === -1) aktifKameraIndex = kameraListesi.length - 1 >= 0 ? kameraListesi.length - 1 : 0;

    if(kameraListesi.length > 1){
        document.getElementById('kamera-degistir-btn').style.display = 'flex';
    }

    kameraBaslat();
}

function kameraBaslat(){
    if(codeReader) { try { codeReader.reset(); } catch(e){} }
    codeReader = new ZXing.BrowserMultiFormatReader(hintOlustur());

    const secilenId = kameraListesi.length ? kameraListesi[aktifKameraIndex].deviceId : null;

    const kisitlar = {
        video: secilenId
            ? { deviceId: { exact: secilenId }, width: { ideal: 1280 }, height: { ideal: 720 } }
            : { facingMode: { ideal: "environment" }, width: { ideal: 1280 }, height: { ideal: 720 } }
    };

    codeReader.decodeFromConstraints(kisitlar, 'video', (result, err) => {
        if (result && !kilit) {
            kilit = true;
            basariliGoruntu();
            isleGonder(result.text);
        }
        if (err && !(err instanceof ZXing.NotFoundException)) {
            console.log(err);
        }
    }).then(() => {
        odakAyarla();
        fenerKontrolEt();
    }).catch(e => {
        console.log(e);
        document.getElementById('sonuc-alan').innerHTML = `
          <div class="sonuc-kart sonuc-hata">
            <div class="sonuc-ust"><span class="sonuc-ikon">📵</span><span class="sonuc-ad">Kamera açılamadı</span></div>
            <div class="sonuc-satirlar">Tarayıcının kamera iznini kontrol et, ya da elle barkod gir.</div>
          </div>`;
    });
}

function odakAyarla(){
    const video = document.getElementById('video');
    const stream = video.srcObject;
    if(!stream) return;
    const track = stream.getVideoTracks()[0];
    if(!track || !track.getCapabilities) return;
    const yetenekler = track.getCapabilities();
    if(yetenekler.focusMode && yetenekler.focusMode.includes('continuous')){
        track.applyConstraints({ advanced: [{ focusMode: 'continuous' }] }).catch(e => {});
    }
}

function basariliGoruntu(){
    const kutu = document.querySelector('.tarama-kutu');
    if(!kutu) return;
    kutu.classList.add('basarili');
    setTimeout(() => kutu.classList.remove('basarili'), 500);
}

function kameraDegistir(){
    if(kameraListesi.length < 2) return;
    aktifKameraIndex = (aktifKameraIndex + 1) % kameraListesi.length;
    document.getElementById('flash-btn').style.display = 'none';
    flashAcik = false;
    kameraBaslat();
}

function fenerKontrolEt(){
    const video = document.getElementById('video');
    const stream = video.srcObject;
    if(!stream) return;
    const track = stream.getVideoTracks()[0];
    if(!track || !track.getCapabilities) return;
    const yetenekler = track.getCapabilities();
    if(yetenekler.torch){
        document.getElementById('flash-btn').style.display = 'flex';
    }
}

function flashDegistir(){
    const video = document.getElementById('video');
    const stream = video.srcObject;
    if(!stream) return;
    const track = stream.getVideoTracks()[0];
    flashAcik = !flashAcik;
    track.applyConstraints({ advanced: [{ torch: flashAcik }] }).catch(e => console.log(e));
    document.getElementById('flash-btn').classList.toggle('aktif', flashAcik);
}

// Tıklanan noktaya odaklanmayı dene (destekleyen cihazlarda)
document.addEventListener('DOMContentLoaded', () => {
    const cerceve = document.getElementById('kamera-cerceve');
    if(!cerceve) return;
    cerceve.addEventListener('click', (e) => {
        const video = document.getElementById('video');
        const stream = video.srcObject;
        if(!stream) return;
        const track = stream.getVideoTracks()[0];
        if(!track || !track.getCapabilities) return;
        const yetenekler = track.getCapabilities();
        if(!yetenekler.pointsOfInterest) return;
        const rect = cerceve.getBoundingClientRect();
        const x = (e.clientX - rect.left) / rect.width;
        const y = (e.clientY - rect.top) / rect.height;
        track.applyConstraints({ advanced: [{ pointsOfInterest: [{x, y}] }] }).catch(err => {});
    });
});

function elleGonder(){
    const kutu = document.getElementById('elle-barkod');
    const kod = kutu.value.trim();
    if(!kod) return;
    kutu.value = '';
    isleGonder(kod);
}

function isleGonder(barkod){
    if(bipSes){ bipSes.currentTime = 0; bipSes.play().catch(e => {}); }
    if(navigator.vibrate) navigator.vibrate(70);

    fetch("/hizli_islem", {
        method: "POST",
        headers: {"Content-Type": "application/json"},
        body: JSON.stringify({ barkod: barkod, tip: aktifTip, siparis_id: siparisId })
    })
    .then(r => r.json())
    .then(d => sonucGoster(d, barkod))
    .finally(() => { setTimeout(() => { kilit = false; }, 900); });
}

function sonucGoster(d, barkod){
    const alan = document.getElementById('sonuc-alan');

    if (d.ok) {
        sonIslem = { barkod: barkod, tip: aktifTip, siparis_id: siparisId };
        const baslikRengi = aktifTip === 'giris' ? '✅' : '📤';
        let siparisSatiri = '';
        if(d.siparis_ilerleme){
            siparisSatiri = `<br>🧾 Sipariş toplam: <b>${d.siparis_ilerleme.verilen} / ${d.siparis_ilerleme.istenen}</b>` +
                (d.siparis_ilerleme.tamamlandi ? ' — 🎉 Sipariş tamamlandı!' : '');
        }
        alan.innerHTML = `
          <div class="sonuc-kart sonuc-basari">
            <div class="sonuc-ust"><span class="sonuc-ikon">${baslikRengi}</span><span class="sonuc-ad">${d.ad}</span></div>
            <div class="sonuc-satirlar">
              🏷️ Cins: <b>${d.cins || '-'}</b> &nbsp; 🔖 Sınıf: <b>${d.sinif || '-'}</b><br>
              ✨ Yüzey: <b>${d.yuzey || '-'}</b> &nbsp; 🎨 Renk: <b>${d.renk || '-'}</b><br>
              📏 Ebat: <b>${d.ebat || '-'}</b> &nbsp; 🏭 Depo: <b>${d.depo || '-'}</b><br>
              📦 Kalan Stok: <b>${d.adet}</b> &nbsp; 📊 Bugünkü Toplam: <b>${d.toplam}</b>${siparisSatiri}
            </div>
            <button class="geri-al-btn" onclick="geriAl()">↩️ Bu işlemi geri al</button>
          </div>`;
    } else if (d.yeni) {
        alan.innerHTML = `
          <div class="sonuc-kart sonuc-yeni">
            <div class="sonuc-ust"><span class="sonuc-ikon">🆕</span><span class="sonuc-ad">Bu barkod stokta yok</span></div>
            <div class="sonuc-satirlar">Yeni ürün ekleme sayfasına yönlendiriliyorsunuz...</div>
          </div>`;
        setTimeout(() => { window.location.href = "/ekle?barkod=" + encodeURIComponent(barkod); }, 1000);
    } else if (d.yanlis_urun) {
        if(navigator.vibrate) navigator.vibrate([120,80,120,80,120]);
        alan.innerHTML = `
          <div class="sonuc-kart sonuc-yanlis">
            <div class="sonuc-ust"><span class="sonuc-ikon">🚫</span><span class="sonuc-ad">${d.msg}</span></div>
            <div class="sonuc-satirlar">"${d.urun_adi || ''}" bu siparişte listelenmiyor. Lütfen sipariş listesindeki ürünleri kontrol et.</div>
          </div>`;
    } else {
        if(navigator.vibrate) navigator.vibrate([60,60,60]);
        alan.innerHTML = `
          <div class="sonuc-kart sonuc-hata">
            <div class="sonuc-ust"><span class="sonuc-ikon">❌</span><span class="sonuc-ad">${d.msg || 'Bulunamadı'}</span></div>
          </div>`;
    }
}

function geriAl(){
    if(!sonIslem) return;
    fetch("/geri_al", {
        method: "POST",
        headers: {"Content-Type": "application/json"},
        body: JSON.stringify(sonIslem)
    })
    .then(r => r.json())
    .then(d => {
        if(d.ok){
            document.getElementById('sonuc-alan').innerHTML =
              '<div class="sonuc-kart"><div class="sonuc-ust"><span class="sonuc-ikon">↩️</span><span class="sonuc-ad">İşlem geri alındı</span></div></div>';
            sonIslem = null;
        }
    });
}
</script>
  
    """
    return render_template_string(
        sayfa(icerik, "Barkod Okut"),
        tip=tip, kullanici=kullanici,
        siparis_id=(int(siparis_id) if siparis_id else None),
        siparis_bilgi_html=siparis_bilgi_html,
    )

@app.route("/ping")
def ping():
    return "ok"
import traceback

@app.errorhandler(500)
def sunucu_hatasi(e):
    hata_metni = traceback.format_exc()
    print("=== SUNUCU HATASI ===")
    print(hata_metni)
    print("======================")
    icerik = f"""
    <h2>⚠️ Bir Şeyler Ters Gitti</h2>
    <p style="text-align:center;color:var(--muted);">
    İşlem tamamlanamadı. Lütfen tekrar dene, sorun devam ederse
    yöneticine haber ver.
    </p>
    <a class="btn mavi" href="/">🏠 Ana Sayfaya Dön</a>
    {'<pre style="font-size:10px;color:#FF6B6B;white-space:pre-wrap;overflow-x:auto;">' + hata_metni + '</pre>' if session.get("rol") == "patron" else ''}
    """
    return sayfa(icerik, "Hata"), 500
if __name__ == "__main__":
    app.run(debug=True)
