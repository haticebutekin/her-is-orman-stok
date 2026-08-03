# -*- coding: utf-8 -*-
"""
TOPLU ÜRÜN EKLEME SCRIPTİ
--------------------------
Bu dosyayı app.py ile AYNI klasöre koy (stok.db orada oluşuyor).
Aynı klasörde "urun_listesi.xlsx" (Depo ve Adet sütunlarını doldurduğun) olmalı.

Çalıştırma:
    pip install openpyxl python-barcode qrcode --break-system-packages   (zaten kuruluysa gerek yok)
    python toplu_ekle.py

Ne yapar:
- urun_listesi.xlsx dosyasındaki her satırı okur
- Barkod boşsa otomatik, benzersiz barkod üretir (app.py ile aynı yöntem)
- urun tablosuna INSERT eder
- adet > 0 ise hareket tablosuna 'giris' kaydı da düşer (stok geçmişi tutarlı olsun diye)
- static/ klasörüne barkod + QR görselini üretir (app.py'nin liste ekranı bunları gösteriyor)
- Barkodu zaten stokta olan (aynı barkod) ürünleri ATLAR, tekrar eklemez
"""

import sqlite3
import random
import os
from openpyxl import load_workbook

try:
    import barcode
    from barcode.writer import ImageWriter
    import qrcode
except ImportError:
    raise SystemExit(
        "Eksik paket var. Şunu çalıştır:\n"
        "pip install python-barcode qrcode --break-system-packages"
    )

DB = "stok.db"
EXCEL = "urun_listesi.xlsx"


def db():
    return sqlite3.connect(DB)


def barkod_uret():
    while True:
        kod = str(random.randint(100000000000, 999999999999))
        with db() as con:
            var = con.execute("SELECT barkod FROM urun WHERE barkod=?", (kod,)).fetchone()
        if not var:
            return kod


def barkod_resim(kod):
    if not os.path.exists("static"):
        os.makedirs("static")
    CODE128 = barcode.get_barcode_class("code128")
    img = CODE128(kod, writer=ImageWriter())
    img.save(os.path.join("static", kod))


def qr_uret(kod):
    if not os.path.exists("static"):
        os.makedirs("static")
    img = qrcode.make(kod)
    img.save(os.path.join("static", kod + "_qr.png"))


def main():
    if not os.path.exists(DB):
        raise SystemExit(f"{DB} bulunamadı. Bu scripti app.py ile aynı klasörde çalıştır (app.py'yi bir kere çalıştırıp veritabanını oluşturman gerekebilir).")
    if not os.path.exists(EXCEL):
        raise SystemExit(f"{EXCEL} bulunamadı. Excel dosyasını bu klasöre kopyala.")

    wb = load_workbook(EXCEL)
    ws = wb["Urunler"]

    eklenen = 0
    atlanan_barkod_var = 0
    atlanan_bos = 0

    with db() as con:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or not row[0]:
                atlanan_bos += 1
                continue

            ad, cins, seri, yuzey, ebat, renk, adet, depo, barkod = (list(row) + [None] * 9)[:9]

            ad = str(ad).strip() if ad else ""
            if not ad:
                atlanan_bos += 1
                continue

            cins = str(cins).strip() if cins else ""
            seri = str(seri).strip() if seri else ""
            yuzey = str(yuzey).strip() if yuzey else ""
            ebat = str(ebat).strip() if ebat else ""
            renk = str(renk).strip() if renk else ""
            depo = str(depo).strip() if depo else ""
            adet = int(adet) if adet not in (None, "") else 0
            barkod = str(barkod).strip() if barkod else ""

            if barkod:
                var = con.execute("SELECT id FROM urun WHERE barkod=?", (barkod,)).fetchone()
                if var:
                    atlanan_barkod_var += 1
                    continue
            else:
                barkod = barkod_uret()

            con.execute("""
                INSERT INTO urun(ad, cins, ebat, kalinlik, yuzey, sinif, renk, adet, depo, barkod)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (ad, cins, ebat, "", yuzey, seri, renk, adet, depo, barkod))

            if adet > 0:
                con.execute("""
                    INSERT INTO hareket (barkod, ad, tip, adet, kullanici)
                    VALUES (?, ?, 'giris', ?, 'Toplu Yukleme')
                """, (barkod, ad, adet))

            barkod_resim(barkod)
            qr_uret(barkod)

            eklenen += 1

        con.commit()

    print(f"Eklenen ürün: {eklenen}")
    print(f"Barkodu zaten kayıtlı olduğu için atlanan: {atlanan_barkod_var}")
    print(f"Boş satır olduğu için atlanan: {atlanan_bos}")


if __name__ == "__main__":
    main()
